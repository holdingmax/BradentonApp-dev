"""
Chase bank activity engine: keyword-to-Detalle categorization rules
(persisted + hardcoded) and the full read/categorize/write pipeline.
"""

import json
import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

RULES_FILENAME = "chase_rules.json"  # Personalizadas -- added by an admin via the UI.
MASTER_RULES_FILENAME = "chase_master_rules.json"  # Maestra -- seeded once from Alfonso's original hardcoded rules, then admin-editable.


def _rules_file_path(filename):
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


def normalize_rule_text(value):
    """Lowercase, strip accents, collapse whitespace; CHECK/CHEQUE -> CHEK."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\b(check|cheque)\b", "chek", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_rule_entry(keyword, detail):
    keyword_text = str(keyword).strip()
    detail_text = str(detail).strip()
    if not keyword_text or not detail_text:
        raise ValueError("La palabra clave y el detalle son obligatorios.")
    return {"keyword": keyword_text, "detail": detail_text}


def _load_rules_file(filename, seed=None):
    """
    Read a keyword/detail rules JSON file, shared by the Maestra and
    Personalizada storage below. If the file doesn't exist yet and `seed`
    is given, it's created with that seed data first -- used once, the
    first time chase_master_rules.json is read.
    """
    path = _rules_file_path(filename)
    if not os.path.isfile(path):
        if seed is None:
            return []
        _save_rules_file(filename, list(seed))

    try:
        with open(path, encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, json.JSONDecodeError):
        return []

    if isinstance(payload, dict):
        rules = payload.get("rules", [])
    elif isinstance(payload, list):
        rules = payload
    else:
        return []

    cleaned = []
    for item in rules:
        if not isinstance(item, dict):
            continue
        keyword = str(item.get("keyword", "")).strip()
        detail = str(item.get("detail", "")).strip()
        if keyword and detail:
            cleaned.append({"keyword": keyword, "detail": detail})
    return cleaned


def _save_rules_file(filename, rules):
    """Persist a full rules list atomically without dropping unrelated entries."""
    if not isinstance(rules, list):
        raise TypeError("Rules must be a list.")

    path = _rules_file_path(filename)
    directory = os.path.dirname(path)
    os.makedirs(directory, exist_ok=True)

    payload = {"rules": rules}
    temp_path = f"{path}.tmp"
    with open(temp_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, path)


def _add_rule(filename, keyword, detail, seed=None):
    """Append or update one rule (by keyword collision) while preserving every other saved rule."""
    entry = _normalize_rule_entry(keyword, detail)
    rules = _load_rules_file(filename, seed=seed)
    keyword_key = entry["keyword"].lower()
    rules = [rule for rule in rules if rule["keyword"].lower() != keyword_key]
    rules.append(entry)
    _save_rules_file(filename, rules)
    return entry


def _edit_rule_by_index(filename, index, keyword, detail, seed=None):
    """Replace one persisted rule in place, keeping its position in the list."""
    entry = _normalize_rule_entry(keyword, detail)
    rules = _load_rules_file(filename, seed=seed)
    try:
        idx = int(index)
    except (TypeError, ValueError):
        raise ValueError("Índice de regla inválido.") from None
    if idx < 0 or idx >= len(rules):
        raise ValueError("La regla seleccionada no se encontró en el almacenamiento.")
    rules[idx] = entry
    _save_rules_file(filename, rules)
    return entry


def _delete_rule_by_index(filename, index, seed=None):
    """Remove one persisted rule by index."""
    rules = _load_rules_file(filename, seed=seed)
    try:
        idx = int(index)
    except (TypeError, ValueError):
        raise ValueError("Índice de regla inválido.") from None
    if idx < 0 or idx >= len(rules):
        raise ValueError("La regla seleccionada no se encontró en el almacenamiento.")
    removed = rules.pop(idx)
    _save_rules_file(filename, rules)
    return removed


def load_dynamic_rules():
    """Return all persisted Personalizada rule dicts: {"keyword": str, "detail": str}."""
    return _load_rules_file(RULES_FILENAME)


def save_dynamic_rules(rules):
    _save_rules_file(RULES_FILENAME, rules)


def add_dynamic_rule(keyword, detail):
    """Append or update one Personalizada rule while preserving every other saved rule."""
    return _add_rule(RULES_FILENAME, keyword, detail)


def edit_dynamic_rule_by_index(index, keyword, detail):
    """Replace one Personalizada rule in place by index in chase_rules.json."""
    return _edit_rule_by_index(RULES_FILENAME, index, keyword, detail)


def delete_dynamic_rule_by_index(index):
    """Remove one Personalizada rule by index in chase_rules.json."""
    return _delete_rule_by_index(RULES_FILENAME, index)


def load_master_rules():
    """
    Return all Maestra rule dicts: {"keyword": str, "detail": str}.

    Seeded once from Alfonso's original hardcoded categorization rules (see
    _DEFAULT_MASTER_RULES below) the first time chase_master_rules.json is
    read -- from then on this file is the live source
    categorize_chase_description reads from, so an admin edit takes effect
    immediately.
    """
    return _load_rules_file(MASTER_RULES_FILENAME, seed=_DEFAULT_MASTER_RULES)


def edit_master_rule_by_index(index, keyword, detail):
    """Replace one Maestra rule in place by index in chase_master_rules.json."""
    return _edit_rule_by_index(
        MASTER_RULES_FILENAME, index, keyword, detail, seed=_DEFAULT_MASTER_RULES
    )


def delete_master_rule_by_index(index):
    """Remove one Maestra rule by index in chase_master_rules.json."""
    return _delete_rule_by_index(MASTER_RULES_FILENAME, index, seed=_DEFAULT_MASTER_RULES)


def _match_longest_keyword(desc, rules):
    """
    Among every rule whose keyword appears in the (already normalized) desc,
    the LONGEST keyword wins -- a generic rule ("KOOLER") never permanently
    shadows a more specific one ("KOOLER FARMS LLC") just because it was
    added earlier or lives in a different rules file.

    Returns Detail text or None.
    """
    best_detail = None
    best_len = 0
    for rule in rules:
        keyword = normalize_rule_text(rule.get("keyword", ""))
        if keyword and keyword in desc and len(keyword) > best_len:
            best_detail = rule.get("detail")
            best_len = len(keyword)
    return best_detail


def match_dynamic_detalle(description):
    """Longest-keyword match against saved Personalizada rules only. Returns Detail text or None."""
    desc = normalize_rule_text(description)
    if not desc:
        return None
    return _match_longest_keyword(desc, load_dynamic_rules())


def list_display_rules():
    """
    Full rule list for the UI grid: Maestra (chase_master_rules.json,
    admin-editable) + Personalizada (chase_rules.json, admin-editable) +
    Combinada (hardcoded multi-condition rules, read-only -- see
    ALFONSO_COMPOUND_DISPLAY_RULES). Maestra/Personalizada rows carry
    rule_type + index so the caller can edit/delete them; Combinada rows
    carry neither since they aren't stored as a simple keyword.
    """
    merged = []
    for idx, rule in enumerate(load_master_rules()):
        merged.append(
            {
                "keyword": rule["keyword"],
                "detail": rule["detail"],
                "source": "Maestra",
                "rule_type": "master",
                "index": idx,
            }
        )
    for idx, rule in enumerate(load_dynamic_rules()):
        merged.append(
            {
                "keyword": rule["keyword"],
                "detail": rule["detail"],
                "source": "Personalizada",
                "rule_type": "custom",
                "index": idx,
            }
        )
    for keyword, detail in ALFONSO_COMPOUND_DISPLAY_RULES:
        merged.append(
            {
                "keyword": keyword,
                "detail": detail,
                "source": "Combinada",
                "rule_type": None,
                "index": None,
            }
        )
    return merged


# ---------------------------------------------------------------------------
# Categorization engine
# ---------------------------------------------------------------------------

EFT_RCV_EXACT_LABEL = "EFT RCV-"
EFT_RCV_RED_FONT = Font(color="FF0000", bold=True)

# Alfonso's original hardcoded rules, flattened into one seed list the first
# time chase_master_rules.json gets created (see load_master_rules above) --
# kept here as frozen historical data; the categorization engine never reads
# these tuples directly after that first seed, only the JSON file itself, so
# an admin edit made afterward always takes effect.
_DEFAULT_MASTER_RULES = (
    {"keyword": "frito-la", "detail": "PROVEEDORES"},
    {"keyword": "hackneyrectampa", "detail": "PROVEEDORES"},
    {"keyword": "cec distributing", "detail": "PROVEEDORES"},
    {"keyword": "gold coast eagle", "detail": "PROVEEDORES"},
    {"keyword": "jj taylor distri", "detail": "PROVEEDORES"},
    {"keyword": "pbg", "detail": "PROVEEDORES"},
    {"keyword": "colonial", "detail": "PROVEEDORES"},
    {"keyword": "redbull", "detail": "PROVEEDORES"},
    {"keyword": "airgas", "detail": "PROVEEDORES"},
    {"keyword": "johnson brothers", "detail": "PROVEEDORES"},
    {"keyword": "low value", "detail": "GASTOS BANCARIOS"},
    {"keyword": "initial fee", "detail": "GASTOS BANCARIOS"},
    {"keyword": "cash deposit immediate", "detail": "GASTOS BANCARIOS"},
    {"keyword": "monthly service fee", "detail": "GASTOS BANCARIOS"},
    {"keyword": "helix ucp", "detail": "REBATE"},
    {"keyword": "ussmokless", "detail": "REBATE"},
    {"keyword": "njoy", "detail": "REBATE"},
    {"keyword": "itg brands", "detail": "REBATE"},
    {"keyword": "john middleton", "detail": "REBATE"},
    {"keyword": "mucs", "detail": "AGUA"},
    {"keyword": "manatee", "detail": "AGUA"},
    {"keyword": "slomin's", "detail": "ALARMA"},
    {"keyword": "slomins", "detail": "ALARMA"},
    {"keyword": "slomin", "detail": "ALARMA"},
    {"keyword": "fpl", "detail": "ENERGIA ELECTRICA"},
    {"keyword": "text me", "detail": "TELEFONO"},
    {"keyword": "innov", "detail": "ELISTAR"},
    {"keyword": "ipf", "detail": "SEGURO"},
    {"keyword": "fla dept", "detail": "SALE TAX"},
    {"keyword": "alg distr", "detail": "REBATE"},
    {"keyword": "finova", "detail": "ADMINISTRATION FEE"},
    {"keyword": "orig co name:mvnt", "detail": "REBATE"},
    {"keyword": "orig co name:fla lottery", "detail": "LOTTERY"},
    {"keyword": "orig co name:cantaloupe", "detail": "VENTA ICE"},
    {"keyword": "online realtime vendor payment", "detail": "SUELDOS"},
    {"keyword": "online realtime payroll payment", "detail": "SUELDOS"},
    {"keyword": "deposit  id number", "detail": "DEPOSITO"},
    {"keyword": "spectrum", "detail": "INTERNET"},
    {"keyword": "jeffrey's lawn", "detail": "REPARACION Y MANTENIMIENTO"},
    {"keyword": "jeffreys lawn", "detail": "REPARACION Y MANTENIMIENTO"},
    {"keyword": "merchant bank", "detail": "COMISIONES Y GASTOS BANCARIOS"},
    {"keyword": "reynolds", "detail": "REBATE"},
    {"keyword": "fla lottery", "detail": "LOTTERY"},
    {"keyword": "cantaloupe", "detail": "VENTA ICE"},
    {"keyword": "mvnt", "detail": "REBATE"},
)

# Keywords used only inside the CHECK/CHEQUE/CHEK + vendor compound rule
# below -- not a flat keyword->detail rule, so it can't live in
# chase_master_rules.json / be edited from the UI.
CHASE_CHECK_PROVEEDORES_KEYWORDS = (
    "coca",
    "coke",
    "midtown",
    "king",
    "liu",
    "icecream",
    "ice cream",
)

# Multi-condition rules (AND'd together) -- can't be expressed as a single
# keyword, so they stay hardcoded and read-only. Shown in the UI grid as
# "Combinada" purely for visibility; matched procedurally in
# categorize_chase_description below, never through the flat rules list.
ALFONSO_COMPOUND_DISPLAY_RULES = (
    ("OPERATING ACCT + PAYMENT", "EFT RCV-"),
    ("OPERATING ACCT + CHEVRON", "REBATE COMBUSTIBLE"),
    ("OPERATING ACCT + MONTHLY", "REBATE COMBUSTIBLE"),
    (
        "CHECK/CHEQUE/CHEK + COCA|COKE|MIDTOWN|KING|LIU|ICECREAM",
        "PROVEEDORES",
    ),
    ("CHECK/CHEQUE/CHEK + FLORIDA", "ADMINISTRATION FEE"),
    ("CHECK - FLORI", "PROVEEDORES"),
)


def _desc_has_check_keyword(desc):
    """True when description contains CHECK, CHEQUE, or CHEK (normalized to chek)."""
    return "chek" in desc


def _is_check_proveedores(desc):
    """CHECK/CHEQUE/CHEK combined with vendor keywords -> PROVEEDORES."""
    return _desc_has_check_keyword(desc) and any(
        keyword in desc for keyword in CHASE_CHECK_PROVEEDORES_KEYWORDS
    )


def _is_deposit_id_number(desc):
    """DEPOSIT ID NUMBER with flexible spacing."""
    return "deposit  id number" in desc or (
        "deposit" in desc and "id number" in desc
    )


def categorize_chase_description(description):
    """
    Map Chase Description text to Detalle category.

    Multi-condition rules (OPERATING ACCT/CHECK combos) are checked first --
    they can't be expressed as a flat keyword. Everything else is a single
    longest-keyword-wins match across Maestra (chase_master_rules.json) and
    Personalizada (chase_rules.json) rules together, so an admin-edited or
    newly added rule can override a broader one regardless of which file it
    lives in.

    Returns category string or None when no rule matches.
    """
    desc = normalize_rule_text(description)
    if not desc:
        return None

    if "operating acct" in desc and "chevron" in desc:
        return "REBATE COMBUSTIBLE"
    if "operating acct" in desc and "monthly" in desc:
        return "REBATE COMBUSTIBLE"
    if "operating acct" in desc and "payment" in desc:
        return "EFT RCV-"

    if _desc_has_check_keyword(desc) and "florida" in desc:
        return "ADMINISTRATION FEE"

    if "chek - flori" in desc:
        return "PROVEEDORES"

    if _is_check_proveedores(desc):
        return "PROVEEDORES"

    matched_detail = _match_longest_keyword(desc, load_master_rules() + load_dynamic_rules())
    if matched_detail:
        return matched_detail

    if _is_deposit_id_number(desc):
        return "DEPOSITO"

    return None


# ---------------------------------------------------------------------------
# Read / categorize / write pipeline
# ---------------------------------------------------------------------------

CHASE_COL_POSTING_DATE = 2  # B
CHASE_COL_AMOUNT = 4  # D
CHASE_COL_BALANCE = 6  # F
CHASE_DATE_NUMBER_FORMAT = "dd/mm/yyyy"
CHASE_INTEGER_NUMBER_FORMAT = "0"
CHASE_DECIMAL_NUMBER_FORMAT = "0.00"
CHASE_DATA_START_ROW = 2


def detalle_cell_is_empty(value):
    """True when Chase Detalle (Column H) has no pre-existing content."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    return str(value).strip() == ""


def find_chase_column(df, name_hint, fallback_index):
    """Locate a column by header name or fixed Excel column index."""
    target = name_hint.strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == target:
            return col
    for col in df.columns:
        if target in str(col).strip().lower():
            return col
    if 0 <= fallback_index < len(df.columns):
        return df.columns[fallback_index]
    return None


def ensure_detalle_column(df):
    """Ensure Detalle exists as column H (8th column); create if missing."""
    detalle_col = find_chase_column(df, "Detalle", 7)
    if detalle_col is not None:
        return detalle_col

    if len(df.columns) >= 8:
        return df.columns[7]

    while len(df.columns) < 7:
        df[f"__col_{len(df.columns)}__"] = ""
    df["Detalle"] = ""
    return "Detalle"


def apply_chase_amount_cell(cell, amount):
    """Apply integer or Spanish-decimal Excel format based on the amount value."""
    amount_float = float(amount)
    if amount_float.is_integer():
        cell.value = int(amount_float)
        cell.number_format = CHASE_INTEGER_NUMBER_FORMAT
    else:
        cell.value = amount_float
        cell.number_format = CHASE_DECIMAL_NUMBER_FORMAT


def read_chase_activity_file(file_path):
    """Read Chase bank activity from CSV or Excel."""
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        return pd.read_csv(file_path, dtype=str, keep_default_na=False)
    if extension in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(file_path, dtype=str, keep_default_na=False)
    raise ValueError(
        f"Tipo de archivo no soportado '{extension}'. Use CSV o Excel (.csv, .xlsx, .xlsm)."
    )


def parse_amount_signed(value):
    """Parse amount text into a signed native float."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    if text.startswith("-"):
        negative = True
        text = text[1:].strip()

    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.rsplit(",", 1)
        if len(parts) == 2 and len(parts[1]) <= 2:
            text = parts[0].replace(".", "") + "." + parts[1]
        else:
            text = text.replace(",", "")

    try:
        amount = float(text)
    except ValueError:
        return 0.0

    return -abs(amount) if negative else amount


def parse_posting_date_value(value):
    """Return datetime for Column B or None when parsing fails."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def sync_detalle_from_excel_workbook(df, file_path, detalle_col):
    """Load pre-existing Column H values from Excel before categorization."""
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        return

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    try:
        for row_offset, index in enumerate(df.index):
            excel_row = CHASE_DATA_START_ROW + row_offset
            cell_value = worksheet.cell(row=excel_row, column=8).value
            if not detalle_cell_is_empty(cell_value):
                df.at[index, detalle_col] = cell_value
    finally:
        workbook.close()


def apply_chase_detalle_categorization(df, description_col, detalle_col):
    """
    Apply keyword rules to Detalle (Column H) only for empty cells.

    Rows with any pre-existing Detalle value are left untouched.
    """
    updated_count = 0
    for index in df.index:
        current = df.at[index, detalle_col]
        if not detalle_cell_is_empty(current):
            continue

        description = df.at[index, description_col]
        category = categorize_chase_description(description)
        if category:
            df.at[index, detalle_col] = category
            updated_count += 1
    return updated_count


def write_chase_excel_file(df, file_path, column_map):
    """Write Chase data to Excel with dates, amounts, balance formulas, and formats."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    posting_idx = column_map["posting"]
    amount_idx = column_map["amount"]

    posting_col_name = df.columns[posting_idx]
    amount_col_name = df.columns[amount_idx]
    detalle_col_name = df.columns[column_map["detalle"]]

    protected_detalle = {}
    for row_offset, index in enumerate(df.index):
        excel_row = CHASE_DATA_START_ROW + row_offset
        existing_h = worksheet.cell(row=excel_row, column=8).value
        if not detalle_cell_is_empty(existing_h):
            protected_detalle[excel_row] = existing_h

    for row_offset, index in enumerate(df.index):
        excel_row = CHASE_DATA_START_ROW + row_offset

        for col_num, col_name in enumerate(df.columns, start=1):
            if col_num == 8 and excel_row in protected_detalle:
                continue
            worksheet.cell(row=excel_row, column=col_num, value=df.at[index, col_name])

        posting_dt = parse_posting_date_value(df.at[index, posting_col_name])
        cell_b = worksheet.cell(row=excel_row, column=CHASE_COL_POSTING_DATE)
        if posting_dt:
            cell_b.value = posting_dt
            cell_b.number_format = CHASE_DATE_NUMBER_FORMAT
        else:
            cell_b.value = df.at[index, posting_col_name]

        cell_d = worksheet.cell(row=excel_row, column=CHASE_COL_AMOUNT)
        apply_chase_amount_cell(
            cell_d, parse_amount_signed(df.at[index, amount_col_name])
        )

        cell_f = worksheet.cell(row=excel_row, column=CHASE_COL_BALANCE)
        if excel_row == CHASE_DATA_START_ROW:
            cell_f.value = f"=+F1+D{excel_row}"
        else:
            cell_f.value = f"=+F{excel_row - 1}+D{excel_row}"
        cell_f.number_format = CHASE_DECIMAL_NUMBER_FORMAT

        if excel_row not in protected_detalle:
            detalle_value = str(df.at[index, detalle_col_name]).strip()
            cell_h = worksheet.cell(row=excel_row, column=8)
            cell_h.value = df.at[index, detalle_col_name]
            if detalle_value == EFT_RCV_EXACT_LABEL:
                cell_h.font = EFT_RCV_RED_FONT

    workbook.save(file_path)


def write_chase_csv_file(df, file_path, column_map):
    """Write Chase CSV with formatted dates, floats, and balance formulas."""
    posting_idx = column_map["posting"]
    amount_idx = column_map["amount"]
    balance_idx = column_map["balance"]
    detalle_idx = column_map["detalle"]

    output = df.copy()
    for row_offset, index in enumerate(output.index):
        posting_dt = parse_posting_date_value(output.at[index, output.columns[posting_idx]])
        if posting_dt:
            output.at[index, output.columns[posting_idx]] = posting_dt.strftime("%d/%m/%Y")

        output.at[index, output.columns[amount_idx]] = parse_amount_signed(
            output.at[index, output.columns[amount_idx]]
        )

        excel_row = CHASE_DATA_START_ROW + row_offset
        if excel_row == CHASE_DATA_START_ROW:
            output.at[index, output.columns[balance_idx]] = f"=+F1+D{excel_row}"
        else:
            output.at[index, output.columns[balance_idx]] = (
                f"=+F{excel_row - 1}+D{excel_row}"
            )

    output.to_csv(file_path, index=False)


def process_chase_categorization(file_path):
    """
    Categorize Chase activity, format columns B/D/F, and save the workbook.

    Column B: DD/MM/YYYY dates
    Column D: signed floats with comma decimal display
    Column F: F1 static baseline; F2 =+F1+D2; row 3+ =+F{r-1}+D{r}
    Column H: Detalle keyword categorization
    """
    df = read_chase_activity_file(file_path)

    description_col = find_chase_column(df, "Description", 2)
    if description_col is None:
        raise ValueError(
            "No se encontró la columna Description (se esperaba en la Columna C)."
        )

    posting_col = find_chase_column(df, "Posting Date", 1)
    if posting_col is None:
        posting_col = find_chase_column(df, "Posting", 1)
    amount_col = find_chase_column(df, "Amount", 3)
    balance_col = find_chase_column(df, "Balance", 5)

    if posting_col is None or amount_col is None or balance_col is None:
        raise ValueError(
            "No se encontraron las columnas requeridas (B: Posting Date, D: Amount, F: Balance)."
        )

    detalle_col = ensure_detalle_column(df)
    if detalle_col not in df.columns:
        df[detalle_col] = ""

    sync_detalle_from_excel_workbook(df, file_path, detalle_col)

    updated_count = apply_chase_detalle_categorization(
        df, description_col, detalle_col
    )

    column_map = {
        "posting": list(df.columns).index(posting_col),
        "amount": list(df.columns).index(amount_col),
        "balance": list(df.columns).index(balance_col),
        "detalle": list(df.columns).index(detalle_col),
    }

    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        write_chase_csv_file(df, file_path, column_map)
    elif extension in {".xlsx", ".xlsm"}:
        write_chase_excel_file(df, file_path, column_map)
    elif extension == ".xls":
        raise ValueError(
            "El formato .xls antiguo no soporta fórmulas. Guarde como .xlsx e intente de nuevo."
        )
    else:
        raise ValueError(f"No se puede guardar un tipo de archivo no soportado '{extension}'.")

    return updated_count, len(df)
