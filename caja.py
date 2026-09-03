"""
Módulo Caja: actualiza las columnas K, N, S y T de la hoja "CAJA" del Excel
Cierre mensual con datos ya cargados en otros dos Excel (Chase y Lottery).

- K (depósitos Chase): filas con Detalle "DEPOSITO" o "DEPOSITO GETTEL" (esta
  última se marca además con un comentario en la celda). No se filtra por el
  Type del banco — DEPOSIT/MISC_CREDIT son inconsistentes para un mismo tipo
  de depósito físico real, el Detalle ya categorizado por el módulo Chase
  Bank es la única fuente confiable.
- S/T (Food Truck / máquina de hielo): filas con Detalle "FOOD TRUCK" o
  "DEPOSITO VENTA ICE" — T queda con la etiqueta ("Food Truck"/"ICE MACHINE",
  o ambas si un día combina las dos). "VENTA ICE" sin el prefijo "DEPOSITO"
  es la venta electrónica de la máquina, no un depósito físico, y no cuenta.
- N (Lottery): columna X ("CUENTA FINAL") del Excel de Lottery, matcheada
  por la fecha de su columna B (la fecha real de cada fila, no la A).

En los tres casos la fila de CAJA se ubica por su columna A (fecha del día
de negocio) — nunca por la C (que es esa misma fecha +1 día).
"""

import os
import tempfile
from datetime import date, datetime

import openpyxl
from openpyxl.comments import Comment

CAJA_SHEET_NAME = "CAJA"
CAJA_DATA_START_ROW = 4
CAJA_COL_DATE = 1  # A — Fecha del día de negocio (la que se usa para matchear)
CAJA_COL_CHASE_DEPOSITS = 11  # K
CAJA_COL_LOTTERY = 14  # N
CAJA_COL_FOOD_ICE = 19  # S
CAJA_COL_FOOD_ICE_LABEL = 20  # T — de qué se trata el importe de S (Food Truck / ICE MACHINE)

CHASE_COL_POSTING_DATE = 2  # B
CHASE_COL_AMOUNT = 4  # D
CHASE_COL_DETALLE = 8  # H — ya categorizado por el módulo Chase Bank

CHASE_DETALLE_DEPOSITO = "DEPOSITO"
# Depósito de la máquina/casino Gettel: cuenta como depósito normal en K, pero
# se marca con un comentario en la celda para que quede visible qué día es.
CHASE_DETALLE_GETTEL = "DEPOSITO GETTEL"
GETTEL_COMMENT_TEXT = "Este día incluye el depósito de Gettel."
GETTEL_COMMENT_AUTHOR = "Caja"

# Nota: el Detalle real de un depósito físico de hielo es "DEPOSITO VENTA ICE"
# (Type DEPOSIT) — "VENTA ICE" a secas es la venta reportada por la máquina vía
# ACH_CREDIT/MISC_CREDIT, que NO es un depósito físico y no cuenta acá.
CHASE_DETALLE_FOOD_TRUCK = "FOOD TRUCK"
CHASE_DETALLE_ICE = "DEPOSITO VENTA ICE"
FOOD_ICE_LABELS = {
    CHASE_DETALLE_FOOD_TRUCK: "Food Truck",
    CHASE_DETALLE_ICE: "ICE MACHINE",
}
# Orden fijo de presentación cuando un mismo día combina ambos.
FOOD_ICE_LABEL_ORDER = (CHASE_DETALLE_ICE, CHASE_DETALLE_FOOD_TRUCK)

LOTTERY_COL_DATE_B = 2
LOTTERY_COL_CUENTA_FINAL = 24  # X


def _create_temp_workbook_path():
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="caja_")
    os.close(fd)
    return temp_path


def _normalize_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _get_caja_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower() == CAJA_SHEET_NAME.lower():
            return workbook[name]
    raise ValueError(
        f'Hoja "{CAJA_SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _iter_caja_dates(sheet):
    """Fechas (columna A) de CAJA desde la fila 4 hasta que se acaben los datos del mes."""
    row = CAJA_DATA_START_ROW
    while True:
        date_key = _normalize_date(sheet.cell(row=row, column=CAJA_COL_DATE).value)
        if date_key is None:
            return
        yield row, date_key
        row += 1


def _format_food_ice_label(detalle_keys):
    labels = [FOOD_ICE_LABELS[key] for key in FOOD_ICE_LABEL_ORDER if key in detalle_keys]
    return ", ".join(labels)


def _collect_chase_amounts(chase_path):
    workbook = openpyxl.load_workbook(chase_path, data_only=True)
    sheet = workbook.active

    deposits_by_date = {}
    food_ice_by_date = {}
    food_ice_labels_by_date = {}
    gettel_dates = set()

    for row in range(2, sheet.max_row + 1):
        detalle = sheet.cell(row=row, column=CHASE_COL_DETALLE).value
        detalle_norm = str(detalle).strip().upper() if detalle else ""

        # No se filtra por Type ("DEPOSIT" vs "MISC_CREDIT"/etc.): el banco categoriza
        # de forma inconsistente el Type de depósitos físicos reales del mismo día (se
        # confirmó con datos reales que DEPOSITO/DEPOSITO GETTEL/FOOD TRUCK/DEPOSITO
        # VENTA ICE aparecen indistintamente como DEPOSIT o MISC_CREDIT, siempre con la
        # Descripción cruda "DEPOSIT") — el Detalle, ya categorizado por el módulo Chase
        # Bank, es la única fuente confiable acá. "VENTA ICE" sin el prefijo "DEPOSITO"
        # (venta electrónica de la máquina vía ACH_CREDIT, no un depósito físico) queda
        # afuera solo porque no matchea ninguna de las claves de FOOD_ICE_LABELS.
        date_key = _normalize_date(sheet.cell(row=row, column=CHASE_COL_POSTING_DATE).value)
        amount = sheet.cell(row=row, column=CHASE_COL_AMOUNT).value
        if date_key is None or not isinstance(amount, (int, float)):
            continue

        if detalle_norm == CHASE_DETALLE_DEPOSITO:
            deposits_by_date[date_key] = deposits_by_date.get(date_key, 0.0) + float(amount)
        elif detalle_norm == CHASE_DETALLE_GETTEL:
            deposits_by_date[date_key] = deposits_by_date.get(date_key, 0.0) + float(amount)
            gettel_dates.add(date_key)
        elif detalle_norm in FOOD_ICE_LABELS:
            food_ice_by_date[date_key] = food_ice_by_date.get(date_key, 0.0) + float(amount)
            food_ice_labels_by_date.setdefault(date_key, set()).add(detalle_norm)

    return deposits_by_date, food_ice_by_date, food_ice_labels_by_date, gettel_dates


def apply_chase_deposits(cierre_path, chase_path):
    """Escribe K (DEPOSITO/GETTEL) y S (FOOD TRUCK / hielo) en CAJA desde el Chase."""
    deposits_by_date, food_ice_by_date, food_ice_labels_by_date, gettel_dates = _collect_chase_amounts(
        chase_path
    )
    if not deposits_by_date and not food_ice_by_date:
        raise ValueError(
            "No se encontraron depósitos (Detalle DEPOSITO / DEPOSITO GETTEL / FOOD TRUCK / "
            "DEPOSITO VENTA ICE) en el Chase. ¿Ya pasó por el módulo Chase Bank para categorizarse?"
        )

    workbook = openpyxl.load_workbook(cierre_path, data_only=False)
    sheet = _get_caja_sheet(workbook)

    remaining_deposits = dict(deposits_by_date)
    remaining_food_ice = dict(food_ice_by_date)
    deposits_written = {}
    food_ice_written = {}
    gettel_days_written = []

    for row, date_key in _iter_caja_dates(sheet):
        if date_key in remaining_deposits:
            amount = round(remaining_deposits.pop(date_key), 2)
            cell = sheet.cell(row=row, column=CAJA_COL_CHASE_DEPOSITS, value=amount)
            if date_key in gettel_dates:
                cell.comment = Comment(GETTEL_COMMENT_TEXT, GETTEL_COMMENT_AUTHOR)
                gettel_days_written.append(date_key)
            deposits_written[date_key] = amount
        if date_key in remaining_food_ice:
            amount = round(remaining_food_ice.pop(date_key), 2)
            sheet.cell(row=row, column=CAJA_COL_FOOD_ICE, value=amount)
            label_text = _format_food_ice_label(food_ice_labels_by_date.get(date_key, set()))
            if label_text:
                sheet.cell(row=row, column=CAJA_COL_FOOD_ICE_LABEL, value=label_text)
            food_ice_written[date_key] = amount

    temp_path = _create_temp_workbook_path()
    workbook.save(temp_path)

    summary = {
        "deposits_written": deposits_written,
        "food_ice_written": food_ice_written,
        "deposits_unmatched": remaining_deposits,
        "food_ice_unmatched": remaining_food_ice,
        "gettel_days_written": gettel_days_written,
    }
    return temp_path, summary


def _collect_lottery_cuenta_final(lottery_path):
    workbook = openpyxl.load_workbook(lottery_path, data_only=True)
    sheet = workbook.active

    values_by_date = {}
    missing_dates = []

    for row in range(1, sheet.max_row + 1):
        date_key = _normalize_date(sheet.cell(row=row, column=LOTTERY_COL_DATE_B).value)
        if date_key is None:
            continue
        value = sheet.cell(row=row, column=LOTTERY_COL_CUENTA_FINAL).value
        if isinstance(value, (int, float)):
            values_by_date[date_key] = float(value)
        else:
            missing_dates.append(date_key)

    return values_by_date, missing_dates


def apply_lottery_cuenta_final(cierre_path, lottery_path):
    """Escribe N en CAJA con la columna X ("CUENTA FINAL") del Excel de Lottery."""
    values_by_date, missing_dates = _collect_lottery_cuenta_final(lottery_path)
    if not values_by_date and not missing_dates:
        raise ValueError('No se encontraron fechas en la columna B del Excel de Lottery.')

    workbook = openpyxl.load_workbook(cierre_path, data_only=False)
    sheet = _get_caja_sheet(workbook)

    remaining = dict(values_by_date)
    written = {}

    for row, date_key in _iter_caja_dates(sheet):
        if date_key in remaining:
            value = round(remaining.pop(date_key), 2)
            sheet.cell(row=row, column=CAJA_COL_LOTTERY, value=value)
            written[date_key] = value

    temp_path = _create_temp_workbook_path()
    workbook.save(temp_path)

    summary = {
        "written": written,
        "unmatched": remaining,
        "missing_cached_value": missing_dates,
    }
    return temp_path, summary
