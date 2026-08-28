import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD

    TKINTER_DND_AVAILABLE = True
except ImportError:
    TKINTER_DND_AVAILABLE = False

import pandas as pd

try:
    import pdfplumber
except ImportError as _pdf_exc:
    pdfplumber = None
    _PDFPLUMBER_IMPORT_ERROR = _pdf_exc
else:
    _PDFPLUMBER_IMPORT_ERROR = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as _openpyxl_exc:
    load_workbook = None
    _OPENPYXL_IMPORT_ERROR = _openpyxl_exc
else:
    _OPENPYXL_IMPORT_ERROR = None

try:
    from cmv_costo import (
        join_paths,
        split_paths,
        update_master_costo_todos_bulk,
        workbook_has_costo_todos_sheet,
    )
except ImportError as _cmv_exc:
    join_paths = None
    split_paths = None
    update_master_costo_todos_bulk = None
    workbook_has_costo_todos_sheet = None
    _CMV_IMPORT_ERROR = _cmv_exc
else:
    _CMV_IMPORT_ERROR = None

try:
    from chase_rules import (
        add_dynamic_rule,
        delete_dynamic_rule,
        delete_dynamic_rule_by_index,
        load_dynamic_rules,
        match_dynamic_detalle,
    )
except ImportError as _chase_rules_exc:
    add_dynamic_rule = None
    delete_dynamic_rule = None
    delete_dynamic_rule_by_index = None
    load_dynamic_rules = None
    match_dynamic_detalle = None
    _CHASE_RULES_IMPORT_ERROR = _chase_rules_exc
else:
    _CHASE_RULES_IMPORT_ERROR = None

try:
    from monthly_sales import process_monthly_sales
    from monthly_sales import join_paths as _monthly_sales_join_paths
    from monthly_sales import split_paths as _monthly_sales_split_paths
except ImportError as _sales_exc:
    process_monthly_sales = None
    _SALES_IMPORT_ERROR = _sales_exc
else:
    _SALES_IMPORT_ERROR = None
    # join_paths/split_paths are identical in cmv_costo and monthly_sales; keep
    # whichever import already succeeded instead of letting this one clobber it.
    if join_paths is None:
        join_paths = _monthly_sales_join_paths
    if split_paths is None:
        split_paths = _monthly_sales_split_paths

try:
    from reporte_diario import process_reporte_diario, process_store_info, process_lottery
except ImportError as _reporte_exc:
    process_reporte_diario = None
    process_store_info = None
    process_lottery = None
    _REPORTE_IMPORT_ERROR = _reporte_exc
else:
    _REPORTE_IMPORT_ERROR = None

try:
    from cupones_append import (
        MonthlyReportFullyDuplicateError,
        NoPendingCouponsError,
        append_monthly_cupones,
        resync_cupones_only,
    )
except ImportError as _cupones_append_exc:
    append_monthly_cupones = None
    resync_cupones_only = None  # type: ignore[assignment,misc]
    MonthlyReportFullyDuplicateError = None  # type: ignore[assignment,misc]
    NoPendingCouponsError = None  # type: ignore[assignment,misc]
    _CUPONES_APPEND_IMPORT_ERROR = _cupones_append_exc
else:
    _CUPONES_APPEND_IMPORT_ERROR = None

try:
    from gettel_toyota_parser import (
        merge_gettel_toyota_into_master,
        merge_gettel_toyota_pdf_into_master,
        process_gettel_pagos,
    )
except ImportError as _gettel_toyota_exc:
    merge_gettel_toyota_into_master = None
    merge_gettel_toyota_pdf_into_master = None
    process_gettel_pagos = None
    _GETTEL_TOYOTA_IMPORT_ERROR = _gettel_toyota_exc
else:
    _GETTEL_TOYOTA_IMPORT_ERROR = None

try:
    from proveedores import append_supplier_invoices, append_supplier_payments
except ImportError as _proveedores_exc:
    append_supplier_invoices = None
    append_supplier_payments = None
    _PROVEEDORES_IMPORT_ERROR = _proveedores_exc
else:
    _PROVEEDORES_IMPORT_ERROR = None

try:
    from proveedores_pago_rules import (
        add_dynamic_rule as add_proveedores_pago_rule,
        delete_dynamic_rule_by_index as delete_proveedores_pago_rule_by_index,
        load_dynamic_rules as load_proveedores_pago_rules,
    )
except ImportError as _proveedores_pago_rules_exc:
    add_proveedores_pago_rule = None
    delete_proveedores_pago_rule_by_index = None
    load_proveedores_pago_rules = None
    _PROVEEDORES_PAGO_RULES_IMPORT_ERROR = _proveedores_pago_rules_exc
else:
    _PROVEEDORES_PAGO_RULES_IMPORT_ERROR = None

from ui_theme import (
    ANALISIS_MASTER_FILETYPES,
    CHASE_THEME,
    CMV_THEME,
    DEPT_FILETYPES,
    EFT_THEME,
    EXCEL_FILETYPES_MASTER,
    FONT,
    GETTEL_THEME,
    PDF_DAILY_FILETYPES,
    PROVEEDORES_THEME,
    REPORTE_DIARIO_THEME,
    SALES_FILETYPES,
    SALES_THEME,
    THEME,
    WINDOW_GEOMETRY,
    WINDOW_MINSIZE,
    apply_notebook_style,
    apply_root_style,
    create_card,
    create_compact_entry,
    create_compact_section_header,
    create_dual_column_tab,
    create_file_row,
    create_header_banner,
    create_info_panel,
    create_log_panel,
    create_panel_label,
    create_primary_button,
    create_scrollable_body,
    create_scrollable_tab_frame,
    create_secondary_button,
    create_status_bar,
    make_tab_icon,
    set_status_style,
)

SHEET_NAME = "Cta Cte J.H.Williams"
LEDGER_START_ROW = 5
BOTTOM_SCAN_START_ROW = 5000
BLOCK_FIRST_COLUMN = 1
PERIMETER_LAST_COLUMN = 13  # A through M (boxed perimeter)
COLUMN_N = 14
COUPON_LAST_COLUMN = 6
EFT_DATE_NUMBER_FORMAT = "dd-mmm"  # Column N — Spanish day-month display (e.g. 10-jun)
CURRENCY_NUMBER_FORMAT = (
    '_(* #,##0.00" $"_);_(* \\(#,##0.00" $"\\);_(* "-"?? " $"_);_(@_)'
)
CURRENCY_COLUMNS_COUPON = (4, 5, 6)  # D, E, F
CURRENCY_COLUMNS_SUMMARY = (10, 12)  # J, L

THIN_SIDE = Side(style="thin", color="000000")
THICK_SIDE = Side(style="medium", color="000000")
CLEAR_FILL = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
NO_FILL = PatternFill(fill_type=None)
DEFAULT_FONT = Font()
BOLD_FONT = Font(bold=True)
EFT_RCV_EXACT_LABEL = "EFT RCV-"
EFT_RCV_RED_FONT = Font(color="FF0000", bold=True)
SUMMARY_ALIGNMENT = Alignment(vertical="center", horizontal="left")
SUMMARY_FIRST_COLUMN = 7
SUMMARY_LAST_COLUMN = 14  # G through N on middle_row
COUPON_DATE_TEXT_PATTERN = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
EFT_DUPLICATE_ALERT = (
    "Alerta: Este EFT ya fue cargado anteriormente con los mismos datos."
)
SUM_FORMULA_RANGE_PATTERN = re.compile(
    r"=\+?\s*SUM\s*\(\s*F(\d+)\s*:\s*F(\d+)\s*\)",
    re.IGNORECASE,
)

CREDIT_COUPON_PATTERN = re.compile(
    r"\b(SI-\d+)/(DDC-\d+)\b", re.IGNORECASE
)
SI_INVOICE_PATTERN = re.compile(r"\b(SI-\d+)\b", re.IGNORECASE)
DDC_COUPON_PATTERN = re.compile(r"\b(DDC-\d+)\b", re.IGNORECASE)
DRAFT_NO_PATTERN = re.compile(r"\b(RCV-\d+)\b", re.IGNORECASE)
DATE_PATTERN = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
EFT_DATE_LABEL_PATTERN = re.compile(
    r"EFT\s*Date\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{4})",
    re.IGNORECASE,
)
CURRENCY_DECIMAL_PATTERN = re.compile(r"\.\d{1,2}\b|,\d{2}\b")


def parse_amount(value):
    """Parse a currency string into a positive float using abs()."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return abs(float(value))
        except (TypeError, ValueError):
            return None

    text = str(value).strip()
    if not text or text.upper() in {"-", "--", "N/A", "NA"}:
        return None

    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1].strip()

    # European decimal comma (14,33) -> 14.33
    if re.search(r",\d{2}\b", text) and "." not in text.split(",")[-1][:3]:
        text = text.replace(".", "").replace(",", ".")

    text = text.replace("$", "").replace(" ", "")
    if text.count(",") > 0 and "." not in text:
        parts = text.rsplit(",", 1)
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) <= 2:
            text = parts[0].replace(",", "") + "." + parts[1]
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", "")

    text = text.strip()
    if text.startswith("-"):
        text = text[1:].strip()
    elif text.endswith("-"):
        text = text[:-1].strip()

    if not text:
        return None

    try:
        return abs(float(text))
    except ValueError:
        return None


def parse_date_to_datetime(date_str):
    """Parse a date string into datetime (tries common PDF formats)."""
    if not date_str:
        return None
    text = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_coupon_date_us(date_str):
    """Column A: strict US format MM/DD/YYYY (e.g. 05/08/2026)."""
    parsed = parse_date_to_datetime(date_str)
    if parsed:
        return parsed.strftime("%m/%d/%Y")
    return date_str


def parse_eft_pdf_date_us(date_str):
    """Parse an EFT PDF date token strictly as MM/DD/YYYY (American)."""
    if not date_str:
        return None
    text = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_elapsed_duration(seconds):
    """Elapsed time as '45 seg' or '6:30 min', for a process-completed status."""
    total_seconds = int(round(seconds))
    if total_seconds < 60:
        return f"{total_seconds} seg"
    minutes, secs = divmod(total_seconds, 60)
    return f"{minutes}:{secs:02d} min"


def format_eft_date_us(date_str):
    """Header EFT date as MM/DD/YYYY string, matching the source PDF."""
    parsed = parse_eft_pdf_date_us(date_str)
    if parsed:
        return parsed.strftime("%m/%d/%Y")
    return date_str


def parse_date_cell(cell):
    """Extract date from a PDF cell; returns MM/DD/YYYY for coupon rows."""
    if not cell:
        return None
    match = DATE_PATTERN.search(str(cell).strip())
    if match:
        return format_coupon_date_us(match.group(1))
    return None


def normalize_row_cells(cells):
    """Return a safe list of stripped string cell values."""
    try:
        return [str(cell).strip() if cell else "" for cell in cells]
    except Exception:
        return []


def expand_row_components(cells):
    """
    Flatten row cells into scan components.

    Splits comma-separated reference blobs (e.g. 0764,0765,0766,9133)
    so they do not shift currency column positions.
    """
    components = []
    for cell in cells:
        if not cell:
            continue
        text = str(cell).strip()
        if not text:
            continue
        if is_reference_blob(text):
            for part in text.split(","):
                part = part.strip()
                if part:
                    components.append(part)
        else:
            components.append(text)
    return components


def is_reference_blob(text):
    """True for comma-separated tracking refs without currency markers."""
    if not text or "$" in text or "(" in text or ")" in text:
        return False
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return False
    if "," not in text:
        return False
    parts = [p.strip() for p in text.split(",")]
    if len(parts) < 2:
        return False
    return all(
        p.replace(" ", "").isdigit() and 1 <= len(p.replace(" ", "")) <= 8
        for p in parts
    )


def is_tracking_reference(value):
    """Detect single tracking numbers (not currency)."""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if is_reference_blob(text):
        return True
    if "$" in text or "(" in text or ")" in text:
        return False
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return False
    cleaned = text.replace(",", "").replace(" ", "")
    if not cleaned.isdigit():
        return False
    return len(cleaned) >= 4


def is_currency_component(value):
    """
    True when a token looks like Gross/Fees/Paid money.

    Uses $, parentheses, or decimal patterns (dot or comma cents).
    """
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if is_reference_blob(text) or is_tracking_reference(text):
        return False
    if "$" in text or "(" in text or ")" in text:
        return True
    if re.search(r"-\s*\$|\$\s*-", text):
        return True
    if text.startswith("-") and re.search(r"\d", text):
        return True
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return True
    if re.fullmatch(r"-?\d+[.,]\d{2}", text.replace("$", "").strip()):
        return True
    return False


def extract_last_three_financial_amounts(components):
    """
    Extract Gross, Fees, and Paid from currency tokens at the row tail.

    Variable reference columns are already removed via expand_row_components().
    When four amounts exist (Gross, Fees, Net, Paid), Net is skipped by taking
    the 4th-from-last, 3rd-from-last, and last currency tokens.
    """
    currency_tokens = [c for c in components if is_currency_component(c)]
    if len(currency_tokens) < 3:
        return None, None, None

    if len(currency_tokens) >= 4:
        selected = [currency_tokens[-4], currency_tokens[-3], currency_tokens[-1]]
    else:
        selected = currency_tokens[-3:]

    amounts = [parse_amount(token) for token in selected]
    if any(amount is None for amount in amounts):
        return None, None, None
    return amounts[0], amounts[1], amounts[2]


def row_contains_credit_coupon(cells):
    """True when the row is a coupon table row."""
    try:
        row_text = " ".join(cells)
        if CREDIT_COUPON_PATTERN.search(row_text):
            return True
        has_si = any(SI_INVOICE_PATTERN.search(c) for c in cells)
        has_ddc = any(DDC_COUPON_PATTERN.search(c) for c in cells)
        return has_si and has_ddc
    except Exception:
        return False


def locate_coupon_combo(cells):
    """Find invoice, coupon, and combo cell index."""
    for index, cell in enumerate(cells):
        match = CREDIT_COUPON_PATTERN.search(cell)
        if match:
            return match.group(1).upper(), match.group(2).upper(), index

    invoice = coupon = None
    si_index = ddc_index = None
    for index, cell in enumerate(cells):
        si_match = SI_INVOICE_PATTERN.search(cell)
        if si_match and "/" not in cell:
            invoice = si_match.group(1).upper()
            si_index = index
        ddc_match = DDC_COUPON_PATTERN.search(cell)
        if ddc_match:
            coupon = ddc_match.group(0).upper()
            ddc_index = index

    if invoice and coupon:
        return invoice, coupon, max(si_index or 0, ddc_index or 0)

    row_text = " ".join(cells)
    match = CREDIT_COUPON_PATTERN.search(row_text)
    if match:
        return match.group(1).upper(), match.group(2).upper(), 1

    return None, None, None


def extract_coupon_columns(cells, fallback_date=None):
    """Parse coupon row using last-three currency detection."""
    cells = normalize_row_cells(cells)
    if not cells or not row_contains_credit_coupon(cells):
        return None

    try:
        invoice, coupon, combo_index = locate_coupon_combo(cells)
        if not invoice or not coupon:
            return None

        components = expand_row_components(cells)
        gross, fees, paid = extract_last_three_financial_amounts(components)
        if gross is None:
            return None

        row_date = None
        if cells:
            row_date = parse_date_cell(cells[0])
        if not row_date and combo_index > 0:
            row_date = parse_date_cell(cells[combo_index - 1])
        if not row_date:
            for cell in cells:
                row_date = parse_date_cell(cell)
                if row_date:
                    break
        if not row_date:
            match = DATE_PATTERN.search(" ".join(cells))
            if match:
                row_date = format_coupon_date_us(match.group(1))
        if not row_date and fallback_date:
            row_date = format_coupon_date_us(fallback_date)

        fees_value = fees if fees is not None else 0.0

        return {
            "date": row_date,
            "invoice": invoice,
            "coupon": coupon,
            "gross_amount": gross,
            "fees_amount": fees_value,
            "paid_amount": paid if paid is not None else 0.0,
        }
    except Exception:
        return None


def _si_match_inside_credit_combo(text, match):
    """True when an SI token belongs to an SI/DDC combo reference."""
    span_start = max(0, match.start() - 12)
    span_end = min(len(text), match.end() + 24)
    return bool(CREDIT_COUPON_PATTERN.search(text[span_start:span_end]))


def extract_paid_invoice_entries(cells):
    """
    Parse one PDF row for all paid invoices at the EFT header (ordered list).

    Collects every SI- reference on the row (not part of an SI/DDC coupon combo)
    and pairs each with a currency amount when available.
    """
    cells = normalize_row_cells(cells)
    if not cells or row_contains_credit_coupon(cells):
        return []

    try:
        invoices_ordered = []
        seen_invoices = set()
        for cell in cells:
            if CREDIT_COUPON_PATTERN.search(cell):
                continue
            for match in SI_INVOICE_PATTERN.finditer(cell):
                if _si_match_inside_credit_combo(cell, match):
                    continue
                invoice = match.group(1).upper()
                if invoice in seen_invoices:
                    continue
                seen_invoices.add(invoice)
                invoices_ordered.append(invoice)

        if not invoices_ordered:
            row_text = " ".join(cells)
            if not SI_INVOICE_PATTERN.search(row_text):
                return []
            for match in SI_INVOICE_PATTERN.finditer(row_text):
                if _si_match_inside_credit_combo(row_text, match):
                    continue
                invoice = match.group(1).upper()
                if invoice in seen_invoices:
                    continue
                seen_invoices.add(invoice)
                invoices_ordered.append(invoice)

        if not invoices_ordered:
            return []

        components = expand_row_components(cells)
        amounts = []
        for token in components:
            if not is_currency_component(token):
                continue
            paid_amount = parse_amount(token)
            if paid_amount is not None:
                amounts.append(paid_amount)
        if not amounts:
            return []

        if len(amounts) >= len(invoices_ordered):
            paired_amounts = amounts[-len(invoices_ordered) :]
        elif len(invoices_ordered) == 1:
            paired_amounts = amounts[-1:]
        else:
            return []

        return [
            {"invoice": invoice, "paid_amount": paid_amount}
            for invoice, paid_amount in zip(invoices_ordered, paired_amounts)
        ]
    except Exception:
        return []


def collect_pdf_rows(pdf_path):
    """Extract table rows and text lines from every PDF page."""
    full_text_parts = []
    rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                full_text_parts.append(page_text)

            for table in page.extract_tables() or []:
                for row in table:
                    if not row or not any(cell for cell in row):
                        continue
                    rows.append(normalize_row_cells(row))

            for line in page_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    if "\t" in line:
                        cells = [p.strip() for p in line.split("\t") if p.strip()]
                    else:
                        cells = re.split(r"\s{2,}", line)
                        if len(cells) <= 1:
                            cells = line.split()
                    if cells:
                        rows.append(cells)
                except Exception:
                    continue

    return "\n".join(full_text_parts), rows


def extract_header(full_text):
    """Extract EFT header date (MM/DD/YYYY) and draft number."""
    header_data = {"eft_date": None, "draft_no": None}
    try:
        draft_match = DRAFT_NO_PATTERN.search(full_text)
        if draft_match:
            header_data["draft_no"] = draft_match.group(1).upper()

        date_match = EFT_DATE_LABEL_PATTERN.search(full_text)
        if not date_match:
            date_match = DATE_PATTERN.search(full_text)
        if date_match:
            header_data["eft_date"] = format_eft_date_us(date_match.group(1))
    except Exception:
        pass
    return header_data


def extract_eft_data(pdf_path):
    """Read EFT PDF and return (header_data, paid_invoices, credit_coupons)."""
    full_text, rows = collect_pdf_rows(pdf_path)
    header_data = extract_header(full_text)
    fallback_date = header_data.get("eft_date")

    paid_invoices = []
    credit_coupons = []
    seen_invoices = set()
    seen_coupons = set()
    coupon_section_started = False

    for cells in rows:
        try:
            if row_contains_credit_coupon(cells):
                coupon_section_started = True

            if coupon_section_started:
                coupon_row = extract_coupon_columns(cells, fallback_date=fallback_date)
                if coupon_row:
                    key = (
                        coupon_row["date"],
                        coupon_row["invoice"],
                        coupon_row["coupon"],
                        coupon_row["gross_amount"],
                        coupon_row["fees_amount"],
                        coupon_row["paid_amount"],
                    )
                    if key not in seen_coupons:
                        seen_coupons.add(key)
                        credit_coupons.append(coupon_row)
                continue

            for paid_entry in extract_paid_invoice_entries(cells):
                key = (paid_entry["invoice"], paid_entry["paid_amount"])
                if key not in seen_invoices:
                    seen_invoices.add(key)
                    paid_invoices.append(paid_entry)
        except Exception:
            continue

    return header_data, paid_invoices, credit_coupons


def is_valid_historical_coupon_date(value):
    """True when column A holds a historical coupon date."""
    if value is None or value == "":
        return False
    if isinstance(value, datetime):
        return True
    text = str(value).strip()
    if COUPON_DATE_TEXT_PATTERN.match(text):
        return True
    if DATE_PATTERN.search(text):
        return True
    return False


def find_last_active_row_bottom_up(worksheet):
    """
    Scan column A from the bottom upward to find the last historical coupon date.

    Avoids mistaking empty middle rows for the ledger end.
    """
    scan_start = max(worksheet.max_row, BOTTOM_SCAN_START_ROW)

    for row in range(scan_start, LEDGER_START_ROW - 1, -1):
        cell_a = worksheet.cell(row=row, column=1).value
        if is_valid_historical_coupon_date(cell_a):
            return row

    raise ValueError(
        "No se encontró una fecha de cupón histórica en la columna A. "
        f"Se revisaron las filas {scan_start} hasta {LEDGER_START_ROW}."
    )


def calculate_middle_row(start_row, end_row):
    """Return the exact middle row of the inserted coupon block."""
    return start_row + ((end_row - start_row) // 2)


def build_balance_formula(middle_row, paid_invoice_count=1):
    """
    Build Column M balance: parent EFT amount minus the sum of all linked invoices.

    Invoice amounts live in Column L from middle_row downward (one per invoice).
    """
    count = max(1, paid_invoice_count)
    if count == 1:
        return f"=+J{middle_row}-L{middle_row}"
    last_invoice_row = middle_row + count - 1
    return f"=+J{middle_row}-SUM(L{middle_row}:L{last_invoice_row})"


def _write_paid_invoices_to_columns_kl(worksheet, middle_row, paid_invoices):
    """Write ordered paid invoices down Columns K and L on existing rows only."""
    for index, invoice_entry in enumerate(paid_invoices):
        row = middle_row + index
        cell_k = worksheet.cell(row=row, column=11, value=invoice_entry["invoice"])
        cell_l = worksheet.cell(
            row=row,
            column=12,
            value=float(invoice_entry["paid_amount"]),
        )
        apply_currency_format(cell_l)
        cell_k.font = BOLD_FONT
        cell_k.alignment = SUMMARY_ALIGNMENT
        cell_l.font = BOLD_FONT
        cell_l.alignment = SUMMARY_ALIGNMENT


def apply_currency_format(cell):
    """Apply regional currency display format to a numeric cell."""
    cell.number_format = CURRENCY_NUMBER_FORMAT


def clear_summary_columns(worksheet, start_row, end_row, middle_row):
    """Clear values in G-N on non-middle rows; column N stays borderless for native gridlines."""
    for row in range(start_row, end_row + 1):
        if row == middle_row:
            continue
        for col in range(SUMMARY_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = None
            cell.font = DEFAULT_FONT
            cell.number_format = "General"
        cell_n = worksheet.cell(row=row, column=COLUMN_N)
        cell_n.value = None
        cell_n.fill = NO_FILL
        cell_n.border = Border()


def apply_eft_block_styling(worksheet, start_row, end_row):
    """
    Two-step border grid for columns A-M (1-13). Column N is excluded entirely.
    """
    thin_all_sides = Border(
        top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE
    )

    # Loop 1: thin borders on all 4 sides of every A-M cell (soft internal gridlines)
    for row in range(start_row, end_row + 1):
        for col in range(BLOCK_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = CLEAR_FILL
            cell.font = DEFAULT_FONT
            cell.border = thin_all_sides

    # Loop 2: medium perimeter on outer bounds only (top, bottom, left A, right M)
    for row in range(start_row, end_row + 1):
        for col in range(BLOCK_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            sides = cell.border
            top = THICK_SIDE if row == start_row else sides.top
            bottom = THICK_SIDE if row == end_row else sides.bottom
            left = THICK_SIDE if col == BLOCK_FIRST_COLUMN else sides.left
            right = THICK_SIDE if col == PERIMETER_LAST_COLUMN else sides.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def apply_summary_row_formatting(worksheet, middle_row):
    """Bold and vertically center summary cells G through N on middle_row only."""
    currency_summary_cols = {8, 9, 10, 12, 13}
    for col in range(SUMMARY_FIRST_COLUMN, SUMMARY_LAST_COLUMN + 1):
        cell = worksheet.cell(row=middle_row, column=col)
        cell.font = BOLD_FONT
        cell.alignment = SUMMARY_ALIGNMENT
        if col in currency_summary_cols:
            apply_currency_format(cell)


def apply_column_n_date(worksheet, middle_row, eft_date_str):
    """Write EFT date in N{middle_row} — bold, dd-mmm display, no borders."""
    cell = worksheet.cell(row=middle_row, column=COLUMN_N)
    cell.font = BOLD_FONT
    cell.border = Border()
    cell.alignment = SUMMARY_ALIGNMENT

    if not eft_date_str:
        return

    eft_dt = parse_eft_pdf_date_us(eft_date_str)
    if eft_dt is None:
        eft_dt = parse_date_to_datetime(eft_date_str)
    if eft_dt:
        cell.value = eft_dt
    else:
        cell.value = eft_date_str
    cell.number_format = EFT_DATE_NUMBER_FORMAT


def open_excel_workbook(path):
    """Open the saved Excel file with the system default application."""
    if sys.platform == "win32":
        time.sleep(0.35)
        os.startfile(path)
    elif sys.platform == "darwin":
        os.system(f'open "{path}"')
    else:
        os.system(f'xdg-open "{path}"')


WORKSPACE_DIR_NAME = "BradentonApp_workspace"


def get_local_workspace_dir():
    """Local temp directory for writable copies of network/locked files."""
    workspace = os.path.join(tempfile.gettempdir(), WORKSPACE_DIR_NAME)
    os.makedirs(workspace, exist_ok=True)
    return workspace


def create_temp_work_copy(source_path):
    """Copy source spreadsheet to local workspace; return absolute temp path."""
    source_path = os.path.abspath(source_path)
    if not os.path.isfile(source_path):
        raise FileNotFoundError(f"Archivo no encontrado: {source_path}")

    base_name = os.path.basename(source_path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    temp_path = os.path.join(get_local_workspace_dir(), f"{stamp}_{base_name}")
    shutil.copy2(source_path, temp_path)
    return temp_path


def detalle_cell_is_empty(value):
    """True when Chase Detalle (Column H) has no pre-existing content."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    return str(value).strip() == ""


def normalize_chase_text(value):
    """
    Normalize Description text for case-insensitive Chase keyword matching.

    Lowercases, collapses whitespace, strips accents, and maps CHECK/CHEQUE
    to CHEK so all cheque variants behave identically.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\b(check|cheque)\b", "chek", text)
    return re.sub(r"\s+", " ", text).strip()


CHASE_CHECK_WORDS = ("chek",)


def _desc_has_check_keyword(desc):
    """True when description contains CHECK, CHEQUE, or CHEK (normalized to chek)."""
    return "chek" in desc


def _desc_contains_any(desc, keywords):
    """Return True if any keyword appears in normalized description text."""
    return any(keyword in desc for keyword in keywords)


# Alfonso master keyword groups (case-insensitive via normalize_chase_text).
CHASE_PROVEEDORES_KEYWORDS = (
    "frito-la",
    "hackneyrectampa",
    "cec distributing",
    "gold coast eagle",
    "jj taylor distri",
    "pbg",
    "colonial",
    "redbull",
    "airgas",
    "johnson brothers",
)
CHASE_GASTOS_BANCARIOS_KEYWORDS = (
    "low value",
    "initial fee",
    "cash deposit immediate",
    "monthly service fee",
)
CHASE_REBATE_KEYWORDS = (
    "helix ucp",
    "ussmokless",
    "njoy",
    "itg brands",
    "john middleton",
)
CHASE_AGUA_KEYWORDS = ("mucs", "manatee")
CHASE_ALARMA_KEYWORDS = ("slomin's", "slomins", "slomin")
CHASE_ENERGIA_KEYWORDS = ("fpl",)
CHASE_TELEFONO_KEYWORDS = ("text me",)
CHASE_ELISTAR_KEYWORDS = ("innov",)
CHASE_CHECK_PROVEEDORES_KEYWORDS = (
    "coca",
    "coke",
    "midtown",
    "king",
    "liu",
    "icecream",
    "ice cream",
)
CHASE_SEGURO_KEYWORDS = ("ipf",)
CHASE_SALE_TAX_KEYWORDS = ("fla dept",)
CHASE_ALG_DISTR_KEYWORDS = ("alg distr",)
CHASE_ADMIN_FEE_KEYWORDS = ("finova",)

# Alfonso single-string master rules (keyword substring -> Detalle).
ALFONSO_MASTER_SINGLE_RULES = (
    ("orig co name:mvnt", "REBATE"),
    ("orig co name:fla lottery", "LOTTERY"),
    ("orig co name:cantaloupe", "VENTA ICE"),
    ("online realtime vendor payment", "SUELDOS"),
    ("online realtime payroll payment", "SUELDOS"),
    ("deposit  id number", "DEPOSITO"),
    ("spectrum", "INTERNET"),
    ("jeffrey's lawn", "REPARACION Y MANTENIMIENTO"),
    ("jeffreys lawn", "REPARACION Y MANTENIMIENTO"),
    ("merchant bank", "COMISIONES Y GASTOS BANCARIOS"),
    ("reynolds", "REBATE"),
    ("fla lottery", "LOTTERY"),
    ("cantaloupe", "VENTA ICE"),
    ("mvnt", "REBATE"),
)

# Alfonso compound rules for UI display (read-only baseline).
ALFONSO_COMPOUND_DISPLAY_RULES = (
    ("OPERATING ACCT + PAYMENT", "EFT RCV-"),
    ("OPERATING ACCT + CHEVRON", "REBATE COMBUSTIBLE"),
    ("OPERATING ACCT + MONTHLY", "REBATE COMBUSTIBLE"),
    (
        "CHECK/CHEQUE/CHEK + COCA|COKE|MIDTOWN|KING|LIU|ICECREAM",
        "PROVEEDORES",
    ),
    ("CHECK/CHEQUE/CHEK + FLORIDA", "ADMINISTRATION FEE"),
    ("CHECK - FLORI", "PROVEEDORES"),
)


def _is_check_proveedores(desc):
    """CHECK/CHEQUE/CHEK combined with vendor keywords -> PROVEEDORES."""
    return _desc_has_check_keyword(desc) and _desc_contains_any(
        desc, CHASE_CHECK_PROVEEDORES_KEYWORDS
    )


def _is_deposit_id_number(desc):
    """DEPOSIT ID NUMBER with flexible spacing."""
    return "deposit  id number" in desc or (
        "deposit" in desc and "id number" in desc
    )


def categorize_chase_description(description):
    """
    Map Chase Description text to Detalle category using Alfonso's keyword rules.

    Returns category string or None when no rule matches.
    """
    desc = normalize_chase_text(description)
    if not desc:
        return None

    # Double Rule 1: OPERATING ACCT + PAYMENT (after Chevron/Monthly checks).
    if "operating acct" in desc and "chevron" in desc:
        return "REBATE COMBUSTIBLE"
    if "operating acct" in desc and "monthly" in desc:
        return "REBATE COMBUSTIBLE"
    if "operating acct" in desc and "payment" in desc:
        return "EFT RCV-"

    # Double Rule 5: CHECK/CHEQUE/CHEK + FLORIDA.
    if _desc_has_check_keyword(desc) and "florida" in desc:
        return "ADMINISTRATION FEE"

    # Single Rule: CHECK - FLORI.
    if "chek - flori" in desc:
        return "PROVEEDORES"

    # Double Rule 4: CHECK/CHEQUE/CHEK + vendor keywords.
    if _is_check_proveedores(desc):
        return "PROVEEDORES"

    if _desc_contains_any(desc, CHASE_GASTOS_BANCARIOS_KEYWORDS):
        return "GASTOS BANCARIOS"
    if _desc_contains_any(desc, CHASE_PROVEEDORES_KEYWORDS):
        return "PROVEEDORES"
    if _desc_contains_any(desc, CHASE_REBATE_KEYWORDS):
        return "REBATE"
    if _desc_contains_any(desc, CHASE_AGUA_KEYWORDS):
        return "AGUA"
    if _desc_contains_any(desc, CHASE_ALARMA_KEYWORDS):
        return "ALARMA"
    if _desc_contains_any(desc, CHASE_ENERGIA_KEYWORDS):
        return "ENERGIA ELECTRICA"
    if _desc_contains_any(desc, CHASE_TELEFONO_KEYWORDS):
        return "TELEFONO"
    if _desc_contains_any(desc, CHASE_ELISTAR_KEYWORDS):
        return "ELISTAR"
    if _desc_contains_any(desc, CHASE_SEGURO_KEYWORDS):
        return "SEGURO"
    if _desc_contains_any(desc, CHASE_SALE_TAX_KEYWORDS):
        return "SALE TAX"
    if _desc_contains_any(desc, CHASE_ALG_DISTR_KEYWORDS):
        return "REBATE"
    if _desc_contains_any(desc, CHASE_ADMIN_FEE_KEYWORDS):
        return "ADMINISTRATION FEE"

    for keyword, detail in ALFONSO_MASTER_SINGLE_RULES:
        if keyword in desc:
            return detail
    if _is_deposit_id_number(desc):
        return "DEPOSITO"

    dynamic_detail = match_dynamic_detalle(description) if match_dynamic_detalle else None
    if dynamic_detail:
        return dynamic_detail

    return None


def find_chase_column(df, name_hint, fallback_index):
    """Locate a column by header name or fixed Excel column index."""
    target = name_hint.strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == target:
            return col
    for col in df.columns:
        if target in str(col).strip().lower():
            return col
    if 0 <= fallback_index < len(df.columns):
        return df.columns[fallback_index]
    return None


def ensure_detalle_column(df):
    """Ensure Detalle exists as column H (8th column); create if missing."""
    detalle_col = find_chase_column(df, "Detalle", 7)
    if detalle_col is not None:
        return detalle_col

    if len(df.columns) >= 8:
        return df.columns[7]

    while len(df.columns) < 7:
        df[f"__col_{len(df.columns)}__"] = ""
    df["Detalle"] = ""
    return "Detalle"


CHASE_COL_POSTING_DATE = 2  # B
CHASE_COL_AMOUNT = 4  # D
CHASE_COL_BALANCE = 6  # F
CHASE_DATE_NUMBER_FORMAT = "dd/mm/yyyy"
CHASE_INTEGER_NUMBER_FORMAT = "0"
CHASE_DECIMAL_NUMBER_FORMAT = "0.00"
CHASE_DATA_START_ROW = 2


def apply_chase_amount_cell(cell, amount):
    """Apply integer or Spanish-decimal Excel format based on the amount value."""
    amount_float = float(amount)
    if amount_float.is_integer():
        cell.value = int(amount_float)
        cell.number_format = CHASE_INTEGER_NUMBER_FORMAT
    else:
        cell.value = amount_float
        cell.number_format = CHASE_DECIMAL_NUMBER_FORMAT


def read_chase_activity_file(file_path):
    """Read Chase bank activity from CSV or Excel."""
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        return pd.read_csv(file_path, dtype=str, keep_default_na=False)
    if extension in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(file_path, dtype=str, keep_default_na=False)
    raise ValueError(
        f"Tipo de archivo no soportado '{extension}'. Use CSV o Excel (.csv, .xlsx, .xlsm)."
    )


def parse_amount_signed(value):
    """Parse amount text into a signed native float."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    if text.startswith("-"):
        negative = True
        text = text[1:].strip()

    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.rsplit(",", 1)
        if len(parts) == 2 and len(parts[1]) <= 2:
            text = parts[0].replace(".", "") + "." + parts[1]
        else:
            text = text.replace(",", "")

    try:
        amount = float(text)
    except ValueError:
        return 0.0

    return -abs(amount) if negative else amount


def parse_posting_date_value(value):
    """Return datetime for Column B or None when parsing fails."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def sync_detalle_from_excel_workbook(df, file_path, detalle_col):
    """Load pre-existing Column H values from Excel before categorization."""
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        return

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    try:
        for row_offset, index in enumerate(df.index):
            excel_row = CHASE_DATA_START_ROW + row_offset
            cell_value = worksheet.cell(row=excel_row, column=8).value
            if not detalle_cell_is_empty(cell_value):
                df.at[index, detalle_col] = cell_value
    finally:
        workbook.close()


def apply_chase_detalle_categorization(df, description_col, detalle_col):
    """
    Apply keyword rules to Detalle (Column H) only for empty cells.

    Rows with any pre-existing Detalle value are left untouched.
    """
    updated_count = 0
    for index in df.index:
        current = df.at[index, detalle_col]
        if not detalle_cell_is_empty(current):
            continue

        description = df.at[index, description_col]
        category = categorize_chase_description(description)
        if category:
            df.at[index, detalle_col] = category
            updated_count += 1
    return updated_count


def write_chase_excel_file(df, file_path, column_map):
    """Write Chase data to Excel with dates, amounts, balance formulas, and formats."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    posting_idx = column_map["posting"]
    amount_idx = column_map["amount"]

    posting_col_name = df.columns[posting_idx]
    amount_col_name = df.columns[amount_idx]
    detalle_col_name = df.columns[column_map["detalle"]]

    protected_detalle = {}
    for row_offset, index in enumerate(df.index):
        excel_row = CHASE_DATA_START_ROW + row_offset
        existing_h = worksheet.cell(row=excel_row, column=8).value
        if not detalle_cell_is_empty(existing_h):
            protected_detalle[excel_row] = existing_h

    for row_offset, index in enumerate(df.index):
        excel_row = CHASE_DATA_START_ROW + row_offset

        for col_num, col_name in enumerate(df.columns, start=1):
            if col_num == 8 and excel_row in protected_detalle:
                continue
            worksheet.cell(row=excel_row, column=col_num, value=df.at[index, col_name])

        posting_dt = parse_posting_date_value(df.at[index, posting_col_name])
        cell_b = worksheet.cell(row=excel_row, column=CHASE_COL_POSTING_DATE)
        if posting_dt:
            cell_b.value = posting_dt
            cell_b.number_format = CHASE_DATE_NUMBER_FORMAT
        else:
            cell_b.value = df.at[index, posting_col_name]

        cell_d = worksheet.cell(row=excel_row, column=CHASE_COL_AMOUNT)
        apply_chase_amount_cell(
            cell_d, parse_amount_signed(df.at[index, amount_col_name])
        )

        cell_f = worksheet.cell(row=excel_row, column=CHASE_COL_BALANCE)
        if excel_row == CHASE_DATA_START_ROW:
            cell_f.value = f"=+F1+D{excel_row}"
        else:
            cell_f.value = f"=+F{excel_row - 1}+D{excel_row}"
        cell_f.number_format = CHASE_DECIMAL_NUMBER_FORMAT

        if excel_row not in protected_detalle:
            detalle_value = str(df.at[index, detalle_col_name]).strip()
            cell_h = worksheet.cell(row=excel_row, column=8)
            cell_h.value = df.at[index, detalle_col_name]
            if detalle_value == EFT_RCV_EXACT_LABEL:
                cell_h.font = EFT_RCV_RED_FONT

    workbook.save(file_path)


def write_chase_csv_file(df, file_path, column_map):
    """Write Chase CSV with formatted dates, floats, and balance formulas."""
    posting_idx = column_map["posting"]
    amount_idx = column_map["amount"]
    balance_idx = column_map["balance"]
    detalle_idx = column_map["detalle"]

    output = df.copy()
    for row_offset, index in enumerate(output.index):
        posting_dt = parse_posting_date_value(output.at[index, output.columns[posting_idx]])
        if posting_dt:
            output.at[index, output.columns[posting_idx]] = posting_dt.strftime("%d/%m/%Y")

        output.at[index, output.columns[amount_idx]] = parse_amount_signed(
            output.at[index, output.columns[amount_idx]]
        )

        excel_row = CHASE_DATA_START_ROW + row_offset
        if excel_row == CHASE_DATA_START_ROW:
            output.at[index, output.columns[balance_idx]] = f"=+F1+D{excel_row}"
        else:
            output.at[index, output.columns[balance_idx]] = (
                f"=+F{excel_row - 1}+D{excel_row}"
            )

    output.to_csv(file_path, index=False)


def process_chase_categorization(file_path):
    """
    Categorize Chase activity, format columns B/D/F, and save the workbook.

    Column B: DD/MM/YYYY dates
    Column D: signed floats with comma decimal display
    Column F: F1 static baseline; F2 =+F1+D2; row 3+ =+F{r-1}+D{r}
    Column H: Detalle keyword categorization
    """
    df = read_chase_activity_file(file_path)

    description_col = find_chase_column(df, "Description", 2)
    if description_col is None:
        raise ValueError(
            "No se encontró la columna Description (se esperaba en la Columna C)."
        )

    posting_col = find_chase_column(df, "Posting Date", 1)
    if posting_col is None:
        posting_col = find_chase_column(df, "Posting", 1)
    amount_col = find_chase_column(df, "Amount", 3)
    balance_col = find_chase_column(df, "Balance", 5)

    if posting_col is None or amount_col is None or balance_col is None:
        raise ValueError(
            "No se encontraron las columnas requeridas (B: Posting Date, D: Amount, F: Balance)."
        )

    detalle_col = ensure_detalle_column(df)
    if detalle_col not in df.columns:
        df[detalle_col] = ""

    sync_detalle_from_excel_workbook(df, file_path, detalle_col)

    updated_count = apply_chase_detalle_categorization(
        df, description_col, detalle_col
    )

    column_map = {
        "posting": list(df.columns).index(posting_col),
        "amount": list(df.columns).index(amount_col),
        "balance": list(df.columns).index(balance_col),
        "detalle": list(df.columns).index(detalle_col),
    }

    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        write_chase_csv_file(df, file_path, column_map)
    elif extension in {".xlsx", ".xlsm"}:
        write_chase_excel_file(df, file_path, column_map)
    elif extension == ".xls":
        raise ValueError(
            "El formato .xls antiguo no soporta fórmulas. Guarde como .xlsx e intente de nuevo."
        )
    else:
        raise ValueError(f"No se puede guardar un tipo de archivo no soportado '{extension}'.")

    return updated_count, len(df)


def _normalize_eft_rcv_number(value):
    """Extract canonical RCV-##### token from draft text or summary cell."""
    if not value:
        return None
    match = DRAFT_NO_PATTERN.search(str(value))
    if match:
        return match.group(1).upper()
    text = str(value).strip().upper()
    return text if text else None


def _normalize_eft_date_key(value):
    """Normalize any EFT date cell to a calendar date for duplicate checks."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    parsed = parse_eft_pdf_date_us(value)
    if parsed:
        return parsed.date()
    parsed = parse_date_to_datetime(value)
    if parsed:
        return parsed.date()
    return None


def _amounts_close(left, right, tolerance=0.01):
    try:
        return abs(float(left) - float(right)) <= tolerance
    except (TypeError, ValueError):
        return False


def _read_cta_summary_net_total(worksheet, row):
    """Read column J net total from cached value or SUM(Fx:Fy) formula text."""
    cell = worksheet.cell(row=row, column=10)
    value = cell.value
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        formula_match = SUM_FORMULA_RANGE_PATTERN.search(value.replace(" ", ""))
        if formula_match:
            start_row = int(formula_match.group(1))
            end_row = int(formula_match.group(2))
            total = 0.0
            for scan_row in range(start_row, end_row + 1):
                total += parse_amount(worksheet.cell(row=scan_row, column=6).value) or 0.0
            return total
        parsed = parse_amount(value)
        if parsed is not None:
            return parsed
    return None


def eft_already_loaded_in_workbook(excel_path, header_data, credit_coupons):
    """
    Return True when Cta Cte already contains the same EFT Nro., date, and net total.
    """
    draft_no = header_data.get("draft_no")
    eft_date_str = header_data.get("eft_date")
    if not draft_no or not eft_date_str or not credit_coupons:
        return False

    incoming_nro = _normalize_eft_rcv_number(draft_no)
    incoming_date = _normalize_eft_date_key(eft_date_str)
    incoming_net = sum(
        parse_amount(coupon.get("paid_amount")) or 0.0 for coupon in credit_coupons
    )
    if not incoming_nro or not incoming_date:
        return False

    for data_only in (True, False):
        workbook = load_workbook(excel_path, data_only=data_only)
        try:
            worksheet = get_worksheet(workbook)
            max_row = max(worksheet.max_row, LEDGER_START_ROW)
            for row in range(LEDGER_START_ROW, max_row + 1):
                summary_text = worksheet.cell(row=row, column=7).value
                if summary_text is None:
                    continue
                summary_upper = str(summary_text).upper()
                if "EFT" not in summary_upper and "RCV-" not in summary_upper:
                    continue

                existing_nro = _normalize_eft_rcv_number(summary_text)
                if existing_nro != incoming_nro:
                    continue

                existing_date = _normalize_eft_date_key(
                    worksheet.cell(row=row, column=COLUMN_N).value
                )
                if existing_date != incoming_date:
                    continue

                existing_net = _read_cta_summary_net_total(worksheet, row)
                if existing_net is not None and _amounts_close(
                    existing_net, incoming_net
                ):
                    return True
        finally:
            workbook.close()

    return False


def get_worksheet(workbook):
    """Return target worksheet, matching name with flexible spelling."""
    if SHEET_NAME in workbook.sheetnames:
        return workbook[SHEET_NAME]
    for name in workbook.sheetnames:
        if name.strip().lower() == SHEET_NAME.strip().lower():
            return workbook[name]
    raise ValueError(
        f'Hoja "{SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def update_excel_workbook(excel_path, header_data, paid_invoices, credit_coupons):
    """
    Insert coupon rows after the last historical entry and apply EFT block styling.

    Returns:
        tuple: (start_row, end_row)
    """
    if not credit_coupons:
        raise ValueError("No se extrajeron cupones de tarjeta de crédito del PDF.")

    workbook = load_workbook(excel_path, data_only=False)
    worksheet = get_worksheet(workbook)

    last_active_row = find_last_active_row_bottom_up(worksheet)
    target_row = last_active_row + 1
    num_rows = len(credit_coupons)

    worksheet.insert_rows(target_row, amount=num_rows)

    start_row = target_row
    end_row = target_row + num_rows - 1
    middle_row = calculate_middle_row(start_row, end_row)
    draft_no = header_data.get("draft_no") or "UNKNOWN"
    eft_date_str = header_data.get("eft_date")

    for offset, coupon in enumerate(credit_coupons):
        row = start_row + offset
        coupon_date = coupon.get("date")
        if coupon_date:
            worksheet.cell(
                row=row, column=1, value=format_coupon_date_us(coupon_date)
            )

        worksheet.cell(row=row, column=2, value=coupon["invoice"])
        worksheet.cell(row=row, column=3, value=coupon["coupon"])

        gross_val = float(coupon["gross_amount"])
        fees_val = float(coupon.get("fees_amount") or 0.0)
        paid_val = float(coupon.get("paid_amount") or 0.0)

        cell_d = worksheet.cell(row=row, column=4, value=gross_val)
        cell_e = worksheet.cell(row=row, column=5, value=fees_val)
        cell_f = worksheet.cell(row=row, column=6, value=paid_val)
        for cell in (cell_d, cell_e, cell_f):
            apply_currency_format(cell)

    worksheet.cell(row=middle_row, column=7, value=f"EFT  {draft_no}")
    cell_h = worksheet.cell(
        row=middle_row, column=8, value=f"=SUM(D{start_row}:D{end_row})"
    )
    cell_i = worksheet.cell(
        row=middle_row, column=9, value=f"=SUM(E{start_row}:E{end_row})"
    )
    cell_j = worksheet.cell(
        row=middle_row, column=10, value=f"=SUM(F{start_row}:F{end_row})"
    )
    for cell in (cell_h, cell_i, cell_j):
        apply_currency_format(cell)

    paid_invoice_count = len(paid_invoices) if paid_invoices else 1
    balance_formula = build_balance_formula(middle_row, paid_invoice_count)
    cell_m = worksheet.cell(row=middle_row, column=13, value=balance_formula)
    apply_currency_format(cell_m)

    clear_summary_columns(worksheet, start_row, end_row, middle_row)
    apply_eft_block_styling(worksheet, start_row, end_row)
    apply_summary_row_formatting(worksheet, middle_row)
    _write_paid_invoices_to_columns_kl(worksheet, middle_row, paid_invoices)
    apply_column_n_date(worksheet, middle_row, eft_date_str)

    workbook.save(excel_path)
    return start_row, end_row, middle_row


class EFTExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Financial Automation Suite")
        apply_root_style(self.root)
        self.root.minsize(*WINDOW_MINSIZE)
        self.root.geometry(WINDOW_GEOMETRY)

        self.pdf_path = tk.StringVar(value="")
        self.excel_path = tk.StringVar(value="")
        self.monthly_coupon_path = tk.StringVar(value="")
        self.gettel_source_excel_path = tk.StringVar(value="")
        self.gettel_destination_excel_path = tk.StringVar(value="")
        self.gettel_pagos_pdf_paths = tk.StringVar(value="")
        self.proveedores_ledger_path = tk.StringVar(value="")
        self.proveedores_invoice_pdf_paths = tk.StringVar(value="")
        self.proveedores_bank_path = tk.StringVar(value="")
        self.proveedores_pago_rule_keyword = tk.StringVar(value="")
        self.proveedores_pago_rule_sheet = tk.StringVar(value="")
        self.chase_path = tk.StringVar(value="")
        self.chase_rule_keyword = tk.StringVar(value="")
        self.chase_rule_detail = tk.StringVar(value="")
        self.cmv_path = tk.StringVar(value="")
        self.cmv_master_path = tk.StringVar(value="")
        self.sales_path = tk.StringVar(value="")
        self.sales_master_path = tk.StringVar(value="")
        self.reporte_master_path = tk.StringVar(value="")
        self.reporte_pdf_path = tk.StringVar(value="")
        self.store_info_master_path = tk.StringVar(value="")
        self.lottery_master_path = tk.StringVar(value="")
        self.lottery_department_pdf_paths = tk.StringVar(value="")
        self.lottery_sales_report_paths = tk.StringVar(value="")
        self.status_text = tk.StringVar(value="Listo — seleccione los archivos para comenzar.")
        self.gettel_status_text = tk.StringVar(
            value="Listo — seleccione Excel de origen y master de destino."
        )
        self.gettel_pagos_status_text = tk.StringVar(
            value="Listo — seleccione el master Cierre y el/los PDF de Pagos."
        )
        self.proveedores_status_text = tk.StringVar(
            value="Listo — seleccione el Excel Ledger y el/los PDF de facturas."
        )
        self.proveedores_pagos_status_text = tk.StringVar(
            value="Listo — seleccione el Excel Ledger y el extracto de Chase ya categorizado."
        )
        self.chase_status_text = tk.StringVar(value="Listo — seleccione un extracto de Chase.")
        self.cmv_status_text = tk.StringVar(value="Listo — seleccione los archivos de departamento y el maestro CMV.")
        self.sales_status_text = tk.StringVar(
            value="Listo — seleccione el Excel maestro CMV y el CSV de ventas del POS."
        )
        self.reporte_status_text = tk.StringVar(
            value="Listo — seleccione el master Bradenton Análisis C-Store y el PDF diario."
        )
        self.store_info_status_text = tk.StringVar(
            value="Listo — seleccione el Excel de Cierre (hoja Store Info) y el PDF diario."
        )
        self.lottery_sales_report_status_text = tk.StringVar(
            value="Listo — seleccione el Excel de Lottery y el/los Daily Sales Report."
        )
        self.lottery_department_status_text = tk.StringVar(
            value="Listo — seleccione el Excel de Lottery y el/los PDF diario(s)."
        )
        self._cmv_dnd_card = None
        self._chase_display_rules_cache = []
        self._build_ui()
        self._setup_drag_and_drop()
        self.root.after_idle(self._initialize_chase_rules_display)

    def _build_ui(self):
        shell = tk.Frame(self.root, bg=THEME.BG)
        shell.pack(fill=tk.BOTH, expand=True)

        create_header_banner(shell)
        body = create_scrollable_body(shell)

        style = ttk.Style(self.root)
        apply_notebook_style(
            style,
            [
                EFT_THEME,
                GETTEL_THEME,
                CHASE_THEME,
                REPORTE_DIARIO_THEME,
                CMV_THEME,
                SALES_THEME,
            ],
        )

        notebook = ttk.Notebook(body, style="Premium.TNotebook")
        notebook.pack(fill=tk.BOTH, expand=True)
        self.notebook = notebook
        self.root.bind_all("<Control-Next>", self._select_next_tab)
        self.root.bind_all("<Control-Prior>", self._select_previous_tab)

        # These three tabs' content can run taller than the window's minimum
        # height (Chase Bank and Cupones y EFT already did; Reporte Diario
        # joined them once Store Info added a second card), so they scroll.
        # The rest fit comfortably and stay as plain frames.
        eft_tab, eft_content = create_scrollable_tab_frame(notebook)
        gettel_tab = tk.Frame(notebook, bg=THEME.BG)
        proveedores_tab = tk.Frame(notebook, bg=THEME.BG)
        chase_tab, chase_content = create_scrollable_tab_frame(notebook)
        reporte_tab, reporte_content = create_scrollable_tab_frame(notebook)
        lottery_tab, lottery_content = create_scrollable_tab_frame(notebook)
        cmv_tab = tk.Frame(notebook, bg=THEME.BG)
        sales_tab = tk.Frame(notebook, bg=THEME.BG)

        # Kept alive on the instance — Tkinter drops PhotoImages with no
        # surviving Python reference, which would blank the tab icons.
        self._tab_icons = [
            make_tab_icon(EFT_THEME.accent),
            make_tab_icon(GETTEL_THEME.accent),
            make_tab_icon(PROVEEDORES_THEME.accent),
            make_tab_icon(CHASE_THEME.accent),
            make_tab_icon(REPORTE_DIARIO_THEME.accent),
            make_tab_icon(REPORTE_DIARIO_THEME.accent),
            make_tab_icon(CMV_THEME.accent),
            make_tab_icon(SALES_THEME.accent),
        ]
        tabs = [
            (eft_tab, "  Cupones y EFT  "),
            (gettel_tab, "  Gettel / Toyota  "),
            (proveedores_tab, "  Proveedores  "),
            (chase_tab, "  Chase Bank  "),
            (reporte_tab, "  Reporte Diario  "),
            (lottery_tab, "  Lottery  "),
            (cmv_tab, "  CMV Costo  "),
            (sales_tab, "  CMV Ventas  "),
        ]
        for (tab, label), icon in zip(tabs, self._tab_icons):
            notebook.add(tab, text=label, image=icon, compound=tk.LEFT)

        self._build_eft_tab(eft_content)
        self._build_gettel_tab(gettel_tab)
        self._build_proveedores_tab(proveedores_tab)
        self._build_chase_tab(chase_content)
        self._build_reporte_diario_tab(reporte_content)
        self._build_lottery_tab(lottery_content)
        self._build_cmv_tab(cmv_tab)
        self._build_sales_tab(sales_tab)

    def _select_next_tab(self, event=None):
        self._cycle_tab(1)

    def _select_previous_tab(self, event=None):
        self._cycle_tab(-1)

    def _cycle_tab(self, direction):
        tabs = self.notebook.tabs()
        if not tabs:
            return
        current = self.notebook.index(self.notebook.select())
        self.notebook.select(tabs[(current + direction) % len(tabs)])

    def _build_eft_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Gestión de Cupones y EFT",
            "Dos flujos independientes sobre el mismo Excel Ledger — (1) PDF de EFT bancario → actualiza Cta Cte, (2) Reporte mensual de J.H. → agrega y cruza Cupones.",
            EFT_THEME,
        )

        ledger_card = create_card(left, section_theme=EFT_THEME)
        create_panel_label(ledger_card, "Excel Ledger", EFT_THEME)
        self.excel_entry = create_file_row(
            ledger_card,
            "Excel Ledger",
            self.excel_path,
            self.select_excel,
            section_theme=EFT_THEME,
            label_width=11,
        )

        cta_card = create_card(left, section_theme=EFT_THEME)
        create_panel_label(cta_card, "Paso 1 — Pipeline Cta Cte (PDF EFT)", EFT_THEME)
        self.pdf_entry = create_file_row(
            cta_card,
            "EFT PDF",
            self.pdf_path,
            self.select_pdf,
            section_theme=EFT_THEME,
            label_width=11,
        )

        cta_actions = tk.Frame(cta_card, bg=EFT_THEME.card_tint)
        cta_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            cta_actions,
            "Actualizar Cta Cte (Procesar PDF EFT)",
            self.process_and_update,
            section_theme=EFT_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            cta_actions,
            "Limpiar PDF",
            lambda: self._clear_field(self.pdf_path, "pdf_entry"),
            section_theme=EFT_THEME,
        ).pack(side=tk.LEFT)

        coupon_card = create_card(left, section_theme=EFT_THEME)
        create_panel_label(
            coupon_card,
            "Paso 2 — Pipeline Cupones (Excel J.H. Mensual)",
            EFT_THEME,
        )
        self.monthly_coupon_entry = create_file_row(
            coupon_card,
            "Reporte Mensual",
            self.monthly_coupon_path,
            self.select_monthly_coupon_report,
            section_theme=EFT_THEME,
            label_width=11,
            browse_label="Examinar…",
        )

        actions = tk.Frame(coupon_card, bg=EFT_THEME.card_tint)
        actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            actions,
            "Cargar Reporte Mensual / Actualizar Cupones",
            self.process_monthly_coupon_append,
            section_theme=EFT_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            actions,
            "Limpiar Reporte",
            lambda: self._clear_field(self.monthly_coupon_path, "monthly_coupon_entry"),
            section_theme=EFT_THEME,
        ).pack(side=tk.LEFT)
        tk.Label(
            actions,
            text='Deja "Reporte Mensual" vacío para solo actualizar pendientes.',
            font=(FONT, 8),
            fg=THEME.TEXT_MUTED_ON_DARK,
            bg=EFT_THEME.card_tint,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(2, 0))

        self.status_label, self.status_dot = create_status_bar(
            left, self.status_text, section_theme=EFT_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "El PDF de EFT actualiza la hoja Cta Cte J.H.Williams",
                "El reporte mensual descarta el N° de bache y agrega A:E limpio",
                "Las columnas F:I se enlazan dinámicamente con Cta Cte",
                "Sin Reporte Mensual: solo actualiza cupones pendientes/$0",
            ],
            section_theme=EFT_THEME,
        )

    def _build_gettel_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Módulo Gettel / Toyota",
            "Resume el Excel diario de Gettel/Toyota (hoja de Rick) por día y lo fusiona en la hoja Gettel-Toyota del master.",
            GETTEL_THEME,
        )

        master_card = create_card(left, section_theme=GETTEL_THEME)
        create_panel_label(master_card, "Excel Maestro (Master Cierre)", GETTEL_THEME)
        self.gettel_destination_excel_entry = create_file_row(
            master_card,
            "Excel Maestro",
            self.gettel_destination_excel_path,
            self.select_gettel_destination_excel,
            section_theme=GETTEL_THEME,
            label_width=13,
            browse_label="Examinar…",
        )

        cupones_card = create_card(left, section_theme=GETTEL_THEME)
        create_panel_label(cupones_card, "Cupones Diarios → hoja Gettel-Toyota", GETTEL_THEME)
        self.gettel_source_excel_entry = create_file_row(
            cupones_card,
            "Excel Cupones",
            self.gettel_source_excel_path,
            self.select_gettel_source_excel,
            section_theme=GETTEL_THEME,
            label_width=13,
            browse_label="Examinar…",
        )
        gettel_actions = tk.Frame(cupones_card, bg=GETTEL_THEME.card_tint)
        gettel_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            gettel_actions,
            "Procesar Reporte Gettel/Toyota",
            self.process_gettel_toyota_report,
            section_theme=GETTEL_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            gettel_actions,
            "Limpiar",
            lambda: self._clear_field(self.gettel_source_excel_path, "gettel_source_excel_entry"),
            section_theme=GETTEL_THEME,
        ).pack(side=tk.LEFT)

        self.gettel_status_label, self.gettel_status_dot = create_status_bar(
            left, self.gettel_status_text, section_theme=GETTEL_THEME
        )

        pagos_card = create_card(left, section_theme=GETTEL_THEME)
        create_panel_label(pagos_card, "Pagos (Cupones) → hoja PAGO Cupones", GETTEL_THEME)
        self.gettel_pagos_entry = create_file_row(
            pagos_card,
            "PDF(s) de Pagos",
            self.gettel_pagos_pdf_paths,
            self.select_gettel_pagos_files,
            section_theme=GETTEL_THEME,
            label_width=32,
            browse_label="Examinar…",
        )
        pagos_actions = tk.Frame(pagos_card, bg=GETTEL_THEME.card_tint)
        pagos_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            pagos_actions,
            "Procesar Pagos",
            self.process_gettel_pagos_file,
            section_theme=GETTEL_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            pagos_actions,
            "Limpiar",
            lambda: self._clear_field(self.gettel_pagos_pdf_paths, "gettel_pagos_entry"),
            section_theme=GETTEL_THEME,
        ).pack(side=tk.LEFT)

        self.gettel_pagos_status_label, self.gettel_pagos_status_dot = create_status_bar(
            left, self.gettel_pagos_status_text, section_theme=GETTEL_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Origen: Excel (hojas GETTEL/GETTLE y TOYOTA) o PDF/foto escaneada",
                "PDF/foto: lee con OCR, un proveedor por archivo (según el nombre)",
                "Suma montos y galones por día calendario (GETTEL y TOYOTA por separado)",
                "Destino: hoja «Gettel-Toyota MM.YYYY» — columnas E-H",
                "Ambos proveedores se emparejan por la Columna A (misma fila que Local Account)",
                "PDF/foto: si el reporte trae un subtotal impreso, se compara automáticamente",
                "Pagos: usa el mismo master Cierre — cada PDF \"PagosN (Empresa).pdf\" es un recibo por página",
                "Pagos: fecha y N° de pago van combinados (A/B); Trans # y Total van por fila (C/D)",
                "Pagos: agranda o achica el bloque de filas para que coincida con la cantidad de recibos",
                "Abre copia temp del master — guarde manualmente con Guardar como",
            ],
            section_theme=GETTEL_THEME,
        )

    def _build_proveedores_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Proveedores",
            "Lee facturas de compra (PDF) y agrega la fila correspondiente en la hoja "
            "de cada proveedor del libro de cuentas corrientes.",
            PROVEEDORES_THEME,
        )

        ledger_card = create_card(left, section_theme=PROVEEDORES_THEME)
        create_panel_label(ledger_card, "Excel Ledger (Cta Cte Proveedores)", PROVEEDORES_THEME)
        self.proveedores_ledger_entry = create_file_row(
            ledger_card,
            "Excel Ledger",
            self.proveedores_ledger_path,
            self.select_proveedores_ledger,
            section_theme=PROVEEDORES_THEME,
            label_width=13,
        )

        invoices_card = create_card(left, section_theme=PROVEEDORES_THEME)
        create_panel_label(invoices_card, "Cargar Facturas → fila \"invoice\" por proveedor", PROVEEDORES_THEME)
        self.proveedores_invoice_entry = create_file_row(
            invoices_card,
            "PDF(s)",
            self.proveedores_invoice_pdf_paths,
            self.select_proveedores_invoice_files,
            section_theme=PROVEEDORES_THEME,
            label_width=13,
            browse_label="Examinar…",
        )
        invoices_actions = tk.Frame(invoices_card, bg=PROVEEDORES_THEME.card_tint)
        invoices_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            invoices_actions,
            "Procesar Facturas",
            self.process_proveedores_invoices,
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            invoices_actions,
            "Limpiar",
            lambda: self._clear_field(self.proveedores_invoice_pdf_paths, "proveedores_invoice_entry"),
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT)

        self.proveedores_status_label, self.proveedores_status_dot = create_status_bar(
            left, self.proveedores_status_text, section_theme=PROVEEDORES_THEME
        )

        payments_card = create_card(left, section_theme=PROVEEDORES_THEME)
        create_panel_label(payments_card, "Cargar Pagos (Chase) → fila \"OP\" por proveedor", PROVEEDORES_THEME)
        self.proveedores_bank_entry = create_file_row(
            payments_card,
            "Extracto Chase",
            self.proveedores_bank_path,
            self.select_proveedores_bank_file,
            section_theme=PROVEEDORES_THEME,
            label_width=13,
            browse_label="Examinar…",
        )
        payments_actions = tk.Frame(payments_card, bg=PROVEEDORES_THEME.card_tint)
        payments_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            payments_actions,
            "Procesar Pagos",
            self.process_proveedores_payments,
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            payments_actions,
            "Limpiar",
            lambda: self._clear_field(self.proveedores_bank_path, "proveedores_bank_entry"),
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT)

        self.proveedores_pagos_status_label, self.proveedores_pagos_status_dot = create_status_bar(
            left, self.proveedores_pagos_status_text, section_theme=PROVEEDORES_THEME
        )

        self._build_proveedores_pago_rules_manager(right)

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Se puede cargar más de un PDF a la vez, incluso de distintos proveedores",
                "El proveedor se detecta automáticamente por el contenido del PDF",
                "Se van sumando proveedores de a poco — si no lo reconoce, el aviso dice cuáles ya están",
                "Se ordenan por fecha y se agregan encadenadas a la última fila real de esa hoja",
                "Si cambia el mes, respeta la fila separadora y alterna el color verde/amarillo",
                "Una factura ya cargada (mismo N°) se omite automáticamente",
                "El BALANCE se recalcula solo — no hace falta tocarlo",
                "Abre copia temp del Excel — guarde manualmente con Guardar como",
                "Pagos: el extracto de Chase debe venir ya categorizado (columna Detalle) "
                "desde la pestaña Chase Bank",
                "Un pago se ubica cronológicamente en su hoja, aunque caiga antes de una "
                "factura ya cargada",
                "Un pago sin proveedor identificado se omite y se avisa al final — agregue "
                "la regla que falta y reprocese",
            ],
            section_theme=PROVEEDORES_THEME,
        )

    def _build_proveedores_pago_rules_manager(self, parent):
        """Keyword-to-sheet-name rule manager for routing Chase payments (persistent JSON + grid UI)."""
        rules_card = create_card(
            parent,
            section_theme=PROVEEDORES_THEME,
            padx=6,
            pady=6,
        )
        create_panel_label(rules_card, "Reglas de Proveedores (Pagos)", PROVEEDORES_THEME)

        create_compact_entry(
            rules_card,
            self.proveedores_pago_rule_keyword,
            section_theme=PROVEEDORES_THEME,
            label="Palabra clave a detectar",
        )
        create_compact_entry(
            rules_card,
            self.proveedores_pago_rule_sheet,
            section_theme=PROVEEDORES_THEME,
            label="Hoja de proveedor",
        )

        btn_row = tk.Frame(rules_card, bg=PROVEEDORES_THEME.card_tint)
        btn_row.pack(fill=tk.X, pady=(4, 4))
        create_primary_button(
            btn_row,
            "Crear Regla",
            self._add_proveedores_pago_rule,
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            btn_row,
            "Eliminar Regla",
            self._delete_proveedores_pago_rule,
            section_theme=PROVEEDORES_THEME,
        ).pack(side=tk.LEFT)

        tk.Label(
            rules_card,
            text="Reglas Cargadas",
            font=(FONT, 8, "bold"),
            fg=THEME.TEXT_SOFT,
            bg=PROVEEDORES_THEME.card_tint,
            anchor=tk.W,
        ).pack(anchor=tk.W, pady=(2, 2))

        list_host = tk.Frame(
            rules_card,
            bg=THEME.SURFACE_ALT,
            highlightbackground=PROVEEDORES_THEME.accent_soft,
            highlightthickness=1,
        )
        list_host.pack(fill=tk.BOTH, expand=True, pady=(0, 4))
        list_host.grid_columnconfigure(0, weight=1)
        list_host.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.configure(
            "ProveedoresPagoRules.Treeview",
            font=(FONT, 9),
            rowheight=22,
            background=THEME.SURFACE_ALT,
            fieldbackground=THEME.SURFACE_ALT,
            foreground=THEME.TEXT,
        )
        style.configure(
            "ProveedoresPagoRules.Treeview.Heading",
            font=(FONT, 9, "bold"),
            background=PROVEEDORES_THEME.card_tint,
            foreground=THEME.TEXT,
        )
        style.map(
            "ProveedoresPagoRules.Treeview",
            background=[("selected", PROVEEDORES_THEME.accent)],
            foreground=[("selected", "#FFFFFF")],
        )

        columns = ("keyword", "sheet_name")
        self.proveedores_pago_rules_tree = ttk.Treeview(
            list_host,
            columns=columns,
            show="headings",
            style="ProveedoresPagoRules.Treeview",
            height=8,
            selectmode="browse",
        )
        self.proveedores_pago_rules_tree.heading("keyword", text="Palabra Clave")
        self.proveedores_pago_rules_tree.heading("sheet_name", text="Hoja de Proveedor")
        self.proveedores_pago_rules_tree.column("keyword", width=180, anchor=tk.W, stretch=True)
        self.proveedores_pago_rules_tree.column("sheet_name", width=140, anchor=tk.W, stretch=True)

        scrollbar = ttk.Scrollbar(
            list_host, orient=tk.VERTICAL, command=self.proveedores_pago_rules_tree.yview
        )
        self.proveedores_pago_rules_tree.configure(yscrollcommand=scrollbar.set)
        self.proveedores_pago_rules_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        self._refresh_proveedores_pago_rules_listbox()

    def _refresh_proveedores_pago_rules_listbox(self):
        if not hasattr(self, "proveedores_pago_rules_tree"):
            return
        for item in self.proveedores_pago_rules_tree.get_children():
            self.proveedores_pago_rules_tree.delete(item)
        if load_proveedores_pago_rules is None:
            self._proveedores_pago_rules_cache = []
            return
        self._proveedores_pago_rules_cache = load_proveedores_pago_rules()
        for rule in self._proveedores_pago_rules_cache:
            self.proveedores_pago_rules_tree.insert(
                "", tk.END, values=(rule["keyword"], rule["sheet_name"])
            )

    def _add_proveedores_pago_rule(self):
        if add_proveedores_pago_rule is None:
            self._set_proveedores_pagos_status(
                f"Reglas de pagos no disponibles: {_PROVEEDORES_PAGO_RULES_IMPORT_ERROR}",
                is_error=True,
            )
            return
        keyword = self.proveedores_pago_rule_keyword.get().strip()
        sheet_name = self.proveedores_pago_rule_sheet.get().strip()
        try:
            add_proveedores_pago_rule(keyword, sheet_name)
            self.proveedores_pago_rule_keyword.set("")
            self.proveedores_pago_rule_sheet.set("")
            self._refresh_proveedores_pago_rules_listbox()
            self._set_proveedores_pagos_status(f"Regla creada: \"{keyword}\" → \"{sheet_name}\".")
        except ValueError as exc:
            messagebox.showwarning("Regla incompleta", str(exc))
            self._set_proveedores_pagos_status(str(exc), is_error=True)

    def _delete_proveedores_pago_rule(self):
        if delete_proveedores_pago_rule_by_index is None:
            self._set_proveedores_pagos_status(
                f"Reglas de pagos no disponibles: {_PROVEEDORES_PAGO_RULES_IMPORT_ERROR}",
                is_error=True,
            )
            return
        selection = self.proveedores_pago_rules_tree.selection()
        if not selection:
            messagebox.showwarning(
                "Sin selección",
                "Seleccione una regla de la lista antes de eliminar.",
            )
            return
        item_id = selection[0]
        index = self.proveedores_pago_rules_tree.index(item_id)
        if index >= len(self._proveedores_pago_rules_cache):
            messagebox.showwarning(
                "Selección inválida",
                "La regla seleccionada ya no existe. La lista se actualizará.",
            )
            self._refresh_proveedores_pago_rules_listbox()
            return
        rule = self._proveedores_pago_rules_cache[index]
        try:
            delete_proveedores_pago_rule_by_index(index)
            self._refresh_proveedores_pago_rules_listbox()
            self._set_proveedores_pagos_status(
                f"Regla eliminada: \"{rule['keyword']}\" → \"{rule['sheet_name']}\"."
            )
        except ValueError as exc:
            messagebox.showwarning("No se pudo eliminar", str(exc))
            self._set_proveedores_pagos_status(str(exc), is_error=True)

    def _build_chase_rules_manager(self, parent):
        """Keyword-to-Detalle rule manager (persistent JSON + grid UI)."""
        rules_card = create_card(
            parent,
            section_theme=CHASE_THEME,
            padx=6,
            pady=6,
            fill=tk.BOTH,
            expand=True,
        )
        create_panel_label(rules_card, "Motor de Reglas Chase", CHASE_THEME)

        create_compact_entry(
            rules_card,
            self.chase_rule_keyword,
            section_theme=CHASE_THEME,
            label="Palabra clave a detectar",
        )
        create_compact_entry(
            rules_card,
            self.chase_rule_detail,
            section_theme=CHASE_THEME,
            label="Detalle a asignar",
        )

        btn_row = tk.Frame(rules_card, bg=CHASE_THEME.card_tint)
        btn_row.pack(fill=tk.X, pady=(4, 4))
        create_primary_button(
            btn_row,
            "Crear Regla",
            self._add_chase_keyword_rule,
            section_theme=CHASE_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            btn_row,
            "Eliminar Regla",
            self._delete_chase_keyword_rule,
            section_theme=CHASE_THEME,
        ).pack(side=tk.LEFT)

        tk.Label(
            rules_card,
            text="Reglas Cargadas",
            font=(FONT, 8, "bold"),
            fg=THEME.TEXT_SOFT,
            bg=CHASE_THEME.card_tint,
            anchor=tk.W,
        ).pack(anchor=tk.W, pady=(2, 2))

        list_host = tk.Frame(
            rules_card,
            bg=THEME.SURFACE_ALT,
            highlightbackground=CHASE_THEME.accent_soft,
            highlightthickness=1,
        )
        list_host.pack(fill=tk.BOTH, expand=True, pady=(0, 4))
        list_host.grid_columnconfigure(0, weight=1)
        list_host.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.configure(
            "ChaseRules.Treeview",
            font=(FONT, 9),
            rowheight=22,
            background=THEME.SURFACE_ALT,
            fieldbackground=THEME.SURFACE_ALT,
            foreground=THEME.TEXT,
        )
        style.configure(
            "ChaseRules.Treeview.Heading",
            font=(FONT, 9, "bold"),
            background=CHASE_THEME.card_tint,
            foreground=THEME.TEXT,
        )
        style.map(
            "ChaseRules.Treeview",
            background=[("selected", CHASE_THEME.accent)],
            foreground=[("selected", "#FFFFFF")],
        )

        columns = ("keyword", "detail", "source")
        self.chase_rules_tree = ttk.Treeview(
            list_host,
            columns=columns,
            show="headings",
            style="ChaseRules.Treeview",
            height=10,
            selectmode="browse",
        )
        self.chase_rules_tree.heading("keyword", text="Palabra Clave")
        self.chase_rules_tree.heading("detail", text="Detalle")
        self.chase_rules_tree.heading("source", text="Tipo")
        self.chase_rules_tree.column("keyword", width=180, anchor=tk.W, stretch=True)
        self.chase_rules_tree.column("detail", width=140, anchor=tk.W, stretch=True)
        self.chase_rules_tree.column("source", width=72, anchor=tk.CENTER, stretch=False)

        scrollbar = ttk.Scrollbar(list_host, orient=tk.VERTICAL, command=self.chase_rules_tree.yview)
        self.chase_rules_tree.configure(yscrollcommand=scrollbar.set)
        self.chase_rules_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        self._reload_chase_rules_from_storage()
        self._refresh_chase_rules_listbox()

    def _builtin_chase_display_rules(self):
        """Hardcoded Alfonso master rules mirrored from categorization logic."""
        rules = []

        def add_keywords(keywords, detail):
            for keyword in keywords:
                rules.append(
                    {"keyword": keyword, "detail": detail, "source": "Maestra"}
                )

        add_keywords(CHASE_PROVEEDORES_KEYWORDS, "PROVEEDORES")
        add_keywords(CHASE_GASTOS_BANCARIOS_KEYWORDS, "GASTOS BANCARIOS")
        add_keywords(CHASE_REBATE_KEYWORDS, "REBATE")
        add_keywords(CHASE_AGUA_KEYWORDS, "AGUA")
        add_keywords(CHASE_ALARMA_KEYWORDS, "ALARMA")
        add_keywords(CHASE_ENERGIA_KEYWORDS, "ENERGIA ELECTRICA")
        add_keywords(CHASE_TELEFONO_KEYWORDS, "TELEFONO")
        add_keywords(CHASE_ELISTAR_KEYWORDS, "ELISTAR")
        add_keywords(CHASE_SEGURO_KEYWORDS, "SEGURO")
        add_keywords(CHASE_SALE_TAX_KEYWORDS, "SALE TAX")
        add_keywords(CHASE_ALG_DISTR_KEYWORDS, "REBATE")
        add_keywords(CHASE_ADMIN_FEE_KEYWORDS, "ADMINISTRATION FEE")

        for keyword, detail in ALFONSO_MASTER_SINGLE_RULES:
            rules.append({"keyword": keyword, "detail": detail, "source": "Maestra"})

        for keyword, detail in ALFONSO_COMPOUND_DISPLAY_RULES:
            rules.append({"keyword": keyword, "detail": detail, "source": "Maestra"})

        return rules

    def _all_chase_display_rules(self):
        """Merge built-in Alfonso rules and persisted JSON rules for the grid."""
        merged = []
        seen = set()

        for rule in self._builtin_chase_display_rules():
            key = (rule["keyword"].lower(), rule["detail"])
            if key not in seen:
                seen.add(key)
                merged.append(rule)

        if load_dynamic_rules is not None:
            for idx, rule in enumerate(load_dynamic_rules()):
                entry = {
                    "keyword": rule["keyword"],
                    "detail": rule["detail"],
                    "source": "Personalizada",
                    "dynamic_index": idx,
                }
                key = (entry["keyword"].lower(), entry["detail"])
                if key not in seen:
                    seen.add(key)
                    merged.append(entry)

        return merged

    def _reload_chase_rules_from_storage(self):
        """Read all persisted rules from chase_rules.json before UI updates."""
        if load_dynamic_rules is None:
            self._chase_saved_rules_cache = []
            return
        self._chase_saved_rules_cache = load_dynamic_rules()

    def _initialize_chase_rules_display(self):
        """Populate Chase rules grid immediately after widgets are realized."""
        self._reload_chase_rules_from_storage()
        self._refresh_chase_rules_listbox()

    def _refresh_chase_rules_listbox(self):
        if not hasattr(self, "chase_rules_tree"):
            return
        self._reload_chase_rules_from_storage()
        for item in self.chase_rules_tree.get_children():
            self.chase_rules_tree.delete(item)
        self._chase_display_rules_cache = self._all_chase_display_rules()
        for rule in self._chase_display_rules_cache:
            self.chase_rules_tree.insert(
                "",
                tk.END,
                values=(rule["keyword"], rule["detail"], rule.get("source", "")),
            )

    def _add_chase_keyword_rule(self):
        if add_dynamic_rule is None:
            self._set_chase_status(
                f"Chase rules module unavailable: {_CHASE_RULES_IMPORT_ERROR}",
                is_error=True,
            )
            return
        keyword = self.chase_rule_keyword.get().strip()
        detail = self.chase_rule_detail.get().strip()
        try:
            add_dynamic_rule(keyword, detail)
            self.chase_rule_keyword.set("")
            self.chase_rule_detail.set("")
            self._refresh_chase_rules_listbox()
            self._set_chase_status(f"Regla creada: \"{keyword}\" → \"{detail}\".")
        except ValueError as exc:
            messagebox.showwarning("Regla incompleta", str(exc))
            self._set_chase_status(str(exc), is_error=True)

    def _delete_chase_keyword_rule(self):
        if delete_dynamic_rule_by_index is None or load_dynamic_rules is None:
            self._set_chase_status(
                f"Chase rules module unavailable: {_CHASE_RULES_IMPORT_ERROR}",
                is_error=True,
            )
            return
        selection = self.chase_rules_tree.selection()
        if not selection:
            messagebox.showwarning(
                "Sin selección",
                "Seleccione una regla de la lista antes de eliminar.",
            )
            return
        item_id = selection[0]
        index = self.chase_rules_tree.index(item_id)
        if index >= len(self._chase_display_rules_cache):
            messagebox.showwarning(
                "Selección inválida",
                "La regla seleccionada ya no existe. La lista se actualizará.",
            )
            self._refresh_chase_rules_listbox()
            return
        rule = self._chase_display_rules_cache[index]
        if rule.get("source") == "Maestra":
            messagebox.showwarning(
                "Regla protegida",
                "Las reglas maestras de Alfonso están bloqueadas y no se pueden eliminar.",
            )
            return
        dynamic_index = rule.get("dynamic_index")
        if dynamic_index is None:
            dynamic_index = next(
                (
                    idx
                    for idx, saved in enumerate(self._chase_saved_rules_cache)
                    if saved["keyword"] == rule["keyword"]
                    and saved["detail"] == rule["detail"]
                ),
                None,
            )
        if dynamic_index is None:
            messagebox.showwarning(
                "No se pudo eliminar",
                "La regla personalizada no se encontró en el archivo.",
            )
            self._refresh_chase_rules_listbox()
            return
        try:
            delete_dynamic_rule_by_index(dynamic_index)
            self._refresh_chase_rules_listbox()
            self._set_chase_status(
                f"Regla eliminada: \"{rule['keyword']}\" → \"{rule['detail']}\"."
            )
        except ValueError as exc:
            messagebox.showwarning("No se pudo eliminar", str(exc))
            self._set_chase_status(str(exc), is_error=True)

    def _build_chase_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Chase Bank",
            "Clasifica los movimientos bancarios comparando palabras clave de la Descripción contra el Detalle, sin tocar las fórmulas existentes.",
            CHASE_THEME,
        )

        card = create_card(left, section_theme=CHASE_THEME)
        self.chase_entry = create_file_row(
            card,
            "Archivo Chase",
            self.chase_path,
            self.select_chase_file,
            section_theme=CHASE_THEME,
            label_width=11,
        )

        actions = tk.Frame(left, bg=THEME.BG)
        actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            actions,
            "Procesar y Categorizar Chase",
            self.process_chase_data,
            section_theme=CHASE_THEME,
        ).pack(anchor=tk.W)

        self.chase_status_label, self.chase_status_dot = create_status_bar(
            left, self.chase_status_text, section_theme=CHASE_THEME
        )

        self._build_chase_rules_manager(right)

    def _build_reporte_diario_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Reporte Diario",
            "Un mismo PDF de cierre diario alimenta varios destinos — Ventas por "
            "Departamento y Store Info.",
            REPORTE_DIARIO_THEME,
        )

        pdf_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(pdf_card, "PDF Diario", REPORTE_DIARIO_THEME)
        self.reporte_pdf_entry = create_file_row(
            pdf_card,
            "PDF(s)",
            self.reporte_pdf_path,
            self.select_reporte_pdf_file,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=11,
            browse_label="Examinar…",
        )
        pdf_actions = tk.Frame(pdf_card, bg=REPORTE_DIARIO_THEME.card_tint)
        pdf_actions.pack(fill=tk.X, pady=(4, 0))
        create_secondary_button(
            pdf_actions,
            "Limpiar PDF",
            self.clear_reporte_pdf,
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(anchor=tk.W)

        ventas_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(ventas_card, "Ventas por Departamento → CARGA AQUI", REPORTE_DIARIO_THEME)
        self.reporte_master_entry = create_file_row(
            ventas_card,
            "Excel Ventas",
            self.reporte_master_path,
            self.select_reporte_master_file,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=11,
        )
        ventas_actions = tk.Frame(ventas_card, bg=REPORTE_DIARIO_THEME.card_tint)
        ventas_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            ventas_actions,
            "Procesar Ventas",
            self.process_reporte_diario_file,
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(anchor=tk.W)

        self.reporte_status_label, self.reporte_status_dot = create_status_bar(
            left, self.reporte_status_text, section_theme=REPORTE_DIARIO_THEME
        )

        store_info_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(store_info_card, "Store Info (Fechas, Fuel, Cash/TC)", REPORTE_DIARIO_THEME)
        self.store_info_master_entry = create_file_row(
            store_info_card,
            "Excel Store Info",
            self.store_info_master_path,
            self.select_store_info_master_file,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=11,
        )
        store_info_actions = tk.Frame(store_info_card, bg=REPORTE_DIARIO_THEME.card_tint)
        store_info_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            store_info_actions,
            "Procesar Store Info",
            self.process_store_info_file,
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(anchor=tk.W)

        self.store_info_status_label, self.store_info_status_dot = create_status_bar(
            left, self.store_info_status_text, section_theme=REPORTE_DIARIO_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Un solo PDF alimenta los dos destinos de abajo",
                "El PDF queda cargado entre procesos — se limpia con \"Limpiar PDF\"",
                "Cada Excel se limpia solo apenas termina de procesarse",
                "Podés elegir varios PDF a la vez (Ctrl/Shift + clic en el diálogo)",
                "Cada PDF se ubica por su propia fecha — no hace falta que sean días seguidos",
                "Ventas: detecta \"Department Sales Report\" automáticamente, sin indicar página",
                "Si el PDF es una foto/escaneo, lee los datos por OCR (corrige giros de página)",
                "Fila 3 de CARGA AQUI: detecta los encabezados desde la columna C",
                "Las columnas con fórmula (GIFT CARD / VARIOS/BOLSA) nunca se tocan",
                "Store Info: lee PERIOD FROM/TO, Fuel, Cash/TC y Network Revenue "
                "(páginas 3-4) y agrega una fila nueva en la hoja Store Info",
                "Abre una copia temporal automáticamente al terminar",
            ],
            section_theme=REPORTE_DIARIO_THEME,
        )

    def _build_lottery_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "Lottery",
            "Dos flujos independientes sobre el mismo Excel de Lottery — cada uno con su "
            "propio PDF y su propio botón, no hace falta cargar los dos juntos.",
            REPORTE_DIARIO_THEME,
        )

        master_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(master_card, "Excel Lottery", REPORTE_DIARIO_THEME)
        self.lottery_master_entry = create_file_row(
            master_card,
            "Excel Lottery",
            self.lottery_master_path,
            self.select_lottery_master_file,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=13,
        )

        sales_report_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(
            sales_report_card, "Daily Sales Report → F/G/H/I/K, P/Q/R/S", REPORTE_DIARIO_THEME
        )
        self.lottery_sales_report_entry = create_file_row(
            sales_report_card,
            "PDF(s)",
            self.lottery_sales_report_paths,
            self.select_lottery_sales_report_files,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=13,
            browse_label="Examinar…",
        )
        sales_report_actions = tk.Frame(sales_report_card, bg=REPORTE_DIARIO_THEME.card_tint)
        sales_report_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            sales_report_actions,
            "Procesar Daily Sales Report",
            self.process_lottery_sales_report_file,
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            sales_report_actions,
            "Limpiar",
            lambda: self._clear_field(self.lottery_sales_report_paths, "lottery_sales_report_entry"),
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(side=tk.LEFT)

        self.lottery_sales_report_status_label, self.lottery_sales_report_status_dot = create_status_bar(
            left, self.lottery_sales_report_status_text, section_theme=REPORTE_DIARIO_THEME
        )

        department_card = create_card(left, section_theme=REPORTE_DIARIO_THEME, padx=4, pady=4)
        create_panel_label(department_card, "PDF Diario → D/E/N/O", REPORTE_DIARIO_THEME)
        self.lottery_department_pdf_entry = create_file_row(
            department_card,
            "PDF(s)",
            self.lottery_department_pdf_paths,
            self.select_lottery_department_pdf_files,
            section_theme=REPORTE_DIARIO_THEME,
            label_width=13,
            browse_label="Examinar…",
        )
        department_actions = tk.Frame(department_card, bg=REPORTE_DIARIO_THEME.card_tint)
        department_actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            department_actions,
            "Procesar PDF Diario",
            self.process_lottery_department_file,
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            department_actions,
            "Limpiar",
            lambda: self._clear_field(self.lottery_department_pdf_paths, "lottery_department_pdf_entry"),
            section_theme=REPORTE_DIARIO_THEME,
        ).pack(side=tk.LEFT)

        self.lottery_department_status_label, self.lottery_department_status_dot = create_status_bar(
            left, self.lottery_department_status_text, section_theme=REPORTE_DIARIO_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Daily Sales Report → completa F/G/H/I/K (Online) y P/Q/R/S (Scratch-Off)",
                "PDF Diario (Department Sales Report) → completa D/E/N/O",
                "Cada botón procesa solo su propia fuente — no hace falta cargar ambas juntas",
                "El Excel Lottery se limpia solo apenas termina de procesarse (cualquiera de los dos botones)",
                "Cada PDF queda cargado hasta que uses su propio botón \"Limpiar\"",
                "Podés cargar varios PDF a la vez en cada campo (Ctrl/Shift + clic)",
                "Cada PDF se ubica por su propia fecha — no hace falta que sean días seguidos",
                "La comisión ONLINE (columna I) se calcula como -6.00% de F (columna J), "
                "y lo que sobra del total se completa en K",
                "Si algún dato no se pudo leer, avisa cuál revisar a mano",
                "Cada Procesar abre su propia copia temporal al terminar",
            ],
            section_theme=REPORTE_DIARIO_THEME,
        )

    def _build_cmv_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "CMV Costo",
            "Selecciona uno o varios archivos de departamento de Elistar; se emparejan por UPC "
            "en la hoja COSTO.TODOS, actualizando solo Costo y Precio.",
            CMV_THEME,
        )

        card = create_card(left, section_theme=CMV_THEME)
        self._cmv_dnd_card = card
        self.cmv_entry = create_file_row(
            card,
            "Archivos Depto.",
            self.cmv_path,
            self.select_cmv_file,
            section_theme=CMV_THEME,
            label_width=11,
            browse_label="Examinar…",
        )
        self.cmv_master_entry = create_file_row(
            card,
            "Maestro CMV",
            self.cmv_master_path,
            self.select_cmv_master_file,
            section_theme=CMV_THEME,
            label_width=11,
        )

        actions = tk.Frame(left, bg=THEME.BG)
        actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            actions,
            "Transformar Excel",
            self.process_cmv_department,
            section_theme=CMV_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            actions,
            "Limpiar Archivos",
            lambda: self._clear_field(self.cmv_path, "cmv_entry"),
            section_theme=CMV_THEME,
        ).pack(side=tk.LEFT)

        self.cmv_status_label, self.cmv_status_dot = create_status_bar(
            left, self.cmv_status_text, section_theme=CMV_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Varios archivos de departamento → se fusionan en una sola COSTO.TODOS",
                "El emparejamiento por UPC actualiza solo Costo (D) y Precio (E)",
                "Abre automáticamente una copia temporal al terminar",
            ],
            section_theme=CMV_THEME,
        )

        if TKINTER_DND_AVAILABLE and self._cmv_dnd_card is not None:
            self._cmv_dnd_card.drop_target_register(DND_FILES)
            self._cmv_dnd_card.dnd_bind("<<Drop>>", self._on_cmv_file_drop)

    def _build_sales_tab(self, parent):
        header, left, right = create_dual_column_tab(parent)
        create_compact_section_header(
            header,
            "CMV Ventas",
            "Selecciona el Excel maestro CMV y uno o más reportes de ventas del POS (CSV/Excel) para actualizar las hojas de departamento.",
            SALES_THEME,
        )

        card = create_card(left, section_theme=SALES_THEME)
        self.sales_master_entry = create_file_row(
            card,
            "Maestro CMV",
            self.sales_master_path,
            self.select_sales_master_file,
            section_theme=SALES_THEME,
            label_width=11,
        )
        self.sales_entry = create_file_row(
            card,
            "POS Reports",
            self.sales_path,
            self.select_sales_file,
            section_theme=SALES_THEME,
            label_width=11,
            browse_label="Examinar…",
        )

        actions = tk.Frame(left, bg=THEME.BG)
        actions.pack(fill=tk.X, pady=(4, 0))
        create_primary_button(
            actions,
            "Procesar Ventas",
            self.process_monthly_sales_file,
            section_theme=SALES_THEME,
        ).pack(side=tk.LEFT, padx=(0, 8))
        create_secondary_button(
            actions,
            "Limpiar POS Reports",
            lambda: self._clear_field(self.sales_path, "sales_entry"),
            section_theme=SALES_THEME,
        ).pack(side=tk.LEFT)

        self.sales_status_label, self.sales_status_dot = create_status_bar(
            left, self.sales_status_text, section_theme=SALES_THEME
        )

        create_info_panel(
            right,
            "Cómo funciona",
            [
                "Enruta por el nombre exacto del departamento (E-GIGARETTE, FOUTAIN, COFFE, etc.)",
                "Las filas de resumen al final del reporte Elistar se excluyen",
                "Desplaza la fila TOTAL y recalcula las fórmulas SUM en E/F/H",
                "RESUMEN A5→Total: enlaza B y E a autosumas E/F de cada hoja",
                "Vista previa temp — guarde el master manualmente con Guardar como",
            ],
            section_theme=SALES_THEME,
        )

    def _setup_drag_and_drop(self):
        """Register the main window as a global drag-and-drop file target."""
        if not TKINTER_DND_AVAILABLE:
            return

        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind("<<Drop>>", self._on_global_file_drop)

    def _on_global_file_drop(self, event):
        """
        Route dropped files by extension: PDF -> Tab 1, spreadsheet -> Chase or CMV tab.
        """
        paths = self.root.tk.splitlist(event.data)
        if not paths:
            return

        normalized = [
            os.path.abspath(str(p).strip()) for p in paths if str(p).strip()
        ]
        if not normalized:
            return

        file_path = normalized[0]
        extension = os.path.splitext(file_path)[1].lower()

        if extension == ".pdf":
            self.pdf_path.set(file_path)
            self._set_status("PDF de EFT seleccionado.")
        elif extension in {".xlsx", ".xls", ".xlsm", ".xls", ".csv"}:
            master_candidates = [
                p
                for p in normalized
                if workbook_has_costo_todos_sheet and workbook_has_costo_todos_sheet(p)
            ]
            dept_candidates = [
                p
                for p in normalized
                if p not in master_candidates
                and self._looks_like_elistars_department_export(p)
            ]
            if master_candidates:
                self.cmv_master_path.set(master_candidates[-1])
            if dept_candidates:
                self._set_cmv_dept_paths(dept_candidates)
                self._set_cmv_status(
                    f"{len(dept_candidates)} archivo(s) de departamento seleccionado(s)."
                )
            elif master_candidates:
                self._set_cmv_status("Excel maestro CMV seleccionado.")
            elif len(normalized) == 1:
                self.chase_path.set(file_path)
                self._set_chase_status("Archivo de Chase seleccionado.")
            else:
                self._set_cmv_status("No se pudieron clasificar los archivos para Chase o CMV.")

    def _looks_like_elistars_department_export(self, file_path):
        """Heuristic: single-column comma blobs or known CMV/department filename tokens."""
        base = os.path.basename(file_path).lower()
        if any(token in base for token in ("elistars", "dept", "auto")):
            return True
        if "cmv" in base and "costo" not in base and "todos" not in base:
            return True
        extension = os.path.splitext(file_path)[1].lower()
        if extension not in {".xlsx", ".xls", ".xlsm", ".csv"}:
            return False
        try:
            if extension == ".csv":
                preview = pd.read_csv(
                    file_path, header=None, nrows=5, dtype=str, keep_default_na=False
                )
            else:
                preview = pd.read_excel(
                    file_path, header=None, nrows=5, dtype=str, keep_default_na=False
                )
        except Exception:
            return False
        if preview.empty:
            return False
        for _, row in preview.iterrows():
            cells = [str(v).strip() for v in row.tolist() if str(v).strip()]
            if len(cells) == 1 and cells[0].count(",") >= 8:
                return True
        return False

    def _on_cmv_file_drop(self, event):
        paths = self.root.tk.splitlist(event.data)
        if not paths:
            return
        dept_files = []
        for raw_path in paths:
            file_path = os.path.abspath(str(raw_path).strip())
            if not file_path:
                continue
            if workbook_has_costo_todos_sheet and workbook_has_costo_todos_sheet(
                file_path
            ):
                self.cmv_master_path.set(file_path)
            else:
                dept_files.append(file_path)
        if dept_files:
            existing = split_paths(self.cmv_path.get()) if split_paths else []
            merged = existing + [p for p in dept_files if p not in existing]
            self._set_cmv_dept_paths(merged)
        count = len(split_paths(self.cmv_path.get())) if split_paths else 0
        self._set_cmv_status(
            f"Selección CMV actualizada ({count} archivo(s) de departamento)."
            if count
            else "Excel maestro CMV seleccionado."
        )

    def _set_status(self, message, is_error=False, completed=False):
        self.status_text.set(message)
        set_status_style(
            self.status_label,
            self.status_dot,
            message,
            section_theme=EFT_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_gettel_status(self, message, is_error=False, completed=False):
        self.gettel_status_text.set(message)
        set_status_style(
            self.gettel_status_label,
            self.gettel_status_dot,
            message,
            section_theme=EFT_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_gettel_pagos_status(self, message, is_error=False, completed=False):
        self.gettel_pagos_status_text.set(message)
        set_status_style(
            self.gettel_pagos_status_label,
            self.gettel_pagos_status_dot,
            message,
            section_theme=GETTEL_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_proveedores_status(self, message, is_error=False, completed=False):
        self.proveedores_status_text.set(message)
        set_status_style(
            self.proveedores_status_label,
            self.proveedores_status_dot,
            message,
            section_theme=PROVEEDORES_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_proveedores_pagos_status(self, message, is_error=False, completed=False):
        self.proveedores_pagos_status_text.set(message)
        set_status_style(
            self.proveedores_pagos_status_label,
            self.proveedores_pagos_status_dot,
            message,
            section_theme=PROVEEDORES_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_chase_status(self, message, is_error=False, completed=False):
        self.chase_status_text.set(message)
        set_status_style(
            self.chase_status_label,
            self.chase_status_dot,
            message,
            section_theme=CHASE_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_cmv_dept_paths(self, paths):
        """Store and display one or more department file paths in the Entry field."""
        if not join_paths or not paths:
            return
        display_value = join_paths(paths)
        self.cmv_path.set(display_value)
        if hasattr(self, "cmv_entry"):
            self.cmv_entry.delete(0, tk.END)
            self.cmv_entry.insert(0, display_value)

    def _set_cmv_status(self, message, is_error=False, completed=False):
        self.cmv_status_text.set(message)
        set_status_style(
            self.cmv_status_label,
            self.cmv_status_dot,
            message,
            section_theme=CMV_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_sales_status(self, message, is_error=False, completed=False):
        self.sales_status_text.set(message)
        set_status_style(
            self.sales_status_label,
            self.sales_status_dot,
            message,
            section_theme=SALES_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_reporte_status(self, message, is_error=False, completed=False):
        self.reporte_status_text.set(message)
        set_status_style(
            self.reporte_status_label,
            self.reporte_status_dot,
            message,
            section_theme=REPORTE_DIARIO_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_store_info_status(self, message, is_error=False, completed=False):
        self.store_info_status_text.set(message)
        set_status_style(
            self.store_info_status_label,
            self.store_info_status_dot,
            message,
            section_theme=REPORTE_DIARIO_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_lottery_sales_report_status(self, message, is_error=False, completed=False):
        self.lottery_sales_report_status_text.set(message)
        set_status_style(
            self.lottery_sales_report_status_label,
            self.lottery_sales_report_status_dot,
            message,
            section_theme=REPORTE_DIARIO_THEME,
            is_error=is_error,
            completed=completed,
        )

    def _set_lottery_department_status(self, message, is_error=False, completed=False):
        self.lottery_department_status_text.set(message)
        set_status_style(
            self.lottery_department_status_label,
            self.lottery_department_status_dot,
            message,
            section_theme=REPORTE_DIARIO_THEME,
            is_error=is_error,
            completed=completed,
        )

    def clear_reporte_pdf(self):
        """
        Reset just the shared PDF Diario box.

        The PDF stays loaded across multiple "Procesar" actions (it feeds
        Ventas and Store Info one after another) — this is the explicit way
        to swap it out. The destination Excels aren't included here since
        each already clears itself right after its own successful process.
        """
        self._clear_field(self.reporte_pdf_path, "reporte_pdf_entry")

    def _clear_field(self, variable, entry_attr):
        """
        Reset one file-path StringVar and its Entry widget.

        The shared primitive behind every per-box "Limpiar" button and every
        automatic post-success master-Excel reset: master/destination Excel
        fields always auto-clear right after a successful process (so the
        same file never gets double-processed by accident), while source
        PDFs/reports are left for the user to clear deliberately, since
        those are often reused across more than one destination.
        """
        variable.set("")
        entry = getattr(self, entry_attr, None)
        if entry is not None:
            entry.delete(0, tk.END)

    def _clear_chase_inputs(self):
        """Clear Chase file path entry widgets after successful processing."""
        self.chase_path.set("")
        self.chase_entry.delete(0, tk.END)

    def select_pdf(self):
        path = filedialog.askopenfilename(
            title="Seleccionar PDF de EFT",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")],
        )
        if path:
            self.pdf_path.set(path)
            self._set_status("PDF de EFT seleccionado.")

    def _ask_excel_ledger_path(self):
        """Shared master workbook file picker for Paso 1 and Paso 2."""
        return filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xlsm"),
                ("Todos los archivos", "*.*"),
            ],
        )

    def select_excel(self):
        path = self._ask_excel_ledger_path()
        if path:
            self.excel_path.set(path)
            self._set_status("Archivo Excel seleccionado.")

    def select_monthly_coupon_report(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Reporte Mensual de Cupones J.H. Williams",
            filetypes=[
                ("Archivos de reporte mensual", "*.xlsx *.xlsm *.xls *.csv"),
                ("Archivos Excel", "*.xlsx *.xlsm *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if path:
            self.monthly_coupon_path.set(path)
            self._set_status("Reporte mensual de cupones seleccionado.")

    def process_and_update(self):
        pdf_path = self.pdf_path.get().strip()
        excel_path = self.excel_path.get().strip()

        if not pdf_path:
            messagebox.showwarning("Archivo faltante", "Seleccione un archivo PDF de EFT.")
            return
        if not excel_path:
            messagebox.showwarning("Archivo faltante", "Seleccione un archivo Excel.")
            return

        self._set_status("Procesando PDF y actualizando Excel...")
        self.root.update_idletasks()

        try:
            header_data, paid_invoices, credit_coupons = extract_eft_data(pdf_path)
            if eft_already_loaded_in_workbook(
                excel_path, header_data, credit_coupons
            ):
                messagebox.showwarning("Alerta", EFT_DUPLICATE_ALERT)
                self._set_status(EFT_DUPLICATE_ALERT, is_error=True)
                return

            temp_excel_path = create_temp_work_copy(excel_path)
            update_excel_workbook(
                temp_excel_path, header_data, paid_invoices, credit_coupons
            )
            open_excel_workbook(temp_excel_path)
            self._clear_field(self.excel_path, "excel_entry")
            self._set_status("¡Proceso completado!", completed=True)
        except FileNotFoundError as exc:
            self._set_status(f"Error: {exc}", is_error=True)
        except ValueError as exc:
            self._set_status(f"Error: {exc}", is_error=True)
        except Exception as exc:
            self._set_status(f"Error: {exc}", is_error=True)

    def process_monthly_coupon_append(self):
        if append_monthly_cupones is None:
            self._set_status(
                f"Cupones append module unavailable: {_CUPONES_APPEND_IMPORT_ERROR}",
                is_error=True,
            )
            return

        excel_path = self.excel_path.get().strip()
        monthly_report_path = self.monthly_coupon_path.get().strip()

        if not excel_path:
            messagebox.showwarning(
                "Archivo faltante",
                "Seleccione el Excel Ledger.",
            )
            return

        if not monthly_report_path:
            self._process_resync_cupones_only(excel_path)
            return

        self._set_status("Appending monthly Cupones report...")
        self.root.update_idletasks()

        try:
            _saved_path, summary = append_monthly_cupones(excel_path, monthly_report_path)
            self._clear_field(self.excel_path, "excel_entry")
            skipped = summary.get("rows_skipped_duplicates", 0)
            resynced = summary.get("rows_resynced_pending", 0)
            status_suffix = (
                f"({summary['rows_appended']} appended, {summary['rows_matched']} matched"
            )
            if skipped:
                status_suffix += f", {skipped} duplicado(s) omitido(s)"
            if resynced:
                status_suffix += f", {resynced} pendiente(s) resincronizado(s)"
            status_suffix += ")."
            self._set_status(
                "Monthly Cupones update completed " + status_suffix,
                completed=True,
            )
        except Exception as exc:
            if (
                MonthlyReportFullyDuplicateError is not None
                and isinstance(exc, MonthlyReportFullyDuplicateError)
            ):
                messagebox.showwarning("Alerta", str(exc))
                self._set_status(str(exc), is_error=True)
                return
            self._set_status(f"Error: {exc}", is_error=True)

    def _process_resync_cupones_only(self, excel_path):
        """
        No monthly report selected: just refresh pending/$0 Cupones rows
        against the current Cta Cte data (e.g. after loading a new EFT with
        no new J.H. Williams coupons to add this time).
        """
        if resync_cupones_only is None:
            self._set_status(
                f"Cupones append module unavailable: {_CUPONES_APPEND_IMPORT_ERROR}",
                is_error=True,
            )
            return

        self._set_status("Actualizando cupones pendientes...")
        self.root.update_idletasks()

        try:
            _saved_path, summary = resync_cupones_only(excel_path)
            self._clear_field(self.excel_path, "excel_entry")
            resynced = summary.get("rows_resynced_pending", 0)
            self._set_status(
                f"Cupones actualizados ({resynced} pendiente(s) resincronizado(s)).",
                completed=True,
            )
        except Exception as exc:
            if (
                NoPendingCouponsError is not None
                and isinstance(exc, NoPendingCouponsError)
            ):
                messagebox.showinfo("Sin pendientes", str(exc))
                self._set_status(str(exc), completed=True)
                return
            self._set_status(f"Error: {exc}", is_error=True)

    def select_gettel_source_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel o PDF/Foto de Origen (Cupones Diarios)",
            filetypes=[
                ("Excel o PDF", "*.xlsx *.pdf"),
                ("Archivos Excel", "*.xlsx"),
                ("PDF (foto escaneada)", "*.pdf"),
            ],
        )
        if not path:
            return
        abs_path = os.path.abspath(path)
        self.gettel_source_excel_path.set(abs_path)
        if hasattr(self, "gettel_source_excel_entry"):
            self.gettel_source_excel_entry.delete(0, tk.END)
            self.gettel_source_excel_entry.insert(0, abs_path)
        self._set_gettel_status(f"Origen: {os.path.basename(abs_path)}")

    def select_gettel_destination_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de Destino (Master Cierre)",
            filetypes=[("Archivos Excel", "*.xlsx")],
        )
        if not path:
            return
        abs_path = os.path.abspath(path)
        self.gettel_destination_excel_path.set(abs_path)
        if hasattr(self, "gettel_destination_excel_entry"):
            self.gettel_destination_excel_entry.delete(0, tk.END)
            self.gettel_destination_excel_entry.insert(0, abs_path)
        self._set_gettel_status(f"Destino: {os.path.basename(abs_path)}")

    def select_gettel_pagos_files(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar PDF(s) de Pagos",
            filetypes=PDF_DAILY_FILETYPES,
        )
        if not filenames:
            return
        paths = [os.path.abspath(path) for path in filenames]
        display_value = join_paths(paths) if join_paths is not None else "; ".join(paths)
        self.gettel_pagos_pdf_paths.set(display_value)
        if hasattr(self, "gettel_pagos_entry"):
            self.gettel_pagos_entry.delete(0, tk.END)
            self.gettel_pagos_entry.insert(0, display_value)
        count = len(paths)
        self._set_gettel_pagos_status(
            f"{count} PDF(s) de Pagos seleccionado(s)." if count != 1 else "1 PDF de Pagos seleccionado."
        )

    def process_gettel_pagos_file(self):
        if process_gettel_pagos is None:
            self._set_gettel_pagos_status(
                f"Gettel/Toyota no disponible: {_GETTEL_TOYOTA_IMPORT_ERROR}",
                is_error=True,
            )
            return

        master_path = self.gettel_destination_excel_path.get().strip()
        pdf_text = self.gettel_pagos_pdf_paths.get().strip()

        if split_paths is not None:
            pdf_paths = split_paths(pdf_text)
        else:
            pdf_paths = [pdf_text] if pdf_text else []

        if not master_path:
            self._set_gettel_pagos_status(
                "Seleccione el Excel de Destino (Master Cierre).", is_error=True
            )
            return
        if not pdf_paths:
            self._set_gettel_pagos_status(
                "Seleccione uno o más PDF de Pagos.", is_error=True
            )
            return

        self._set_gettel_pagos_status(f"Procesando {len(pdf_paths)} PDF(s) de Pagos...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _preview_path, summary = process_gettel_pagos(master_path, pdf_paths)
            self._clear_field(self.gettel_destination_excel_path, "gettel_destination_excel_entry")
            warnings = [
                f"Pago {batch['payment_number']} ({batch['company']}, {batch['date']}): {batch['warning']}"
                for batch in summary.get("batch_results", [])
                if batch.get("warning")
            ]
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_gettel_pagos_status(f"Completado en {elapsed}.", completed=True)
            if warnings:
                messagebox.showwarning(
                    "Revisar Pagos",
                    "Se completó el proceso, pero hay datos que no se pudieron leer "
                    "(revisar manualmente):\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_gettel_pagos_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def select_proveedores_ledger(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel Ledger (Cta Cte Proveedores)",
            filetypes=[("Archivos Excel", "*.xlsx")],
        )
        if not path:
            return
        abs_path = os.path.abspath(path)
        self.proveedores_ledger_path.set(abs_path)
        if hasattr(self, "proveedores_ledger_entry"):
            self.proveedores_ledger_entry.delete(0, tk.END)
            self.proveedores_ledger_entry.insert(0, abs_path)
        self._set_proveedores_status(f"Excel Ledger: {os.path.basename(abs_path)}")

    def select_proveedores_invoice_files(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar PDF(s) de Facturas",
            filetypes=PDF_DAILY_FILETYPES,
        )
        if not filenames:
            return

        # Cada proveedor vive en su propia carpeta del Drive, así que el
        # diálogo nativo (que solo multi-selecciona dentro de UNA carpeta)
        # no alcanza para armar un lote de varios proveedores de una vez.
        # Se acumula lo ya elegido con lo nuevo -- "Limpiar" sigue
        # vaciando el campo para arrancar un lote nuevo de cero.
        existing_text = self.proveedores_invoice_pdf_paths.get().strip()
        if split_paths is not None:
            existing_paths = split_paths(existing_text)
        else:
            existing_paths = [existing_text] if existing_text else []

        combined_paths = list(existing_paths)
        for path in filenames:
            abs_path = os.path.abspath(path)
            if abs_path not in combined_paths:
                combined_paths.append(abs_path)

        display_value = (
            join_paths(combined_paths) if join_paths is not None else "; ".join(combined_paths)
        )
        self.proveedores_invoice_pdf_paths.set(display_value)
        if hasattr(self, "proveedores_invoice_entry"):
            self.proveedores_invoice_entry.delete(0, tk.END)
            self.proveedores_invoice_entry.insert(0, display_value)
        count = len(combined_paths)
        self._set_proveedores_status(
            f"{count} PDF(s) de facturas seleccionado(s)." if count != 1 else "1 PDF de factura seleccionado."
        )

    def process_proveedores_invoices(self):
        if append_supplier_invoices is None:
            self._set_proveedores_status(
                f"Proveedores no disponible: {_PROVEEDORES_IMPORT_ERROR}",
                is_error=True,
            )
            return

        ledger_path = self.proveedores_ledger_path.get().strip()
        pdf_text = self.proveedores_invoice_pdf_paths.get().strip()

        if split_paths is not None:
            pdf_paths = split_paths(pdf_text)
        else:
            pdf_paths = [pdf_text] if pdf_text else []

        if not ledger_path:
            self._set_proveedores_status(
                "Seleccione el Excel Ledger.", is_error=True
            )
            return
        if not pdf_paths:
            self._set_proveedores_status(
                "Seleccione uno o más PDF de facturas.", is_error=True
            )
            return

        self._set_proveedores_status(f"Procesando {len(pdf_paths)} PDF(s) de facturas...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _preview_path, summary = append_supplier_invoices(ledger_path, pdf_paths)
            self._clear_field(self.proveedores_ledger_path, "proveedores_ledger_entry")
            warnings = [
                f"{batch['supplier']}: {len(batch['duplicates_skipped'])} factura(s) ya "
                f"cargada(s), omitida(s) — {', '.join(batch['duplicates_skipped'])}"
                for batch in summary.get("batch_results", [])
                if batch.get("duplicates_skipped")
            ]
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_proveedores_status(
                f"Completado en {elapsed} ({summary['invoices_appended']} factura(s) agregada(s)).",
                completed=True,
            )
            if warnings:
                messagebox.showwarning(
                    "Revisar Proveedores",
                    "Se completó el proceso, pero hay avisos:\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_proveedores_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def select_proveedores_bank_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Extracto de Chase (ya categorizado)",
            filetypes=[
                ("Archivos Chase", "*.csv *.xlsx *.xlsm *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Archivos Excel", "*.xlsx *.xlsm"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if path:
            self.proveedores_bank_path.set(os.path.abspath(path))
            if hasattr(self, "proveedores_bank_entry"):
                self.proveedores_bank_entry.delete(0, tk.END)
                self.proveedores_bank_entry.insert(0, os.path.abspath(path))
            self._set_proveedores_pagos_status("Extracto de Chase seleccionado.")

    def process_proveedores_payments(self):
        if append_supplier_payments is None:
            self._set_proveedores_pagos_status(
                f"Proveedores no disponible: {_PROVEEDORES_IMPORT_ERROR}",
                is_error=True,
            )
            return

        ledger_path = self.proveedores_ledger_path.get().strip()
        bank_path = self.proveedores_bank_path.get().strip()

        if not ledger_path:
            self._set_proveedores_pagos_status(
                "Seleccione el Excel Ledger.", is_error=True
            )
            return
        if not bank_path:
            self._set_proveedores_pagos_status(
                "Seleccione el extracto de Chase.", is_error=True
            )
            return

        self._set_proveedores_pagos_status("Procesando pagos del extracto de Chase...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _preview_path, summary = append_supplier_payments(ledger_path, bank_path)
            self._clear_field(self.proveedores_ledger_path, "proveedores_ledger_entry")
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_proveedores_pagos_status(
                f"Completado en {elapsed} ({summary['payments_appended']} pago(s) agregado(s)).",
                completed=True,
            )
            warnings = [
                f"{batch['sheet_name']}: {len(batch['duplicates_skipped'])} pago(s) ya "
                f"cargado(s), omitido(s) — {', '.join(batch['duplicates_skipped'])}"
                for batch in summary.get("batch_results", [])
                if batch.get("duplicates_skipped")
            ]
            unmatched = summary.get("unmatched") or []
            if unmatched:
                warnings.append(
                    f"{len(unmatched)} pago(s) sin proveedor identificado:\n"
                    + "\n".join(unmatched)
                )
            if warnings:
                messagebox.showwarning(
                    "Revisar Pagos de Proveedores",
                    "Se completó el proceso, pero hay avisos:\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_proveedores_pagos_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def process_gettel_toyota_report(self):
        source_path = self.gettel_source_excel_path.get().strip()
        destination_path = self.gettel_destination_excel_path.get().strip()
        if not source_path or not os.path.isfile(source_path):
            messagebox.showwarning(
                "Archivo faltante",
                "Seleccione el Excel o PDF/foto de origen (cupones diarios).",
            )
            return
        if not destination_path or not os.path.isfile(destination_path):
            messagebox.showwarning(
                "Archivo faltante",
                "Seleccione el Excel de destino (master Cierre).",
            )
            return

        is_pdf = os.path.splitext(source_path)[1].lower() == ".pdf"
        if is_pdf:
            self._process_gettel_toyota_pdf(source_path, destination_path)
        else:
            self._process_gettel_toyota_excel(source_path, destination_path)

    def _process_gettel_toyota_excel(self, source_path, destination_path):
        if merge_gettel_toyota_into_master is None:
            self._set_gettel_status(
                f"Gettel/Toyota no disponible: {_GETTEL_TOYOTA_IMPORT_ERROR}",
                is_error=True,
            )
            return

        self._set_gettel_status("Resumiendo cupones y fusionando en el master...")
        self.root.update_idletasks()

        try:
            _preview_path, rows_matched, gettel_days, toyota_days, unmatched_days = (
                merge_gettel_toyota_into_master(
                    source_path, destination_path, launch=True
                )
            )
            self._clear_field(self.gettel_destination_excel_path, "gettel_destination_excel_entry")
            status_suffix = (
                f"({rows_matched} fila(s) actualizadas; "
                f"{gettel_days} día(s) GETTEL, {toyota_days} día(s) TOYOTA)"
            )
            if unmatched_days:
                status_suffix += (
                    f" — {len(unmatched_days)} día(s) del origen sin fila en el "
                    "master, ignorado(s)"
                )
            self._set_gettel_status(
                "Master abierto en Excel " + status_suffix + " — use Guardar como.",
                completed=True,
            )
        except Exception as exc:
            self._set_gettel_status(f"Error: {exc}", is_error=True)

    def _process_gettel_toyota_pdf(self, source_path, destination_path):
        if merge_gettel_toyota_pdf_into_master is None:
            self._set_gettel_status(
                f"Gettel/Toyota no disponible: {_GETTEL_TOYOTA_IMPORT_ERROR}",
                is_error=True,
            )
            return

        self._set_gettel_status("Leyendo la foto/PDF con OCR (puede tardar unos segundos)...")
        self.root.update_idletasks()

        try:
            _preview_path, rows_matched, vendor, days, diagnostics = (
                merge_gettel_toyota_pdf_into_master(
                    source_path, destination_path, launch=True
                )
            )
            self._clear_field(self.gettel_destination_excel_path, "gettel_destination_excel_entry")
            status_suffix = (
                f"({rows_matched} día(s) actualizados, proveedor {vendor})"
            )
            unmatched_days = diagnostics.get("unmatched_days") or []
            if unmatched_days:
                status_suffix += (
                    f" — {len(unmatched_days)} día(s) del reporte sin fila en el "
                    "master, ignorado(s)"
                )
            if diagnostics.get("printed_subtotal_found"):
                amount_ok = diagnostics.get("amount_matches_subtotal")
                gallons_ok = diagnostics.get("gallons_matches_subtotal")
                if amount_ok and gallons_ok:
                    status_suffix += " — coincide con el subtotal impreso."
                else:
                    status_suffix += (
                        " — ALERTA: no coincide exacto con el subtotal impreso "
                        f"(calculado ${diagnostics['computed_amount']:.2f} vs "
                        f"impreso ${diagnostics['printed_subtotal_amount']:.2f}); "
                        "revise la foto antes de guardar."
                    )
            self._set_gettel_status(
                "Master abierto en Excel " + status_suffix,
                completed=True,
                is_error=not (
                    diagnostics.get("amount_matches_subtotal", True)
                    and diagnostics.get("gallons_matches_subtotal", True)
                ),
            )
        except Exception as exc:
            self._set_gettel_status(f"Error: {exc}", is_error=True)

    def select_chase_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar CSV/Excel de Chase",
            filetypes=[
                ("Archivos Chase", "*.csv *.xlsx *.xlsm *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Archivos Excel", "*.xlsx *.xlsm"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if path:
            self.chase_path.set(path)
            self._set_chase_status("Archivo de Chase seleccionado.")

    def process_chase_data(self):
        file_path = self.chase_path.get().strip()
        if not file_path:
            self._set_chase_status("Seleccione un archivo CSV o Excel de Chase.", is_error=True)
            return

        self._set_chase_status("Procesando movimientos bancarios de Chase...")
        self.root.update_idletasks()

        try:
            temp_chase_path = create_temp_work_copy(file_path)
            updated_count, total_rows = process_chase_categorization(temp_chase_path)
            open_excel_workbook(temp_chase_path)
            self._clear_chase_inputs()
            self._set_chase_status(
                f"¡Procesamiento de Chase completado! ({updated_count} de {total_rows} filas categorizadas.)",
                completed=True,
            )
        except FileNotFoundError as exc:
            self._set_chase_status(f"Error: {exc}", is_error=True)
        except ValueError as exc:
            self._set_chase_status(f"Error: {exc}", is_error=True)
        except Exception as exc:
            self._set_chase_status(f"Error: {exc}", is_error=True)

    def select_cmv_file(self):
        paths = filedialog.askopenfilenames(
            title="Seleccionar Archivos de Departamento de Elistar",
            filetypes=DEPT_FILETYPES,
        )
        if not paths:
            return
        normalized = [os.path.abspath(str(p).strip()) for p in paths if str(p).strip()]
        if not normalized:
            return
        # Igual que _on_cmv_file_drop: se fusiona con lo ya elegido en vez de
        # reemplazarlo, porque el diálogo nativo solo multi-selecciona dentro
        # de UNA carpeta y los exports de distintos departamentos de Elistar
        # pueden vivir en carpetas separadas.
        existing = split_paths(self.cmv_path.get()) if split_paths else []
        merged = existing + [p for p in normalized if p not in existing]
        self._set_cmv_dept_paths(merged)
        self._set_cmv_status(
            f"{len(merged)} archivo(s) de departamento seleccionado(s) para procesar."
        )

    def select_sales_master_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar Excel Maestro CMV",
            filetypes=EXCEL_FILETYPES_MASTER,
        )
        if not filename:
            return
        self.sales_master_path.set(os.path.abspath(filename))
        self._set_sales_status("Excel maestro CMV seleccionado.")

    def select_reporte_master_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar Master Bradenton Análisis C-Store",
            filetypes=ANALISIS_MASTER_FILETYPES,
        )
        if not filename:
            return
        self.reporte_master_path.set(os.path.abspath(filename))
        self._set_reporte_status("Master Bradenton Análisis C-Store seleccionado.")

    def select_reporte_pdf_file(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar Reportes PDF Diarios",
            filetypes=PDF_DAILY_FILETYPES,
        )
        if not filenames:
            return
        paths = [os.path.abspath(path) for path in filenames]
        if join_paths is not None:
            display_value = join_paths(paths)
        else:
            display_value = "; ".join(paths)
        self.reporte_pdf_path.set(display_value)
        if hasattr(self, "reporte_pdf_entry"):
            self.reporte_pdf_entry.delete(0, tk.END)
            self.reporte_pdf_entry.insert(0, display_value)
        count = len(paths)
        self._set_reporte_status(
            f"{count} PDF(s) diario(s) seleccionado(s)."
            if count != 1
            else "1 PDF diario seleccionado."
        )

    def process_reporte_diario_file(self):
        if process_reporte_diario is None:
            self._set_reporte_status(
                f"Módulo de Reporte Diario no disponible: {_REPORTE_IMPORT_ERROR}",
                is_error=True,
            )
            return

        master_path = self.reporte_master_path.get().strip()
        pdf_path_text = self.reporte_pdf_path.get().strip()

        if split_paths is not None:
            pdf_paths = split_paths(pdf_path_text)
        elif pdf_path_text:
            pdf_paths = [pdf_path_text]
        else:
            pdf_paths = []

        if not master_path:
            self._set_reporte_status(
                "Seleccione el master Bradenton Análisis C-Store.",
                is_error=True,
            )
            return
        if not pdf_paths:
            self._set_reporte_status(
                "Seleccione uno o más reportes PDF diarios.",
                is_error=True,
            )
            return

        self._set_reporte_status(
            f"Procesando {len(pdf_paths)} PDF(s) diario(s)..."
        )
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _temp_path, summary = process_reporte_diario(master_path, pdf_paths)
            self._clear_field(self.reporte_master_path, "reporte_master_entry")
            warnings = []
            for batch in summary.get("batch_results", []):
                if batch.get("warning"):
                    warnings.append(
                        f"Día {batch['calendar_day']} ({batch['filename']}): {batch['warning']}"
                    )
                skipped = batch.get("skipped") or []
                if skipped:
                    warnings.append(
                        f"Día {batch['calendar_day']} ({batch['filename']}): "
                        f"omitidos — {', '.join(skipped)}"
                    )

            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_reporte_status(f"Completado en {elapsed}.", completed=True)
            if warnings:
                messagebox.showwarning(
                    "Revisar Reporte Diario",
                    "Se completó el proceso, pero hay avisos:\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_reporte_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def select_store_info_master_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar Excel de Cierre (hoja Store Info)",
            filetypes=ANALISIS_MASTER_FILETYPES,
        )
        if not filename:
            return
        self.store_info_master_path.set(os.path.abspath(filename))
        self._set_store_info_status("Excel de Cierre (Store Info) seleccionado.")

    def select_lottery_master_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar Excel de Lottery",
            filetypes=ANALISIS_MASTER_FILETYPES,
        )
        if not filename:
            return
        self.lottery_master_path.set(os.path.abspath(filename))
        self._set_lottery_sales_report_status("Excel de Lottery seleccionado.")
        self._set_lottery_department_status("Excel de Lottery seleccionado.")

    def select_lottery_sales_report_files(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar Daily Sales Report(s) de Lottery",
            filetypes=PDF_DAILY_FILETYPES,
        )
        if not filenames:
            return
        paths = [os.path.abspath(path) for path in filenames]
        display_value = join_paths(paths) if join_paths is not None else "; ".join(paths)
        self.lottery_sales_report_paths.set(display_value)
        if hasattr(self, "lottery_sales_report_entry"):
            self.lottery_sales_report_entry.delete(0, tk.END)
            self.lottery_sales_report_entry.insert(0, display_value)
        count = len(paths)
        self._set_lottery_sales_report_status(
            f"{count} Daily Sales Report(s) seleccionado(s)."
            if count != 1
            else "1 Daily Sales Report seleccionado."
        )

    def select_lottery_department_pdf_files(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar PDF(s) Diario(s) — Department Sales Report",
            filetypes=PDF_DAILY_FILETYPES,
        )
        if not filenames:
            return
        paths = [os.path.abspath(path) for path in filenames]
        display_value = join_paths(paths) if join_paths is not None else "; ".join(paths)
        self.lottery_department_pdf_paths.set(display_value)
        if hasattr(self, "lottery_department_pdf_entry"):
            self.lottery_department_pdf_entry.delete(0, tk.END)
            self.lottery_department_pdf_entry.insert(0, display_value)
        count = len(paths)
        self._set_lottery_department_status(
            f"{count} PDF(s) diario(s) seleccionado(s)."
            if count != 1
            else "1 PDF diario seleccionado."
        )

    def process_store_info_file(self):
        if process_store_info is None:
            self._set_store_info_status(
                f"Módulo de Reporte Diario no disponible: {_REPORTE_IMPORT_ERROR}",
                is_error=True,
            )
            return

        master_path = self.store_info_master_path.get().strip()
        pdf_path_text = self.reporte_pdf_path.get().strip()

        if split_paths is not None:
            pdf_paths = split_paths(pdf_path_text)
        elif pdf_path_text:
            pdf_paths = [pdf_path_text]
        else:
            pdf_paths = []

        if not master_path:
            self._set_store_info_status(
                "Seleccione el Excel de Cierre (hoja Store Info).",
                is_error=True,
            )
            return
        if not pdf_paths:
            self._set_store_info_status(
                "Seleccione uno o más reportes PDF diarios.",
                is_error=True,
            )
            return

        self._set_store_info_status(f"Procesando {len(pdf_paths)} PDF(s) diario(s)...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _temp_path, summary = process_store_info(master_path, pdf_paths)
            self._clear_field(self.store_info_master_path, "store_info_master_entry")
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_store_info_status(f"Completado en {elapsed}.", completed=True)
        except Exception as exc:
            self._set_store_info_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def process_lottery_sales_report_file(self):
        """Update only F/G/H/I/K (Online) and P/Q/R/S (Scratch-Off) from the Daily Sales Report PDF(s)."""
        if process_lottery is None:
            self._set_lottery_sales_report_status(
                f"Módulo de Reporte Diario no disponible: {_REPORTE_IMPORT_ERROR}",
                is_error=True,
            )
            return

        master_path = self.lottery_master_path.get().strip()
        sales_report_pdf_text = self.lottery_sales_report_paths.get().strip()

        if split_paths is not None:
            sales_report_pdf_paths = split_paths(sales_report_pdf_text)
        else:
            sales_report_pdf_paths = [sales_report_pdf_text] if sales_report_pdf_text else []

        if not master_path:
            self._set_lottery_sales_report_status("Seleccione el Excel de Lottery.", is_error=True)
            return
        if not sales_report_pdf_paths:
            self._set_lottery_sales_report_status(
                "Seleccione uno o más Daily Sales Report de Lottery.", is_error=True
            )
            return

        self._set_lottery_sales_report_status(f"Procesando {len(sales_report_pdf_paths)} PDF(s)...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _temp_path, summary = process_lottery(master_path, [], sales_report_pdf_paths)
            self._clear_field(self.lottery_master_path, "lottery_master_entry")
            warnings = [
                f"{batch['report_date']} ({batch['filename']}): {batch['warning']}"
                for batch in summary.get("batch_results", [])
                if batch.get("warning")
            ]
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_lottery_sales_report_status(f"Completado en {elapsed}.", completed=True)
            if warnings:
                messagebox.showwarning(
                    "Revisar Lottery",
                    "Se completó el proceso, pero hay datos que no se pudieron leer "
                    "(quedaron con el valor que ya tenía la plantilla):\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_lottery_sales_report_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def process_lottery_department_file(self):
        """Update only D/E/N/O from the shared daily PDF's Department Sales Report."""
        if process_lottery is None:
            self._set_lottery_department_status(
                f"Módulo de Reporte Diario no disponible: {_REPORTE_IMPORT_ERROR}",
                is_error=True,
            )
            return

        master_path = self.lottery_master_path.get().strip()
        department_pdf_text = self.lottery_department_pdf_paths.get().strip()

        if split_paths is not None:
            department_pdf_paths = split_paths(department_pdf_text)
        else:
            department_pdf_paths = [department_pdf_text] if department_pdf_text else []

        if not master_path:
            self._set_lottery_department_status("Seleccione el Excel de Lottery.", is_error=True)
            return
        if not department_pdf_paths:
            self._set_lottery_department_status(
                "Seleccione uno o más PDF diarios.", is_error=True
            )
            return

        self._set_lottery_department_status(f"Procesando {len(department_pdf_paths)} PDF(s)...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        start_time = time.time()

        try:
            _temp_path, summary = process_lottery(master_path, department_pdf_paths, [])
            self._clear_field(self.lottery_master_path, "lottery_master_entry")
            warnings = [
                f"{batch['report_date']} ({batch['filename']}): {batch['warning']}"
                for batch in summary.get("batch_results", [])
                if batch.get("warning")
            ]
            elapsed = format_elapsed_duration(time.time() - start_time)
            self._set_lottery_department_status(f"Completado en {elapsed}.", completed=True)
            if warnings:
                messagebox.showwarning(
                    "Revisar Lottery",
                    "Se completó el proceso, pero hay datos que no se pudieron leer "
                    "(quedaron con el valor que ya tenía la plantilla):\n\n" + "\n\n".join(warnings),
                )
        except Exception as exc:
            self._set_lottery_department_status(f"Error: {exc}", is_error=True)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def select_sales_file(self):
        filenames = filedialog.askopenfilenames(
            title="Seleccionar Reportes de Ventas del POS",
            filetypes=SALES_FILETYPES,
        )
        if not filenames:
            return
        paths = [os.path.abspath(path) for path in filenames]
        if join_paths is not None:
            display_value = join_paths(paths)
        else:
            display_value = "; ".join(paths)
        self.sales_path.set(display_value)
        if hasattr(self, "sales_entry"):
            self.sales_entry.delete(0, tk.END)
            self.sales_entry.insert(0, display_value)
        count = len(paths)
        self._set_sales_status(
            f"{count} archivo(s) de ventas seleccionado(s)."
            if count != 1
            else "1 archivo de ventas seleccionado."
        )

    def process_monthly_sales_file(self):
        if process_monthly_sales is None or split_paths is None:
            messagebox.showerror(
                "Error",
                f"Módulo de Ventas Mensuales no disponible: {_SALES_IMPORT_ERROR}",
            )
            self._set_sales_status(
                f"Módulo de Ventas Mensuales no disponible: {_SALES_IMPORT_ERROR}",
                is_error=True,
            )
            return

        sales_files = split_paths(self.sales_path.get())
        master_path = self.sales_master_path.get().strip()
        if not master_path:
            messagebox.showwarning(
                "Archivo requerido",
                "Seleccione el archivo maestro CMV (.xlsx / .xlsm).",
            )
            self._set_sales_status(
                "Seleccione el Excel maestro CMV.", is_error=True
            )
            return
        if not sales_files:
            messagebox.showwarning(
                "Archivo requerido",
                "Seleccione uno o más archivos de ventas (CSV / Excel).",
            )
            self._set_sales_status(
                "Seleccione uno o más reportes de ventas del POS.", is_error=True
            )
            return

        self._set_sales_status(
            f"Procesando {len(sales_files)} archivo(s) de ventas y actualizando hojas de departamento..."
        )
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        try:
            frame, _preview_path = process_monthly_sales(sales_files, master_path)
            dept_count = frame["Dept Name"].nunique()
            self._clear_field(self.sales_master_path, "sales_master_entry")
            self._set_sales_status(
                f"Se actualizaron {len(frame)} fila(s) de {len(sales_files)} archivo(s) "
                f"en {dept_count} departamento(s). Vista previa del master abierta.",
                completed=True,
            )
        except Exception as exc:
            logger_msg = str(exc)
            self._set_sales_status(f"Error: {logger_msg}", is_error=True)
            messagebox.showerror("Error", logger_msg)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()

    def select_cmv_master_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar Excel Maestro CMV",
            filetypes=EXCEL_FILETYPES_MASTER,
        )
        if not filename:
            return
        self.cmv_master_path.set(os.path.abspath(filename))
        self._set_cmv_status("Excel maestro CMV seleccionado.")

    def process_cmv_department(self):
        if update_master_costo_todos_bulk is None or split_paths is None:
            messagebox.showerror(
                "Error",
                f"Módulo de CMV no disponible: {_CMV_IMPORT_ERROR}",
            )
            self._set_cmv_status(
                f"Módulo de CMV no disponible: {_CMV_IMPORT_ERROR}", is_error=True
            )
            return

        dept_paths = split_paths(self.cmv_path.get())
        master_path = self.cmv_master_path.get().strip()

        if not dept_paths:
            messagebox.showwarning(
                "Archivos requeridos",
                "Seleccione uno o más archivos de departamento (CSV/Excel).",
            )
            self._set_cmv_status(
                "Seleccione uno o más archivos de departamento de Elistar.", is_error=True
            )
            return
        if not master_path:
            messagebox.showwarning(
                "Archivo requerido",
                "Seleccione el archivo maestro CMV (.xls / .xlsx / .xlsm).",
            )
            self._set_cmv_status(
                "Seleccione el Excel maestro CMV.", is_error=True
            )
            return

        self._set_cmv_status(
            f"Actualizando COSTO.TODOS con {len(dept_paths)} archivo(s) de departamento..."
        )
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        try:
            master_abs = os.path.abspath(master_path)
            (
                saved_path,
                file_results,
                total_parsed,
                total_rows_updated,
                upcs_not_in_master,
                _master_upc_count,
            ) = update_master_costo_todos_bulk(master_abs, dept_paths)

            self._clear_field(self.cmv_master_path, "cmv_master_entry")
            self._set_cmv_status(
                f"Vista previa abierta en Excel ({total_rows_updated} fila(s) añadidas).",
                completed=True,
            )
        except Exception as exc:
            logger_msg = str(exc)
            self._set_cmv_status(f"Error: {logger_msg}", is_error=True)
            messagebox.showerror("Error", logger_msg)
        finally:
            self.root.config(cursor="")
            self.root.update_idletasks()


def ensure_runtime_dependencies():
    """Validate required packages before launching the UI."""
    missing = []
    if _PDFPLUMBER_IMPORT_ERROR is not None:
        missing.append("pdfplumber")
    if _OPENPYXL_IMPORT_ERROR is not None:
        missing.append("openpyxl")
    try:
        import pandas  # noqa: F401
    except ImportError:
        missing.append("pandas")
    if _CMV_IMPORT_ERROR is not None:
        missing.append("cmv_costo (módulo local)")
    if _SALES_IMPORT_ERROR is not None:
        missing.append("monthly_sales (módulo local)")
    if _REPORTE_IMPORT_ERROR is not None:
        missing.append("reporte_diario (módulo local)")
    if sys.platform == "win32":
        try:
            import win32com.client  # noqa: F401
        except ImportError:
            missing.append("pywin32")
    if missing:
        packages = " ".join(
            pkg
            for pkg in missing
            if pkg
            not in {
                "cmv_costo (módulo local)",
                "monthly_sales (módulo local)",
                "reporte_diario (módulo local)",
            }
        )
        hint = (
            f"Faltan dependencias: {', '.join(missing)}.\n\n"
            f"Instale con:\n  pip install -r requirements.txt"
        )
        if packages:
            hint += f"\n  pip install {packages}"
        raise SystemExit(hint)


def main():
    ensure_runtime_dependencies()
    root = TkinterDnD.Tk() if TKINTER_DND_AVAILABLE else tk.Tk()
    EFTExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
