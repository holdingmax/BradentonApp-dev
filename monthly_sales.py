"""
Monthly Sales Update — POS CSV parsing and Master CMV department sheet injection.

Phase 1: Parse Top-Selling CSV into five normalized columns.
Phase 2: Inject sales rows into department sheets on the Master CMV workbook.
"""

import csv
import os
import re
import sys
import tempfile
from copy import copy

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.formula.translate import Translator
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
    THIN_BORDER = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    EMPTY_BORDER = Border()
    EMPTY_FILL = PatternFill(fill_type=None)
    LEFT_ALIGNMENT = Alignment(horizontal="left")
    CENTER_ALIGNMENT = Alignment(horizontal="center")
    RIGHT_ALIGNMENT = Alignment(horizontal="right")
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False
    THIN_BORDER = None  # type: ignore[assignment,misc]
    EMPTY_BORDER = None  # type: ignore[assignment,misc]
    EMPTY_FILL = None  # type: ignore[assignment,misc]
    LEFT_ALIGNMENT = None  # type: ignore[assignment,misc]
    CENTER_ALIGNMENT = None  # type: ignore[assignment,misc]
    RIGHT_ALIGNMENT = None  # type: ignore[assignment,misc]

SALES_COLUMNS = ["UPC", "Name", "Dept Name", "Count", "Retail/Amount"]
SALES_DATA_START_ROW = 5
SALES_DATA_COLUMNS = 5  # A through E only
SALES_FORMULA_COLUMNS = (6, 7, 8, 9)  # F through I
SALES_SHEET_LAST_COLUMN = 9  # Column I — full closing row width
TOTAL_LABEL_COLUMNS = (2, 3, 4)  # B, C, D — merged label / TOTAL text
TOTAL_BOUNDARY_MARK = "TOTAL"
TOTAL_SUM_COLUMNS = (5, 6, 8)  # E, F, H — dynamic =SUM targets on TOTAL row
SALES_SPACER_ROWS = 1  # exactly one blank row above the TOTAL row
DELIMITER_BLOCK_TAIL_ROWS = 2  # closing row + trailing footer rows

RESUMEN_SHEET_NAME = "RESUMEN"
RESUMEN_FIRST_DATA_ROW = 5
RESUMEN_DRIVER_COLUMN = 1  # A — department name driver
RESUMEN_SALES_TOTAL_COLUMN = 2  # B — link to department column E autosum
RESUMEN_SECONDARY_TOTAL_COLUMN = 5  # E — link to department column F autosum
RESUMEN_TOTAL_STOP_LABELS = frozenset({"Total", "TOTAL"})
DECIMAL_TWO_FORMAT = "0.00"

PATH_SEPARATORS = (";", "|")

DEPARTMENT_SHEETS = (
    "AUTO",
    "BEER",
    "SODA",
    "SNUFF",
    "GEN-CTN",
    "GEN-PAK",
    "CIGARS",
    "MAJPAK",
    "MAJCR",
    "E-GIGARETTE",
    "TAXABLE",
    "SNACK",
    "CANDY",
    "BOILED PEANUTS",
    "FLOWER",
    "ICE CREAM",
    "GROCERIES",
    "FOUTAIN",
    "NONTAX",
    "JUICE",
    "WATER",
    "MILK",
    "COFFE",
    "HOT DOG",
    "HBA",
    "PROPANE",
)
FIRST_DEPARTMENT_SHEET_INDEX = 4  # 0-based index of first department sheet

ELISTAR_SUMMARY_MARKERS = (
    "dept ",
    "qty sold",
    "total sales",
    "total cost",
    "gross profit",
    "total retail",
    "total amount",
    "total qty",
    "department total",
)

ELISTAR_SUMMARY_EXACT_LABELS = frozenset(
    {
        "qty sold",
        "total sales",
        "total cost",
        "gross profit",
        "total retail",
        "total amount",
        "total qty",
        "department total",
    }
)


def join_paths(paths):
    """Join absolute sales file paths for multi-select display in the Entry widget."""
    normalized = [
        os.path.abspath(str(path).strip())
        for path in paths
        if path is not None and str(path).strip()
    ]
    return "; ".join(normalized)


def split_paths(text):
    if not text or not str(text).strip():
        return []
    raw = str(text).strip()
    for separator in PATH_SEPARATORS:
        if separator in raw:
            return [part.strip() for part in raw.split(separator) if part.strip()]
    return [raw]


def _normalize_sales_file_paths(file_paths):
    if isinstance(file_paths, (list, tuple)):
        return [os.path.abspath(str(path).strip()) for path in file_paths if str(path).strip()]
    if not file_paths or not str(file_paths).strip():
        return []
    return split_paths(str(file_paths))


# Normalized CSV department labels -> canonical master workbook tab name
CSV_DEPARTMENT_TO_SHEET = {
    "auto": "AUTO",
    "automotive": "AUTO",
    "beer": "BEER",
    "beer/wine": "BEER",
    "beer wine": "BEER",
    "beer-wine": "BEER",
    "soda": "SODA",
    "snuff": "SNUFF",
    "gen-ctn": "GEN-CTN",
    "gen ctn": "GEN-CTN",
    "genctn": "GEN-CTN",
    "gen-pak": "GEN-PAK",
    "gen pak": "GEN-PAK",
    "genpak": "GEN-PAK",
    "cigars": "CIGARS",
    "majpak": "MAJPAK",
    "majcr": "MAJCR",
    "e-cig": "E-GIGARETTE",
    "e cig": "E-GIGARETTE",
    "ecig": "E-GIGARETTE",
    "e-cigarette": "E-GIGARETTE",
    "e cigarette": "E-GIGARETTE",
    "e-gigarette": "E-GIGARETTE",
    "taxable": "TAXABLE",
    "nontax": "NONTAX",
    "non-tax": "NONTAX",
    "non tax": "NONTAX",
    "fountain": "FOUTAIN",
    "foutain": "FOUTAIN",
    "coffee": "COFFE",
    "coffe": "COFFE",
    "hot dog": "HOT DOG",
    "hotdog": "HOT DOG",
    "hba": "HBA",
    "propane": "PROPANE",
    "prop hd": "PROPANE",
    "prop-hd": "PROPANE",
    "snack": "SNACK",
    "candy": "CANDY",
    "boiled peanuts": "BOILED PEANUTS",
    "boiled-peanuts": "BOILED PEANUTS",
    "flower": "FLOWER",
    "ice cream": "ICE CREAM",
    "ice-cream": "ICE CREAM",
    "groceries": "GROCERIES",
    "grocery": "GROCERIES",
    "juice": "JUICE",
    "water": "WATER",
    "milk": "MILK",
}

# Longest-token-first substring routes when the CSV label is not an exact alias
ORDERED_DEPARTMENT_TOKEN_ROUTES = (
    ("beer/wine", "BEER"),
    ("beer wine", "BEER"),
    ("boiled peanuts", "BOILED PEANUTS"),
    ("boiled-peanuts", "BOILED PEANUTS"),
    ("ice cream", "ICE CREAM"),
    ("ice-cream", "ICE CREAM"),
    ("gen-pak", "GEN-PAK"),
    ("gen pak", "GEN-PAK"),
    ("gen-ctn", "GEN-CTN"),
    ("gen ctn", "GEN-CTN"),
    ("e-gigarette", "E-GIGARETTE"),
    ("e-cigarette", "E-GIGARETTE"),
    ("e-cig", "E-GIGARETTE"),
    ("e cig", "E-GIGARETTE"),
    ("hot dog", "HOT DOG"),
    ("prop hd", "PROPANE"),
    ("prop-hd", "PROPANE"),
    ("propane", "PROPANE"),
    ("groceries", "GROCERIES"),
    ("majpak", "MAJPAK"),
    ("majcr", "MAJCR"),
    ("cigars", "CIGARS"),
    ("nontax", "NONTAX"),
    ("non-tax", "NONTAX"),
    ("non tax", "NONTAX"),
    ("taxable", "TAXABLE"),
    ("fountain", "FOUTAIN"),
    ("foutain", "FOUTAIN"),
    ("coffee", "COFFE"),
    ("coffe", "COFFE"),
    ("flower", "FLOWER"),
    ("snack", "SNACK"),
    ("candy", "CANDY"),
    ("juice", "JUICE"),
    ("water", "WATER"),
    ("milk", "MILK"),
    ("snuff", "SNUFF"),
    ("soda", "SODA"),
    ("beer", "BEER"),
    ("auto", "AUTO"),
    ("hba", "HBA"),
)


def _strip_cell(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _sniff_delimiter(file_path, sample_size=65536):
    try:
        with open(file_path, encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(sample_size)
    except OSError:
        return None
    if not sample.strip():
        return None
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        counts = {sep: sample.count(sep) for sep in (";", ",", "\t", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else None


def _read_csv_with_separator(file_path, separator):
    engine = "python" if separator is None else "c"
    return pd.read_csv(
        file_path,
        sep=separator,
        engine=engine,
        header=None,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
        quoting=csv.QUOTE_MINIMAL,
    )


def _expand_single_column_blob(frame):
    """Split a one-column export that still contains delimiter characters."""
    series = frame.iloc[:, 0].astype(str).map(_strip_cell)
    best_split = frame
    best_cols = 1

    for separator in (";", ",", "\t", "|"):
        split = series.str.split(separator, expand=True)
        split = split.apply(lambda column: column.map(_strip_cell))
        if split.shape[1] > best_cols:
            best_cols = split.shape[1]
            best_split = split

    return best_split


def _read_delimited_csv(file_path):
    """
    Load a POS CSV with dynamic delimiter detection (comma, semicolon, tab, etc.).
    """
    last_error = None
    candidates = []
    separators = []

    sniffed = _sniff_delimiter(file_path)
    if sniffed:
        separators.append(sniffed)

    separators.extend([None, ";", ",", "\t", "|"])
    seen = set()
    ordered_separators = []
    for separator in separators:
        key = repr(separator)
        if key not in seen:
            seen.add(key)
            ordered_separators.append(separator)

    for separator in ordered_separators:
        try:
            frame = _read_csv_with_separator(file_path, separator)
            column_count = frame.shape[1]
            candidates.append((column_count, separator, frame))
            if column_count >= 5:
                return frame.iloc[:, :].copy()
        except Exception as exc:
            last_error = exc

    if not candidates:
        raise ValueError(
            f"Could not read sales CSV: {os.path.basename(file_path)}"
        ) from last_error

    column_count, _separator, frame = max(candidates, key=lambda item: item[0])

    if column_count <= 1:
        frame = _expand_single_column_blob(frame)
        column_count = frame.shape[1]

    if column_count <= 1:
        raise ValueError(
            "Sales CSV could not be split into columns. "
            "Check delimiter formatting (comma/semicolon/tab)."
        )

    return frame


def _drop_header_row(frame):
    if frame.empty:
        return frame

    first_row = [_strip_cell(value).lower() for value in frame.iloc[0].tolist()]
    joined = " ".join(first_row)
    header_markers = ("upc", "name", "dept", "count", "retail", "amount", "qty")
    if any(marker in joined for marker in header_markers):
        return frame.iloc[1:].reset_index(drop=True)
    return frame


def _field_matches_elistar_summary_marker(text):
    """True when a UPC, Name, or Dept cell marks the Elistar bottom summary block."""
    normalized = _normalize_label(text)
    if not normalized:
        return False
    if normalized.startswith("dept "):
        return True
    if normalized in ELISTAR_SUMMARY_EXACT_LABELS:
        return True
    for marker in ELISTAR_SUMMARY_MARKERS:
        if marker in normalized:
            return True
    return False


def _is_elistar_summary_boundary_row(row_values):
    """
    Detect the Elistar department summary footer; product import must stop here.
    """
    cells = [_strip_cell(value) for value in row_values]
    upc_text = cells[0] if len(cells) > 0 else ""
    name_text = cells[1] if len(cells) > 1 else ""
    dept_text = cells[2] if len(cells) > 2 else ""

    for field_text in (upc_text, name_text, dept_text):
        if _field_matches_elistar_summary_marker(field_text):
            return True

    row_blob = _normalize_label(" ".join(cell for cell in cells if cell))
    if not row_blob:
        return False

    if row_blob.startswith("dept "):
        return True

    summary_hits = sum(
        1 for marker in ELISTAR_SUMMARY_MARKERS if marker in row_blob
    )
    if summary_hits >= 2:
        return True

    if "qty sold" in row_blob and any(
        token in row_blob for token in ("total sales", "total cost", "gross profit")
    ):
        return True

    return False


def _trim_elistar_summary_rows(raw):
    """Keep only product rows above the Elistar bottom summary section."""
    if raw.empty:
        return raw

    kept_rows = []
    for row_idx in range(len(raw)):
        row_values = raw.iloc[row_idx].tolist()
        if _is_elistar_summary_boundary_row(row_values):
            break
        kept_rows.append(row_values)

    if not kept_rows:
        return raw.iloc[0:0].copy()

    return pd.DataFrame(kept_rows, columns=raw.columns).reset_index(drop=True)


def _coerce_upc_int(value):
    text = _strip_cell(value)
    if not text or text.lower() in {"nan", "none"}:
        return None

    if "." in text:
        whole, fractional = text.split(".", 1)
        if fractional == "" or set(fractional) <= {"0"}:
            text = whole

    digits = re.sub(r"\D", "", text)
    if not digits:
        return None
    return int(digits)


def _coerce_count_int(value):
    text = _strip_cell(value)
    if not text:
        return None
    text = text.replace(",", "")
    if "." in text:
        whole, fractional = text.split(".", 1)
        if fractional == "" or set(fractional) <= {"0"}:
            text = whole
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _coerce_retail_float(value):
    text = _strip_cell(value)
    if not text:
        return None
    text = text.replace("$", "").replace(" ", "")
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _ensure_openpyxl_available():
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "Monthly sales master injection requires openpyxl. "
            "Install with: pip install openpyxl"
        )


def _normalize_label(value):
    return re.sub(r"\s+", " ", _strip_cell(value)).lower()


def _validate_master_path(master_path):
    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError(
            "Master CMV workbook must be .xlsx or .xlsm for sales injection."
        )
    return extension


def _create_temp_master_path():
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="sales_master_preview_")
    os.close(fd)
    return temp_path


def _build_sheet_lookup(workbook):
    lookup = {}
    for index, name in enumerate(workbook.sheetnames):
        lookup[_normalize_label(name)] = (workbook[name], index, name)
    return lookup


def _is_nontax_department(key):
    return (
        "nontax" in key
        or "non-tax" in key
        or "non tax" in key
        or key in {"nontax", "non-tax", "non tax"}
    )


def _is_taxable_department(key):
    if _is_nontax_department(key):
        return False
    return "taxable" in key or key == "taxable"


def _resolve_title_sheet_aliases(key):
    """
    Map department titles to exact workbook tab spellings (name-based, not tab index).
    """
    if "e-gigarette" in key or "e-cigarette" in key or "e-cig" in key or key in {
        "e cig",
        "ecig",
    }:
        return "E-GIGARETTE"
    if "foutain" in key or "fountain" in key:
        return "FOUTAIN"
    if "coffee" in key or key == "coffe" or key.startswith("coffe "):
        return "COFFE"
    if "hot dog" in key or key.replace("-", "") == "hotdog":
        return "HOT DOG"
    if "propane" in key or "prop hd" in key or "prop-hd" in key:
        return "PROPANE"
    return None


def _resolve_sheet_name(dept_name):
    """
    Map parsed Dept Name text to a canonical Master CMV department worksheet.

    Uses explicit alias rules first, then exact sheet-name matching, then
    controlled substring routes (longest token first). Sheets resolve by tab
    title string via wb[sheet_name], never by physical tab order.
    """
    key = _normalize_label(dept_name)
    if not key:
        return None

    if _is_nontax_department(key):
        return "NONTAX"
    if _is_taxable_department(key):
        return "TAXABLE"

    title_sheet = _resolve_title_sheet_aliases(key)
    if title_sheet is not None:
        return title_sheet

    if key in CSV_DEPARTMENT_TO_SHEET:
        return CSV_DEPARTMENT_TO_SHEET[key]

    if "beer/wine" in key or key.startswith("beer/") or key.endswith("/wine"):
        return "BEER"
    if key == "auto" or key.startswith("auto/") or key.startswith("auto "):
        return "AUTO"

    for sheet_name in DEPARTMENT_SHEETS:
        sheet_key = _normalize_label(sheet_name)
        if key == sheet_key:
            return sheet_name
        if key.replace("-", "").replace(" ", "") == sheet_key.replace("-", "").replace(
            " ", ""
        ):
            return sheet_name

    if "beer" in key:
        return "BEER"

    for token, sheet_name in ORDERED_DEPARTMENT_TOKEN_ROUTES:
        if token in key:
            return sheet_name

    return None


def _get_department_sheet(workbook, sheet_lookup, canonical_name):
    """Resolve the openpyxl worksheet for a canonical department tab name."""
    target = _normalize_label(canonical_name)

    if target in sheet_lookup:
        return sheet_lookup[target][0]

    for sheet_name in workbook.sheetnames:
        if _normalize_label(sheet_name) == target:
            return workbook[sheet_name]

    target_flat = target.replace("-", "").replace(" ", "")
    for sheet_name in workbook.sheetnames:
        name_flat = _normalize_label(sheet_name).replace("-", "").replace(" ", "")
        if name_flat == target_flat:
            return workbook[sheet_name]

    for index, expected in enumerate(DEPARTMENT_SHEETS):
        if _normalize_label(expected) != target:
            continue
        sheet_index = FIRST_DEPARTMENT_SHEET_INDEX + index
        if sheet_index < len(workbook.worksheets):
            return workbook.worksheets[sheet_index]

    available = ", ".join(workbook.sheetnames)
    raise ValueError(
        f'Department sheet "{canonical_name}" not found. Available: {available}'
    )


def _safe_read_cell_value(sheet, row_idx, col_idx):
    try:
        return sheet.cell(row=row_idx, column=col_idx).value
    except (AttributeError, TypeError, Exception):
        return None


def _safe_clear_cell(sheet, row_idx, col_idx):
    try:
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = None
    except (AttributeError, TypeError, Exception):
        pass


def _safe_write_cell(
    sheet, row_idx, col_idx, value, number_format=None, alignment=None
):
    try:
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        if number_format is not None:
            cell.number_format = number_format
        if alignment is not None:
            cell.alignment = alignment
    except (AttributeError, TypeError, Exception):
        pass


def _safe_apply_border(sheet, row_idx, col_idx):
    try:
        sheet.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    except (AttributeError, TypeError, Exception):
        pass


def _last_populated_row_in_block(sheet):
    last_row = SALES_DATA_START_ROW - 1
    scan_to = max(sheet.max_row or SALES_DATA_START_ROW, SALES_DATA_START_ROW)
    for row_idx in range(SALES_DATA_START_ROW, scan_to + 1):
        for col_idx in range(1, SALES_DATA_COLUMNS + 1):
            if _safe_read_cell_value(sheet, row_idx, col_idx) not in (None, ""):
                last_row = row_idx
                break
    return last_row


def _cell_has_formula(value):
    return isinstance(value, str) and value.startswith("=")


def _last_active_formula_row_fi(sheet, start_row, end_row):
    """Last row between start_row and end_row with active formulas in F:I."""
    if end_row < start_row:
        return start_row - 1

    last_row = start_row - 1
    for row_idx in range(start_row, end_row + 1):
        for col_idx in SALES_FORMULA_COLUMNS:
            try:
                value = _safe_read_cell_value(sheet, row_idx, col_idx)
            except (AttributeError, TypeError, Exception):
                continue
            if value in (None, ""):
                continue
            if _cell_has_formula(value):
                last_row = max(last_row, row_idx)
                break
    return last_row


def _copy_cell_formula_and_style(source_cell, target_cell):
    """Copy value/style into target, translating formulas to the target coordinate."""
    value = source_cell.value
    if _cell_has_formula(value):
        try:
            target_cell.value = Translator(
                value, origin=source_cell.coordinate
            ).translate_formula(target_cell.coordinate)
        except (AttributeError, TypeError, Exception):
            target_cell.value = value
    else:
        target_cell.value = value

    try:
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
    except (AttributeError, TypeError, Exception):
        pass


def _copy_fi_row(sheet, source_row, target_row):
    """Copy columns F:I from source_row into target_row with formula translation."""
    for col_idx in SALES_FORMULA_COLUMNS:
        try:
            source_cell = sheet.cell(row=source_row, column=col_idx)
            target_cell = sheet.cell(row=target_row, column=col_idx)
            _copy_cell_formula_and_style(source_cell, target_cell)
        except (AttributeError, TypeError, Exception):
            continue


def _extend_formula_rows_fi(sheet, template_last_row, sales_zone_end_row):
    """
    Fill new sales-zone rows in F:I by copying/translating from the row above.
    """
    if sales_zone_end_row < SALES_DATA_START_ROW:
        return

    seed_row = template_last_row
    if seed_row < SALES_DATA_START_ROW:
        seed_row = SALES_DATA_START_ROW

    first_new_row = seed_row + 1
    if first_new_row > sales_zone_end_row:
        return

    for target_row in range(first_new_row, sales_zone_end_row + 1):
        try:
            _copy_fi_row(sheet, target_row - 1, target_row)
        except (AttributeError, TypeError, Exception):
            continue


def _cell_contains_total(value):
    """True when cell text starts with or contains the word TOTAL."""
    text = _strip_cell(value).upper()
    if not text:
        return False
    if text.startswith(TOTAL_BOUNDARY_MARK):
        return True
    return bool(re.search(rf"\b{TOTAL_BOUNDARY_MARK}\b", text))


def _detect_total_row_index(sheet):
    """
    Scan column B (and merged B:C:D) from row 5 for the TOTAL closing label.

    Returns total_row_index — the absolute lower boundary of the sales grid.
    """
    scan_to = max(sheet.max_row or SALES_DATA_START_ROW, SALES_DATA_START_ROW)

    for row_idx in range(SALES_DATA_START_ROW, scan_to + 1):
        for col_idx in TOTAL_LABEL_COLUMNS:
            try:
                value = _safe_read_cell_value(sheet, row_idx, col_idx)
            except (AttributeError, TypeError, Exception):
                continue
            if _cell_contains_total(value):
                return row_idx

    try:
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row < SALES_DATA_START_ROW:
                continue
            if merged_range.min_col > 4 or merged_range.max_col < 2:
                continue
            anchor_row = merged_range.min_row
            try:
                value = _safe_read_cell_value(sheet, anchor_row, 2)
            except (AttributeError, TypeError, Exception):
                continue
            if _cell_contains_total(value):
                return anchor_row
    except (AttributeError, TypeError, Exception):
        pass

    last_populated = _last_populated_row_in_block(sheet)
    return max(last_populated + 1, SALES_DATA_START_ROW)


def _total_row_block_bounds(sheet, total_start):
    """Closing block span across A:I anchored on the TOTAL row."""
    block_end = total_start + DELIMITER_BLOCK_TAIL_ROWS

    try:
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= total_start <= merged_range.max_row:
                if (
                    merged_range.min_col <= SALES_SHEET_LAST_COLUMN
                    and merged_range.max_col >= 1
                ):
                    block_end = max(block_end, merged_range.max_row)
    except (AttributeError, TypeError, Exception):
        pass

    scan_to = max(sheet.max_row or block_end, block_end)
    last_row_with_content = total_start
    for row_idx in range(total_start, scan_to + 1):
        has_row_content = False
        for col_idx in range(1, SALES_SHEET_LAST_COLUMN + 1):
            if _safe_read_cell_value(sheet, row_idx, col_idx) not in (None, ""):
                has_row_content = True
                break
        if has_row_content:
            last_row_with_content = row_idx
        elif row_idx > total_start:
            break

    block_end = max(block_end, last_row_with_content)
    return total_start, block_end


def _capture_bcd_merge_bounds(sheet, row_idx):
    """Return (min_row, max_row, min_col, max_col) for a B:C:D merge on row_idx."""
    try:
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row_idx <= merged_range.max_row:
                if merged_range.min_col <= 4 and merged_range.max_col >= 2:
                    return (
                        merged_range.min_row,
                        merged_range.max_row,
                        merged_range.min_col,
                        merged_range.max_col,
                    )
    except (AttributeError, TypeError, Exception):
        pass
    return None


def _ensure_bcd_merge_on_total_row(sheet, total_row_index, prior_bounds=None):
    """Restore B:C:D merge on the TOTAL row after a shift."""
    min_col = 2
    max_col = 4
    if prior_bounds is not None:
        _, _, min_col, max_col = prior_bounds

    merge_ref = (
        f"{get_column_letter(min_col)}{total_row_index}:"
        f"{get_column_letter(max_col)}{total_row_index}"
    )
    try:
        sheet.merge_cells(merge_ref)
    except (AttributeError, TypeError, Exception):
        pass


def _displace_total_row_block(sheet, total_row_index, row_offset):
    """Shift the entire closing row slice (A:I) by row_offset (down or up)."""
    if row_offset == 0:
        return total_row_index

    start_row, end_row = _total_row_block_bounds(sheet, total_row_index)
    cell_range = f"A{start_row}:I{end_row}"

    try:
        sheet.move_range(cell_range, rows=row_offset, cols=0, translate=True)
    except (AttributeError, TypeError, Exception):
        _manual_shift_total_row_block(sheet, start_row, end_row, row_offset)

    return total_row_index + row_offset


def _manual_shift_total_row_block(sheet, start_row, end_row, row_offset):
    """Fallback: copy A:I closing cells downward with formula translation."""
    dest_start = start_row + row_offset

    for source_row in range(end_row, start_row - 1, -1):
        target_row = source_row + row_offset
        for col_idx in range(1, SALES_SHEET_LAST_COLUMN + 1):
            try:
                source_cell = sheet.cell(row=source_row, column=col_idx)
                target_cell = sheet.cell(row=target_row, column=col_idx)
                _copy_cell_formula_and_style(source_cell, target_cell)
                source_cell.value = None
            except (AttributeError, TypeError, Exception):
                continue

    for row_idx in range(start_row, dest_start):
        for col_idx in range(1, SALES_SHEET_LAST_COLUMN + 1):
            _safe_clear_cell(sheet, row_idx, col_idx)


def _ensure_sales_writing_capacity(sheet, record_count):
    """
    Position the TOTAL row so one blank spacer sits above it after data is written.

    Shifts the full A:I closing block to target row (last_data_row + 2) and
    preserves the B:C:D merge on the TOTAL row.
    Returns (total_row_index_after_shift, original_last_formula_row_fi).
    """
    total_row_index = _detect_total_row_index(sheet)
    spacer_row = total_row_index - SALES_SPACER_ROWS
    sales_zone_end = max(spacer_row - 1, SALES_DATA_START_ROW - 1)
    original_formula_last = _last_active_formula_row_fi(
        sheet, SALES_DATA_START_ROW, sales_zone_end
    )

    last_data_row = SALES_DATA_START_ROW + max(record_count, 0) - 1
    if record_count <= 0:
        last_data_row = SALES_DATA_START_ROW - 1
    target_total_row = last_data_row + SALES_SPACER_ROWS + 1

    bcd_merge_bounds = _capture_bcd_merge_bounds(sheet, total_row_index)
    row_offset = target_total_row - total_row_index
    if row_offset != 0:
        total_row_index = _displace_total_row_block(
            sheet, total_row_index, row_offset
        )

    _ensure_bcd_merge_on_total_row(sheet, total_row_index, bcd_merge_bounds)
    return total_row_index, original_formula_last


def _purge_leftover_rows_before_total(sheet, last_data_row, total_row_index):
    """
    Delete structural rows between the spacer and TOTAL so layout collapses cleanly.

    Keeps exactly one blank row (last_data_row + 1) immediately above the TOTAL row.
    Returns the updated total_row_index after any deletions.
    """
    if last_data_row < SALES_DATA_START_ROW:
        return total_row_index

    spacer_row = last_data_row + SALES_SPACER_ROWS
    excess_start = spacer_row + 1
    excess_end = total_row_index - 1

    if excess_start > excess_end:
        return total_row_index

    delete_count = excess_end - excess_start + 1
    try:
        sheet.delete_rows(excess_start, delete_count)
    except (AttributeError, TypeError, Exception):
        return total_row_index

    return total_row_index - delete_count


def _unmerge_sales_grid_ae(sheet, first_row, last_row):
    """Unmerge any merged regions intersecting columns A–E in the sales write range."""
    if last_row < first_row:
        return

    ranges_to_unmerge = []
    try:
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.max_row < first_row or merged_range.min_row > last_row:
                continue
            if merged_range.max_col < 1 or merged_range.min_col > SALES_DATA_COLUMNS:
                continue
            if merged_range.min_col <= SALES_DATA_COLUMNS and merged_range.max_col >= 1:
                ranges_to_unmerge.append(str(merged_range))
    except (AttributeError, TypeError, Exception):
        return

    for range_ref in ranges_to_unmerge:
        try:
            sheet.unmerge_cells(range_ref)
        except (AttributeError, TypeError, Exception):
            continue


def _clear_sales_data_block(sheet, total_row_index):
    """Clear columns A–E from row 5 through the row above the spacer."""
    end_row = total_row_index - SALES_SPACER_ROWS - 1
    if end_row < SALES_DATA_START_ROW:
        return

    for row_idx in range(SALES_DATA_START_ROW, end_row + 1):
        for col_idx in range(1, SALES_DATA_COLUMNS + 1):
            _safe_clear_cell(sheet, row_idx, col_idx)


def _strip_spacer_row(sheet, spacer_row):
    """
    Clear the single blank spacer row (A:I): no values, borders, or background fill.
    """
    if spacer_row < SALES_DATA_START_ROW:
        return

    for col_idx in range(1, SALES_SHEET_LAST_COLUMN + 1):
        try:
            cell = sheet.cell(row=spacer_row, column=col_idx)
            cell.value = None
            cell.border = EMPTY_BORDER
            cell.fill = EMPTY_FILL
        except (AttributeError, TypeError, Exception):
            continue


def _summary_structural_end_row(sheet, total_row_index):
    """Last row of the closing summary block anchored on the TOTAL row."""
    _, block_end = _total_row_block_bounds(sheet, total_row_index)
    return max(total_row_index, block_end)


def _purge_ghost_rows_below_summary(sheet, total_row_index):
    """
    Delete all rows below the summary section to remove stray borders/fills.

    Uses total_row_index (and any trailing footer rows in the same block) as the
    absolute structural end, then removes everything beneath sheet.max_row.
    """
    if total_row_index < SALES_DATA_START_ROW:
        return

    structural_end = _summary_structural_end_row(sheet, total_row_index)
    try:
        max_row = sheet.max_row or structural_end
    except (AttributeError, TypeError, Exception):
        return

    if max_row <= structural_end:
        return

    delete_count = max_row - structural_end
    try:
        sheet.delete_rows(structural_end + 1, delete_count)
    except (AttributeError, TypeError, Exception):
        pass


def _rewrite_total_sum_formulas(sheet, total_row_index, last_data_row):
    """Set native =SUM formulas on the TOTAL row for columns E, F, and H."""
    if last_data_row < SALES_DATA_START_ROW:
        return

    for col_idx in TOTAL_SUM_COLUMNS:
        column_letter = get_column_letter(col_idx)
        formula = (
            f"=SUM({column_letter}{SALES_DATA_START_ROW}:"
            f"{column_letter}{last_data_row})"
        )
        try:
            cell = sheet.cell(row=total_row_index, column=col_idx)
            cell.value = formula
        except (AttributeError, TypeError, Exception):
            continue


def _resumen_driver_is_total(value):
    """True when Column A equals the RESUMEN footer marker (Total / TOTAL)."""
    return _strip_cell(value) in RESUMEN_TOTAL_STOP_LABELS


def _resolve_resumen_department_sheet_name(workbook, department_label):
    """Match a RESUMEN Column A label to an existing workbook tab name."""
    label = _strip_cell(department_label)
    if not label:
        return None
    if label in workbook.sheetnames:
        return label

    normalized = _normalize_label(label)
    for sheet_name in workbook.sheetnames:
        if _normalize_label(sheet_name) == normalized:
            return sheet_name

    label_flat = normalized.replace("-", "").replace(" ", "")
    for sheet_name in workbook.sheetnames:
        name_flat = _normalize_label(sheet_name).replace("-", "").replace(" ", "")
        if name_flat == label_flat:
            return sheet_name

    return None


def _cell_is_autosum_formula(cell):
    """True when a cell holds a live =SUM (or other) autosum formula."""
    value = cell.value
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text.startswith("="):
        return False
    return "SUM" in text.upper()


def _find_department_autosum_row(worksheet, column_index):
    """
    Locate the shifting autosum row in one department column (E or F).

    Scans from row 5 downward for formula cells containing SUM. When multiple
    matches exist, the lowest row wins (monthly TOTAL footer).
    """
    scan_to = max(worksheet.max_row or SALES_DATA_START_ROW, SALES_DATA_START_ROW)
    matches = []
    for row_idx in range(SALES_DATA_START_ROW, scan_to + 1):
        try:
            cell = worksheet.cell(row=row_idx, column=column_index)
        except (AttributeError, TypeError, Exception):
            continue
        if _cell_is_autosum_formula(cell):
            matches.append(row_idx)

    if matches:
        return matches[-1]

    try:
        total_row_index = _detect_total_row_index(worksheet)
        total_cell = worksheet.cell(row=total_row_index, column=column_index)
        if _cell_is_autosum_formula(total_cell):
            return total_row_index
    except (AttributeError, TypeError, Exception):
        pass

    return None


def _excel_external_reference(sheet_name, column_letter, row_index):
    """Build ='Sheet Name'!E42 with Excel-safe single-quote escaping."""
    escaped = sheet_name.replace("'", "''")
    return f"='{escaped}'!{column_letter}{row_index}"


def _write_resumen_link_cell(worksheet, row_index, column_index, formula):
    """Write a cross-sheet formula on RESUMEN with Spanish decimal formatting."""
    try:
        cell = worksheet.cell(row=row_index, column=column_index, value=formula)
        cell.number_format = DECIMAL_TWO_FORMAT
        if RIGHT_ALIGNMENT is not None:
            cell.alignment = RIGHT_ALIGNMENT
    except (AttributeError, TypeError, Exception):
        pass


def _get_resumen_worksheet(workbook):
    if RESUMEN_SHEET_NAME in workbook.sheetnames:
        return workbook[RESUMEN_SHEET_NAME]
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().upper() == RESUMEN_SHEET_NAME:
            return workbook[sheet_name]
    raise ValueError(f'Sheet "{RESUMEN_SHEET_NAME}" not found in the master workbook.')


def update_resumen_department_links(workbook):
    """
    Wire RESUMEN rows to each department sheet's shifting E/F autosum cells.

    Scans Column A from row 5 until a Total/TOTAL marker. For every department
    tab that exists, writes live formulas into RESUMEN columns B and E.
    """
    resumen = _get_resumen_worksheet(workbook)
    links_written = 0
    scan_limit = max(resumen.max_row or RESUMEN_FIRST_DATA_ROW, RESUMEN_FIRST_DATA_ROW)
    scan_limit += 200

    row_idx = RESUMEN_FIRST_DATA_ROW
    while row_idx <= scan_limit:
        driver_value = _safe_read_cell_value(resumen, row_idx, RESUMEN_DRIVER_COLUMN)
        if _resumen_driver_is_total(driver_value):
            break

        department_label = _strip_cell(driver_value)
        if department_label:
            sheet_name = _resolve_resumen_department_sheet_name(
                workbook, department_label
            )
            if sheet_name is not None:
                department_sheet = workbook[sheet_name]
                row_e = _find_department_autosum_row(department_sheet, 5)
                row_f = _find_department_autosum_row(department_sheet, 6)

                row_linked = False
                if row_e is not None:
                    formula_b = _excel_external_reference(sheet_name, "E", row_e)
                    _write_resumen_link_cell(
                        resumen, row_idx, RESUMEN_SALES_TOTAL_COLUMN, formula_b
                    )
                    row_linked = True
                if row_f is not None:
                    formula_e = _excel_external_reference(sheet_name, "F", row_f)
                    _write_resumen_link_cell(
                        resumen,
                        row_idx,
                        RESUMEN_SECONDARY_TOTAL_COLUMN,
                        formula_e,
                    )
                    row_linked = True

                if row_linked:
                    links_written += 1

        row_idx += 1

    return links_written


def _write_sales_rows(sheet, dept_frame):
    row_idx = SALES_DATA_START_ROW
    for _, row in dept_frame.iterrows():
        try:
            _safe_write_cell(sheet, row_idx, 1, int(row["UPC"]), number_format="0")
            _safe_write_cell(
                sheet,
                row_idx,
                2,
                str(row["Name"]),
                alignment=LEFT_ALIGNMENT,
            )
            _safe_write_cell(
                sheet,
                row_idx,
                3,
                str(row["Dept Name"]),
                alignment=CENTER_ALIGNMENT,
            )
            _safe_write_cell(
                sheet,
                row_idx,
                4,
                int(row["Count"]),
                number_format="0",
                alignment=CENTER_ALIGNMENT,
            )
            _safe_write_cell(
                sheet,
                row_idx,
                5,
                float(row["Retail/Amount"]),
                number_format="0.00",
                alignment=CENTER_ALIGNMENT,
            )
            for col_idx in range(1, SALES_DATA_COLUMNS + 1):
                _safe_apply_border(sheet, row_idx, col_idx)
        except (AttributeError, TypeError, Exception):
            pass
        row_idx += 1


def _inject_sales_into_master(master_path, file_paths):
    _ensure_openpyxl_available()
    master_path = os.path.abspath(master_path)
    _validate_master_path(master_path)

    paths = _normalize_sales_file_paths(file_paths)
    if not paths:
        raise ValueError("No sales report files provided.")

    extension = os.path.splitext(master_path)[1].lower()
    workbook = load_workbook(
        master_path,
        data_only=False,
        keep_vba=extension == ".xlsm",
    )

    sheet_lookup = _build_sheet_lookup(workbook)
    sheet_batches = {}
    unmapped = []

    for file_path in paths:
        frame = parse_monthly_sales_file(file_path)
        for dept_name, group in frame.groupby("Dept Name", sort=False):
            sheet_name = _resolve_sheet_name(dept_name)
            if sheet_name is None:
                unmapped.append(_strip_cell(dept_name))
                continue
            sheet_batches.setdefault(sheet_name, []).append(
                group.reset_index(drop=True)
            )

    if unmapped:
        workbook.close()
        raise ValueError(
            "Could not map department(s) to master sheets: "
            + ", ".join(sorted(set(unmapped)))
        )

    if not sheet_batches:
        workbook.close()
        raise ValueError("No department rows could be mapped to master sheets.")

    for sheet_name, batches in sheet_batches.items():
        combined = pd.concat(batches, ignore_index=True)
        sheet = _get_department_sheet(workbook, sheet_lookup, sheet_name)
        record_count = len(combined)
        total_row_index, original_formula_last = _ensure_sales_writing_capacity(
            sheet, record_count
        )
        last_data_row = (
            SALES_DATA_START_ROW + record_count - 1
            if record_count > 0
            else SALES_DATA_START_ROW - 1
        )
        if last_data_row >= SALES_DATA_START_ROW:
            total_row_index = _purge_leftover_rows_before_total(
                sheet, last_data_row, total_row_index
            )
            clear_through_row = max(
                last_data_row,
                total_row_index - SALES_SPACER_ROWS - 1,
            )
            _unmerge_sales_grid_ae(
                sheet, SALES_DATA_START_ROW, clear_through_row
            )
        _clear_sales_data_block(sheet, total_row_index)
        _write_sales_rows(sheet, combined)
        if last_data_row >= SALES_DATA_START_ROW:
            _extend_formula_rows_fi(
                sheet,
                original_formula_last,
                last_data_row,
            )
            _rewrite_total_sum_formulas(sheet, total_row_index, last_data_row)
            _ensure_bcd_merge_on_total_row(sheet, total_row_index)
            _strip_spacer_row(sheet, total_row_index - SALES_SPACER_ROWS)

        _purge_ghost_rows_below_summary(sheet, total_row_index)

    update_resumen_department_links(workbook)

    temp_path = _create_temp_master_path()
    abs_temp_path = os.path.abspath(temp_path)
    workbook.save(abs_temp_path)
    workbook.close()
    _launch_temp_workbook(abs_temp_path)
    return temp_path


def _launch_temp_workbook(temp_path):
    """Open the transient master preview without success message boxes."""
    if sys.platform == "win32":
        os.startfile(temp_path)
    elif sys.platform == "darwin":
        os.system(f'open "{temp_path}"')
    else:
        os.system(f'xdg-open "{temp_path}"')


def _create_preview_csv_path():
    fd, temp_path = tempfile.mkstemp(
        suffix=".csv",
        prefix="preview_ventas_limpias_",
    )
    os.close(fd)
    return temp_path


def _save_and_open_preview(frame):
    preview_path = _create_preview_csv_path()
    frame.to_csv(
        preview_path,
        index=False,
        sep=";",
        decimal=",",
        encoding="utf-8-sig",
    )
    os.startfile(os.path.abspath(preview_path))
    return preview_path


def _load_sales_raw_table(file_path):
    """Load a POS sales CSV or Excel export as a raw columnar table."""
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        return _read_delimited_csv(file_path)
    if extension in {".xlsx", ".xlsm"}:
        return pd.read_excel(
            file_path,
            header=None,
            dtype=str,
            keep_default_na=False,
        )
    if extension == ".xls":
        return pd.read_excel(
            file_path,
            header=None,
            dtype=str,
            keep_default_na=False,
            engine="xlrd",
        )
    raise ValueError(
        f"Unsupported sales file type '{extension}'. Use .csv, .xlsx, .xlsm, or .xls."
    )


def _finalize_sales_frame(frame):
    if frame.empty:
        return frame

    frame["Name"] = frame["Name"].astype(str).str.strip()
    frame["Dept Name"] = frame["Dept Name"].astype(str).str.strip()

    summary_mask = frame.apply(
        lambda row: (
            _field_matches_elistar_summary_marker(str(row["UPC"]))
            or _field_matches_elistar_summary_marker(str(row["Name"]))
            or _field_matches_elistar_summary_marker(str(row["Dept Name"]))
        ),
        axis=1,
    )
    if summary_mask.any():
        first_idx = next(
            (idx for idx, hit in enumerate(summary_mask.tolist()) if hit),
            len(frame),
        )
        frame = frame.iloc[:first_idx].copy()

    frame["UPC"] = frame["UPC"].apply(_coerce_upc_int)
    frame = frame[frame["UPC"].notna()].copy()
    frame["UPC"] = frame["UPC"].astype(int)

    frame["Count"] = frame["Count"].apply(_coerce_count_int)
    frame = frame[frame["Count"].notna()].copy()
    frame["Count"] = frame["Count"].astype(int)

    frame["Retail/Amount"] = frame["Retail/Amount"].apply(_coerce_retail_float)
    frame = frame[frame["Retail/Amount"].notna()].copy()
    frame["Retail/Amount"] = frame["Retail/Amount"].astype(float)

    return frame.reset_index(drop=True)


def parse_monthly_sales_file(file_path):
    """Parse and clean a Top-Selling POS CSV or Excel report into five columns."""
    file_path = os.path.abspath(file_path)
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Sales file not found: {file_path}")

    raw = _load_sales_raw_table(file_path)
    raw = _drop_header_row(raw)
    raw = _trim_elistar_summary_rows(raw)

    if raw.shape[1] < 5:
        raise ValueError(
            f"Expected at least 5 columns in {os.path.basename(file_path)}; "
            f"found {raw.shape[1]}."
        )

    frame = raw.iloc[:, :5].copy()
    frame.columns = SALES_COLUMNS
    frame = _finalize_sales_frame(frame)

    if frame.empty:
        raise ValueError(
            f"No valid sales rows remained after cleaning {os.path.basename(file_path)}."
        )

    return frame


def parse_monthly_sales_csv(file_path):
    """Backward-compatible alias for single CSV parsing."""
    return parse_monthly_sales_file(file_path)


def process_monthly_sales(file_paths, master_path=None):
    """
    Parse one or more POS sales files and inject rows into Master CMV sheets.

    Loads the master workbook once, processes every selected file through the
    routing and sheet pipeline, saves a single temp preview, and opens silently.

    Returns:
        tuple: (combined DataFrame, optional temp master workbook path)
    """
    paths = _normalize_sales_file_paths(file_paths)
    if not paths:
        raise ValueError("No sales report files provided.")

    frames = [parse_monthly_sales_file(path) for path in paths]
    combined = pd.concat(frames, ignore_index=True)
    temp_master_path = None

    if master_path and str(master_path).strip():
        temp_master_path = _inject_sales_into_master(
            str(master_path).strip(), paths
        )
    else:
        _save_and_open_preview(combined)

    return combined, temp_master_path
