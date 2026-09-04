"""
EFT PDF -> Cta Cte J.H.Williams engine: parses the bank EFT PDF (credit-card
coupons and paid-invoice references) and writes the coupon block into the
Cta Cte ledger with its summary row, balance formula, and borders.
"""

import difflib
import os
import re
from datetime import datetime

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SHEET_NAME = "Cta Cte J.H.Williams"
LEDGER_START_ROW = 5
BOTTOM_SCAN_START_ROW = 5000
BLOCK_FIRST_COLUMN = 1
PERIMETER_LAST_COLUMN = 13  # A through M (boxed perimeter)
COLUMN_N = 14
COUPON_LAST_COLUMN = 6
EFT_DATE_NUMBER_FORMAT = "dd-mmm"  # Column N — Spanish day-month display (e.g. 10-jun)
CURRENCY_NUMBER_FORMAT = (
    '_(* #,##0.00" $"_);_(* \\(#,##0.00" $"\\);_(* "-"?? " $"_);_(@_)'
)
CURRENCY_COLUMNS_COUPON = (4, 5, 6)  # D, E, F
CURRENCY_COLUMNS_SUMMARY = (10, 12)  # J, L

THIN_SIDE = Side(style="thin", color="000000")
THICK_SIDE = Side(style="medium", color="000000")
CLEAR_FILL = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
NO_FILL = PatternFill(fill_type=None)
DEFAULT_FONT = Font()
BOLD_FONT = Font(bold=True)
SUMMARY_ALIGNMENT = Alignment(vertical="center", horizontal="left")
SUMMARY_FIRST_COLUMN = 7
SUMMARY_LAST_COLUMN = 14  # G through N on middle_row
COUPON_DATE_TEXT_PATTERN = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
EFT_DUPLICATE_ALERT = (
    "Alerta: Este EFT ya fue cargado anteriormente con los mismos datos."
)
SUM_FORMULA_RANGE_PATTERN = re.compile(
    r"=\+?\s*SUM\s*\(\s*F(\d+)\s*:\s*F(\d+)\s*\)",
    re.IGNORECASE,
)

CREDIT_COUPON_PATTERN = re.compile(
    r"\b(SI-\d+)/(DDC-\d+)\b", re.IGNORECASE
)
SI_INVOICE_PATTERN = re.compile(r"\b(SI-\d+)\b", re.IGNORECASE)
DDC_COUPON_PATTERN = re.compile(r"\b(DDC-\d+)\b", re.IGNORECASE)
DRAFT_NO_PATTERN = re.compile(r"\b(RCV-\d+)\b", re.IGNORECASE)
DATE_PATTERN = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
EFT_DATE_LABEL_PATTERN = re.compile(
    r"EFT\s*Date\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{4})",
    re.IGNORECASE,
)
CURRENCY_DECIMAL_PATTERN = re.compile(r"\.\d{1,2}\b|,\d{2}\b")
CREDIT_TABLE_HEADER_WORDS = {"gross", "fees", "net", "paid"}


def parse_amount(value):
    """Parse a currency string into a positive float using abs()."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return abs(float(value))
        except (TypeError, ValueError):
            return None

    text = str(value).strip()
    if not text or text.upper() in {"-", "--", "N/A", "NA"}:
        return None

    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1].strip()

    # European decimal comma (14,33) -> 14.33
    if re.search(r",\d{2}\b", text) and "." not in text.split(",")[-1][:3]:
        text = text.replace(".", "").replace(",", ".")

    text = text.replace("$", "").replace(" ", "")
    if text.count(",") > 0 and "." not in text:
        parts = text.rsplit(",", 1)
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) <= 2:
            text = parts[0].replace(",", "") + "." + parts[1]
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", "")

    text = text.strip()
    if text.startswith("-"):
        text = text[1:].strip()
    elif text.endswith("-"):
        text = text[:-1].strip()

    if not text:
        return None

    try:
        return abs(float(text))
    except ValueError:
        return None


def parse_date_to_datetime(date_str):
    """Parse a date string into datetime (tries common PDF formats)."""
    if not date_str:
        return None
    text = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_coupon_date_us(date_str):
    """Column A: strict US format MM/DD/YYYY (e.g. 05/08/2026)."""
    parsed = parse_date_to_datetime(date_str)
    if parsed:
        return parsed.strftime("%m/%d/%Y")
    return date_str


def parse_eft_pdf_date_us(date_str):
    """Parse an EFT PDF date token strictly as MM/DD/YYYY (American)."""
    if not date_str:
        return None
    text = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_eft_date_us(date_str):
    """Header EFT date as MM/DD/YYYY string, matching the source PDF."""
    parsed = parse_eft_pdf_date_us(date_str)
    if parsed:
        return parsed.strftime("%m/%d/%Y")
    return date_str


def parse_date_cell(cell):
    """Extract date from a PDF cell; returns MM/DD/YYYY for coupon rows."""
    if not cell:
        return None
    match = DATE_PATTERN.search(str(cell).strip())
    if match:
        return format_coupon_date_us(match.group(1))
    return None


def normalize_row_cells(cells):
    """Return a safe list of stripped string cell values."""
    try:
        return [str(cell).strip() if cell else "" for cell in cells]
    except Exception:
        return []


def expand_row_components(cells):
    """
    Flatten row cells into scan components.

    Splits comma-separated reference blobs (e.g. 0764,0765,0766,9133)
    so they do not shift currency column positions.
    """
    components = []
    for cell in cells:
        if not cell:
            continue
        text = str(cell).strip()
        if not text:
            continue
        if is_reference_blob(text):
            for part in text.split(","):
                part = part.strip()
                if part:
                    components.append(part)
        else:
            components.append(text)
    return components


def is_reference_blob(text):
    """True for comma-separated tracking refs without currency markers."""
    if not text or "$" in text or "(" in text or ")" in text:
        return False
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return False
    if "," not in text:
        return False
    parts = [p.strip() for p in text.split(",")]
    if len(parts) < 2:
        return False
    return all(
        p.replace(" ", "").isdigit() and 1 <= len(p.replace(" ", "")) <= 8
        for p in parts
    )


def is_tracking_reference(value):
    """Detect single tracking numbers (not currency)."""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if is_reference_blob(text):
        return True
    if "$" in text or "(" in text or ")" in text:
        return False
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return False
    cleaned = text.replace(",", "").replace(" ", "")
    if not cleaned.isdigit():
        return False
    return len(cleaned) >= 4


def is_currency_component(value):
    """
    True when a token looks like Gross/Fees/Paid money.

    Uses $, parentheses, or decimal patterns (dot or comma cents).
    """
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if is_reference_blob(text) or is_tracking_reference(text):
        return False
    if "$" in text or "(" in text or ")" in text:
        return True
    if re.search(r"-\s*\$|\$\s*-", text):
        return True
    if text.startswith("-") and re.search(r"\d", text):
        return True
    if CURRENCY_DECIMAL_PATTERN.search(text):
        return True
    if re.fullmatch(r"-?\d+[.,]\d{2}", text.replace("$", "").strip()):
        return True
    return False


def extract_last_three_financial_amounts(components):
    """
    Extract Gross, Fees, and Paid from currency tokens at the row tail.

    Variable reference columns are already removed via expand_row_components().
    When four amounts exist (Gross, Fees, Net, Paid), Net is skipped by taking
    the 4th-from-last, 3rd-from-last, and last currency tokens.
    """
    currency_tokens = [c for c in components if is_currency_component(c)]
    if len(currency_tokens) < 3:
        return None, None, None

    if len(currency_tokens) >= 4:
        selected = [currency_tokens[-4], currency_tokens[-3], currency_tokens[-1]]
    else:
        selected = currency_tokens[-3:]

    amounts = [parse_amount(token) for token in selected]
    if any(amount is None for amount in amounts):
        return None, None, None
    return amounts[0], amounts[1], amounts[2]


def row_is_credit_table_header(cells):
    """True for the 'Gross Fees Net Paid' header row that opens the credit table."""
    try:
        words = {cell.strip().lower() for cell in cells if cell}
        return CREDIT_TABLE_HEADER_WORDS.issubset(words)
    except Exception:
        return False


def row_contains_credit_coupon(cells):
    """True when the row is a coupon table row."""
    try:
        row_text = " ".join(cells)
        if CREDIT_COUPON_PATTERN.search(row_text):
            return True
        has_si = any(SI_INVOICE_PATTERN.search(c) for c in cells)
        has_ddc = any(DDC_COUPON_PATTERN.search(c) for c in cells)
        return has_si and has_ddc
    except Exception:
        return False


def locate_coupon_combo(cells):
    """Find invoice, coupon, and combo cell index."""
    for index, cell in enumerate(cells):
        match = CREDIT_COUPON_PATTERN.search(cell)
        if match:
            return match.group(1).upper(), match.group(2).upper(), index

    invoice = coupon = None
    si_index = ddc_index = None
    for index, cell in enumerate(cells):
        si_match = SI_INVOICE_PATTERN.search(cell)
        if si_match and "/" not in cell:
            invoice = si_match.group(1).upper()
            si_index = index
        ddc_match = DDC_COUPON_PATTERN.search(cell)
        if ddc_match:
            coupon = ddc_match.group(0).upper()
            ddc_index = index

    if invoice and coupon:
        return invoice, coupon, max(si_index or 0, ddc_index or 0)

    row_text = " ".join(cells)
    match = CREDIT_COUPON_PATTERN.search(row_text)
    if match:
        return match.group(1).upper(), match.group(2).upper(), 1

    # Fila dentro de la tabla de créditos sin DDC- (no todos los pagos
    # traen coupon combo) — se conserva igual, con coupon vacío.
    if invoice:
        return invoice, None, si_index or 0

    return None, None, None


def extract_coupon_columns(cells, fallback_date=None):
    """Parse coupon row using last-three currency detection."""
    cells = normalize_row_cells(cells)
    if not cells:
        return None

    try:
        invoice, coupon, combo_index = locate_coupon_combo(cells)
        if not invoice:
            return None

        components = expand_row_components(cells)
        gross, fees, paid = extract_last_three_financial_amounts(components)
        if gross is None:
            return None

        row_date = None
        if cells:
            row_date = parse_date_cell(cells[0])
        if not row_date and combo_index > 0:
            row_date = parse_date_cell(cells[combo_index - 1])
        if not row_date:
            for cell in cells:
                row_date = parse_date_cell(cell)
                if row_date:
                    break
        if not row_date:
            match = DATE_PATTERN.search(" ".join(cells))
            if match:
                row_date = format_coupon_date_us(match.group(1))
        if not row_date and fallback_date:
            row_date = format_coupon_date_us(fallback_date)

        fees_value = fees if fees is not None else 0.0

        return {
            "date": row_date,
            "invoice": invoice,
            "coupon": coupon,
            "gross_amount": gross,
            "fees_amount": fees_value,
            "paid_amount": paid if paid is not None else 0.0,
        }
    except Exception:
        return None


def _si_match_inside_credit_combo(text, match):
    """True when an SI token belongs to an SI/DDC combo reference."""
    span_start = max(0, match.start() - 12)
    span_end = min(len(text), match.end() + 24)
    return bool(CREDIT_COUPON_PATTERN.search(text[span_start:span_end]))


def extract_paid_invoice_entries(cells):
    """
    Parse one PDF row for all paid invoices at the EFT header (ordered list).

    Collects every SI- reference on the row (not part of an SI/DDC coupon combo)
    and pairs each with a currency amount when available.
    """
    cells = normalize_row_cells(cells)
    if not cells or row_contains_credit_coupon(cells):
        return []

    try:
        invoices_ordered = []
        seen_invoices = set()
        for cell in cells:
            if CREDIT_COUPON_PATTERN.search(cell):
                continue
            for match in SI_INVOICE_PATTERN.finditer(cell):
                if _si_match_inside_credit_combo(cell, match):
                    continue
                invoice = match.group(1).upper()
                if invoice in seen_invoices:
                    continue
                seen_invoices.add(invoice)
                invoices_ordered.append(invoice)

        if not invoices_ordered:
            row_text = " ".join(cells)
            if not SI_INVOICE_PATTERN.search(row_text):
                return []
            for match in SI_INVOICE_PATTERN.finditer(row_text):
                if _si_match_inside_credit_combo(row_text, match):
                    continue
                invoice = match.group(1).upper()
                if invoice in seen_invoices:
                    continue
                seen_invoices.add(invoice)
                invoices_ordered.append(invoice)

        if not invoices_ordered:
            return []

        components = expand_row_components(cells)
        amounts = []
        for token in components:
            if not is_currency_component(token):
                continue
            paid_amount = parse_amount(token)
            if paid_amount is not None:
                amounts.append(paid_amount)
        if not amounts:
            return []

        if len(amounts) >= len(invoices_ordered):
            paired_amounts = amounts[-len(invoices_ordered) :]
        elif len(invoices_ordered) == 1:
            paired_amounts = amounts[-1:]
        else:
            return []

        return [
            {"invoice": invoice, "paid_amount": paid_amount}
            for invoice, paid_amount in zip(invoices_ordered, paired_amounts)
        ]
    except Exception:
        return []


def collect_pdf_rows(pdf_path):
    """Extract table rows and text lines from every PDF page."""
    full_text_parts = []
    rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                full_text_parts.append(page_text)

            for table in page.extract_tables() or []:
                for row in table:
                    if not row or not any(cell for cell in row):
                        continue
                    rows.append(normalize_row_cells(row))

            for line in page_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    if "\t" in line:
                        cells = [p.strip() for p in line.split("\t") if p.strip()]
                    else:
                        cells = re.split(r"\s{2,}", line)
                        if len(cells) <= 1:
                            cells = line.split()
                    if cells:
                        rows.append(cells)
                except Exception:
                    continue

    return "\n".join(full_text_parts), rows


def extract_header(full_text):
    """Extract EFT header date (MM/DD/YYYY) and draft number."""
    header_data = {"eft_date": None, "draft_no": None}
    try:
        draft_match = DRAFT_NO_PATTERN.search(full_text)
        if draft_match:
            header_data["draft_no"] = draft_match.group(1).upper()

        date_match = EFT_DATE_LABEL_PATTERN.search(full_text)
        if not date_match:
            date_match = DATE_PATTERN.search(full_text)
        if date_match:
            header_data["eft_date"] = format_eft_date_us(date_match.group(1))
    except Exception:
        pass
    return header_data


def extract_eft_data(pdf_path):
    """
    Read EFT PDF and return (header_data, paid_invoices, credit_coupons, skipped_coupon_rows).

    `skipped_coupon_rows` counts rows that visibly look like a coupon line
    (an SI-/DDC- combo, per row_contains_credit_coupon) but that
    extract_coupon_columns still couldn't turn into a usable row (typically
    because it couldn't find 3+ legible currency tokens) -- previously these
    were dropped with zero tracking, so a partially-corrupted table silently
    lost coupons with no way to know without recounting the PDF by hand.
    Rows that simply aren't coupon rows at all (footer text, boilerplate)
    are NOT counted here, only ones that positively matched the coupon
    pattern and still failed to parse.
    """
    full_text, rows = collect_pdf_rows(pdf_path)
    header_data = extract_header(full_text)
    fallback_date = header_data.get("eft_date")

    paid_invoices = []
    credit_coupons = []
    seen_invoices = set()
    seen_coupons = set()
    coupon_section_started = False
    skipped_coupon_rows = 0

    for cells in rows:
        try:
            if row_is_credit_table_header(cells):
                coupon_section_started = True
                continue

            looks_like_coupon_row = row_contains_credit_coupon(cells)
            if looks_like_coupon_row:
                coupon_section_started = True

            if coupon_section_started:
                coupon_row = extract_coupon_columns(cells, fallback_date=fallback_date)
                if coupon_row:
                    key = (
                        coupon_row["date"],
                        coupon_row["invoice"],
                        coupon_row["coupon"],
                        coupon_row["gross_amount"],
                        coupon_row["fees_amount"],
                        coupon_row["paid_amount"],
                    )
                    if key not in seen_coupons:
                        seen_coupons.add(key)
                        credit_coupons.append(coupon_row)
                elif looks_like_coupon_row:
                    skipped_coupon_rows += 1
                continue

            for paid_entry in extract_paid_invoice_entries(cells):
                key = (paid_entry["invoice"], paid_entry["paid_amount"])
                if key not in seen_invoices:
                    seen_invoices.add(key)
                    paid_invoices.append(paid_entry)
        except Exception:
            continue

    return header_data, paid_invoices, credit_coupons, skipped_coupon_rows


def is_valid_historical_coupon_date(value):
    """True when column A holds a historical coupon date."""
    if value is None or value == "":
        return False
    if isinstance(value, datetime):
        return True
    text = str(value).strip()
    if COUPON_DATE_TEXT_PATTERN.match(text):
        return True
    if DATE_PATTERN.search(text):
        return True
    return False


def find_last_active_row_bottom_up(worksheet):
    """
    Scan column A from the bottom upward to find the last historical coupon date.

    Avoids mistaking empty middle rows for the ledger end.
    """
    scan_start = max(worksheet.max_row, BOTTOM_SCAN_START_ROW)

    for row in range(scan_start, LEDGER_START_ROW - 1, -1):
        cell_a = worksheet.cell(row=row, column=1).value
        if is_valid_historical_coupon_date(cell_a):
            return row

    raise ValueError(
        "No se encontró una fecha de cupón histórica en la columna A. "
        f"Se revisaron las filas {scan_start} hasta {LEDGER_START_ROW}."
    )


def calculate_middle_row(start_row, end_row):
    """Return the exact middle row of the inserted coupon block."""
    return start_row + ((end_row - start_row) // 2)


def build_balance_formula(middle_row, paid_invoice_count=1):
    """
    Build Column M balance: parent EFT amount minus the sum of all linked invoices.

    Invoice amounts live in Column L from middle_row downward (one per invoice).
    """
    count = max(1, paid_invoice_count)
    if count == 1:
        return f"=+J{middle_row}-L{middle_row}"
    last_invoice_row = middle_row + count - 1
    return f"=+J{middle_row}-SUM(L{middle_row}:L{last_invoice_row})"


def _write_paid_invoices_to_columns_kl(worksheet, middle_row, paid_invoices):
    """Write ordered paid invoices down Columns K and L on existing rows only."""
    for index, invoice_entry in enumerate(paid_invoices):
        row = middle_row + index
        cell_k = worksheet.cell(row=row, column=11, value=invoice_entry["invoice"])
        cell_l = worksheet.cell(
            row=row,
            column=12,
            value=float(invoice_entry["paid_amount"]),
        )
        apply_currency_format(cell_l)
        cell_k.font = BOLD_FONT
        cell_k.alignment = SUMMARY_ALIGNMENT
        cell_l.font = BOLD_FONT
        cell_l.alignment = SUMMARY_ALIGNMENT


def apply_currency_format(cell):
    """Apply regional currency display format to a numeric cell."""
    cell.number_format = CURRENCY_NUMBER_FORMAT


def clear_summary_columns(worksheet, start_row, end_row, middle_row):
    """Clear values in G-N on non-middle rows; column N stays borderless for native gridlines."""
    for row in range(start_row, end_row + 1):
        if row == middle_row:
            continue
        for col in range(SUMMARY_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = None
            cell.font = DEFAULT_FONT
            cell.number_format = "General"
        cell_n = worksheet.cell(row=row, column=COLUMN_N)
        cell_n.value = None
        cell_n.fill = NO_FILL
        cell_n.border = Border()


def apply_eft_block_styling(worksheet, start_row, end_row):
    """
    Two-step border grid for columns A-M (1-13). Column N is excluded entirely.
    """
    thin_all_sides = Border(
        top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE
    )

    # Loop 1: thin borders on all 4 sides of every A-M cell (soft internal gridlines)
    for row in range(start_row, end_row + 1):
        for col in range(BLOCK_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = CLEAR_FILL
            cell.font = DEFAULT_FONT
            cell.border = thin_all_sides

    # Loop 2: medium perimeter on outer bounds only (top, bottom, left A, right M)
    for row in range(start_row, end_row + 1):
        for col in range(BLOCK_FIRST_COLUMN, PERIMETER_LAST_COLUMN + 1):
            cell = worksheet.cell(row=row, column=col)
            sides = cell.border
            top = THICK_SIDE if row == start_row else sides.top
            bottom = THICK_SIDE if row == end_row else sides.bottom
            left = THICK_SIDE if col == BLOCK_FIRST_COLUMN else sides.left
            right = THICK_SIDE if col == PERIMETER_LAST_COLUMN else sides.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def apply_summary_row_formatting(worksheet, middle_row):
    """Bold and vertically center summary cells G through N on middle_row only."""
    currency_summary_cols = {8, 9, 10, 12, 13}
    for col in range(SUMMARY_FIRST_COLUMN, SUMMARY_LAST_COLUMN + 1):
        cell = worksheet.cell(row=middle_row, column=col)
        cell.font = BOLD_FONT
        cell.alignment = SUMMARY_ALIGNMENT
        if col in currency_summary_cols:
            apply_currency_format(cell)


def apply_column_n_date(worksheet, middle_row, eft_date_str):
    """Write EFT date in N{middle_row} — bold, dd-mmm display, no borders."""
    cell = worksheet.cell(row=middle_row, column=COLUMN_N)
    cell.font = BOLD_FONT
    cell.border = Border()
    cell.alignment = SUMMARY_ALIGNMENT

    if not eft_date_str:
        return

    eft_dt = parse_eft_pdf_date_us(eft_date_str)
    if eft_dt is None:
        eft_dt = parse_date_to_datetime(eft_date_str)
    if eft_dt:
        cell.value = eft_dt
    else:
        cell.value = eft_date_str
    cell.number_format = EFT_DATE_NUMBER_FORMAT


def _normalize_eft_rcv_number(value):
    """Extract canonical RCV-##### token from draft text or summary cell."""
    if not value:
        return None
    match = DRAFT_NO_PATTERN.search(str(value))
    if match:
        return match.group(1).upper()
    text = str(value).strip().upper()
    return text if text else None


def _normalize_eft_date_key(value):
    """Normalize any EFT date cell to a calendar date for duplicate checks."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    parsed = parse_eft_pdf_date_us(value)
    if parsed:
        return parsed.date()
    parsed = parse_date_to_datetime(value)
    if parsed:
        return parsed.date()
    return None


def _amounts_close(left, right, tolerance=0.01):
    try:
        return abs(float(left) - float(right)) <= tolerance
    except (TypeError, ValueError):
        return False


def _read_cta_summary_net_total(worksheet, row):
    """Read column J net total from cached value or SUM(Fx:Fy) formula text."""
    cell = worksheet.cell(row=row, column=10)
    value = cell.value
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        formula_match = SUM_FORMULA_RANGE_PATTERN.search(value.replace(" ", ""))
        if formula_match:
            start_row = int(formula_match.group(1))
            end_row = int(formula_match.group(2))
            total = 0.0
            for scan_row in range(start_row, end_row + 1):
                total += parse_amount(worksheet.cell(row=scan_row, column=6).value) or 0.0
            return total
        parsed = parse_amount(value)
        if parsed is not None:
            return parsed
    return None


def eft_already_loaded_in_workbook(excel_path, header_data, credit_coupons):
    """
    Return True when Cta Cte already contains the same EFT Nro., date, and net total.
    """
    draft_no = header_data.get("draft_no")
    eft_date_str = header_data.get("eft_date")
    if not draft_no or not eft_date_str or not credit_coupons:
        return False

    extension = os.path.splitext(str(excel_path))[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel Ledger debe ser .xlsx o .xlsm.")

    incoming_nro = _normalize_eft_rcv_number(draft_no)
    incoming_date = _normalize_eft_date_key(eft_date_str)
    incoming_net = sum(
        parse_amount(coupon.get("paid_amount")) or 0.0 for coupon in credit_coupons
    )
    if not incoming_nro or not incoming_date:
        return False

    for data_only in (True, False):
        workbook = load_workbook(excel_path, data_only=data_only)
        try:
            worksheet = get_worksheet(workbook)
            max_row = max(worksheet.max_row, LEDGER_START_ROW)
            for row in range(LEDGER_START_ROW, max_row + 1):
                summary_text = worksheet.cell(row=row, column=7).value
                if summary_text is None:
                    continue
                summary_upper = str(summary_text).upper()
                if "EFT" not in summary_upper and "RCV-" not in summary_upper:
                    continue

                existing_nro = _normalize_eft_rcv_number(summary_text)
                if existing_nro != incoming_nro:
                    continue

                existing_date = _normalize_eft_date_key(
                    worksheet.cell(row=row, column=COLUMN_N).value
                )
                if existing_date != incoming_date:
                    continue

                existing_net = _read_cta_summary_net_total(worksheet, row)
                if existing_net is not None and _amounts_close(
                    existing_net, incoming_net
                ):
                    return True
        finally:
            workbook.close()

    return False


def get_worksheet(workbook):
    """Return target worksheet, matching name with flexible spelling."""
    if SHEET_NAME in workbook.sheetnames:
        return workbook[SHEET_NAME]
    for name in workbook.sheetnames:
        if name.strip().lower() == SHEET_NAME.strip().lower():
            return workbook[name]
    # Tolera variaciones menores de tipeo en el nombre real de la hoja (ej.
    # "J.H.Williams" vs "J.H.Wiliams", con una sola "l") — nunca asumir que
    # el nombre va a quedar exactamente igual para siempre. Cutoff alto para
    # no matchear por error una hoja sin relación.
    close = difflib.get_close_matches(SHEET_NAME, workbook.sheetnames, n=1, cutoff=0.85)
    if close:
        return workbook[close[0]]
    raise ValueError(
        f'Hoja "{SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def update_excel_workbook(excel_path, header_data, paid_invoices, credit_coupons):
    """
    Insert coupon rows after the last historical entry and apply EFT block styling.

    Returns:
        tuple: (start_row, end_row)
    """
    if not credit_coupons:
        raise ValueError("No se extrajeron cupones de tarjeta de crédito del PDF.")

    extension = os.path.splitext(str(excel_path))[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel Ledger debe ser .xlsx o .xlsm.")

    workbook = load_workbook(excel_path, data_only=False)
    worksheet = get_worksheet(workbook)

    last_active_row = find_last_active_row_bottom_up(worksheet)
    target_row = last_active_row + 1
    paid_invoice_count = len(paid_invoices) if paid_invoices else 1

    # El bloque insertado tiene que tener lugar para escribir TODAS las
    # facturas pagadas en K/L a partir de middle_row -- si no, la escritura
    # sigue más allá de end_row y pisa en silencio filas del ledger que
    # insert_rows ya corrió hacia abajo, pero que no son parte de este
    # bloque (bug real encontrado en la auditoría de 2026-09-03: un EFT con
    # más facturas pagadas que cupones corrompía filas ajenas sin ningún
    # aviso). Con num_rows = 2*paid_invoice_count-1, el bloque queda
    # exactamente centrado en middle_row con lugar para las N facturas
    # debajo -- ver calculate_middle_row: rows desde middle_row hasta
    # end_row = num_rows - (num_rows-1)//2, que con este tamaño da
    # exactamente paid_invoice_count.
    min_rows_for_paid_invoices = 2 * paid_invoice_count - 1
    num_rows = max(len(credit_coupons), min_rows_for_paid_invoices)

    worksheet.insert_rows(target_row, amount=num_rows)

    start_row = target_row
    end_row = target_row + num_rows - 1
    middle_row = calculate_middle_row(start_row, end_row)
    assert end_row - middle_row + 1 >= paid_invoice_count, (
        "El bloque de EFT quedó más chico de lo necesario para las facturas pagadas -- "
        "esto sería un bug en el cálculo de tamaño de arriba, no un dato del usuario."
    )
    draft_no = header_data.get("draft_no") or "UNKNOWN"
    eft_date_str = header_data.get("eft_date")

    for offset, coupon in enumerate(credit_coupons):
        row = start_row + offset
        coupon_date = coupon.get("date")
        if coupon_date:
            worksheet.cell(
                row=row, column=1, value=format_coupon_date_us(coupon_date)
            )

        worksheet.cell(row=row, column=2, value=coupon["invoice"])
        worksheet.cell(row=row, column=3, value=coupon["coupon"])

        gross_val = float(coupon["gross_amount"])
        fees_val = float(coupon.get("fees_amount") or 0.0)
        paid_val = float(coupon.get("paid_amount") or 0.0)

        cell_d = worksheet.cell(row=row, column=4, value=gross_val)
        cell_e = worksheet.cell(row=row, column=5, value=fees_val)
        cell_f = worksheet.cell(row=row, column=6, value=paid_val)
        for cell in (cell_d, cell_e, cell_f):
            apply_currency_format(cell)

    worksheet.cell(row=middle_row, column=7, value=f"EFT  {draft_no}")
    cell_h = worksheet.cell(
        row=middle_row, column=8, value=f"=SUM(D{start_row}:D{end_row})"
    )
    cell_i = worksheet.cell(
        row=middle_row, column=9, value=f"=SUM(E{start_row}:E{end_row})"
    )
    cell_j = worksheet.cell(
        row=middle_row, column=10, value=f"=SUM(F{start_row}:F{end_row})"
    )
    for cell in (cell_h, cell_i, cell_j):
        apply_currency_format(cell)

    balance_formula = build_balance_formula(middle_row, paid_invoice_count)
    cell_m = worksheet.cell(row=middle_row, column=13, value=balance_formula)
    apply_currency_format(cell_m)

    clear_summary_columns(worksheet, start_row, end_row, middle_row)
    apply_eft_block_styling(worksheet, start_row, end_row)
    apply_summary_row_formatting(worksheet, middle_row)
    _write_paid_invoices_to_columns_kl(worksheet, middle_row, paid_invoices)
    apply_column_n_date(worksheet, middle_row, eft_date_str)

    workbook.save(excel_path)
    return start_row, end_row, middle_row
