"""Append monthly J.H. Williams coupon reports to the Cupones worksheet."""

import gc
import os
import re
import sys
import tempfile
import time
from datetime import datetime

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]
    MergedCell = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False

MONTHLY_HEADER_ROW = 2
MONTHLY_DATA_START_ROW = 3

CUPONES_SHEET = "Cupones"
CTA_CTE_SHEET = "Cta Cte J.H.Williams"

CUPONES_COL_DATE = 1
CUPONES_COL_COUPON = 2
CUPONES_COL_GROSS = 3
CUPONES_COL_FEES = 4
CUPONES_COL_NET = 5
CUPONES_COL_EFT_FORMULA = 6
CUPONES_COL_DIF = 7
CUPONES_COL_EFT_NO = 8
CUPONES_COL_MES_EFT = 9

CTA_COL_INVOICE = 3
CTA_COL_GROSS = 4
CTA_COL_FEES = 5
CTA_COL_NET = 6
CTA_COL_NRO = 7
CTA_COL_NETO_FINAL = 14

CTA_SCAN_START_ROW = 5
CUPONES_SCAN_START_ROW = 2
DECIMAL_NUMBER_FORMAT = "0.00"
EFT_DATE_NUMBER_FORMAT = "dd-mmm"
COUPON_DATE_TEXT_PATTERN = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
DDC_COUPON_PATTERN = re.compile(r"\b(DDC-\d+)\b", re.IGNORECASE)
RCV_PATTERN = re.compile(r"\b(RCV-\d+)\b", re.IGNORECASE)

LEFT_ALIGNMENT = Alignment(horizontal="left") if Alignment is not None else None
RIGHT_ALIGNMENT = Alignment(horizontal="right") if Alignment is not None else None
PRIMARY_SPLIT_RED_FONT = Font(color="FF0000") if Font is not None else None
SPLIT_GROUP_FILL_COLORS = (
    "FFFDE9D9",  # peach
    "FFDCE6F1",  # blue
    "FFE2EFDA",  # green
    "FFFFF2CC",  # yellow
    "FFF2DCDB",  # rose
    "FFD9D2E9",  # purple
    "FFD8E4BC",  # olive
    "FFFCE4D6",  # orange
    "FFDDEBF7",  # pale blue
    "FFEAD1DC",  # pale pink
    "FFD0E0E3",  # pale teal
    "FFFCE9DB",  # cream
)
CUPONES_LEFT_COLUMNS = (
    CUPONES_COL_DATE,
    CUPONES_COL_COUPON,
    CUPONES_COL_EFT_NO,
)
CUPONES_RIGHT_COLUMNS = (
    CUPONES_COL_GROSS,
    CUPONES_COL_FEES,
    CUPONES_COL_NET,
    CUPONES_COL_EFT_FORMULA,
    CUPONES_COL_DIF,
    CUPONES_COL_MES_EFT,
)

MONTHLY_REPORT_FULL_DUPLICATE_ALERT = (
    "Alerta: Este reporte mensual ya fue cargado en su totalidad anteriormente."
)
NO_PENDING_COUPONS_ALERT = (
    "No hay cupones pendientes para actualizar: todo lo que hay en Cupones ya "
    "está completamente reconciliado."
)
AMOUNT_MATCH_TOLERANCE = 0.01


class MonthlyReportFullyDuplicateError(Exception):
    """Raised when every incoming coupon row already exists on Cupones."""


class NoPendingCouponsError(Exception):
    """Raised by resync_cupones_only when there's nothing left to update."""


def _ensure_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "Agregar cupones requiere openpyxl. Instale con: pip install openpyxl"
        )


def _cell_for_write(worksheet, row, column):
    """Return the writable top-left cell when the target is inside a merge."""
    cell = worksheet.cell(row=row, column=column)
    if MergedCell is not None and isinstance(cell, MergedCell):
        for merged in worksheet.merged_cells.ranges:
            if (
                merged.min_row <= row <= merged.max_row
                and merged.min_col <= column <= merged.max_col
            ):
                return worksheet.cell(row=merged.min_row, column=merged.min_col)
    return cell


def _coerce_primitive_cell_value(value):
    """Return the underlying cell value, never an openpyxl cell object."""
    if value is None:
        return None
    type_name = type(value).__name__
    if type_name in {"ReadOnlyCell", "Cell", "MergedCell"}:
        return value.value
    if hasattr(value, "value") and not isinstance(
        value, (str, bytes, int, float, bool, datetime)
    ):
        try:
            return value.value
        except Exception:
            return value
    return value


def _safe_string_from_value(value):
    value = _coerce_primitive_cell_value(value)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    return str(value).strip()


def _strip_cell(value):
    value = _coerce_primitive_cell_value(value)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    return str(value).strip()


def _parse_amount(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _strip_cell(value).replace("$", "").replace(",", "").strip()
    if not text:
        return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_date_to_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = _strip_cell(value)
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def _format_coupon_date_us(value):
    parsed = _parse_date_to_datetime(value)
    if parsed:
        return parsed.strftime("%m/%d/%Y")
    text = _strip_cell(value)
    return text if text else None


def _get_sheet(workbook, target_name):
    if target_name in workbook.sheetnames:
        return workbook[target_name]
    lowered = target_name.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == lowered:
            return workbook[name]
    raise ValueError(
        f'Hoja "{target_name}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _create_temp_workbook_path(suffix=".xlsx"):
    fd, temp_path = tempfile.mkstemp(suffix=suffix, prefix="cupones_append_")
    os.close(fd)
    return temp_path


def _launch_temp_workbook(temp_path):
    abs_path = os.path.abspath(temp_path)
    if sys.platform == "win32":
        time.sleep(0.35)
        os.startfile(abs_path)
    elif sys.platform == "darwin":
        os.system(f'open "{abs_path}"')
    else:
        os.system(f'xdg-open "{abs_path}"')


def _apply_decimal_format(worksheet, row, column, value):
    cell = _cell_for_write(worksheet, row, column)
    cell.value = float(value)
    cell.number_format = DECIMAL_NUMBER_FORMAT


def _record_sort_key(record):
    parsed = _parse_date_to_datetime(record.get("date"))
    if parsed:
        return parsed
    return datetime.max


def _sort_records_by_date(records):
    return sorted(records, key=_record_sort_key)


def _expand_records_by_coupon_split(records):
    """
    Split comma-separated Coupon IDs into one record per DDC code.

    First split row keeps report amounts; subsequent split rows are zeroed.
    """
    expanded = []
    split_group_counter = 0
    for record in records:
        coupon_ids = _split_coupon_ids(record.get("coupon"))
        if not coupon_ids:
            coupon_text = _strip_cell(record.get("coupon"))
            if coupon_text:
                expanded.append(
                    {
                        "date": record["date"],
                        "coupon": coupon_text,
                        "gross": 0.0,
                        "fees": 0.0,
                        "net": 0.0,
                        "is_primary_split": False,
                        "split_group_id": None,
                        "split_index": 0,
                    }
                )
            continue
        split_group_id = None
        if len(coupon_ids) > 1:
            split_group_counter += 1
            split_group_id = f"group_{split_group_counter}"
        is_multi_coupon = len(coupon_ids) > 1
        for idx, coupon_id in enumerate(coupon_ids):
            expanded.append(
                {
                    "date": record["date"],
                    "coupon": coupon_id,
                    "gross": _parse_amount(record.get("gross", 0)) if idx == 0 else 0.0,
                    "fees": _parse_amount(record.get("fees", 0)) if idx == 0 else 0.0,
                    "net": _parse_amount(record.get("net", 0)) if idx == 0 else 0.0,
                    "is_primary_split": is_multi_coupon and idx == 0,
                    "split_group_id": split_group_id,
                    "split_index": idx,
                }
            )
    return expanded


def _read_cta_financial_breakdown(worksheet, row):
    """
    Read Gross (D), Disc/Fees (E), and Net Amt (F) from one Cta Cte coupon row.
    """
    return (
        _parse_amount(worksheet.cell(row=row, column=CTA_COL_GROSS).value),
        _parse_amount(worksheet.cell(row=row, column=CTA_COL_FEES).value),
        _parse_amount(worksheet.cell(row=row, column=CTA_COL_NET).value),
    )


def _resolve_single_coupon_id(coupon_text):
    """Return the one Coupon ID for the current Cupones row (Column B)."""
    coupon_ids = _split_coupon_ids(coupon_text)
    if coupon_ids:
        return coupon_ids[0]
    text = _strip_cell(coupon_text).upper()
    return text if text else None


def _write_cupones_date_cell(worksheet, row, record):
    """Write Column A as Excel datetime (mm/dd/yyyy) or normalized US date text."""
    cell_a = _cell_for_write(worksheet, row, CUPONES_COL_DATE)
    raw = record.get("date")
    parsed = raw if isinstance(raw, datetime) else _parse_date_to_datetime(raw)
    if parsed:
        cell_a.value = parsed
        cell_a.number_format = "mm/dd/yyyy"
        return

    formatted = _format_coupon_date_us(raw)
    if formatted:
        retry = _parse_date_to_datetime(formatted)
        if retry:
            cell_a.value = retry
        else:
            cell_a.value = formatted
        cell_a.number_format = "mm/dd/yyyy"


def _write_cupones_base_columns(worksheet, row, record):
    _write_cupones_date_cell(worksheet, row, record)

    coupon_text = _strip_cell(record["coupon"])
    _cell_for_write(worksheet, row, CUPONES_COL_COUPON).value = coupon_text

    _apply_decimal_format(worksheet, row, CUPONES_COL_GROSS, record["gross"])
    _apply_decimal_format(worksheet, row, CUPONES_COL_FEES, record["fees"])
    _apply_decimal_format(worksheet, row, CUPONES_COL_NET, record["net"])
    return coupon_text


def _clear_cupones_control_columns(worksheet, row):
    """Initialize F-I control columns before Cta Cte cross-reference."""
    cell_f = _cell_for_write(worksheet, row, CUPONES_COL_EFT_FORMULA)
    cell_f.value = ""
    cell_f.alignment = RIGHT_ALIGNMENT
    cell_f.number_format = DECIMAL_NUMBER_FORMAT

    cell_g = _cell_for_write(worksheet, row, CUPONES_COL_DIF)
    cell_g.value = f"=E{row}-F{row}"
    cell_g.alignment = RIGHT_ALIGNMENT
    cell_g.number_format = DECIMAL_NUMBER_FORMAT

    cell_h = _cell_for_write(worksheet, row, CUPONES_COL_EFT_NO)
    cell_h.value = ""
    cell_h.alignment = LEFT_ALIGNMENT

    cell_i = _cell_for_write(worksheet, row, CUPONES_COL_MES_EFT)
    cell_i.value = ""
    cell_i.alignment = RIGHT_ALIGNMENT


def _amounts_close(left, right, tolerance=AMOUNT_MATCH_TOLERANCE):
    try:
        return abs(float(left) - float(right)) <= tolerance
    except (TypeError, ValueError):
        return False


def _record_amount_tuple(record):
    return (
        _parse_amount(record.get("gross")),
        _parse_amount(record.get("fees")),
        _parse_amount(record.get("net")),
    )


def _build_existing_cupones_entries(worksheet):
    """Index existing Cupones coupon IDs with their C/D/E metrics."""
    entries = []
    max_row = max(worksheet.max_row, CUPONES_SCAN_START_ROW)
    for row in range(CUPONES_SCAN_START_ROW, max_row + 1):
        coupon_text = _strip_cell(worksheet.cell(row=row, column=CUPONES_COL_COUPON).value)
        if not coupon_text:
            continue
        coupon_ids = _split_coupon_ids(coupon_text)
        if not coupon_ids:
            single = _strip_cell(coupon_text).upper()
            if single:
                coupon_ids = [single]
        metrics = (
            _parse_amount(worksheet.cell(row=row, column=CUPONES_COL_GROSS).value),
            _parse_amount(worksheet.cell(row=row, column=CUPONES_COL_FEES).value),
            _parse_amount(worksheet.cell(row=row, column=CUPONES_COL_NET).value),
        )
        for coupon_id in coupon_ids:
            entries.append({"coupon_id": coupon_id, "metrics": metrics})
    return entries


def _cupones_record_is_duplicate(existing_entries, record):
    """
    True when coupon ID already exists on Cupones with matching metrics.

    Split child rows that arrive as 0.00 are treated as duplicates when the
    coupon ID is already present on the sheet.
    """
    coupon_id = _strip_cell(record.get("coupon")).upper()
    if not coupon_id:
        return False

    incoming_metrics = _record_amount_tuple(record)
    incoming_all_zero = all(abs(value) < 1e-9 for value in incoming_metrics)

    for entry in existing_entries:
        if entry["coupon_id"] != coupon_id:
            continue
        if incoming_all_zero:
            return True
        if all(
            _amounts_close(incoming_value, existing_value)
            for incoming_value, existing_value in zip(
                incoming_metrics, entry["metrics"]
            )
        ):
            return True
    return False


def _filter_new_cupones_records(cupones_sheet, monthly_rows):
    """Return only incoming rows that are not already present on Cupones."""
    existing_entries = _build_existing_cupones_entries(cupones_sheet)
    filtered = []
    skipped = 0
    for record in monthly_rows:
        if _cupones_record_is_duplicate(existing_entries, record):
            skipped += 1
            continue
        filtered.append(record)
    return filtered, skipped


def _is_zero_amount(value):
    try:
        return abs(float(value)) < 1e-9
    except (TypeError, ValueError):
        return False


def _row_amounts_are_zero(cupones_sheet, row):
    gross = cupones_sheet.cell(row=row, column=CUPONES_COL_GROSS).value
    fees = cupones_sheet.cell(row=row, column=CUPONES_COL_FEES).value
    net = cupones_sheet.cell(row=row, column=CUPONES_COL_NET).value
    return _is_zero_amount(gross) and _is_zero_amount(fees) and _is_zero_amount(net)


def _aggregate_cta_totals_for_coupon_id(cta_sheet, coupon_id):
    """Sum all Cta Cte D/E/F amounts for one coupon ID (scan row 2..max)."""
    target = _strip_cell(coupon_id).upper()
    if not target:
        return 0.0, 0.0, 0.0

    total_gross = 0.0
    total_fees = 0.0
    total_net = 0.0
    max_row = max(cta_sheet.max_row, 2)

    for scan_row in range(2, max_row + 1):
        invoice_text = cta_sheet.cell(row=scan_row, column=CTA_COL_INVOICE).value
        row_coupon_ids = _extract_ddc_ids(invoice_text)
        if target not in row_coupon_ids:
            continue
        gross, fees, net = _read_cta_financial_breakdown(cta_sheet, scan_row)
        total_gross += float(gross or 0.0)
        total_fees += float(fees or 0.0)
        total_net += float(net or 0.0)

    return total_gross, total_fees, total_net


def _aggregate_cta_totals_for_coupon_text(cta_sheet, coupon_text):
    """Aggregate totals for all coupon IDs present in one Cupones coupon cell."""
    coupon_ids = _split_coupon_ids(coupon_text)
    if not coupon_ids:
        single = _resolve_single_coupon_id(coupon_text)
        coupon_ids = [single] if single else []

    total_gross = 0.0
    total_fees = 0.0
    total_net = 0.0
    for coupon_id in dict.fromkeys(coupon_ids):
        gross, fees, net = _aggregate_cta_totals_for_coupon_id(cta_sheet, coupon_id)
        total_gross += gross
        total_fees += fees
        total_net += net
    return total_gross, total_fees, total_net


def _fill_zero_amounts_from_cta(cupones_sheet, cta_sheet, row, coupon_text):
    """Inject static C/D/E floats from aggregated Cta Cte matches when C/D/E are zero."""
    if not _row_amounts_are_zero(cupones_sheet, row):
        return
    gross, fees, net = _aggregate_cta_totals_for_coupon_text(cta_sheet, coupon_text)
    _apply_decimal_format(cupones_sheet, row, CUPONES_COL_GROSS, gross)
    _apply_decimal_format(cupones_sheet, row, CUPONES_COL_FEES, fees)
    _apply_decimal_format(cupones_sheet, row, CUPONES_COL_NET, net)


def _highlight_primary_split_amounts_red(cupones_sheet, row):
    if PRIMARY_SPLIT_RED_FONT is None:
        return
    for col in (CUPONES_COL_GROSS, CUPONES_COL_FEES, CUPONES_COL_NET):
        _cell_for_write(cupones_sheet, row, col).font = PRIMARY_SPLIT_RED_FONT


def _apply_child_rows_standard_black(cupones_sheet, row):
    if Font is None:
        return
    black_font = Font(color="000000")
    for col in (CUPONES_COL_GROSS, CUPONES_COL_FEES, CUPONES_COL_NET):
        _cell_for_write(cupones_sheet, row, col).font = black_font


def _read_float_cell(cupones_sheet, row, column):
    return _parse_amount(cupones_sheet.cell(row=row, column=column).value)


def _apply_parent_child_subtractions(cupones_sheet, split_group_rows):
    """
    Phase 2: subtract all child C/D/E totals from each split-group parent row.
    """
    for rows in split_group_rows.values():
        if len(rows) <= 1:
            continue
        parent_row = rows[0]
        child_rows = rows[1:]

        parent_gross = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_GROSS)
        parent_fees = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_FEES)
        parent_net = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_NET)

        child_gross = sum(
            _read_float_cell(cupones_sheet, child_row, CUPONES_COL_GROSS)
            for child_row in child_rows
        )
        child_fees = sum(
            _read_float_cell(cupones_sheet, child_row, CUPONES_COL_FEES)
            for child_row in child_rows
        )
        child_net = sum(
            _read_float_cell(cupones_sheet, child_row, CUPONES_COL_NET)
            for child_row in child_rows
        )

        _apply_decimal_format(
            cupones_sheet, parent_row, CUPONES_COL_GROSS, parent_gross - child_gross
        )
        _apply_decimal_format(
            cupones_sheet, parent_row, CUPONES_COL_FEES, parent_fees - child_fees
        )
        _apply_decimal_format(cupones_sheet, parent_row, CUPONES_COL_NET, parent_net - child_net)
        _highlight_primary_split_amounts_red(cupones_sheet, parent_row)


def _apply_split_group_colors(cupones_sheet, split_group_rows):
    """
    Fill the Coupon/Reference cell (Column B) of every row in a split group
    with a color unique to that group, cycling through a fixed palette so
    consecutive batches never share the same color.
    """
    if PatternFill is None:
        return
    for color_index, rows in enumerate(split_group_rows.values()):
        if len(rows) <= 1:
            continue
        color = SPLIT_GROUP_FILL_COLORS[color_index % len(SPLIT_GROUP_FILL_COLORS)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in rows:
            _cell_for_write(cupones_sheet, row, CUPONES_COL_COUPON).fill = fill


_PENDING_PLACEHOLDER_VALUES = {"RCV-MISSING", "MES-MISSING"}


def _cupones_row_is_fully_reconciled(cupones_sheet, row):
    """
    True only when F (EFT formula), H (EFT No.) and I (Mes EFT) are all
    already filled with a real value. These three are only ever written
    together once a matching Cta Cte invoice is found, so this marks a row
    as fully done and untouchable. "RCV-MISSING"/"MES-MISSING" placeholders
    (written when the invoice exists but isn't inside any EFT box yet) do
    NOT count as reconciled, so the row keeps getting retried until a real
    EFT box shows up around it.
    """
    f_value = _strip_cell(cupones_sheet.cell(row=row, column=CUPONES_COL_EFT_FORMULA).value)
    h_value = _strip_cell(cupones_sheet.cell(row=row, column=CUPONES_COL_EFT_NO).value)
    i_value = _strip_cell(cupones_sheet.cell(row=row, column=CUPONES_COL_MES_EFT).value)
    if h_value.upper() in _PENDING_PLACEHOLDER_VALUES or i_value.upper() in _PENDING_PLACEHOLDER_VALUES:
        return False
    return bool(f_value) and bool(h_value) and bool(i_value)


def _cell_fill_color_key(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    rgb = getattr(fill.start_color, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def _reconstruct_split_groups_by_fill(cupones_sheet, start_row=CUPONES_SCAN_START_ROW):
    """
    Rebuild split-group row membership from the per-batch fill color applied
    by _apply_split_group_colors, by collapsing consecutive rows that share
    the exact same solid fill color into one group (first row = parent).

    This lets a later resync find a split group's parent even for groups
    written in a previous run, since the fill color is the only persisted
    marker of "these rows came from the same J.H. Williams batch."
    """
    max_row = max(cupones_sheet.max_row, start_row)
    groups = []
    current_key = None
    current_rows = []
    for row in range(start_row, max_row + 1):
        coupon_text = _strip_cell(cupones_sheet.cell(row=row, column=CUPONES_COL_COUPON).value)
        key = (
            _cell_fill_color_key(cupones_sheet.cell(row=row, column=CUPONES_COL_COUPON))
            if coupon_text
            else None
        )
        if key is not None and key == current_key:
            current_rows.append(row)
            continue
        if current_key is not None and len(current_rows) > 1:
            groups.append(current_rows)
        current_key = key
        current_rows = [row] if key is not None else []
    if current_key is not None and len(current_rows) > 1:
        groups.append(current_rows)
    return groups


def _resync_pending_cupones_rows(cupones_sheet, cta_sheet, coupon_index):
    """
    Revisit every Cupones row that is not fully reconciled (F/H/I incomplete)
    and retry the Cta Cte cross-reference, so coupons that were pending when
    first pasted (or split-children that started at $0.00) get filled in once
    a matching Cta Cte entry shows up in a later EFT load. Rows already fully
    reconciled (F, H and I all filled) are left completely untouched.

    When a split-group child gets filled in this way, the exact amount just
    discovered is subtracted from that group's parent row (identified via
    its shared fill color), so the parent always ends up holding only the
    portion of the original batch total that hasn't been individually
    matched yet. This step runs even if the parent row itself already looks
    "reconciled," since a parent's amount can still change as siblings
    resolve later.
    """
    resynced = 0
    row_to_parent = {}
    for group in _reconstruct_split_groups_by_fill(cupones_sheet):
        parent_row = group[0]
        for child_row in group[1:]:
            row_to_parent[child_row] = parent_row

    cta_eft_boxes = build_cta_eft_boxes(cta_sheet)
    max_row = max(cupones_sheet.max_row, CUPONES_SCAN_START_ROW)
    for row in range(CUPONES_SCAN_START_ROW, max_row + 1):
        coupon_text = _strip_cell(cupones_sheet.cell(row=row, column=CUPONES_COL_COUPON).value)
        if not coupon_text:
            continue
        if _cupones_row_is_fully_reconciled(cupones_sheet, row):
            continue

        before = (
            _read_float_cell(cupones_sheet, row, CUPONES_COL_GROSS),
            _read_float_cell(cupones_sheet, row, CUPONES_COL_FEES),
            _read_float_cell(cupones_sheet, row, CUPONES_COL_NET),
        )
        if _apply_cta_truth_to_cupones_row(cupones_sheet, cta_sheet, row, coupon_text, coupon_index, cta_eft_boxes):
            resynced += 1
        _apply_cupones_row_alignment(cupones_sheet, row)

        parent_row = row_to_parent.get(row)
        if parent_row is None:
            continue
        after = (
            _read_float_cell(cupones_sheet, row, CUPONES_COL_GROSS),
            _read_float_cell(cupones_sheet, row, CUPONES_COL_FEES),
            _read_float_cell(cupones_sheet, row, CUPONES_COL_NET),
        )
        delta_gross, delta_fees, delta_net = (a - b for a, b in zip(after, before))
        if abs(delta_gross) < 1e-9 and abs(delta_fees) < 1e-9 and abs(delta_net) < 1e-9:
            continue

        parent_gross = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_GROSS)
        parent_fees = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_FEES)
        parent_net = _read_float_cell(cupones_sheet, parent_row, CUPONES_COL_NET)
        _apply_decimal_format(cupones_sheet, parent_row, CUPONES_COL_GROSS, parent_gross - delta_gross)
        _apply_decimal_format(cupones_sheet, parent_row, CUPONES_COL_FEES, parent_fees - delta_fees)
        _apply_decimal_format(cupones_sheet, parent_row, CUPONES_COL_NET, parent_net - delta_net)
        _highlight_primary_split_amounts_red(cupones_sheet, parent_row)

    return resynced


def _apply_cta_truth_to_cupones_row(
    cupones_sheet, cta_sheet, row, coupon_text, coupon_index, cta_eft_boxes
):
    """
    Apply cross-reference controls for one Cupones row from matched Cta Cte row.

    F/G formulas; H/I <- the RCV number and applied date found anywhere
    inside the matched invoice's EFT box (see build_cta_eft_boxes).
    """
    matched_rows = _find_matching_cta_rows(coupon_text, coupon_index)
    if not matched_rows:
        return False

    formula = _build_eft_formula(CTA_CTE_SHEET, matched_rows)
    if formula:
        _cell_for_write(cupones_sheet, row, CUPONES_COL_EFT_FORMULA).value = formula
    matched_cta_row = matched_rows[0]
    _fill_zero_amounts_from_cta(cupones_sheet, cta_sheet, row, coupon_text)

    box = _find_box_for_row(cta_eft_boxes, matched_cta_row)
    nro_value, mes_eft = _find_eft_marker_in_box(cta_sheet, box)
    if not nro_value:
        nro_value = "RCV-MISSING"
    _cell_for_write(cupones_sheet, row, CUPONES_COL_EFT_NO).value = nro_value

    if mes_eft:
        cell_i = _cell_for_write(cupones_sheet, row, CUPONES_COL_MES_EFT)
        cell_i.value = mes_eft
        cell_i.number_format = EFT_DATE_NUMBER_FORMAT
    else:
        _cell_for_write(cupones_sheet, row, CUPONES_COL_MES_EFT).value = "MES-MISSING"

    return True


def _apply_cupones_row_alignment(worksheet, row):
    if Alignment is None:
        return
    for col in CUPONES_LEFT_COLUMNS:
        _cell_for_write(worksheet, row, col).alignment = LEFT_ALIGNMENT
    for col in CUPONES_RIGHT_COLUMNS:
        _cell_for_write(worksheet, row, col).alignment = RIGHT_ALIGNMENT


def _stretch_column_b_width(worksheet, max_text_length):
    if get_column_letter is None or max_text_length <= 0:
        return
    letter = get_column_letter(CUPONES_COL_COUPON)
    target_width = float(max_text_length + 3)
    current = worksheet.column_dimensions[letter].width
    if current is None or target_width > current:
        worksheet.column_dimensions[letter].width = target_width


def find_last_cupones_row(worksheet, start_row=CUPONES_SCAN_START_ROW):
    """
    Return the true last occupied row scanning bottom-up.

    Checks column A (date) first, but falls back to column B (coupon) so a
    row whose date failed to parse and was left blank still counts as
    occupied -- otherwise the next append starts one row too early and
    overwrites that row's already-written coupon/gross/fees/net data.
    """
    max_row = max(worksheet.max_row, start_row)
    for row in range(max_row, start_row - 1, -1):
        value_a = worksheet.cell(row=row, column=CUPONES_COL_DATE).value
        value_b = worksheet.cell(row=row, column=CUPONES_COL_COUPON).value
        if (value_a is not None and _strip_cell(value_a) != "") or (
            value_b is not None and _strip_cell(value_b) != ""
        ):
            return row
    return start_row - 1


def _extract_ddc_ids(text):
    if not text:
        return []
    return [match.group(1).upper() for match in DDC_COUPON_PATTERN.finditer(str(text))]


def _split_coupon_ids(value):
    text = _strip_cell(value)
    if not text:
        return []
    parts = [part.strip().upper() for part in text.split(",")]
    coupons = []
    for part in parts:
        if not part:
            continue
        if DDC_COUPON_PATTERN.fullmatch(part):
            coupons.append(part.upper())
            continue
        coupons.extend(_extract_ddc_ids(part))
    unique = []
    seen = set()
    for coupon in coupons:
        if coupon and coupon not in seen:
            seen.add(coupon)
            unique.append(coupon)
    return unique


def build_cta_coupon_index(worksheet, start_row=CTA_SCAN_START_ROW):
    """
    Map DDC coupon IDs to Cta Cte row numbers by scanning column C (Invoice/Coupon).
    """
    index = {}
    max_row = max(worksheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        cell_value = worksheet.cell(row=row, column=CTA_COL_INVOICE).value
        for coupon_id in _extract_ddc_ids(cell_value):
            index.setdefault(coupon_id, []).append(row)
    for coupon_id in index:
        index[coupon_id] = sorted(set(index[coupon_id]))
    return index


CTA_BOX_BORDER_COLUMN = 1  # Column A carries the hand-drawn EFT box border.
_BOX_BORDER_STYLES = {"medium", "thick"}


def _cta_border_style(cell, side):
    edge = getattr(cell.border, side, None)
    return edge.style if edge is not None else None


def build_cta_eft_boxes(worksheet):
    """
    Reconstruct the EFT "boxes" drawn by hand in Cta Cte: a medium/thick
    bordered rectangle enclosing every invoice row that one EFT payment
    settled. The RCV number and applied date can be annotated on ANY row
    inside that box (often the row a specific invoice was matched against),
    not necessarily at the box's top or bottom edge — so the only reliable
    way to attribute an invoice row to its EFT is to find the box it falls
    inside (by border), then look for the annotation anywhere within that
    same box. A plain "nearest annotated row" search in either direction
    can cross into a neighboring EFT's box and misattribute the coupon.

    Returns a sorted list of (start_row, end_row) tuples.
    """
    boxes = []
    box_start = None
    max_row = worksheet.max_row
    for row in range(1, max_row + 1):
        cell = worksheet.cell(row=row, column=CTA_BOX_BORDER_COLUMN)
        if _cta_border_style(cell, "top") in _BOX_BORDER_STYLES and box_start is None:
            box_start = row
        if _cta_border_style(cell, "bottom") in _BOX_BORDER_STYLES and box_start is not None:
            boxes.append((box_start, row))
            box_start = None
    return boxes


def _find_box_for_row(cta_eft_boxes, row):
    for start, end in cta_eft_boxes:
        if start <= row <= end:
            return start, end
    return None


def _find_eft_marker_in_box(worksheet, box):
    """Return (rcv_text, applied_date) found anywhere inside one EFT box."""
    if box is None:
        return None, None
    start, end = box
    rcv_value = None
    date_value = None
    for scan_row in range(start, end + 1):
        if rcv_value is None:
            text = _strip_cell(worksheet.cell(row=scan_row, column=CTA_COL_NRO).value)
            match = RCV_PATTERN.search(text)
            if match:
                rcv_value = match.group(1).upper()
        if date_value is None:
            value = worksheet.cell(row=scan_row, column=CTA_COL_NETO_FINAL).value
            if isinstance(value, datetime):
                date_value = value
            elif value not in (None, ""):
                parsed = _parse_date_to_datetime(value)
                if parsed:
                    date_value = parsed
        if rcv_value is not None and date_value is not None:
            break
    return rcv_value, date_value


def _build_eft_formula(sheet_name, cta_rows):
    if not cta_rows:
        return None
    parts = [f"+'{sheet_name}'!F{cta_row}" for cta_row in cta_rows]
    return "=" + "".join(parts)


def _find_matching_cta_rows(coupon_text, coupon_index):
    """Collect all matching Cta Cte rows for one Cupones coupon cell."""
    coupon_ids = _split_coupon_ids(coupon_text)
    if not coupon_ids:
        single = _resolve_single_coupon_id(coupon_text)
        coupon_ids = [single] if single else []

    seen = set()
    matches = []
    for coupon_id in coupon_ids:
        for row in coupon_index.get(coupon_id, []):
            if row in seen:
                continue
            seen.add(row)
            matches.append(row)
    return matches


def _is_batch_header(label):
    token = re.sub(r"\s+", " ", _strip_cell(label)).upper()
    return "BATCH" in token


def _read_monthly_excel_row_values(worksheet, row_number):
    """Read one report row via ws.cell(...).value (never raw cell objects)."""
    max_col = max(worksheet.max_column or 0, 12)
    values = []
    for col in range(1, max_col + 1):
        raw = _coerce_primitive_cell_value(
            worksheet.cell(row=row_number, column=col).value
        )
        if isinstance(raw, datetime):
            values.append(raw)
        elif isinstance(raw, (int, float)):
            values.append(raw)
        else:
            values.append(_safe_string_from_value(raw))

    while values:
        tail = values[-1]
        if isinstance(tail, (datetime, int, float)):
            break
        if _strip_cell(tail):
            break
        values.pop()
    return values


def _find_batch_column_index(header_cells):
    batch_idx = None
    try:
        for idx, label in enumerate(header_cells):
            if _is_batch_header(label):
                batch_idx = idx
                break
    except Exception:
        batch_idx = None
    return batch_idx


def _resolve_batch_column_index(header_cells):
    """
    Drop J.H. Williams \"Batch No(s)\" column (typically source column B / index 1).
    """
    batch_idx = _find_batch_column_index(header_cells)
    if batch_idx is None and len(header_cells) > 1:
        batch_idx = 1
    return batch_idx


def _build_keep_indices(header_cells, batch_idx):
    try:
        return [idx for idx in range(len(header_cells)) if idx != batch_idx]
    except Exception:
        return list(range(len(header_cells)))


def _record_from_filtered_row(filtered):
    if len(filtered) < 2:
        return None
    coupon_raw = _strip_cell(filtered[1] if len(filtered) > 1 else "")
    if not coupon_raw:
        return None
    if coupon_raw.lower() in {"coupon", "cupon", "cupón", "coupon id", "id"}:
        return None
    return {
        "date": filtered[0] if filtered else "",
        "coupon": coupon_raw,
        "gross": _parse_amount(filtered[2] if len(filtered) > 2 else 0),
        "fees": _parse_amount(filtered[3] if len(filtered) > 3 else 0),
        "net": _parse_amount(filtered[4] if len(filtered) > 4 else 0),
    }


def _read_monthly_coupon_rows_from_table(header_cells, data_rows):
    """
    Row 1 skipped upstream; header_cells are from row 2; data_rows start at row 3.
    """
    batch_idx = _resolve_batch_column_index(header_cells)
    keep_indices = _build_keep_indices(header_cells, batch_idx)
    if not keep_indices:
        return []

    records = []
    for row_values in data_rows:
        filtered = [
            row_values[idx] if idx < len(row_values) else ""
            for idx in keep_indices
        ]
        record = _record_from_filtered_row(filtered)
        if record:
            records.append(record)
    return records


def _read_monthly_coupon_rows_excel(monthly_path):
    _ensure_openpyxl()
    workbook = load_workbook(monthly_path, data_only=True)
    try:
        worksheet = workbook.active
        header_cells = _read_monthly_excel_row_values(worksheet, MONTHLY_HEADER_ROW)

        data_rows = []
        max_row = worksheet.max_row or MONTHLY_DATA_START_ROW
        for row_number in range(MONTHLY_DATA_START_ROW, max_row + 1):
            row_values = _read_monthly_excel_row_values(worksheet, row_number)
            if not row_values:
                continue
            has_content = any(
                isinstance(item, (datetime, int, float)) or _strip_cell(item)
                for item in row_values
            )
            if has_content:
                data_rows.append(row_values)
    finally:
        workbook.close()

    return _read_monthly_coupon_rows_from_table(header_cells, data_rows)


def _read_monthly_coupon_rows_csv(monthly_path):
    raw = pd.read_csv(monthly_path, header=None, dtype=str, keep_default_na=False)
    if len(raw) < MONTHLY_DATA_START_ROW:
        return []

    header_cells = [_strip_cell(value) for value in raw.iloc[MONTHLY_HEADER_ROW - 1].tolist()]
    data_rows = [
        [_strip_cell(value) for value in row.tolist()]
        for _, row in raw.iloc[MONTHLY_DATA_START_ROW - 1 :].iterrows()
    ]
    return _read_monthly_coupon_rows_from_table(header_cells, data_rows)


def _read_monthly_coupon_rows(monthly_path):
    """
    Parse J.H. Williams monthly coupon export.

    Row 1 (merged title block) is skipped. Row 2 provides headers; any column
    containing \"BATCH\" is dropped before mapping Date/Coupon/Gross/Fees/Net.
    """
    extension = os.path.splitext(monthly_path)[1].lower()
    if extension == ".csv":
        return _read_monthly_coupon_rows_csv(monthly_path)
    if extension in {".xlsx", ".xlsm", ".xls"}:
        return _read_monthly_coupon_rows_excel(monthly_path)
    raise ValueError(
        f"Tipo de reporte mensual no soportado '{extension}'. Use CSV o Excel."
    )


def resync_cupones_only(master_path):
    """
    Re-run the Cta Cte cross-reference over the whole Cupones sheet without
    appending any new coupons. For when the user has just loaded a new EFT
    (Pipeline 1) and wants pending or $0 split coupons refreshed, but has no
    new monthly J.H. Williams report to add this time.

    Returns:
        tuple: (saved preview path, summary dict)
    """
    _ensure_openpyxl()
    master_path = os.path.abspath(str(master_path).strip())
    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel maestro no encontrado: {master_path}")

    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel maestro debe ser .xlsx o .xlsm.")

    keep_vba = extension == ".xlsm"
    workbook = None
    rows_resynced_pending = 0
    try:
        workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
        cupones_sheet = _get_sheet(workbook, CUPONES_SHEET)
        cta_sheet = _get_sheet(workbook, CTA_CTE_SHEET)
        coupon_index = build_cta_coupon_index(cta_sheet)

        rows_resynced_pending = _resync_pending_cupones_rows(
            cupones_sheet, cta_sheet, coupon_index
        )
        if not rows_resynced_pending:
            raise NoPendingCouponsError(NO_PENDING_COUPONS_ALERT)

        save_path = _create_temp_workbook_path(suffix=extension or ".xlsx")
        workbook.save(save_path)
    except PermissionError as exc:
        raise PermissionError("ERROR: Cierra el archivo Excel antes de continuar") from exc
    finally:
        if workbook is not None:
            workbook.close()
        gc.collect()

    _launch_temp_workbook(save_path)

    summary = {"rows_resynced_pending": rows_resynced_pending}
    return save_path, summary


def append_monthly_cupones(master_path, monthly_path):
    """
    Append monthly coupon rows to Cupones and populate F-I from Cta Cte.

    Returns:
        tuple: (saved master path, summary dict)
    """
    _ensure_openpyxl()
    master_path = os.path.abspath(str(master_path).strip())
    monthly_path = os.path.abspath(str(monthly_path).strip())

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel maestro no encontrado: {master_path}")
    if not os.path.isfile(monthly_path):
        raise FileNotFoundError(f"Reporte mensual de cupones no encontrado: {monthly_path}")

    raw_monthly_rows = _read_monthly_coupon_rows(monthly_path)
    if not raw_monthly_rows:
        raise ValueError("No se encontraron filas de cupones en el reporte mensual.")

    # Enforce chronological append order (earliest date first).
    monthly_rows = _sort_records_by_date(_expand_records_by_coupon_split(raw_monthly_rows))
    del raw_monthly_rows

    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel maestro debe ser .xlsx o .xlsm.")

    keep_vba = extension == ".xlsm"
    rows_skipped_duplicates = 0
    start_row = 0
    appended = 0
    workbook = None
    try:
        workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
        cupones_sheet = _get_sheet(workbook, CUPONES_SHEET)
        monthly_rows, rows_skipped_duplicates = _filter_new_cupones_records(
            cupones_sheet, monthly_rows
        )
        if not monthly_rows:
            raise MonthlyReportFullyDuplicateError(MONTHLY_REPORT_FULL_DUPLICATE_ALERT)

        cta_sheet = _get_sheet(workbook, CTA_CTE_SHEET)
        coupon_index = build_cta_coupon_index(cta_sheet)
        cta_eft_boxes = build_cta_eft_boxes(cta_sheet)

        last_row = find_last_cupones_row(cupones_sheet)
        start_row = last_row + 1

        save_path = _create_temp_workbook_path(suffix=extension or ".xlsx")
        appended = 0
        rows_matched = 0
        unmatched_coupons = []
        max_coupon_text_len = 0
        split_group_rows = {}

        for offset, record in enumerate(monthly_rows):
            row = start_row + offset
            coupon_text = _write_cupones_base_columns(cupones_sheet, row, record)
            max_coupon_text_len = max(max_coupon_text_len, len(coupon_text or ""))
            _clear_cupones_control_columns(cupones_sheet, row)
            if _apply_cta_truth_to_cupones_row(
                cupones_sheet, cta_sheet, row, coupon_text, coupon_index, cta_eft_boxes
            ):
                rows_matched += 1
            else:
                unmatched_coupons.append(coupon_text)
            split_group_id = record.get("split_group_id")
            if split_group_id:
                split_group_rows.setdefault(split_group_id, []).append(row)
                if not record.get("is_primary_split"):
                    _apply_child_rows_standard_black(cupones_sheet, row)
            _apply_cupones_row_alignment(cupones_sheet, row)
            appended += 1

        _apply_parent_child_subtractions(cupones_sheet, split_group_rows)
        _apply_split_group_colors(cupones_sheet, split_group_rows)
        rows_resynced_pending = _resync_pending_cupones_rows(
            cupones_sheet, cta_sheet, coupon_index
        )
        for rows in split_group_rows.values():
            for r in rows:
                _apply_cupones_row_alignment(cupones_sheet, r)
        _stretch_column_b_width(cupones_sheet, max_coupon_text_len)
        workbook.save(save_path)
    except PermissionError as exc:
        raise PermissionError("ERROR: Cierra el archivo Excel antes de continuar") from exc
    finally:
        if workbook is not None:
            workbook.close()
        if "monthly_rows" in locals():
            monthly_rows = None
        gc.collect()

    _launch_temp_workbook(save_path)

    summary = {
        "rows_appended": appended,
        "start_row": start_row,
        "end_row": start_row + appended - 1 if appended else start_row - 1,
        "rows_matched": rows_matched,
        "rows_skipped_duplicates": rows_skipped_duplicates,
        "rows_resynced_pending": rows_resynced_pending,
        "unmatched_coupons": unmatched_coupons,
    }
    return save_path, summary
