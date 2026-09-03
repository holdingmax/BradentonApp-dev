"""
Reporte Diario — daily closure PDF extraction into Bradenton C-Store master.

Reads department totals from a user-selected PDF page of a daily closure
report, maps headers dynamically on sheet \"CARGA AQUI\" (row 3, column C onward),
and injects count/amount pairs on the first eligible operational row.
"""

import io
import os
import re
import tempfile
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher, get_close_matches
from datetime import date, datetime, time, timedelta

try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # type: ignore[assignment]

try:
    import pytesseract
    from PIL import Image
except ImportError:
    pytesseract = None  # type: ignore[assignment]
    Image = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
    COUNT_CELL_ALIGNMENT = Alignment(horizontal="center")
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment,misc]
    COUNT_CELL_ALIGNMENT = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False

SHEET_NAME = "CARGA AQUI"
HEADER_ROW = 3
HEADER_START_COLUMN = 3  # Column C
DATA_START_ROW = 5
DATE_SCAN_COLUMN = 1  # Column A only — calendar day matching
MANUAL_REPORT_COUNT_COLUMN = 71  # Column BS — total de unidades del reporte, cargado a mano

DEFAULT_PDF_PAGE_INDEX = 3  # Fourth page (0-based) for full un-cropped daily PDF

# Fixed 1-based PDF column layout (daily closure report)
PDF_DEPT_NAME_COL = 0  # 1st column — Dept.Name
PDF_NET_COUNT_COL = 4  # 5th column — Net Count
PDF_NET_SALES_COL = 7  # 8th column — Net Sales $
PDF_MIN_COLUMNS = 8

PDF_HEADER_DEPT_TOKENS = ("dept", "name")
PDF_HEADER_NET_COUNT_TOKENS = ("net", "count")
PDF_HEADER_NET_SALES_TOKENS = ("net", "sales")

DEPARTMENT_SALES_REPORT_ANCHOR = "Department Sales Report"
SAFE_DROP_REPORT_ANCHOR = "Safe Drop Report"

PROTECTED_DEPARTMENT_LABELS = frozenset(
    {
        "gift card",
        "varios/bolsa",
        "varios / bolsa",
        "varios",
        "bolsa",
    }
)

SUMMARY_STOP_MARKERS = (
    "total sales",
    "total cost",
    "gross profit",
    "grand total",
    "qty sold",
    "total qty",
    "department total",
    "net sales total",
)

SUB_HEADER_LABELS = frozenset(
    {
        "count",
        "qty",
        "units",
        "amount",
        "sales",
        "net",
        "net sales",
        "units sold",
        "qty sold",
        "cant",
        "cantidad",
        "importe",
    }
)

DEPARTMENT_ALIASES = {
    "hot dog": "hot dogs",
    "hotdog": "hot dogs",
    "hot dogs": "hot dogs",
    "beer/wine": "beer/wine",
    "beer wine": "beer/wine",
    "beer-wine": "beer/wine",
    "beer": "beer/wine",
    "e-cig": "e-gigarette",
    "e cig": "e-gigarette",
    "ecig": "e-gigarette",
    "e-cigarette": "e-gigarette",
    "e cigarette": "e-gigarette",
    "e-gigarette": "e-gigarette",
    "coffee": "coffe",
    "coffe": "coffe",
    "fountain": "foutain",
    "foutain": "foutain",
    "boiled peanuts": "boiled peanuts",
    "boiled-peanuts": "boiled peanuts",
    "automotive": "auto",
    "auto": "auto",
    "ice cream": "ice cream",
    "ice-cream": "ice cream",
    "prop hd": "propane",
    "prop-hd": "propane",
    "propane": "propane",
}

# Parsed PDF labels -> exact CARGA AQUI row 3 worksheet titles
DEPARTMENT_NAME_NORMALIZATION = {
    "e-gigarette": "E-GIGARETTE",
    "e-cig": "E-GIGARETTE",
    "e cig": "E-GIGARETTE",
    "ecig": "E-GIGARETTE",
    "e-cigarette": "E-GIGARETTE",
    "e cigarette": "E-GIGARETTE",
    "foutain": "FOUTAIN",
    "fountain": "FOUTAIN",
    "coffe": "COFFE",
    "coffee": "COFFE",
    "beer/wine": "BEER/WINE",
    "beer wine": "BEER/WINE",
    "beer-wine": "BEER/WINE",
    "beer": "BEER/WINE",
    "hot dog": "HOT DOGS",
    "hotdog": "HOT DOGS",
    "hot dogs": "HOT DOGS",
    "boiled peanuts": "BOILED PEANUTS",
    "boiled-peanuts": "BOILED PEANUTS",
    "automotive": "AUTO",
    "auto": "AUTO",
    "ice cream": "ICE CREAM",
    "ice-cream": "ICE CREAM",
    "propane": "PROPANE",
    "prop hd": "PROPANE",
    "prop-hd": "PROPANE",
    "local acct": "GETTEL/TOYOTA",
}

# Department row layout (right to left): ... | Net Count | ... | Net Sales $ | % of sales
NET_COUNT_REVERSE_INDEX = -5
NET_SALES_REVERSE_INDEX = -2
MIN_ROW_SPLIT_PARTS = 5

SALES_ALERT_FONT_COLOR = "FF0000"
GROUP_500_DEPARTMENTS = frozenset(
    {
        "AUTO",
        "BOILED PEANUTS",
        "HBA",
        "ICECREAM",
        "MILK",
        "PROPANE",
        "GROCERIES",
        "NONTAX",
        "CANDY",
        "COFFE",
        "JUICE",
        "SNACK",
        "FOUTAIN",
        "WATER",
    }
)
GROUP_1400_DEPARTMENTS = frozenset(
    {
        "BEER/WINE",
        "CIGARS",
        "E-GIGARETTE",
        "GEN-CTN",
        "GEN-PAK",
        "MAJ CR",
        "MAJ PAK",
        "SNUFF",
        "SODA",
        "ONLINE",
        "SKOFF",
    }
)
GROUP_500_THRESHOLD = 500.00
GROUP_1400_THRESHOLD = 1400.00
GETTEL_TOYOTA_THRESHOLD = 17000.00

FUSED_DEPT_TRAILING_COUNT_RE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z0-9\s/\-.'&]*?)\s*(?P<count>\d+)$"
)


from ocr_utils import (
    ensure_pytesseract as _ensure_pytesseract,
    extract_largest_page_image as _extract_page_image,
)


def _ensure_pdfplumber():
    if pdfplumber is None:
        raise ImportError(
            "Reporte Diario requiere pdfplumber. Instale con: pip install pdfplumber"
        )


def _ensure_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "Reporte Diario requiere openpyxl. Instale con: pip install openpyxl"
        )


class _LazyPdfPageImages:
    """
    Decode and orient a PDF page's embedded photo only the first time it's
    actually requested, caching the result for any later reuse.

    A daily closure PDF can run 50+ pages, but every report only ever needs
    2-3 of them (the anchor page found on the first or second try, plus one
    continuation page) — decoding and running Tesseract's orientation check
    on every other page was pure wasted work driving up processing time,
    especially across a multi-PDF batch. This makes cost proportional to
    pages actually read instead of pages in the file.
    """

    def __init__(self, pdf_path):
        _ensure_pdfplumber()
        _ensure_pytesseract()
        self._pdf = pdfplumber.open(pdf_path)
        self._cache = {}

    def __len__(self):
        return len(self._pdf.pages)

    def __getitem__(self, index):
        if index not in self._cache:
            self._cache[index] = self._load(index)
        return self._cache[index]

    def _load(self, index):
        return _extract_page_image(self._pdf.pages[index])

    def close(self):
        self._pdf.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()


def _strip_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _clean_dept_name(value):
    """Col 1 — strip trailing spaces, newlines, and collapse internal whitespace."""
    text = _strip_cell(value)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _sanitize_parsed_dept_name(value):
    """Strip phantom leading/trailing dots, hyphens, and spaces from Dept.Name."""
    text = _clean_dept_name(value)
    text = re.sub(r"^[.\-\s]+|[.\-\s]+$", "", text)
    return text


def _normalize_department_spacing(value):
    """
    Collapse mixed whitespace inside department labels to one space.

    Handles repeated spaces, tabs, and embedded newlines from un-cropped PDF text.
    """
    text = _strip_cell(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _fallback_department_alias(dept_name):
    """
    Force-map partially corrupted labels to known worksheet departments.

    Used by the never-skip policy to avoid dropping rows due to OCR spacing shifts.
    """
    raw = _normalize_department_spacing(dept_name).upper()
    key = _normalize_department_label(raw)
    if not raw:
        return ""
    if "LOCAL" in raw:
        return "GETTEL/TOYOTA"
    if "BEER" in raw:
        return "BEER/WINE"
    if "GIGARETTE" in raw or "CIGARETTE" in raw:
        return "E-GIGARETTE"
    if "MAJ" in raw and "PAK" in raw:
        return "MAJ PAK"
    if "MAJ" in raw and "CR" in raw:
        return "MAJ CR"
    if "GEN" in raw and "PAK" in raw:
        return "GEN-PAK"
    if "GEN" in raw and "CTN" in raw:
        return "GEN-CTN"
    if "FLOWER" in raw:
        return "FLOWERS"
    if "FOUNTAIN" in raw:
        return "FOUTAIN"
    if key in DEPARTMENT_NAME_NORMALIZATION:
        return DEPARTMENT_NAME_NORMALIZATION[key]
    return raw


def normalize_parsed_department_name(value):
    """
    Map parsed PDF department text to exact CARGA AQUI row 3 titles.
    """
    text = _normalize_department_spacing(_sanitize_parsed_dept_name(value)).upper()
    if not text:
        return ""
    forced = _fallback_department_alias(text)
    if forced:
        text = forced
    key = _normalize_department_label(text)
    if key in DEPARTMENT_NAME_NORMALIZATION:
        return DEPARTMENT_NAME_NORMALIZATION[key]
    if key in DEPARTMENT_ALIASES:
        alias_target = _normalize_department_label(DEPARTMENT_ALIASES[key])
        if alias_target in DEPARTMENT_NAME_NORMALIZATION:
            return DEPARTMENT_NAME_NORMALIZATION[alias_target]
    return text


def _row_cells_to_line(cells):
    """Join table cells into one line for reverse-index whitespace parsing."""
    return " ".join(_strip_cell(cell) for cell in (cells or []) if _strip_cell(cell))


def _strip_discount_tokens(text):
    """
    Remove negative Discount $ tokens (e.g. -$5.26) from a raw line.

    Ensures discount amounts never interfere with Net Sales $ / Net Count parsing.
    """
    if not text:
        return text
    tokens = str(text).split()
    kept = []
    for token in tokens:
        plain = token.replace(",", "")
        if "-$" in plain or re.search(r"-\$\s*\d", plain):
            continue
        kept.append(token)
    return " ".join(kept)


def _is_pdf_header_parts(parts):
    joined = " ".join(parts).lower()
    return "dept" in joined and "name" in joined


def _is_alphabetic_dept_token(token):
    if not token:
        return False
    cleaned = token.replace(",", "")
    if re.fullmatch(r"[\d.$%-]+", cleaned):
        return False
    return bool(re.search(r"[A-Za-z]", token))


def _split_dept_name_from_fused_tail(text):
    """Isolate GEN-PAK from GEN-PAK31 when count was not a separate token."""
    dept_text = _sanitize_parsed_dept_name(text)
    if not dept_text:
        return "", None
    match = FUSED_DEPT_TRAILING_COUNT_RE.match(dept_text)
    if match:
        return (
            _sanitize_parsed_dept_name(match.group("name")),
            match.group("count"),
        )
    compact = dept_text.replace(" ", "")
    match = re.match(r"^(.+?)(\d+)$", compact)
    if match and re.search(r"[A-Za-z/]", match.group(1)):
        return (
            _sanitize_parsed_dept_name(match.group(1)),
            match.group(2),
        )
    return dept_text, None


def _resolve_reverse_indices(part_count):
    """Return (net_count_index, net_sales_index) for the row width."""
    if part_count >= 8:
        return NET_COUNT_REVERSE_INDEX, NET_SALES_REVERSE_INDEX
    if part_count >= 5:
        return -3, NET_SALES_REVERSE_INDEX
    return None, None


def _parse_row_by_reverse_index(line):
    """
    Stable reverse-index row parser after Department Sales Report.

    Whitespace split, then:
      [-2] = Net Sales $ (skip [-1] % of sales)
      [-5] = Net Count on full-width rows (or [-3] on compact rows)
      left tokens = Dept.Name (alphabetic parts only)
    """
    text = _normalize_department_spacing(_strip_cell(line))
    text = _strip_discount_tokens(text)
    # Keep department labels detached from trailing numeric tokens on dense PDF rows.
    text = re.sub(r"(?i)(BEER/WINE|E-?CIGARETTE)(?=\d)", r"\1 ", text)
    if not text:
        return None

    parts = text.split()
    filtered_parts = [
        token
        for token in parts
        if token and not token.startswith("-$") and "-$" not in token
    ]
    if len(filtered_parts) < MIN_ROW_SPLIT_PARTS or _is_pdf_header_parts(filtered_parts):
        return None

    tail_tokens = (
        filtered_parts[:-1]
        if filtered_parts and "%" in filtered_parts[-1]
        else list(filtered_parts)
    )

    amount_index = None
    amount = 0.0
    count = 0
    count_index = None
    # Stable reverse-index parse after dynamic discount-token removal.
    # Net Sales: [-2] (skip trailing percent), Net Count: [-5] on full rows.
    try:
        if len(tail_tokens) >= 5:
            amount_index = len(tail_tokens) - 2
            amount = float(_sanitize_sales_float(tail_tokens[amount_index]))
            if len(tail_tokens) >= 8:
                count_index = len(tail_tokens) - 5
            else:
                count_index = len(tail_tokens) - 3
            count = int(_safe_parse_count(tail_tokens[count_index]))
        else:
            count_idx, sales_index = _resolve_reverse_indices(len(tail_tokens))
            if sales_index is not None:
                amount_raw = tail_tokens[sales_index]
                amount = float(_sanitize_sales_float(amount_raw))
                amount_index = (
                    sales_index
                    if sales_index >= 0
                    else len(tail_tokens) + sales_index
                )
            if count_idx is not None and -len(tail_tokens) <= count_idx < len(tail_tokens):
                count = int(_safe_parse_count(tail_tokens[count_idx]))
                count_index = (
                    count_idx if count_idx >= 0 else len(tail_tokens) + count_idx
                )
    except Exception:
        amount = float(_safe_parse_amount(tail_tokens[-1] if tail_tokens else 0.0))
        count = int(_safe_parse_count(tail_tokens[-2] if len(tail_tokens) >= 2 else 0))
        amount_index = max(len(tail_tokens) - 1, 1)
        count_index = max(amount_index - 1, 0)

    if count_index is None:
        count_index = max(amount_index - 1, 0)

    left_end = count_index
    if left_end <= 0:
        return None

    split_tokens = [chunk.strip() for chunk in re.split(r"\s{2,}", text) if chunk.strip()]
    cleaned_from_split = (
        _normalize_department_spacing(split_tokens[0]).upper() if split_tokens else ""
    )
    dept_tokens = [
        token for token in tail_tokens[:left_end] if _is_alphabetic_dept_token(token)
    ]
    dept_raw = cleaned_from_split or _normalize_department_spacing(" ".join(dept_tokens))
    dept_raw = _normalize_department_spacing(dept_raw).upper()
    dept_text, fused_count = _split_dept_name_from_fused_tail(dept_raw)
    if fused_count is not None and count == 0:
        count = int(_safe_parse_count(fused_count))

    department = normalize_parsed_department_name(_fallback_department_alias(dept_text))
    if not department or _is_summary_row(department):
        return None
    if _is_protected_department(department):
        return None

    return {
        "department": department,
        "count": count,
        "amount": amount,
    }


def _safe_parse_count(value):
    try:
        return _parse_count(value)
    except (TypeError, ValueError, OverflowError):
        return 0


def _safe_parse_amount(value):
    try:
        return _parse_amount(value)
    except (TypeError, ValueError, OverflowError):
        return 0.0


def _normalize_department_label(value):
    text = _clean_dept_name(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _canonical_department_key(value):
    key = _normalize_department_label(value)
    if not key:
        return ""
    if key in DEPARTMENT_ALIASES:
        return _normalize_department_label(DEPARTMENT_ALIASES[key])
    for alias, canonical in DEPARTMENT_ALIASES.items():
        if key == alias or key.startswith(alias + " ") or key.endswith(" " + alias):
            return _normalize_department_label(canonical)
    return key


def _department_keys_match(pdf_key, header_key):
    if not pdf_key or not header_key:
        return False
    if pdf_key == header_key:
        return True
    pdf_canon = _canonical_department_key(pdf_key)
    header_canon = _canonical_department_key(header_key)
    if pdf_canon and pdf_canon == header_canon:
        return True
    if pdf_canon in header_canon or header_canon in pdf_canon:
        return True
    pdf_compact = pdf_canon.replace(" ", "").replace("-", "").replace("/", "")
    header_compact = header_canon.replace(" ", "").replace("-", "").replace("/", "")
    return pdf_compact == header_compact


def _sanitize_numeric_text(value):
    """Strip whitespace, currency symbols, and grouping separators."""
    text = _strip_cell(value)
    if not text:
        return ""
    text = text.replace("$", "").replace("€", "").replace("£", "")
    text = text.replace(" ", "").replace("\u00a0", "").replace("\n", "").replace("\r", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    text = text.replace(",", "")
    return text


def _parse_count(value):
    """Col 5 — Net Count: strip symbols, default empty to 0, return int."""
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return 0
    text = _sanitize_numeric_text(value)
    if not text:
        return 0
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    try:
        number = float(text)
        if number.is_integer():
            return int(number)
    except ValueError:
        return 0
    return 0


def _parse_amount(value):
    """Col 8 — Net Sales $: strip currency, handle negatives, return float."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _sanitize_numeric_text(value)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_sub_header(label):
    return _normalize_department_label(label) in SUB_HEADER_LABELS


def _is_summary_row(label):
    key = _normalize_department_label(label)
    if not key:
        return True
    if key.startswith("dept "):
        return True
    if key in PROTECTED_DEPARTMENT_LABELS:
        return False
    return any(marker in key for marker in SUMMARY_STOP_MARKERS)


def _is_protected_department(label):
    key = _normalize_department_label(label)
    return key in PROTECTED_DEPARTMENT_LABELS


def _cell_has_formula(cell):
    if getattr(cell, "data_type", None) == "f":
        return True
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def _create_temp_workbook_path():
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="reporte_diario_")
    os.close(fd)
    return temp_path



def _get_carga_aqui_sheet(workbook):
    target = SHEET_NAME.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    raise ValueError(
        f'Hoja "{SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _pad_row(row, min_columns=PDF_MIN_COLUMNS):
    cells = [_strip_cell(cell) for cell in (row or [])]
    if len(cells) < min_columns:
        cells.extend([""] * (min_columns - len(cells)))
    return cells


def _header_cell_matches(label, required_tokens):
    key = _normalize_department_label(label).replace(".", " ")
    return all(token in key for token in required_tokens)


def _is_pdf_department_header_row(cells):
    padded = _pad_row(cells)
    return (
        _header_cell_matches(padded[PDF_DEPT_NAME_COL], PDF_HEADER_DEPT_TOKENS)
        and _header_cell_matches(padded[PDF_NET_COUNT_COL], PDF_HEADER_NET_COUNT_TOKENS)
        and _header_cell_matches(padded[PDF_NET_SALES_COL], PDF_HEADER_NET_SALES_TOKENS)
    )


def _extract_tables_from_page(page):
    """Extract tables from a PDF page (silent — no console output)."""
    tables = page.extract_tables() or []
    if tables:
        return tables

    line_settings = {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 4,
        "join_tolerance": 4,
    }
    tables = page.extract_tables(line_settings) or []
    if tables:
        return tables

    text_settings = {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_tolerance": 4,
    }
    return page.extract_tables(text_settings) or []


FILENAME_DAY_MONTH_PATTERN = re.compile(
    r"(?<!\d)(\d{1,2})[-/.](\d{1,2})(?!\d)"
)


def extract_day_from_filename(file_path):
    """
    Extract the calendar day-of-month from a daily PDF filename.

    Examples: \"Close Store 01-05.pdf\" -> 1, \"Close Store 07-05.pdf\" -> 7.
    """
    base = os.path.splitext(os.path.basename(file_path))[0]
    for match in FILENAME_DAY_MONTH_PATTERN.finditer(base):
        target_day = int(match.group(1))
        month = int(match.group(2))
        if 1 <= target_day <= 31 and 1 <= month <= 12:
            return target_day

    trailing = re.search(r"[^\d](\d{1,2})\s*$", base)
    if trailing:
        target_day = int(trailing.group(1))
        if 1 <= target_day <= 31:
            return target_day

    raise ValueError(
        f"No se pudo extraer el día del nombre de archivo: {os.path.basename(file_path)}. "
        "Se esperaba un patrón como 'Close Store 01-05.pdf'."
    )


def _cell_calendar_day(value):
    """
    Return the day-of-month (1-31) from a Column A cell value, or None.

    datetime -> .day; plain integers 1-31 -> day-of-month; larger numbers -> Excel serial.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return int(value.day)
    if isinstance(value, int):
        day = int(value)
        if 1 <= day <= 31:
            return day
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(day)
            if isinstance(converted, datetime):
                return int(converted.day)
        except (ValueError, TypeError, OverflowError):
            return None
        return None
    if isinstance(value, float):
        if value.is_integer():
            whole = int(value)
            if 1 <= whole <= 31:
                return whole
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value)
            if isinstance(converted, datetime):
                return int(converted.day)
        except (ValueError, TypeError, OverflowError):
            return None
        return None
    text = _strip_cell(value)
    if not text:
        return None
    if re.fullmatch(r"\d{1,2}", text):
        day = int(text)
        if 1 <= day <= 31:
            return day
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return int(datetime.strptime(text, fmt).day)
        except ValueError:
            continue
    if re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$", text):
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
            try:
                return int(datetime.strptime(text, fmt).day)
            except ValueError:
                continue
    month_match = re.match(r"^(\d{1,2})[-/]([a-zA-Z]{3,9})(?:[-/]\d{2,4})?$", text)
    if month_match:
        return int(month_match.group(1))
    if re.match(r"^\d{1,2}-[a-zA-Z]{3}(-\d{2,4})?$", text):
        return int(text.split("-", 1)[0])
    return None


def find_row_for_calendar_day(sheet, target_day, start_row=DATA_START_ROW):
    """
    Locate the row whose Column A date matches target_day (strict integer compare).

    Scans column A only from start_row downward so each PDF locks one unique row.
    """
    target_day = int(target_day)
    max_row = max(sheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        cell_value = sheet.cell(row=row, column=DATE_SCAN_COLUMN).value
        cell_day = _cell_calendar_day(cell_value)
        if cell_day is not None and int(cell_day) == target_day:
            return row
    raise ValueError(
        f"No se encontró una fila con el día {target_day} en la columna A de la hoja {SHEET_NAME} "
        f"(desde la fila {start_row})."
    )


def _parse_pdf_data_row(cells):
    """
    Extract one department row using reverse-index whitespace splitting.

    Col layout: Dept.Name | ... | Net Count | ... | Net Sales $ | % of sales
    """
    try:
        line = _row_cells_to_line(cells)
        if not line:
            return None
        parsed = _parse_row_by_reverse_index(line)
        if parsed is None:
            return None
        return {
            "department": parsed["department"],
            "count": int(parsed["count"]),
            "amount": float(parsed["amount"]),
        }
    except (TypeError, ValueError, OverflowError, IndexError, AttributeError):
        return None


def _line_contains_anchor(line):
    return DEPARTMENT_SALES_REPORT_ANCHOR.lower() in _strip_cell(line).lower()


def _line_contains_stop_marker(line):
    """True at 'Safe Drop Report' — everything from there on is out of scope."""
    text = _strip_cell(line).lower()
    return "safe drop" in text


_OCR_HEADER_COMPANION_TOKENS = ("sales", "count", "refund", "item")


def _is_ocr_header_line(text):
    """
    True for the "Dept. Name ... Net Count ... Net Sales $ ..." header row,
    which OCR sometimes repeats mid-table as its own smashed-together line —
    including right before a same-day Refund sub-table, which reprints its
    own header. "dept" is checked as normal, but OCR also regularly drops
    its leading "D" ("Dept." -> "Ept."), so "name" alongside any other
    column-header word is treated as the same header line too — no real
    department is ever literally named "Name".
    """
    key = _normalize_department_label(text)
    if "dept" in key and "name" in key:
        return True
    return "name" in key and any(token in key for token in _OCR_HEADER_COMPANION_TOKENS)


def _anchor_passed_in_page_text(page):
    """True once 'Department Sales Report' appears in line-by-line page text."""
    for line in (page.extract_text() or "").splitlines():
        if _line_contains_anchor(line):
            return True
    return False


def _crop_page_below_anchor(page):
    """
    Return a pdfplumber page cropped to content below the anchor phrase.

    Falls back to the full page when search geometry is unavailable.
    """
    hits = page.search(DEPARTMENT_SALES_REPORT_ANCHOR, case=False) or []
    if hits:
        anchor_bottom = max(hit["bottom"] for hit in hits)
        return page.crop((0, anchor_bottom, page.width, page.height))
    return page


def _parse_department_table_rows(rows, require_header=True):
    if not rows:
        return []

    header_index = None
    if require_header:
        for idx, row in enumerate(rows):
            if _is_pdf_department_header_row(row):
                header_index = idx
                break
        start = (header_index + 1) if header_index is not None else 0
    else:
        start = 0

    records = []
    for row in rows[start:]:
        if not row:
            continue
        record = _parse_pdf_data_row(row)
        if record is None:
            if records and _clean_dept_name(_pad_row(row)[PDF_DEPT_NAME_COL]):
                break
            continue
        records.append(record)
    return records


def _parse_tables_with_line_anchor(page):
    """
    Parse department rows only after the Department Sales Report anchor.

    Scans raw text line-by-line to locate the anchor, then reads table rows
    from the cropped region below it (or skips pre-anchor table rows).
    """
    if not _anchor_passed_in_page_text(page):
        raise ValueError(
            f'No se encontró el ancla "{DEPARTMENT_SALES_REPORT_ANCHOR}" en la página del PDF.'
        )

    cropped = _crop_page_below_anchor(page)
    tables = _extract_tables_from_page(cropped)
    records = []
    for table in tables:
        parsed = _parse_department_table_rows(table)
        if parsed:
            records = parsed
            break

    if records:
        return records

    past_anchor = False
    for table in _extract_tables_from_page(page):
        for row in table or []:
            row_text = _row_cells_to_line(row)
            if not past_anchor:
                if _line_contains_anchor(row_text):
                    past_anchor = True
                continue
            if _is_pdf_department_header_row(row):
                continue
            record = _parse_pdf_data_row(row)
            if record is None:
                if records:
                    probe = _row_cells_to_line(row)
                    if probe and (
                        _is_summary_row(probe)
                        or "total sales" in probe.lower()
                    ):
                        break
                continue
            records.append(record)
        if records:
            break

    if not records:
        past_anchor = False
        for line in (page.extract_text() or "").splitlines():
            if not past_anchor:
                if _line_contains_anchor(line):
                    past_anchor = True
                continue
            if not _strip_cell(line):
                continue
            if _is_pdf_department_header_row([line]):
                continue
            parsed = _parse_row_by_reverse_index(line)
            if parsed is None:
                if records:
                    break
                continue
            records.append(parsed)

    return records


# No single Tesseract page-segmentation mode is reliably best across every
# photo, so each page is tried under each of these and scored (see
# _score_ocr_page_text) rather than trusting one fixed mode.
_OCR_TEXT_CONFIGS = ("--psm 6", "--psm 4", "--psm 3", "--psm 11")

# WATER is always the last department printed on a real report; when it was
# a $0.00 day it's omitted and TAXABLE (the second-to-last) becomes the last
# one. Used only as an informational cross-check, never to stop parsing.
LAST_DEPARTMENT_CANDIDATES = ("WATER", "TAXABLE")

# Fields after Dept.Name, right to left: % of Sales (dropped separately),
# Net Sales $, Discount $, Refund $, Net Count, Refund Count, Item Count,
# Gross Sales $ — 7 fields once the trailing % token is stripped.
_OCR_ROW_TAIL_FIELDS = 7
_OCR_NET_SALES_REVERSE_INDEX = -1
_OCR_NET_COUNT_REVERSE_INDEX = -4


def _parse_ocr_department_row(line):
    """
    Parse one OCR'd "Department Sales Report" line.

    Tesseract's image_to_string collapses the original column gaps to single
    spaces, so — unlike the pdfplumber text parser, which relies on
    multi-space gaps to isolate the department name — this indexes strictly
    from the right against the report's fixed layout: Dept.Name | Gross
    Sales $ | Item Count | Refund Count | Net Count | Refund $ | Discount $ |
    Net Sales $ | % of Sales. Returns None for anything that isn't a
    parseable row; the printed grand-total line (no department name) comes
    back with department="" and is_total=True for the OCR subtotal cross-check.
    """
    text = _normalize_department_spacing(_strip_cell(line))
    if not text:
        return None
    if _is_ocr_header_line(text):
        return None

    parts = [token for token in text.split() if token]
    if len(parts) < _OCR_ROW_TAIL_FIELDS + 1:
        return None

    # The trailing "% of Sales" column is always present in the source table
    # even when Tesseract fails to read its literal "%" character (e.g. a
    # garbled "OBI" instead of "0.02%") — so it is always dropped, rather
    # than only when the "%" glyph itself came through.
    tail_tokens = parts[:-1]

    net_sales_raw = tail_tokens[_OCR_NET_SALES_REVERSE_INDEX]
    net_count_raw = tail_tokens[_OCR_NET_COUNT_REVERSE_INDEX]
    dept_tokens = tail_tokens[:-_OCR_ROW_TAIL_FIELDS]

    amount = _sanitize_sales_float(net_sales_raw)
    count = _safe_parse_count(net_count_raw)

    if not dept_tokens:
        return {"department": "", "count": int(count), "amount": float(amount), "is_total": True}

    dept_raw = _normalize_department_spacing(" ".join(dept_tokens)).upper()
    dept_text, _fused_count = _split_dept_name_from_fused_tail(dept_raw)
    department = normalize_parsed_department_name(_fallback_department_alias(dept_text))
    if not department or _is_summary_row(department) or _is_protected_department(department):
        return None

    return {"department": department, "count": int(count), "amount": float(amount), "is_total": False}


def _score_ocr_page_text(text):
    """
    Score one OCR attempt by how many department rows it yields.

    If the anchor line is present, only rows between it and the
    "Safe Drop Report" stop marker count (skips preamble tables read on the
    same page). If not, the whole page is treated as table continuation —
    real reports sometimes spill the department table onto a second page.
    """
    lines = text.splitlines() if text else []
    start = 0
    has_anchor = False
    for index, line in enumerate(lines):
        if _line_contains_anchor(line):
            start = index + 1
            has_anchor = True
            break

    count = 0
    for line in lines[start:]:
        if _line_contains_stop_marker(line):
            break
        parsed = _parse_ocr_department_row(line)
        if parsed is not None and not parsed["is_total"]:
            count += 1
    return (int(has_anchor), count)


def _ocr_page_text(image):
    """Best-effort OCR of a full page photo into plain text."""
    if image is None:
        return ""
    _ensure_pytesseract()
    best_text = ""
    best_score = (-1, -1)
    for config in _OCR_TEXT_CONFIGS:
        try:
            text = pytesseract.image_to_string(image, config=config) or ""
        except Exception:
            continue
        score = _score_ocr_page_text(text)
        if score > best_score:
            best_score = score
            best_text = text
    return best_text


def parse_elistar_daily_pdf_ocr(pdf_path, start_page_index=DEFAULT_PDF_PAGE_INDEX):
    """
    OCR-based extraction for photographed/scanned daily PDFs (no
    extractable text). The "Department Sales Report" table doesn't always
    land on the same page and can spill onto the next one, so this scans
    forward from start_page_index (wrapping around) for the anchor,
    corrects page rotation via Tesseract OSD, and keeps reading department
    rows across pages until the "Safe Drop Report" anchor is reached.

    Returns:
        tuple[list[dict], dict]: (records, diagnostics) — diagnostics has
        keys "pages_used", "last_department" and "subtotal_mismatch" (the
        printed grand-total line is cross-checked against the sum of parsed
        rows as a soft OCR-quality sanity check, never a blocking one).
    """
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    images = _LazyPdfPageImages(pdf_path)
    total_pages = len(images)
    if not (0 <= start_page_index < total_pages):
        start_page_index = 0

    search_order = list(range(start_page_index, total_pages)) + list(
        range(0, start_page_index)
    )

    anchor_page = None
    anchor_page_text = None
    for idx in search_order:
        text = _ocr_page_text(images[idx])
        if _line_contains_anchor(text):
            anchor_page = idx
            anchor_page_text = text
            break

    if anchor_page is None:
        images.close()
        raise ValueError(
            f'No se encontró el ancla "{DEPARTMENT_SALES_REPORT_ANCHOR}" en ninguna '
            "página del PDF (vía OCR). Verifique que el reporte no esté demasiado "
            "borroso o girado."
        )

    records = []
    printed_totals = None
    pages_used = []
    reached_stop = False
    idx = anchor_page
    while idx < total_pages:
        page_text = anchor_page_text if idx == anchor_page else _ocr_page_text(images[idx])
        pages_used.append(idx + 1)
        past_anchor = idx != anchor_page
        for line in page_text.splitlines():
            if not past_anchor:
                if _line_contains_anchor(line):
                    past_anchor = True
                continue
            if _line_contains_stop_marker(line):
                reached_stop = True
                break
            if not _strip_cell(line):
                continue
            parsed = _parse_ocr_department_row(line)
            if parsed is None:
                continue
            if parsed["is_total"]:
                printed_totals = parsed
            else:
                records.append(parsed)
        if reached_stop:
            break
        idx += 1

    images.close()

    if not records:
        raise ValueError(
            f"No se encontraron registros de departamento vía OCR después de "
            f'"{DEPARTMENT_SALES_REPORT_ANCHOR}". Verifique que el reporte no esté '
            "demasiado borroso o girado."
        )

    subtotal_mismatch = None
    if printed_totals is not None:
        computed_amount = round(sum(r["amount"] for r in records), 2)
        computed_count = sum(r["count"] for r in records)
        amount_diff = round(abs(computed_amount - printed_totals["amount"]), 2)
        count_diff = abs(computed_count - printed_totals["count"])
        if amount_diff > 1.00 or count_diff > 2:
            subtotal_mismatch = {
                "computed_amount": computed_amount,
                "printed_amount": printed_totals["amount"],
                "computed_count": computed_count,
                "printed_count": printed_totals["count"],
            }

    diagnostics = {
        "pages_used": pages_used,
        "last_department": records[-1]["department"],
        "subtotal_mismatch": subtotal_mismatch,
    }
    return records, diagnostics


def parse_elistar_daily_pdf_page(pdf_path, page_index=DEFAULT_PDF_PAGE_INDEX):
    """
    Extract department records from the selected PDF page (0-based index).

    Only rows after the \"Department Sales Report\" anchor are parsed.
    Uses fixed columns: Dept.Name (1), Net Count (5), Net Sales $ (8).
    Falls back to OCR automatically when the page has no extractable text
    at all (a photographed/scanned report, as opposed to a digital export).

    Returns:
        tuple[list[dict], dict]: (records, diagnostics) — diagnostics has
        keys "used_ocr", "pages_used" (OCR only) and "last_department"
        (OCR only, may be absent for text-based extraction).
    """
    _ensure_pdfplumber()
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    if page_index < 0:
        raise ValueError("El índice de página del PDF debe ser 0 o mayor.")

    with pdfplumber.open(pdf_path) as pdf:
        if page_index >= len(pdf.pages):
            raise ValueError(
                f"Página {page_index + 1} del PDF no encontrada; "
                f"el documento tiene {len(pdf.pages)} página(s)."
            )
        page = pdf.pages[page_index]
        has_text = bool((page.extract_text() or "").strip())
        records = _parse_tables_with_line_anchor(page) if has_text else []

    if records:
        return [dict(record) for record in records], {"used_ocr": False}

    if not has_text:
        ocr_records, ocr_diagnostics = parse_elistar_daily_pdf_ocr(
            pdf_path, start_page_index=page_index
        )
        diagnostics = {"used_ocr": True}
        diagnostics.update(ocr_diagnostics)
        return [dict(record) for record in ocr_records], diagnostics

    raise ValueError(
        f"No se encontraron registros de departamento en la página {page_index + 1} del PDF después de "
        f'"{DEPARTMENT_SALES_REPORT_ANCHOR}". '
        "Se esperaba Dept.Name (col 1), Net Count (col 5), Net Sales $ (col 8)."
    )


def build_department_column_map(sheet):
    """
    Scan row 3 from column C outward and map department labels to (count_col, amount_col).

    Each department header marks a COUNT | NET SALES column pair (first = Net Count,
    second = Net Sales $). Position-independent — keyed by normalized department label.
    """
    mapping = {}
    protected = set()
    col = HEADER_START_COLUMN
    max_col = max(sheet.max_column, HEADER_START_COLUMN)
    sub_header_row = HEADER_ROW + 1

    while col <= max_col:
        header_value = sheet.cell(row=HEADER_ROW, column=col).value
        label = _strip_cell(header_value)
        if not label or _is_sub_header(label):
            col += 1
            continue

        norm = _normalize_department_label(label)
        count_col = col
        amount_col = col + 1

        sub_count = _normalize_department_label(
            sheet.cell(row=sub_header_row, column=count_col).value
        )
        sub_amount = _normalize_department_label(
            sheet.cell(row=sub_header_row, column=amount_col).value
        )
        if sub_count and sub_amount:
            first_is_sales = "net" in sub_count and "sales" in sub_count
            second_is_count = "count" in sub_amount
            if first_is_sales and second_is_count:
                count_col, amount_col = amount_col, count_col

        if _is_protected_department(label):
            protected.add(norm)
            col += 2
            continue

        mapping[norm] = (count_col, amount_col)
        mapping[_canonical_department_key(label)] = (count_col, amount_col)

        canonical = _canonical_department_key(label)
        if canonical and canonical not in mapping:
            mapping[canonical] = (count_col, amount_col)

        col += 2

    if not mapping:
        raise ValueError(
            f"No se encontraron encabezados de departamento en la fila {HEADER_ROW} desde la columna C."
        )

    return mapping, protected


def _resolve_department_columns(dept_name, column_map):
    key = _normalize_department_label(dept_name)
    if key in column_map:
        return column_map[key]

    canon = _canonical_department_key(dept_name)
    if canon in column_map:
        return column_map[canon]

    for header_key, coords in column_map.items():
        if _department_keys_match(key, header_key):
            return coords
    forced = _fallback_department_alias(dept_name)
    forced_key = _normalize_department_label(forced)
    if forced_key in column_map:
        return column_map[forced_key]
    normalized_header_keys = [k for k in column_map.keys() if k]
    if normalized_header_keys:
        fuzzy = get_close_matches(forced_key or key, normalized_header_keys, n=1, cutoff=0.6)
        if fuzzy:
            return column_map[fuzzy[0]]
    return None


def _department_label_keys(dept_name):
    """Normalized keys used to match worksheet / PDF department labels."""
    keys = set()
    key = _normalize_department_label(dept_name)
    if key:
        keys.add(key)
    canon = _canonical_department_key(dept_name)
    if canon:
        keys.add(canon)
    return keys


def _apply_sales_font_alert(cell):
    """Bright-red font only — never alters cell background fill."""
    if Font is None:
        return
    base = cell.font
    cell.font = Font(
        color=SALES_ALERT_FONT_COLOR,
        name=base.name,
        size=base.size,
    )


def _write_count_cell(sheet, row, column, value):
    cell = sheet.cell(row=row, column=column, value=int(value))
    cell.number_format = "0"
    if COUNT_CELL_ALIGNMENT is not None:
        cell.alignment = COUNT_CELL_ALIGNMENT


def _sanitize_sales_float(raw_value):
    """
    Convert parsed Net Sales token into a native float safely.

    Removes currency symbols/grouping separators and defaults to 0.00 on failure.
    """
    token = _strip_cell(raw_value)
    token = token.replace("$", "").replace(",", "").strip()
    try:
        return float(token)
    except (TypeError, ValueError):
        return 0.00


def _write_amount_cell(sheet, row, column, value, department=None):
    amount = _sanitize_sales_float(value)
    if amount.is_integer():
        cell = sheet.cell(row=row, column=column, value=int(amount))
    else:
        cell = sheet.cell(row=row, column=column, value=amount)
    cell.number_format = "0.00"

    if department is not None:
        dept_label = re.sub(r"\s+", " ", _strip_cell(department)).strip().upper()
        if dept_label == "LOCAL ACCT":
            dept_label = "GETTEL/TOYOTA"

        if dept_label == "GETTEL/TOYOTA":
            if amount >= GETTEL_TOYOTA_THRESHOLD:
                _apply_sales_font_alert(cell)
            return

        if dept_label in GROUP_1400_DEPARTMENTS:
            if amount >= GROUP_1400_THRESHOLD:
                _apply_sales_font_alert(cell)
            return

        if dept_label in GROUP_500_DEPARTMENTS:
            if amount >= GROUP_500_THRESHOLD:
                _apply_sales_font_alert(cell)
            return


def inject_daily_sales(sheet, pdf_records, column_map, target_row):
    """
    Write parsed PDF Net Count / Net Sales $ into the mapped COUNT | NET SALES columns.

    Skips protected/formula-backed columns (e.g. GIFT CARD, VARIOS/BOLSA).
    Only writes to target_row — never clears or shifts other rows.
    """
    target_row = int(target_row)
    written = []
    skipped = []

    for record in pdf_records:
        dept_name = record["department"]
        if _is_protected_department(dept_name):
            skipped.append(dept_name)
            continue

        coords = _resolve_department_columns(dept_name, column_map)
        if coords is None:
            skipped.append(dept_name)
            continue

        count_col, amount_col = coords
        count_cell = sheet.cell(row=target_row, column=count_col)
        amount_cell = sheet.cell(row=target_row, column=amount_col)

        if _cell_has_formula(count_cell) or _cell_has_formula(amount_cell):
            skipped.append(dept_name)
            continue

        amount_value = _sanitize_sales_float(record["amount"])
        _write_count_cell(sheet, target_row, count_col, int(record["count"]))
        _write_amount_cell(
            sheet,
            target_row,
            amount_col,
            amount_value,
            department=dept_name,
        )
        written.append(
            {
                "department": dept_name,
                "count_col": get_column_letter(count_col),
                "amount_col": get_column_letter(amount_col),
                "count": int(record["count"]),
                "amount": float(record["amount"]),
            }
        )

    return written, skipped


def _read_manual_report_count(sheet, row):
    """
    BS — el total de unidades impreso en el reporte, cargado a mano por el
    usuario (BR, al lado, es una fórmula que suma lo ya cargado en la fila
    y no se toca acá).

    Un día real nunca imprime 0 unidades vendidas, así que un 0 literal se
    trata igual que una celda vacía (todavía no se cargó el dato a mano).
    Una fórmula sin calcular también se ignora, por las dudas.
    """
    raw = sheet.cell(row=row, column=MANUAL_REPORT_COUNT_COLUMN).value
    if raw is None or (isinstance(raw, str) and (not raw.strip() or raw.strip().startswith("="))):
        return None
    return _safe_parse_count(raw) or None


def check_loaded_count_against_manual_entry(sheet, target_row, column_map):
    """
    Compare BS (cargado a mano desde el reporte impreso) contra la suma de
    Net Count ya presente en todas las columnas de departamento de esa fila.

    Devuelve None si BS está vacío (todavía no se cargó) o si ambos totales
    coinciden; si no, un dict con los dos totales para armar el aviso.
    """
    manual_count = _read_manual_report_count(sheet, target_row)
    if manual_count is None:
        return None

    seen_columns = set()
    loaded_count = 0
    for count_col, _amount_col in column_map.values():
        if count_col in seen_columns:
            continue
        seen_columns.add(count_col)
        loaded_count += _safe_parse_count(sheet.cell(row=target_row, column=count_col).value)

    if loaded_count == manual_count:
        return None
    return {"manual_count": manual_count, "loaded_count": loaded_count}


def _normalize_pdf_paths(pdf_paths):
    if pdf_paths is None:
        return []
    if isinstance(pdf_paths, str):
        text = pdf_paths.strip()
        if not text:
            return []
        for separator in (";", "|"):
            if separator in text:
                return [
                    os.path.abspath(part.strip())
                    for part in text.split(separator)
                    if part.strip()
                ]
        return [os.path.abspath(text)]
    return [
        os.path.abspath(str(path).strip())
        for path in pdf_paths
        if path is not None and str(path).strip()
    ]


_MAX_CONCURRENT_PDF_WORKERS = 4


def _parse_pdfs_concurrently(paths, parse_fn, **kwargs):
    """
    Run parse_fn(path, **kwargs) for every path, in parallel when there's
    more than one.

    Each call ultimately shells out to Tesseract as its own OS process, so
    this isn't fighting Python's GIL — a batch of daily PDFs really can be
    read at the same time on a multi-core machine instead of strictly one
    after another, with no change to how any single PDF is read. Capped at
    a handful of workers so a big batch doesn't overwhelm the machine.
    """

    def _run(path):
        try:
            return parse_fn(path, **kwargs)
        except Exception as exc:
            raise type(exc)(f"{os.path.basename(path)}: {exc}") from exc

    if len(paths) <= 1:
        return {path: _run(path) for path in paths}

    max_workers = min(len(paths), os.cpu_count() or _MAX_CONCURRENT_PDF_WORKERS, _MAX_CONCURRENT_PDF_WORKERS)
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_run, path): path for path in paths}
        for future in as_completed(futures):
            results[futures[future]] = future.result()
    return results


def process_reporte_diario(
    master_path, pdf_paths, page_index=DEFAULT_PDF_PAGE_INDEX
):
    """
    Parse one or more daily PDFs and inject each into its calendar day row.

    Day-of-month is read from each PDF filename (e.g. Close Store 07-05.pdf -> day 7).
    Each PDF is read (OCR included) in parallel across a small worker pool,
    since that read — not the Excel write that follows — is what makes a
    multi-PDF batch slow. Workbook is saved once after the full batch completes.

    Returns:
        tuple: (temp_path, summary dict)
    """
    _ensure_openpyxl()
    master_path = os.path.abspath(str(master_path).strip())
    paths = _normalize_pdf_paths(pdf_paths)

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel maestro no encontrado: {master_path}")
    if not paths:
        raise ValueError("No se proporcionaron PDF diarios.")

    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel maestro debe ser .xlsx o .xlsm.")

    sorted_paths = sorted(paths, key=extract_day_from_filename)
    for pdf_path in sorted_paths:
        if not os.path.isfile(pdf_path):
            raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    parsed_by_path = _parse_pdfs_concurrently(
        sorted_paths, parse_elistar_daily_pdf_page, page_index=page_index
    )

    keep_vba = extension == ".xlsm"
    workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
    sheet = _get_carga_aqui_sheet(workbook)
    column_map, protected = build_department_column_map(sheet)

    batch_results = []
    total_written = 0
    total_skipped = 0
    total_departments = 0

    for pdf_path in sorted_paths:
        pdf_records, pdf_diagnostics = parsed_by_path[pdf_path]
        target_day = int(extract_day_from_filename(pdf_path))
        target_row = find_row_for_calendar_day(sheet, target_day)

        written, skipped = inject_daily_sales(
            sheet, pdf_records, column_map, target_row
        )

        total_written += len(written)
        total_skipped += len(skipped)
        total_departments += len(pdf_records)

        last_department = pdf_diagnostics.get("last_department")
        warnings = []
        if pdf_diagnostics.get("used_ocr") and last_department not in (
            None,
        ) + LAST_DEPARTMENT_CANDIDATES:
            warnings.append(
                f"El último departamento leído fue \"{last_department}\", no "
                f"{'/'.join(LAST_DEPARTMENT_CANDIDATES)} — revise si la tabla se cortó."
            )
        mismatch = pdf_diagnostics.get("subtotal_mismatch")
        if mismatch:
            warnings.append(
                "El total impreso en el PDF no coincide con lo leído: "
                f"{mismatch['computed_count']} vs {mismatch['printed_count']} unidades, "
                f"${mismatch['computed_amount']:.2f} vs ${mismatch['printed_amount']:.2f} — "
                "revise el OCR."
            )
        manual_mismatch = check_loaded_count_against_manual_entry(
            sheet, target_row, column_map
        )
        if manual_mismatch:
            warnings.append(
                f"La columna BS dice {manual_mismatch['manual_count']} unidades, pero "
                f"en la fila hay {manual_mismatch['loaded_count']} cargadas — revise "
                "los departamentos."
            )
        warning = " | ".join(warnings) if warnings else None

        batch_results.append(
            {
                "pdf_path": pdf_path,
                "filename": os.path.basename(pdf_path),
                "calendar_day": target_day,
                "target_row": target_row,
                "departments_written": len(written),
                "departments_skipped": len(skipped),
                "written": list(written),
                "skipped": list(skipped),
                "used_ocr": bool(pdf_diagnostics.get("used_ocr")),
                "pages_used": pdf_diagnostics.get("pages_used"),
                "warning": warning,
            }
        )
        pdf_records = None

    temp_path = _create_temp_workbook_path()
    workbook.save(os.path.abspath(temp_path))
    workbook.close()

    summary = {
        "files_processed": len(batch_results),
        "departments_written": total_written,
        "departments_skipped": total_skipped,
        "pdf_departments": total_departments,
        "protected_headers": sorted(protected),
        "batch_results": batch_results,
    }
    return temp_path, summary


# ---------------------------------------------------------------------------
# Store Info — a second extraction from the same daily PDF (pages 3
# and the start of 4, up to "Network Revenue") into a different workbook's
# "Store Info" sheet: one summary row per day, appended after the last one.
# ---------------------------------------------------------------------------

STORE_INFO_SHEET_NAME = "Store Info"
STORE_INFO_DATA_START_ROW = 2
DEFAULT_STORE_INFO_PAGE_INDEX = 2  # Third page (0-based) — "PERIOD FROM:" anchor
STORE_INFO_MAX_CONTINUATION_PAGES = 3  # Anchor page + up to 2 more, to reach Network Revenue

PERIOD_FROM_ANCHOR = "period from"
NETWORK_REVENUE_ANCHOR = "network revenue"

# 1-based column numbers for the fields this extraction is allowed to touch.
# Everything else (H, I..N, R, U, W) is either a formula or filled in some
# other way and must never be written here.
STORE_INFO_COL_FROM_DATE = 1  # A — Fecha (FROM date + 1 day)
STORE_INFO_COL_FROM_TIME = 2  # B — hs (FROM time, unchanged)
STORE_INFO_COL_TO_DATE = 3  # C — Fecha (FROM date + 2 days)
STORE_INFO_COL_TO_TIME = 4  # D — hs (TO time, unchanged)
STORE_INFO_COL_VOLUME = 5  # E — Volume
STORE_INFO_COL_SALES_FUEL = 6  # F — SALES FUEL
STORE_INFO_COL_DESC_COMB = 7  # G — Desc. Comb
STORE_INFO_COL_NON_FUEL = 15  # O — C-Store=Total Non Fuel
STORE_INFO_COL_DESC_OTROS = 16  # P — desc otros
STORE_INFO_COL_TAX_COLLECT = 17  # Q — Tax collet
STORE_INFO_COL_CASH = 19  # S — Cash
STORE_INFO_COL_CREDIT = 20  # T — TC (written as a "=a+b+c" formula, like the sheet's own history)
STORE_INFO_COL_LOCAL_ACCOUNTS = 22  # V — Local Account
STORE_INFO_COL_NETWORK_REVENUE = 24  # X — Network Revenue

_MONTH_NAME_TO_NUMBER = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

_PERIOD_FROM_TO_RE = re.compile(
    r"([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})\s+(\d{1,2}):(\d{2})\s*([AP])\.?M\.?"
    r".*?TO:?\s*([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})\s+(\d{1,2}):(\d{2})\s*([AP])\.?M\.?",
    re.IGNORECASE,
)

# A label line is "words... <first numeric-looking token> ...": everything
# before that split point is the label, everything from it on are the values.
_NUMERIC_TOKEN_RE = re.compile(r"^[-~(]*\$?\(?-?[\d,]+\.?\d*\)?%?$")

_STORE_INFO_SCORING_ANCHORS = (
    "period from",
    "total fuel sales",
    "fuel discounts",
    "total non fuel sales",
    "other discounts",
    "total taxes collected",
    "local accounts",
    "network revenue",
)


def _to_24h_time(hour12, minute, ampm):
    hour12 = int(hour12) % 12
    if ampm.upper().startswith("P"):
        return time(hour12 + 12, int(minute))
    return time(hour12, int(minute))


def _parse_period_from_to_line(line):
    """
    Parse "PERIOD FROM: Aug 18, 2026 10:51 PM TO: Aug 19, 2026 10:42 PM".

    Returns a dict with from_date/from_time/to_date/to_time, or None.
    """
    match = _PERIOD_FROM_TO_RE.search(line)
    if not match:
        return None
    (
        from_mon, from_day, from_year, from_hh, from_mm, from_ampm,
        to_mon, to_day, to_year, to_hh, to_mm, to_ampm,
    ) = match.groups()
    from_month = _MONTH_NAME_TO_NUMBER.get(from_mon[:3].lower())
    to_month = _MONTH_NAME_TO_NUMBER.get(to_mon[:3].lower())
    if not from_month or not to_month:
        return None
    try:
        from_date_value = date(int(from_year), from_month, int(from_day))
        to_date_value = date(int(to_year), to_month, int(to_day))
    except ValueError:
        return None
    return {
        "from_date": from_date_value,
        "from_time": _to_24h_time(from_hh, from_mm, from_ampm),
        "to_date": to_date_value,
        "to_time": _to_24h_time(to_hh, to_mm, to_ampm),
    }


def _split_label_and_values(line):
    """
    Split "Total Fuel Sales 1,236.070 $5,152.10" into label + value tokens.

    Strips stray punctuation off the label's edge (e.g. OCR sometimes reads
    a faint column rule next to "Cash" as "Cash :", or a graphic behind the
    text as "Cash |)") so it still matches the real label exactly, without
    loosening the match enough to also catch a longer label like "Cash
    Acceptor Cash". Value tokens are filtered down to ones that actually
    look numeric, so a trailing OCR artifact (e.g. "Total Taxes Collected
    $157.89 ;") doesn't get picked up as the amount.
    """
    tokens = [t for t in _normalize_department_spacing(_strip_cell(line)).split() if t]
    for index, token in enumerate(tokens):
        if _NUMERIC_TOKEN_RE.match(token) and index > 0:
            label = re.sub(r"[^A-Za-z]+$", "", " ".join(tokens[:index])).strip()
            values = [t for t in tokens[index:] if _NUMERIC_TOKEN_RE.match(t)]
            return label, values
    return None, []


def _sanitize_store_info_float(raw_value):
    """Like _sanitize_sales_float, but also treats a leading '~' as a minus sign (OCR glyph noise)."""
    text = _strip_cell(raw_value)
    if text.startswith("~"):
        text = "-" + text[1:]
    return _sanitize_sales_float(text)


def _find_label_values(lines, *target_labels):
    """Return the value tokens of the first line whose label exactly matches one of target_labels."""
    targets = {label.lower() for label in target_labels}
    for line in lines:
        label, values = _split_label_and_values(line)
        if label and label.lower() in targets:
            return values
    return None


def _score_store_info_page_text(text):
    lower = text.lower() if text else ""
    return sum(1 for anchor in _STORE_INFO_SCORING_ANCHORS if anchor in lower)


def _ocr_store_info_page_text(image):
    """Best-effort OCR of a Store Info page — scored by how many known anchors it recovers."""
    if image is None:
        return ""
    _ensure_pytesseract()
    best_text = ""
    best_score = -1
    for config in _OCR_TEXT_CONFIGS:
        try:
            text = pytesseract.image_to_string(image, config=config) or ""
        except Exception:
            continue
        score = _score_store_info_page_text(text)
        if score > best_score:
            best_score = score
            best_text = text
    return best_text


def _force_positive(value):
    return abs(value)


def _force_negative(value):
    return -abs(value)


def _extract_store_info_fields(lines):
    """
    Pull every Store Info value out of the OCR'd lines of pages 3(-4).

    Every field's sign is fixed by its known business meaning rather than
    trusted from the OCR'd "-" glyph, which is one of the easiest characters
    for Tesseract to drop or invent on a blurry photo: Desc. Comb (G) and
    desc otros (P) are always a discount (negative); every other dollar
    figure and the fuel Volume are always positive.
    """
    period = None
    for line in lines:
        if PERIOD_FROM_ANCHOR in line.lower():
            period = _parse_period_from_to_line(line)
            if period:
                break
    if period is None:
        raise ValueError('No se encontró la línea "PERIOD FROM: ... TO: ..." en el PDF.')

    fuel_values = _find_label_values(lines, "Total Fuel Sales")
    if not fuel_values or len(fuel_values) < 2:
        raise ValueError('No se encontró "Total Fuel Sales" con Volume y Sales.')
    volume = _force_positive(_sanitize_store_info_float(fuel_values[0]))
    sales_fuel = _force_positive(_sanitize_store_info_float(fuel_values[-1]))

    fuel_discounts = _find_label_values(lines, "Fuel Discounts")
    desc_comb = (
        _force_negative(_sanitize_store_info_float(fuel_discounts[-1]))
        if fuel_discounts
        else 0.0
    )

    non_fuel = _find_label_values(lines, "Total Non Fuel Sales")
    if not non_fuel:
        raise ValueError('No se encontró "Total Non Fuel Sales".')
    non_fuel_total = _force_positive(_sanitize_store_info_float(non_fuel[-1]))

    other_discounts = _find_label_values(lines, "Other Discounts")
    desc_otros = (
        _force_negative(_sanitize_store_info_float(other_discounts[-1]))
        if other_discounts
        else 0.0
    )

    taxes = _find_label_values(lines, "Total Taxes Collected")
    if not taxes:
        raise ValueError('No se encontró "Total Taxes Collected".')
    tax_collect = _force_positive(_sanitize_store_info_float(taxes[-1]))

    cash_values = _find_label_values(lines, "Cash")
    if not cash_values:
        raise ValueError('No se encontró la fila "Cash" bajo Method of Payment Totals.')
    cash = _force_positive(_sanitize_store_info_float(cash_values[-1]))

    # Every payment-method row strictly between "Cash" and "LOCAL ACCOUNTS"
    # (Credit, Crind CREDIT/DEBIT, CRIND P97, Debit, etc.) gets summed into
    # the credit-card total — zero-valued rows are dropped, same as the
    # sheet's own historical "=a+b+c" formulas.
    credit_terms = []
    local_accounts = None
    in_range = False
    for line in lines:
        label, values = _split_label_and_values(line)
        if label is None:
            continue
        norm_label = label.lower()
        if norm_label == "cash":
            in_range = True
            continue
        if norm_label == "local accounts":
            if values:
                local_accounts = _force_positive(_sanitize_store_info_float(values[-1]))
            break
        if in_range and values:
            amount = _force_positive(_sanitize_store_info_float(values[-1]))
            if amount:
                credit_terms.append(amount)

    if local_accounts is None:
        raise ValueError('No se encontró la fila "LOCAL ACCOUNTS".')

    network_values = _find_label_values(lines, "Network Revenue")
    if not network_values:
        raise ValueError('No se encontró "Network Revenue".')
    network_revenue = _force_positive(_sanitize_store_info_float(network_values[-1]))

    # No se usa para escribir Store Info (esa fila sigue viniendo de
    # cash+credit_terms+local_accounts, columna por columna) -- se agrega
    # solo como dato adicional para el control de Cierre mensual
    # (controles_cierre_mensual.py), que necesita el "Total Revenue" tal
    # como lo imprime el propio POS para cruzarlo contra el total del mes
    # ya cargado en la hoja, sin depender de si esta función terminó
    # sumando exactamente lo mismo (ver _extract_store_info_fields arriba:
    # credit_terms deja de sumar apenas llega a "LOCAL ACCOUNTS", así que
    # cualquier categoría de pago rara que aparezca después -- Loyalty,
    # Other, Overruns, etc. -- no entra ahí, pero sí está reflejada en el
    # "Total Revenue" que imprime el reporte).
    total_revenue_values = _find_label_values(lines, "Total Revenue")
    if not total_revenue_values:
        raise ValueError('No se encontró "Total Revenue".')
    total_revenue = _force_positive(_sanitize_store_info_float(total_revenue_values[-1]))

    return {
        "from_date": period["from_date"],
        "from_time": period["from_time"],
        "to_date": period["to_date"],
        "to_time": period["to_time"],
        "volume": volume,
        "sales_fuel": sales_fuel,
        "desc_comb": desc_comb,
        "non_fuel_total": non_fuel_total,
        "desc_otros": desc_otros,
        "tax_collect": tax_collect,
        "cash": cash,
        "credit_terms": credit_terms,
        "local_accounts": local_accounts,
        "network_revenue": network_revenue,
        "total_revenue": total_revenue,
    }


def extract_store_info_from_pdf(pdf_path, start_page_index=DEFAULT_STORE_INFO_PAGE_INDEX):
    """
    OCR the "PERIOD FROM:" page (generally page 3) plus as many following
    pages as needed to reach "Network Revenue" (generally the start of page
    4), and pull every Store Info field out of the combined text.
    """
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    images = _LazyPdfPageImages(pdf_path)
    try:
        total_pages = len(images)
        if not (0 <= start_page_index < total_pages):
            start_page_index = 0

        search_order = list(range(start_page_index, total_pages)) + list(
            range(0, start_page_index)
        )

        anchor_page = None
        anchor_text = None
        for idx in search_order:
            text = _ocr_store_info_page_text(images[idx])
            if PERIOD_FROM_ANCHOR in text.lower():
                anchor_page = idx
                anchor_text = text
                break

        if anchor_page is None:
            raise ValueError(
                'No se encontró el ancla "PERIOD FROM:" en ninguna página del PDF (vía OCR). '
                "Verifique que el reporte no esté demasiado borroso o girado."
            )

        lines = list(anchor_text.splitlines())
        pages_used = [anchor_page + 1]
        idx = anchor_page + 1
        pages_tried = 1
        while (
            NETWORK_REVENUE_ANCHOR not in "\n".join(lines).lower()
            and idx < total_pages
            and pages_tried < STORE_INFO_MAX_CONTINUATION_PAGES
        ):
            text = _ocr_store_info_page_text(images[idx])
            lines.extend(text.splitlines())
            pages_used.append(idx + 1)
            pages_tried += 1
            idx += 1
    finally:
        images.close()

    fields = _extract_store_info_fields(lines)
    fields["pages_used"] = pages_used
    return fields


def _find_store_info_sheet(workbook):
    target = STORE_INFO_SHEET_NAME.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    raise ValueError(
        f'Hoja "{STORE_INFO_SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _store_info_row_for_day(day_of_month):
    """
    Row N holds day (N-1) of the month — row 2 is always day 1 — matching
    how CARGA AQUI itself pins one calendar day to one fixed row. This is
    what lets a later, out-of-order PDF (day 10 after day 1) land on the row
    that actually corresponds to it instead of just the next blank one,
    leaving days 2-9 correctly blank until their own reports arrive.
    """
    return STORE_INFO_DATA_START_ROW + (day_of_month - 1)


def _build_credit_terms_formula(amounts):
    """Mirror the sheet's own history: '=a+b+c', omitting zero-valued rows."""
    if not amounts:
        return 0.0
    return "=" + "+".join(f"{amount:.2f}" for amount in amounts)


def write_store_info_row(sheet, fields):
    """Write one Store Info row on the row matching this report's calendar day."""
    from_date = fields["from_date"]
    to_date = fields["to_date"]
    col_a_date = datetime(from_date.year, from_date.month, from_date.day) + timedelta(days=1)
    col_c_date = datetime(to_date.year, to_date.month, to_date.day) + timedelta(days=1)
    row = _store_info_row_for_day(col_a_date.day)

    sheet.cell(row=row, column=STORE_INFO_COL_FROM_DATE, value=col_a_date)
    sheet.cell(row=row, column=STORE_INFO_COL_FROM_TIME, value=fields["from_time"])
    sheet.cell(
        row=row,
        column=STORE_INFO_COL_TO_DATE,
        value=col_c_date,
    )
    sheet.cell(row=row, column=STORE_INFO_COL_TO_TIME, value=fields["to_time"])
    sheet.cell(row=row, column=STORE_INFO_COL_VOLUME, value=fields["volume"])
    sheet.cell(row=row, column=STORE_INFO_COL_SALES_FUEL, value=fields["sales_fuel"])
    sheet.cell(row=row, column=STORE_INFO_COL_DESC_COMB, value=fields["desc_comb"])
    sheet.cell(row=row, column=STORE_INFO_COL_NON_FUEL, value=fields["non_fuel_total"])
    sheet.cell(row=row, column=STORE_INFO_COL_DESC_OTROS, value=fields["desc_otros"])
    sheet.cell(row=row, column=STORE_INFO_COL_TAX_COLLECT, value=fields["tax_collect"])
    sheet.cell(row=row, column=STORE_INFO_COL_CASH, value=fields["cash"])
    sheet.cell(
        row=row,
        column=STORE_INFO_COL_CREDIT,
        value=_build_credit_terms_formula(fields["credit_terms"]),
    )
    sheet.cell(row=row, column=STORE_INFO_COL_LOCAL_ACCOUNTS, value=fields["local_accounts"])
    sheet.cell(
        row=row, column=STORE_INFO_COL_NETWORK_REVENUE, value=fields["network_revenue"]
    )
    return row


def process_store_info(master_path, pdf_paths):
    """
    Parse one or more daily PDFs and write one Store Info row per day to the
    "Store Info" sheet of a separate workbook — each on the row matching its
    own calendar day (row 2 = day 1), not just the next blank row, so PDFs
    for non-consecutive days land where they belong and gaps stay blank.

    Each PDF is read (OCR included) in parallel across a small worker pool,
    since that read — not the Excel write that follows — is what makes a
    multi-PDF batch slow.

    Returns:
        tuple: (temp_path, summary dict)
    """
    _ensure_openpyxl()
    master_path = os.path.abspath(str(master_path).strip())
    paths = _normalize_pdf_paths(pdf_paths)

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel de Store Info no encontrado: {master_path}")
    if not paths:
        raise ValueError("No se proporcionaron PDF diarios.")

    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel de Store Info debe ser .xlsx o .xlsm.")

    sorted_paths = sorted(paths, key=extract_day_from_filename)
    for pdf_path in sorted_paths:
        if not os.path.isfile(pdf_path):
            raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    fields_by_path = _parse_pdfs_concurrently(sorted_paths, extract_store_info_from_pdf)

    keep_vba = extension == ".xlsm"
    workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
    sheet = _find_store_info_sheet(workbook)

    batch_results = []

    for pdf_path in sorted_paths:
        fields = fields_by_path[pdf_path]
        row = write_store_info_row(sheet, fields)
        batch_results.append(
            {
                "pdf_path": pdf_path,
                "filename": os.path.basename(pdf_path),
                "target_row": row,
                "from_date": fields["from_date"].strftime("%d/%m/%Y"),
                "pages_used": fields.get("pages_used"),
            }
        )

    temp_path = _create_temp_workbook_path()
    workbook.save(os.path.abspath(temp_path))
    workbook.close()

    summary = {
        "files_processed": len(batch_results),
        "batch_results": batch_results,
    }
    return temp_path, summary


# ---------------------------------------------------------------------------
# Lottery — a third extraction from the same daily PDF, this time from its
# last two pages (Florida Lottery terminal receipts: "Daily Terminal Games
# Sales" for ONLINE and "Daily Scratch-Off Games Sales" for SKOFF), combined
# with the ONLINE/SKOFF rows already read off the Department Sales Report
# for the Ventas pipeline, into one row of the monthly Lottery workbook.
# ---------------------------------------------------------------------------

LOTTERY_DATA_START_ROW = 4  # Row 3 is the header; data starts at row 4
LOTTERY_COL_DATE_A = 1  # A — Fecha (period start, informational only)
LOTTERY_COL_DATE_B = 2  # B — Fecha (the date every row is actually keyed on)
LOTTERY_COL_ONLINE_COUNT = 4  # D — ONLINE Net Count (from Department Sales Report)
LOTTERY_COL_ONLINE_NET_SALES = 5  # E — ONLINE Net Sales (from Department Sales Report)
LOTTERY_COL_SALES = 6  # F — NET SALES off the terminal-games receipt
LOTTERY_COL_PAGOS = 7  # G — PAYS off the terminal-games receipt
LOTTERY_COL_CASH_BALANCE = 8  # H — F - G, pasted as a value, never a formula
LOTTERY_COL_COMIS = 9  # I — the NET SALES nested under TOTAL SALES COMM
LOTTERY_COL_PRIZE_FREE_PLAYS = 11  # K — PRIZE FREE PLAYS (+ PROMO FREE PLAYS if any)
LOTTERY_COL_SKOFF_COUNT = 14  # N — SKOFF Net Count (from Department Sales Report)
LOTTERY_COL_SKOFF_NET_SALES = 15  # O — SKOFF Net Sales (from Department Sales Report)
LOTTERY_COL_SKOFF_PAYS_UNITS = 16  # P — PAYS unit count off the scratch-off receipt
LOTTERY_COL_SKOFF_PAYS_AMOUNT = 17  # Q — PAYS $ amount off the scratch-off receipt
LOTTERY_COL_SKOFF_SALES_AMOUNT = 18  # R — "Instant Sales Amount" off the Daily Sales Report
LOTTERY_COL_SALES_COMM = 19  # S — SALES COMM off the scratch-off receipt

LOTTERY_ONLINE_DEPARTMENT = "ONLINE"
LOTTERY_SKOFF_DEPARTMENT = "SKOFF"


def _find_department_record(records, department_name):
    for record in records:
        if record["department"] == department_name:
            return record
    return None


def extract_lottery_department_fields_from_pdf(pdf_path, page_index=DEFAULT_PDF_PAGE_INDEX):
    """
    D/E/N/O — the ONLINE/SKOFF rows off the Department Sales Report (same
    page the Ventas pipeline reads) — plus this PDF's own business date.

    The date is read the same way Store Info reads it (PERIOD FROM/TO, +1
    day) rather than off the Lottery receipt pages: those are small, dense,
    watermark-obscured text where OCR has repeatedly misread a single digit
    (e.g. "08" -> "09"), while PERIOD FROM/TO is a larger, cleaner line that
    Store Info already reads reliably from this exact same PDF.

    F/G/H/I/K and P/Q/R/S no longer come from here: see
    extract_lottery_receipt_fields_from_sales_report, which reads those
    columns from the Florida Lottery portal's own "Daily Sales Report"
    PDF — real embedded text instead of a photographed receipt.
    """
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    dept_records, _dept_diagnostics = parse_elistar_daily_pdf_page(pdf_path, page_index=page_index)
    online_record = _find_department_record(dept_records, LOTTERY_ONLINE_DEPARTMENT)
    if online_record is None:
        raise ValueError(
            f'No se encontró el departamento "{LOTTERY_ONLINE_DEPARTMENT}" en el Department Sales Report.'
        )
    skoff_record = _find_department_record(dept_records, LOTTERY_SKOFF_DEPARTMENT)
    if skoff_record is None:
        raise ValueError(
            f'No se encontró el departamento "{LOTTERY_SKOFF_DEPARTMENT}" en el Department Sales Report.'
        )

    store_info_fields = extract_store_info_from_pdf(pdf_path)
    from_date = store_info_fields["from_date"]
    report_date = date(from_date.year, from_date.month, from_date.day) + timedelta(days=1)

    return {
        "report_date": report_date,
        "online_count": int(online_record["count"]),
        "online_net_sales": float(online_record["amount"]),
        "skoff_count": int(skoff_record["count"]),
        "skoff_net_sales": float(skoff_record["amount"]),
    }


_LOTTERY_SALES_REPORT_START_DATE_RE = re.compile(r"Start Date:\s*(\d{4})-(\d{2})-(\d{2})")

# The Lottery's own commission rate on ONLINE net sales — column J's own
# "=+I/F" formula is meant to always read -6.00%; confirmed exactly against
# three independent real days (comis/sales = -15.66/261, -15.0/250, -5.82/97).
_LOTTERY_ONLINE_COMMISSION_RATE = 0.06

# label -> (result key, sign rule) for every value this module needs off the
# Florida Lottery portal's "Daily Sales Report" PDF. Unlike the photographed
# terminal receipt, this document has a real embedded text layer, so each
# label is matched once, exactly, with no fuzzy/OCR tolerance needed.
_LOTTERY_SALES_REPORT_FIELDS = (
    ("sales", "Net Terminal Sales Amount", _force_positive),
    ("pagos", "Terminal Pay Amount", _force_negative),
    ("comis", "Terminal Sales Commission", _force_negative),
    ("pays_units", "Instant Tickets Paid", None),
    ("pays_amount", "Instant Pay Amount", _force_negative),
    ("skoff_sales_amount", "Instant Sales Amount", _force_positive),
    ("sales_comm", "Instant Sales Commission", _force_negative),
)


def _parse_lottery_sales_report_date(text):
    match = _LOTTERY_SALES_REPORT_START_DATE_RE.search(text)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _extract_sales_report_value(text, label):
    """Numeric token immediately after an exact label — safe here since the label set has no overlapping substrings."""
    match = re.search(re.escape(label) + r"\s*\$?(-?[\d,]+\.?\d*)", text)
    if not match:
        return None
    return match.group(1).replace(",", "")


def extract_lottery_receipt_fields_from_sales_report(pdf_path):
    """
    F/G/H/I/K (ONLINE) and P/Q/R/S (SKOFF) off the Florida Lottery portal's
    own "Daily Sales Report" PDF — real embedded text, not a photographed
    terminal receipt, so there's no OCR and no ambiguity. This source only
    prints the combined "Terminal Sales Commission" total, not the NET
    SALES / PRIZE FREE PLAYS split the physical receipt shows — but I/F is
    fixed at -6.00% by the Lottery's own commission rate (column J's own
    "=+I/F" formula), so I is rebuilt from that fixed rate against F and K
    takes whatever's left of the total, reproducing the same split the
    receipt shows without needing to read it.
    """
    _ensure_pdfplumber()
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    with pdfplumber.open(pdf_path) as pdf:
        raw_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    text = _normalize_department_spacing(raw_text)

    report_date = _parse_lottery_sales_report_date(text)
    if report_date is None:
        raise ValueError('No se encontró "Start Date" en el Daily Sales Report de Lottery.')

    warnings = []
    values = {}
    for key, label, sign_rule in _LOTTERY_SALES_REPORT_FIELDS:
        raw = _extract_sales_report_value(text, label)
        if raw is None:
            values[key] = None
            warnings.append(f'No se pudo leer "{label}" en el Daily Sales Report de Lottery — revisar manualmente.')
            continue
        number = float(raw)
        values[key] = sign_rule(number) if sign_rule else _safe_parse_count(raw)

    sales = values["sales"]
    pagos = values["pagos"]
    cash_balance = round(sales + pagos, 2) if sales is not None and pagos is not None else None

    total_comm = values["comis"]
    if total_comm is not None and sales is not None:
        comis = round(-_LOTTERY_ONLINE_COMMISSION_RATE * sales, 2)
        prize_free_plays = round(total_comm - comis, 2)
    elif total_comm is not None:
        comis = total_comm
        prize_free_plays = 0.0
        warnings.append(
            'No se pudo leer "Net Terminal Sales Amount" (columna F), así que no se pudo dividir '
            "la comisión entre I y K — se cargó el total completo en I."
        )
    else:
        comis = None
        prize_free_plays = None

    return {
        "report_date": report_date,
        "sales": sales,
        "pagos": pagos,
        "cash_balance": cash_balance,
        "comis": comis,
        "prize_free_plays": prize_free_plays,
        "pays_units": values["pays_units"],
        "pays_amount": values["pays_amount"],
        "skoff_sales_amount": values["skoff_sales_amount"],
        "sales_comm": values["sales_comm"],
        "warning": " | ".join(warnings) if warnings else None,
    }


def _find_lottery_sheet(workbook):
    """
    The Lottery workbook has one sheet per month, named after the month
    (e.g. "08.2026") — so unlike CARGA AQUI or Store Info there's no fixed
    name to look for. Falls back to the sheet matching the known header
    signature (Fecha/Fecha/.../COUNT in row 3) when there's more than one.
    """
    names = workbook.sheetnames
    if len(names) == 1:
        return workbook[names[0]]

    for name in names:
        sheet = workbook[name]
        header_a = _strip_cell(sheet.cell(row=3, column=LOTTERY_COL_DATE_A).value).lower()
        header_d = _strip_cell(sheet.cell(row=3, column=LOTTERY_COL_ONLINE_COUNT).value).lower()
        if header_a == "fecha" and header_d == "count":
            return sheet

    raise ValueError(
        f'No se pudo identificar la hoja de Lottery. Disponibles: {", ".join(names)}'
    )


def _find_lottery_row_for_date(sheet, target_date, start_row=LOTTERY_DATA_START_ROW):
    """
    Exact (day, month, year) match first; if that fails, fall back to
    (day, month) alone. A single Lottery workbook only ever spans one
    month, so a day/month match is already unambiguous — and this recovers
    from a lone OCR-misread year digit (e.g. "26" read as "25"), which
    doesn't affect which real-world day the report is for.
    """
    max_row = max(sheet.max_row, start_row)
    day_month_row = None
    for row in range(start_row, max_row + 1):
        cell_value = sheet.cell(row=row, column=LOTTERY_COL_DATE_B).value
        if not isinstance(cell_value, datetime):
            continue
        if cell_value.date() == target_date:
            return row
        if day_month_row is None and (cell_value.month, cell_value.day) == (
            target_date.month,
            target_date.day,
        ):
            day_month_row = row
    if day_month_row is not None:
        return day_month_row
    raise ValueError(
        f"No se encontró una fila con la fecha {target_date.strftime('%d/%m/%Y')} en la "
        f"columna B de la hoja de Lottery."
    )


def write_lottery_row(sheet, fields):
    row = _find_lottery_row_for_date(sheet, fields["report_date"])

    if fields.get("online_count") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_ONLINE_COUNT, value=fields["online_count"])
    if fields.get("online_net_sales") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_ONLINE_NET_SALES, value=fields["online_net_sales"])
    if fields.get("sales") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SALES, value=fields["sales"])
    if fields.get("pagos") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_PAGOS, value=fields["pagos"])
    if fields.get("cash_balance") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_CASH_BALANCE, value=fields["cash_balance"])
    if fields.get("comis") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_COMIS, value=fields["comis"])
    if fields.get("prize_free_plays") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_PRIZE_FREE_PLAYS, value=fields["prize_free_plays"])
    if fields.get("skoff_count") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SKOFF_COUNT, value=fields["skoff_count"])
    if fields.get("skoff_net_sales") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SKOFF_NET_SALES, value=fields["skoff_net_sales"])
    if fields.get("pays_units") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SKOFF_PAYS_UNITS, value=fields["pays_units"])
    if fields.get("pays_amount") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SKOFF_PAYS_AMOUNT, value=fields["pays_amount"])
    if fields.get("skoff_sales_amount") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SKOFF_SALES_AMOUNT, value=fields["skoff_sales_amount"])
    if fields.get("sales_comm") is not None:
        sheet.cell(row=row, column=LOTTERY_COL_SALES_COMM, value=fields["sales_comm"])
    return row


def process_lottery(master_path, department_pdf_paths, sales_report_pdf_paths):
    """
    Write one Lottery row per business date, merging two independent PDF
    sources by that date: the shared daily PDF's Department Sales Report
    (D/E/N/O — ONLINE/SKOFF counts and sales, same source the Ventas
    pipeline reads) and the Florida Lottery portal's own "Daily Sales
    Report" PDF (F/G/H/I/K, P/Q/R/S). Either source can be given alone —
    a date present in only one still gets a row written with just that
    source's columns, and a warning noting the other is missing — since a
    given batch of uploads won't always include both for every day.

    Each PDF set is parsed in parallel (within its own set) before any
    writing happens, since that read is what makes a multi-PDF batch slow.

    Returns:
        tuple: (temp_path, summary dict)
    """
    _ensure_openpyxl()
    master_path = os.path.abspath(str(master_path).strip())
    department_paths = _normalize_pdf_paths(department_pdf_paths)
    sales_report_paths = _normalize_pdf_paths(sales_report_pdf_paths)

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel de Lottery no encontrado: {master_path}")
    if not department_paths and not sales_report_paths:
        raise ValueError("No se proporcionaron PDFs de Lottery.")

    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel de Lottery debe ser .xlsx o .xlsm.")

    for pdf_path in department_paths + sales_report_paths:
        if not os.path.isfile(pdf_path):
            raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    department_fields_by_path = (
        _parse_pdfs_concurrently(department_paths, extract_lottery_department_fields_from_pdf)
        if department_paths
        else {}
    )
    sales_report_fields_by_path = (
        _parse_pdfs_concurrently(sales_report_paths, extract_lottery_receipt_fields_from_sales_report)
        if sales_report_paths
        else {}
    )

    department_by_date = {
        fields["report_date"]: (path, fields) for path, fields in department_fields_by_path.items()
    }
    sales_report_by_date = {
        fields["report_date"]: (path, fields) for path, fields in sales_report_fields_by_path.items()
    }

    keep_vba = extension == ".xlsm"
    workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
    sheet = _find_lottery_sheet(workbook)

    all_dates = sorted(set(department_by_date) | set(sales_report_by_date))

    batch_results = []
    for report_date in all_dates:
        dept_path, dept_fields = department_by_date.get(report_date, (None, None))
        sales_path, sales_fields = sales_report_by_date.get(report_date, (None, None))

        merged = {"report_date": report_date}
        warnings = []
        if dept_fields is not None:
            merged.update(
                {
                    "online_count": dept_fields["online_count"],
                    "online_net_sales": dept_fields["online_net_sales"],
                    "skoff_count": dept_fields["skoff_count"],
                    "skoff_net_sales": dept_fields["skoff_net_sales"],
                }
            )
        elif department_paths and sales_report_paths:
            # Only worth flagging when this run intentionally supplied both
            # sources and one of them simply has no PDF for this particular
            # date — a single-source run (the normal case: the GUI's two
            # Lottery flows each call this with the other list empty) isn't
            # missing anything, it just never touches the other's columns.
            warnings.append(
                "No se subió el PDF diario con el Department Sales Report para esta fecha "
                "— columnas D/E/N/O sin actualizar."
            )
        if sales_fields is not None:
            for key in (
                "sales", "pagos", "cash_balance", "comis", "prize_free_plays",
                "pays_units", "pays_amount", "skoff_sales_amount", "sales_comm",
            ):
                merged[key] = sales_fields[key]
            if sales_fields.get("warning"):
                warnings.append(sales_fields["warning"])
        elif department_paths and sales_report_paths:
            warnings.append(
                "No se subió el Daily Sales Report de Lottery para esta fecha "
                "— columnas F/G/H/I/K/P/Q/R/S sin actualizar."
            )

        row = write_lottery_row(sheet, merged)
        batch_results.append(
            {
                "filename": os.path.basename(sales_path or dept_path),
                "target_row": row,
                "report_date": report_date.strftime("%d/%m/%Y"),
                "warning": " | ".join(warnings) if warnings else None,
            }
        )

    temp_path = _create_temp_workbook_path()
    workbook.save(os.path.abspath(temp_path))
    workbook.close()

    summary = {
        "files_processed": len(batch_results),
        "batch_results": batch_results,
    }
    return temp_path, summary
