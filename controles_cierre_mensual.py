"""
Controles: Cierre mensual — primer módulo de la sección Controles.

Cruza el PDF "Store Sales Summary Report" mensual (el mismo tipo de reporte
que Reporte Diario ya lee día a día, acá para todo el mes de una sola vez)
contra el total mensual ya cargado en la hoja "Store Info" del Excel Cierre
-- Total Revenue y Network Revenue. A diferencia del chequeo "dif 1"/"dif 2"
que ya vive como fórmula en la propia hoja (compara una fila contra sí
misma, así que un día salteado, cargado dos veces, o mal tipeado de forma
consistente no lo detecta), este control usa una fuente 100% independiente
-- el reporte oficial del POS para el mes completo.

Es de solo lectura: nunca escribe nada en el Excel ni genera un archivo
para descargar -- el resultado se muestra en pantalla (verde/rojo por
chequeo), según lo que ya se decidió para toda la sección Controles.
"""

import calendar
import os
import re
from datetime import datetime

from openpyxl import load_workbook

from reporte_diario import (
    STORE_INFO_COL_CASH,
    STORE_INFO_COL_CREDIT,
    STORE_INFO_COL_DESC_COMB,
    STORE_INFO_COL_DESC_OTROS,
    STORE_INFO_COL_LOCAL_ACCOUNTS,
    STORE_INFO_COL_NETWORK_REVENUE,
    STORE_INFO_COL_NON_FUEL,
    STORE_INFO_COL_SALES_FUEL,
    STORE_INFO_COL_TAX_COLLECT,
    STORE_INFO_COL_VOLUME,
    _find_store_info_sheet,
    _store_info_row_for_day,
    extract_store_info_from_pdf,
)

TOLERANCE = 0.01

# Other (columna U, 21 -- un lugar después de TC) es la única columna de
# pago que entra en el Total Revenue (W=SUM(S:V)) sin tener su propio
# chequeo individual pedido por el usuario -- reporte_diario.py nunca la
# escribe (no tiene nombre propio ahí), se deriva de STORE_INFO_COL_CREDIT
# para no duplicar el número de columna a mano.
STORE_INFO_COL_OTHER = STORE_INFO_COL_CREDIT + 1


def _eval_literal_sum_cell(value):
    """
    Devuelve el número real de una celda de Store Info -- puede ser un
    literal (float/int), una fórmula "=num+num+num..." (como escribe
    reporte_diario.py para TC, o como el usuario carga a mano un desglose
    de pagos), o estar vacía (None -> 0.0). Nunca evalúa una fórmula con
    referencias a otras celdas -- si aparece algo así, falla en vez de
    arriesgar un resultado incorrecto (regla de oro del proyecto: nunca
    un valor de baja confianza en silencio).
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.startswith("="):
        body = value[1:]
        if not re.match(r"^[\d.\-+]+$", body):
            raise ValueError(
                f"No se puede evaluar la fórmula {value!r} de Store Info -- tiene algo "
                "más que números y signos, revisar a mano."
            )
        return sum(float(token) for token in re.findall(r"-?\d+\.?\d*", body))
    raise ValueError(f"Valor inesperado en una celda de Store Info: {value!r}")


# Cada entrada agrupa la columna de Store Info a sumar por día con el
# nombre de campo que usa `_extract_store_info_fields` (reporte_diario.py)
# para ese mismo dato en el PDF -- una sola lista maneja tanto la suma
# mensual del Excel como, más abajo, el armado de los 9 chequeos puntuales
# pedidos por el usuario (Volume, Total Fuel Sales, Fuel Discounts, Total
# Non Fuel Sales, Other Discounts, Total Taxes Collected, Cash, TC, Local
# Accounts), sin duplicar la lista de columnas dos veces.
_MONTHLY_SUM_COLUMNS = {
    "volume": STORE_INFO_COL_VOLUME,
    "sales_fuel": STORE_INFO_COL_SALES_FUEL,
    "desc_comb": STORE_INFO_COL_DESC_COMB,
    "non_fuel_total": STORE_INFO_COL_NON_FUEL,
    "desc_otros": STORE_INFO_COL_DESC_OTROS,
    "tax_collect": STORE_INFO_COL_TAX_COLLECT,
    "cash": STORE_INFO_COL_CASH,
    "credit_terms": STORE_INFO_COL_CREDIT,
    "other": STORE_INFO_COL_OTHER,
    "local_accounts": STORE_INFO_COL_LOCAL_ACCOUNTS,
    "network_revenue": STORE_INFO_COL_NETWORK_REVENUE,
}


def _sum_store_info_month(sheet, year, month):
    """
    Suma, para cada día real del mes (year, month), cada columna de
    Store Info listada en `_MONTHLY_SUM_COLUMNS` -- incluye "other" (U),
    que no tiene su propio chequeo pero hace falta para reconstruir Total
    Revenue (Cash+TC+Other+Local Account, columnas S-V -- misma fórmula
    que la propia plantilla usa en W, =SUM(S:V)).

    Nunca lee con data_only=True: eso dependería de que el archivo haya
    sido recalculado y guardado en Excel real antes de subirlo acá -- en
    vez de eso, recalcula cada total en Python a partir de las mismas
    columnas de entrada, así el control funciona apenas se sube el Excel,
    sin pedirle al usuario un paso extra.

    Devuelve (totales, missing_days, periods_seen) -- un día sin fecha real
    cargada en columna A se reporta en missing_days en vez de aportar 0 en
    silencio, para que el usuario sepa que ese día todavía no se cargó (y
    no confundir "no cargado" con "cargado en cero"). `periods_seen` es el
    conjunto de todos los (año, mes) reales encontrados en columna A de
    las filas revisadas -- como esta función ubica cada fila por posición
    (row = day+1 del mes pedido, no por fecha), un Excel Cierre de OTRO
    mes igual tiene una fecha real en cada una de esas filas -- solo que
    de un período distinto. Detectarlo acá permite que
    `check_store_info_monthly` avise que el Excel cargado no pertenece al
    período del PDF (subir esa combinación no está mal -- sirve para
    comparar -- pero hay que avisarlo).
    """
    days_in_month = calendar.monthrange(year, month)[1]
    totals = {field: 0.0 for field in _MONTHLY_SUM_COLUMNS}
    missing_days = []
    periods_seen = set()
    for day in range(1, days_in_month + 1):
        row = _store_info_row_for_day(day)
        date_value = sheet.cell(row=row, column=1).value
        if not isinstance(date_value, datetime):
            missing_days.append(day)
            continue
        periods_seen.add((date_value.year, date_value.month))
        for field, column in _MONTHLY_SUM_COLUMNS.items():
            totals[field] += _eval_literal_sum_cell(sheet.cell(row=row, column=column).value)
    return totals, missing_days, periods_seen


def _build_check(label, pdf_value, excel_value, unit="$", tolerance=TOLERANCE):
    diff = round(pdf_value - excel_value, 2)
    return {
        "label": label,
        "pdf_value": round(pdf_value, 2),
        "excel_value": round(excel_value, 2),
        "diff": diff,
        "ok": abs(diff) <= tolerance,
        "unit": unit,
    }


def check_store_info_monthly(cierre_path, monthly_pdf_path):
    """
    Cruza el PDF "Store Sales Summary Report" mensual (páginas 1-2, "Store
    Sales Summary Report" + "Method of Payment Totals Report") contra la
    hoja Store Info del Excel Cierre para el mismo mes/año (tomado del
    propio período del PDF, "PERIOD FROM: ... TO: ..."). Devuelve un dict
    con un chequeo por cada dato que imprime esas dos páginas -- Volume,
    Total Fuel Sales, Fuel Discounts, Total Non Fuel Sales, Other
    Discounts, Total Taxes Collected, Cash, tarjetas de crédito (TC),
    Local Accounts, Network Revenue y Total Revenue -- listo para mostrar
    en pantalla. Nunca escribe ni descarga nada.
    """
    cierre_path = os.path.abspath(str(cierre_path).strip())
    if not os.path.isfile(cierre_path):
        raise FileNotFoundError(f"Excel Cierre no encontrado: {cierre_path}")

    pdf_fields = extract_store_info_from_pdf(monthly_pdf_path, start_page_index=0)
    period_from = pdf_fields["from_date"]
    period_to = pdf_fields["to_date"]

    workbook = load_workbook(cierre_path, data_only=False)
    try:
        sheet = _find_store_info_sheet(workbook)
        totals, missing_days, periods_seen = _sum_store_info_month(sheet, period_from.year, period_from.month)
    finally:
        workbook.close()

    other_periods = sorted(periods_seen - {(period_from.year, period_from.month)})
    period_mismatch = [f"{month:02d}/{year}" for year, month in other_periods]

    pdf_credit_terms = sum(pdf_fields["credit_terms"])
    excel_total_revenue = (
        totals["cash"] + totals["credit_terms"] + totals["other"] + totals["local_accounts"]
    )

    checks = [
        # Cada día carga su propio Volume ya redondeado a centésimas de
        # galón (ver Store Info columna E) -- sumar 28-31 valores así
        # redondeados contra el total que imprime el propio POS a fin de
        # mes puede acumular unos pocos centésimos de diferencia sin que
        # haya ningún día salteado o mal cargado (confirmado contra agosto
        # real: 0.04 gal de diferencia en 40,464.81 gal, un mes sin ningún
        # error real). Tolerancia más ancha solo acá -- el resto de los
        # chequeos son montos en dólares y se quedan en la tolerancia
        # estándar de ±$0.01.
        _build_check("Volume", pdf_fields["volume"], totals["volume"], unit="gal", tolerance=0.15),
        _build_check("Total Fuel Sales", pdf_fields["sales_fuel"], totals["sales_fuel"]),
        _build_check("Fuel Discounts", pdf_fields["desc_comb"], totals["desc_comb"]),
        _build_check("Total Non Fuel Sales", pdf_fields["non_fuel_total"], totals["non_fuel_total"]),
        _build_check("Other Discounts", pdf_fields["desc_otros"], totals["desc_otros"]),
        _build_check("Total Taxes Collected", pdf_fields["tax_collect"], totals["tax_collect"]),
        _build_check("Cash", pdf_fields["cash"], totals["cash"]),
        _build_check("Tarjetas de Crédito (TC)", pdf_credit_terms, totals["credit_terms"]),
        _build_check("Local Accounts", pdf_fields["local_accounts"], totals["local_accounts"]),
        _build_check("Network Revenue", pdf_fields["network_revenue"], totals["network_revenue"]),
        _build_check("Total Revenue", pdf_fields["total_revenue"], excel_total_revenue),
    ]

    return {
        "period_from": period_from,
        "period_to": period_to,
        "missing_days": missing_days,
        "period_mismatch": period_mismatch,
        "checks": checks,
        "all_ok": all(check["ok"] for check in checks) and not missing_days,
    }
