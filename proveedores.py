"""
Proveedores — carga de facturas de compra (PDF) al libro de cuentas
corrientes por proveedor ("Bradenton. Cta Cte Proveedores.xlsx"), una hoja
por proveedor.

Cada proveedor tiene su propia forma de facturar (texto digital vs.
escaneado, una o varias páginas, foto de cheque incluida, etc.), así que
la extracción vive en SUPPLIER_REGISTRY: un extractor por proveedor,
detectado automáticamente por el contenido del PDF. Se agregan proveedores
nuevos ahí, sin tocar la lógica de inserción de filas.
"""

import difflib
import io
import os
import re
import tempfile
from copy import copy
from datetime import datetime

try:
    import pdfplumber
except ImportError:  # pragma: no cover - environment guard
    pdfplumber = None  # type: ignore[assignment]

try:
    import pytesseract
    from PIL import Image
except ImportError:  # pragma: no cover - environment guard
    pytesseract = None  # type: ignore[assignment]
    Image = None  # type: ignore[assignment]

try:
    import cv2
    import numpy as np
except ImportError:  # pragma: no cover - environment guard
    cv2 = None  # type: ignore[assignment]
    np = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.cell_range import CellRange
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - environment guard
    load_workbook = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    CellRange = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
except ImportError:  # pragma: no cover - environment guard
    pd = None  # type: ignore[assignment]

from proveedores_pago_rules import match_supplier_sheet
from ocr_utils import (
    ensure_pdfplumber as _ensure_pdfplumber,
    ensure_pytesseract as _ensure_pytesseract,
    ensure_cv2 as _ensure_cv2,
    correct_image_orientation as _correct_image_orientation,
    extract_largest_page_image as _extract_page_image,
    crop_relative as _crop_relative,
    remove_grid_lines as _remove_grid_lines,
)

# Columnas de cada hoja de proveedor (1-based): DATE|COMPROB|N|Amount(DEBE)|
# Amount(HABER)|BALANCE|DETAIL. Las filas de factura solo escriben
# DATE/COMPROB/N/DEBE — HABER queda para los pagos (fase futura, via Chase).
INVOICE_ROW_COLUMNS = (1, 2, 3, 4, 6, 7)
COL_DATE, COL_COMPROB, COL_NUMERO, COL_DEBE, COL_HABER, COL_BALANCE, COL_DETALLE = range(1, 8)

# El color de la columna N (COMPROB) alterna mes a mes; se detecta el último
# usado y se alterna, nunca se hardcodea cuál mes es cuál color.
GREEN_FILL = "FF92D050"
YELLOW_FILL = "FFFFFF00"
MONTH_FILL_COLORS = (GREEN_FILL, YELLOW_FILL)

# Pestaña de la hoja: celeste mientras el proveedor tenga saldo pendiente
# (le debemos), sin color cuando el saldo vuelve a 0. Mismo celeste que el
# usuario ya venía usando a mano en algunas hojas del libro real (ej.
# AIRGAS, KING'S) -- no es un color nuevo, solo se automatiza.
DEBT_TAB_COLOR = "FF00B0F0"


def _create_temp_workbook_path():
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="proveedores_")
    os.close(fd)
    return temp_path


def _parse_mmdd_year_flexible(date_text):
    """
    Parsea "M/D/AA" o "M/D/AAAA" con el mismo patrón -- varios extractores
    permiten año de 2 o 4 dígitos en su regex pero llamaban a strptime con
    "%y" fijo, así que un año de 4 dígitos rompía con un ValueError críptico
    de la librería estándar en vez del mensaje de error propio de la función.
    """
    for fmt in ("%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(date_text, fmt)
        except ValueError:
            continue
    raise ValueError(f'No se pudo interpretar la fecha "{date_text}".')


# ---- Extractores por proveedor ----

def _extract_ht_hackney_invoice(pdf_path):
    """
    Las facturas de H.T. Hackney vienen en varias páginas; el N° de
    invoice y la fecha están en el encabezado de la primera página, y el
    total final real (con fuel surcharge, delivery charge e impuestos) solo
    aparece en la última página.
    """
    _ensure_pdfplumber()
    with pdfplumber.open(pdf_path) as pdf:
        first_text = pdf.pages[0].extract_text() or ""
        last_text = pdf.pages[-1].extract_text() or ""

    invoice_match = re.search(r"Invoice #:\s*(\d+)", first_text)
    date_match = re.search(r"Invoice Date:\s*(\d{1,2}/\d{1,2}/\d{2,4})", first_text)
    totals = re.findall(r"Total:\s*([\d,]+\.\d{2})", last_text)

    if not (invoice_match and date_match and totals):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total "
            "del PDF de H.T. Hackney."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = _parse_mmdd_year_flexible(date_match.group(1))
    amount = float(totals[-1].replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_cec_invoice(pdf_path):
    """
    Las facturas de CEC (Chinook Enterprises Corp.) son un escaneo de una
    sola página, sin capa de texto -- hace falta OCR. El N° de invoice
    puede salir con un espacio de más entre dígitos (ej. "188412 7"), y el
    total final a veces se pierde por el fondo gris de su casillero, así
    que se usa el "Subtotal:" -- numéricamente igual al total en todas las
    facturas vistas -- que siempre se lee bien.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura CEC."
        )
    text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"Inv\s*#\s*([\d\s]+?)\n", text)
    date_match = re.search(r"Order taken on\s*(\d{1,2}/\d{1,2}/\d{4})", text)
    subtotal_match = re.search(r"Subtotal:?\s*\$?\s*([\d,]+\.\d{2})", text)

    if not (invoice_match and date_match and subtotal_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total "
            "del PDF de CEC (Chinook)."
        )

    invoice_no = int(re.sub(r"\s+", "", invoice_match.group(1)))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(subtotal_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_colonial_invoice(pdf_path):
    """
    Colonial Wholesale Dist. LLC -- escaneo de una sola página. El N° y la
    fecha están en la cajita de arriba a la derecha ("INVOICE NO." /
    "INVOICE DATE"); el "BALANCE DUE" en la esquina inferior derecha, sobre
    la línea de la firma.

    Limitación conocida: si el conductor corrigió el total a mano (ej. un
    crédito tachado y reescrito), la letra manuscrita no se puede leer de
    forma confiable y esta función va a fallar con un error claro -- en ese
    caso hay que cargar esa factura a mano.

    2026-09-02: se detectaron facturas reales (ago-2026) donde el recorte
    angosto original no encuentra el BALANCE DUE -- la tabla de totales se
    corrió de posición según la cantidad de ítems, y en facturas de 2
    páginas el total puede caer en la página 2. Si el recorte rápido no
    encuentra nada, se prueba una franja más ancha, agrandada y sin líneas
    de grilla, ancladas al texto "BALANCE DUE", en cada página del PDF --
    con VARIAS alturas de recorte (no una sola): se probaron 3 facturas
    reales de agosto-2026 y cada una necesitó una altura de recorte
    distinta para que la tabla de totales completa entrara en el recorte
    (0.80 vs 0.82), así que una altura fija no alcanza. **Aun así, esto
    NO es un fix completo**: validado contra las 140 facturas del
    historial completo (2023-2026), sigue fallando ~60% -- no es algo
    que empezó en agosto, Colonial viene fallando en la mayoría de sus
    facturas desde siempre por calidad de escaneo real, no por esto en
    particular. Mismo tipo de límite ya aceptado para AZ Southeast.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages = [_extract_page_image(page) for page in pdf.pages]
    pages = [img for img in pages if img is not None]
    if not pages:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura Colonial."
        )

    top_text = pytesseract.image_to_string(_crop_relative(pages[0], 0.68, 0.0, 1.0, 0.22))
    invoice_match = re.search(r"\b(\d{7})\b", top_text)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", top_text)

    bottom_text = pytesseract.image_to_string(_crop_relative(pages[0], 0.50, 0.90, 1.0, 1.0))
    balance_match = re.search(r"([\d,]+\.\d{2})", bottom_text)

    if balance_match is None:
        for image in pages:
            for top in (0.76, 0.78, 0.80, 0.82, 0.84, 0.86):
                wide_crop = _crop_relative(image, 0.0, top, 1.0, 1.0)
                wide_crop = wide_crop.resize((wide_crop.width * 3, wide_crop.height * 3))
                wide_text = pytesseract.image_to_string(_remove_grid_lines(wide_crop))
                balance_match = re.search(
                    r"BALANCE\s*DUE\.?\s*\n*\s*\$?\s*([\d,]+\.\d{2})", wide_text, re.IGNORECASE
                )
                if balance_match:
                    break
            if balance_match:
                break

    if not (invoice_match and date_match and balance_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "Colonial (si tiene una corrección a mano, cargue esta factura manualmente)."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = _parse_mmdd_year_flexible(date_match.group(1))
    amount = float(balance_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_gce_invoice(pdf_path):
    """
    Gold Coast Eagle -- escaneo; con muchos ítems, la factura se corre a
    una segunda página y la línea de confirmación ("Inv# 657065
    $1,381.00", con invoice y total juntos) queda ahí, no en la primera.

    La fecha real vive en una línea de confirmación tipo "Thu Aug 20,
    2026" (día de semana + mes + día + año) -- bug real encontrado
    2026-09-03: el regex viejo (\\w{3} suelto, sin exigir un día de semana
    válido) matcheaba el TEXTO "...Expires Mar 31, 2027" del recuadro de
    licencia que casi todas las facturas traen cerca del encabezado
    ("License: ... Expires Mar 31, 2027" -> "res Mar 31, 2027" matcheaba
    igual), leyendo una fecha completamente ajena a la factura sin ningún
    aviso. Exigir un día de semana real (Mon/Tue/.../Sun) al principio
    descarta ese falso positivo.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    confirm_match = re.search(r"Inv.{0,4}?(\d{5,7})\D{0,3}\$?\s*([\d,]+\.\d{2})", text)
    date_match = re.search(r"\b(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+\w{3}\s+\d{1,2},\s*\d{4}\b", text)

    if not (confirm_match and date_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "Gold Coast Eagle."
        )

    invoice_no = int(confirm_match.group(1))
    invoice_date = datetime.strptime(date_match.group(0), "%a %b %d, %Y")
    amount = float(confirm_match.group(2).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_frito_lay_invoice(pdf_path):
    """
    Frito-Lay -- escaneo; puede traer una foto del cheque en una página
    aparte (se ignora, ninguno de los campos buscados aparece ahí). El
    total real es "TOTAL DUE", no el "GROSS SALES AMOUNT" de la tabla.

    Hay al menos dos plantillas distintas de recibo: "CASH SALE" (con
    "DATE: 27 Dec 2025" y "TOTAL DUE:") y "CHARGE SALES" (con la fecha
    suelta como "11/06/24" sin etiqueta, y "TOTAL DUE =" con igual en vez
    de dos puntos) -- se intentan ambos formatos.

    Bug real encontrado 2026-09-03: en el encabezado, "INVOICE #" a veces
    sale tan garbleado ("rvorce #; 96809721", con un dígito además mal
    leído) que ni el texto ni el número matchean -- la etiqueta "Document
    #:" del pie de página trae el mismo número de forma mucho más
    confiable, se usa como respaldo SOLO cuando "INVOICE #" no matcheó
    nada (se confirmó en una muestra real que "Document #:" puede leer un
    dígito distinto cuando "INVOICE #" sí está limpio -- nunca preferirlo
    por sobre uno que ya funcionó).
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    invoice_match = re.search(r"INVOICE\s*#\s*[:\s]*(\d+)", text, re.IGNORECASE)
    if not invoice_match:
        invoice_match = re.search(r"Document\s*#\s*:?\s*(\d+)", text, re.IGNORECASE)
    total_match = re.search(r"TOTAL DUE\s*[:=]\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    date_match = re.search(r"DATE:\s*(\d{1,2}\s+\w{3}\s+\d{4})", text, re.IGNORECASE)
    if date_match:
        invoice_date = datetime.strptime(date_match.group(1), "%d %b %Y")
    else:
        date_match = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2})\b", text)
        invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%y") if date_match else None

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Frito-Lay."
        )

    invoice_no = int(invoice_match.group(1))
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _kings_find_amount(totals_crop):
    """
    El cuadro de totales de King's tiene bordes que confunden a Tesseract
    -- según el escaneo, el OCR de la caja completa a veces mezcla el
    bloque de etiquetas (Sub Total/Other Charges/.../Total) con el bloque
    de importes en un orden que ya no coincide (columnas leídas por
    separado), y a veces la etiqueta "Total:" ni se lee (tapada por el
    sello de USDA que cae justo en esa fila) -- una búsqueda de texto tipo
    "Total:\\s*(importe)" sobre el string plano falla en ambos casos.

    Se ubica "Total" (no "Sub Total") por posición con
    pytesseract.image_to_data, ignorando basura pegada al texto (acentos
    mal leídos, "|", etc.), y se recorta+amplía esa fila sola para
    releerla limpia -- mismo patrón ya usado para Signarama/Midtown. Si
    ni el crudo ni la versión sin grilla tienen la etiqueta legible, el
    importe de "Total" es siempre el último renglón de la caja -- se cae
    a tomar el importe con mayor coordenada Y (más abajo) entre todos los
    encontrados.
    """
    alpha_re = re.compile(r"[^a-zA-Z]")
    money_re = re.compile(r"^[\d,]+\.\d{2}$")

    for variant in (_remove_grid_lines(totals_crop), totals_crop):
        data = pytesseract.image_to_data(variant, output_type=pytesseract.Output.DICT)
        for i, word in enumerate(data["text"]):
            if alpha_re.sub("", word).lower() != "total":
                continue
            prev_word = data["text"][i - 1] if i > 0 else ""
            if alpha_re.sub("", prev_word).lower() == "sub":
                continue
            top, height, left = data["top"][i], data["height"][i], data["left"][i]
            strip = variant.crop((left, max(0, top - 10), variant.width, top + height + 15))
            strip = strip.resize((strip.width * 3, strip.height * 3), Image.LANCZOS)
            match = re.search(r"([\d,]+\.\d{2})", pytesseract.image_to_string(strip, config="--psm 7"))
            if match:
                return float(match.group(1).replace(",", ""))

    data = pytesseract.image_to_data(_remove_grid_lines(totals_crop), output_type=pytesseract.Output.DICT)
    candidates = []
    for i, word in enumerate(data["text"]):
        cleaned = word.strip().lstrip("[|_").rstrip("])|_")
        if money_re.match(cleaned):
            candidates.append((data["top"][i], cleaned))
    if candidates:
        candidates.sort()
        return float(candidates[-1][1].replace(",", ""))
    return None


def _kings_find_invoice_and_date(info_crop):
    """
    La tabla de arriba (Terms/PO-REF/Ship Via/Salesperson/Invoice Date/
    Invoice #) también tiene grilla, y la fila de datos puede aparecer
    con separadores de fecha perdidos ("8/4/2026" leído "84/2026") si se
    intenta leer toda la franja de una -- pero casi siempre el N° de
    factura (6 dígitos exactos) y la fecha completa salen limpios como
    tokens sueltos en un primer pase de pytesseract.image_to_data. Si
    alguno de los dos falta, se ubica la fila (por el otro dato ya
    encontrado, o por cualquier token con dígitos+"/") y se recorta+
    amplía sola para releerla con más resolución.
    """
    date_full_re = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
    date_loose_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")
    invoice_exact_re = re.compile(r"^\d{6}$")
    invoice_loose_re = re.compile(r"(?<!\d)(\d{6})(?!\d)")

    data = pytesseract.image_to_data(info_crop, output_type=pytesseract.Output.DICT)
    invoice_idx = date_idx = None
    invoice_no = invoice_date = None
    for i, word in enumerate(data["text"]):
        cleaned = word.strip()
        if invoice_idx is None and invoice_exact_re.match(cleaned):
            invoice_idx = i
            invoice_no = int(cleaned)
        if date_idx is None:
            match = date_full_re.match(cleaned)
            if match:
                date_idx = i
                invoice_date = datetime(int(match.group(3)), int(match.group(1)), int(match.group(2)))

    if invoice_no is not None and invoice_date is not None:
        return invoice_no, invoice_date

    # Preferir un token con pinta de fecha (aunque venga con el "/"
    # perdido, ej. "84/2026") como ancla de fila -- su posición/alto
    # suele recortar mejor esa fila que anclar en el N° de factura, que a
    # veces cae en una franja distinta de la tabla.
    anchor_idx = date_idx
    if anchor_idx is None:
        for i, word in enumerate(data["text"]):
            cleaned = word.strip()
            if re.search(r"\d", cleaned) and "/" in cleaned:
                anchor_idx = i
                break
    if anchor_idx is None:
        anchor_idx = invoice_idx
    if anchor_idx is None:
        return invoice_no, invoice_date

    top, height = data["top"][anchor_idx], data["height"][anchor_idx]
    row = info_crop.crop((0, max(0, top - 15), info_crop.width, top + height + 15))
    row = row.resize((row.width * 4, row.height * 4), Image.LANCZOS)
    row_text = pytesseract.image_to_string(row, config="--psm 6")

    if invoice_no is None:
        match = invoice_loose_re.search(row_text)
        if match:
            invoice_no = int(match.group(1))
    if invoice_date is None:
        match = date_loose_re.search(row_text)
        if match:
            invoice_date = datetime(int(match.group(3)), int(match.group(1)), int(match.group(2)))
    return invoice_no, invoice_date


def _extract_kings_invoice(pdf_path):
    """
    King's Wholesale Florists -- tanto la fecha/N° de invoice (tabla de
    arriba) como el total (cuadro "Total:" abajo a la derecha) viven en
    tablas con bordes que Tesseract confunde con el texto, y el orden de
    lectura del cuadro de totales puede descolocar etiquetas de importes
    -- ver _kings_find_amount/_kings_find_invoice_and_date, que anclan
    por posición (pytesseract.image_to_data) en vez de buscar el texto
    en un string plano.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura King's."
        )

    info_crop = _crop_relative(image, 0.0, 0.27, 1.0, 0.40)
    totals_crop = _crop_relative(image, 0.55, 0.70, 1.0, 1.0)

    invoice_no, invoice_date = _kings_find_invoice_and_date(info_crop)
    amount = _kings_find_amount(totals_crop)

    if invoice_no is None or invoice_date is None or amount is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de King's."
        )

    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_red_bull_invoice(pdf_path):
    """
    Red Bull Distribution -- escaneo de una sola página. El renglón
    "TOTAL DUE" final a veces no lo lee Tesseract, pero el mismo número
    aparece en el cuadro TOTALS como "INVOICE" (Deposit/Tax siempre en
    $0.00 en las facturas vistas, así que es el mismo importe).
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura Red Bull."
        )
    text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"\bInv\w{0,6}:\s*(\d{6,})", text)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\s+\d{1,2}:\d{2}\s*[AP]M", text)
    total_match = re.search(r"INVOICE\D*([\d,]+\.\d{2})", text)

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Red Bull."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_sweetheart_invoice(pdf_path):
    """
    Sweetheart Ice Cream -- escaneo de una sola página con la info que hace
    falta; puede traer una foto del cheque en una segunda página (se
    ignora). "BALANCE DUE" a veces sale con un dígito mal leído -- se usa
    "TOTAL SALES", que es el mismo importe y lee mejor.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura Sweetheart."
        )
    text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"INVOICE.{0,4}?(\d{8,})", text, re.IGNORECASE)
    date_match = re.search(r"Date:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
    total_match = re.search(r"TOTAL SALES:\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Sweetheart."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_bimbo_invoice(pdf_path):
    """
    Bimbo Bakeries -- escaneo, puede traer una foto de cheque en otra
    página (se ignora). El total real ("TICKET TOTALS", ya neto de
    devoluciones) vive en una tabla con bordes:

    1) A veces hay una línea de confirmación al pie ("{invoice} {fecha}
       {total}") que trae fecha y total juntos y lee mejor que la tabla.
    2) Si esa línea quedó tapada por el recibo de "Paid Out" grapado
       encima (pasa seguido), se recorta la fila de "TICKET" -- ubicada
       dinámicamente, no a una altura fija, porque varía según la
       cantidad de ítems -- y se le quitan las líneas de grilla.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        pages_images = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_images.append(image)
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    invoice_match = re.search(r"INVOICE#\s*([\d\s]+?)\s*\n", text, re.IGNORECASE)
    if not invoice_match:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer el N° de invoice del PDF de Bimbo."
        )
    invoice_no = int(re.sub(r"\s+", "", invoice_match.group(1)))

    # Anclar la línea de confirmación al pie ("{invoice} {fecha} {total}",
    # ver docstring) al propio N° de invoice ya confirmado -- sin esto, el
    # patrón "fecha seguida de números" no tiene ninguna palabra ancla y
    # puede matchear una fecha/monto de la tabla de ítems antes de llegar
    # a la línea real de confirmación. Se toleran espacios sueltos de OCR
    # entre los dígitos del invoice, como en el resto del texto.
    invoice_digits_pattern = r"\s*".join(re.escape(digit) for digit in str(invoice_no))
    footer_match = re.search(
        invoice_digits_pattern + r"\s*(\d{1,2}/\d{1,2}/\d{4})[:.]*\s*([\d,]+)[.\s]+(\d{2})\b",
        text,
    )
    if footer_match:
        invoice_date = datetime.strptime(footer_match.group(1), "%m/%d/%Y")
        amount = float(f"{footer_match.group(2).replace(',', '')}.{footer_match.group(3)}")
        return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}

    date_match = re.search(r"SDD[:;]?\s*(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE)

    amount = None
    for image in pages_images:
        data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
        y = next(
            (data["top"][i] for i, word in enumerate(data["text"]) if word.strip().upper() == "TICKET"),
            None,
        )
        if y is None:
            continue
        crop = image.crop((0, max(0, y - 10), image.width, min(image.height, y + 220)))
        clean_text = pytesseract.image_to_string(_remove_grid_lines(crop, upscale=2))
        amounts = re.findall(r"([\d,]+\.\d{2})", clean_text)
        if amounts:
            amount = float(amounts[-1].replace(",", ""))
            break

    if not (date_match and amount is not None):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer fecha/total del PDF de Bimbo "
            "(si el recibo de Paid Out tapa el total, cargue esta factura manualmente)."
        )

    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_midtown_invoice(pdf_path):
    """
    Midtown Wholesale -- "Sales Order" de 2 páginas + una foto de cheque,
    en cualquier orden entre las 3 (a veces el cheque va primero, a veces
    último), así que se juntan las 3 antes de buscar los campos.

    El "Notes:" / "Subtotal:" van en dos columnas separadas, y eso
    descoloca el orden de lectura del OCR (separa la etiqueta "Total" de
    su importe si se busca en el texto completo de la página) -- por eso
    se ubica "Subtotal" con coordenadas y se recorta esa columna sola.

    Limitación conocida: si el chofer tachó un ítem a mano y corrigió el
    Total a mano (le pasó al menos una vez), el OCR de esa columna sale
    ilegible (sin el punto decimal) y esta función falla con un error
    claro -- en ese caso hay que cargar la factura manualmente, nunca va
    a devolver el importe viejo por error.

    2026-09-02: Midtown cambió de plantilla en algún momento antes de
    agosto -- la nueva ("INVOICE : {n}", "Date: {fecha}", "Grand Total
    ${monto}") ya no usa "Receipt #"/"Receipt Date" y no descoloca el
    total en columnas separadas, así que no hace falta el recorte de
    Subtotal. Se prueba primero la plantilla vieja (por si el usuario
    todavía recibe facturas viejas) y, si no matchea, la nueva.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        totals_text = ""
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is None:
                continue
            pages_text.append(pytesseract.image_to_string(image))
            if totals_text:
                continue
            data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
            y = next((data["top"][i] for i, w in enumerate(data["text"]) if "Subtotal" in w), None)
            x = next((data["left"][i] for i, w in enumerate(data["text"]) if "Subtotal" in w), None)
            if y is not None:
                crop = image.crop((max(0, x - 50), max(0, y - 10), image.width, min(image.height, y + 250)))
                totals_text = pytesseract.image_to_string(crop)
    text = "\n".join(pages_text)

    invoice_match = re.search(r"Receipt #:?\s*(\d+)", text, re.IGNORECASE)
    date_match = re.search(r"Receipt Date\s*:?\s*(\d{1,2})-(\d{1,2})-(\d{4})", text, re.IGNORECASE)
    total_match = re.search(r"\bTotal\D{0,5}\$?\s*([\d,]+\.\d{2})", totals_text)

    if invoice_match and date_match and total_match:
        invoice_no = int(invoice_match.group(1))
        month, day, year = date_match.groups()
        invoice_date = datetime(int(year), int(month), int(day))
        amount = float(total_match.group(1).replace(",", ""))
        return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}

    invoice_match = re.search(r"\bINVOICE\s*:\s*(\d+)", text, re.IGNORECASE)
    date_match = re.search(r"(?<!Due )\bDate:\s*(\d{1,2})/(\d{1,2})/(\d{4})", text, re.IGNORECASE)
    total_match = re.search(r"Grand Total\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Midtown."
        )

    invoice_no = int(invoice_match.group(1))
    month, day, year = date_match.groups()
    invoice_date = datetime(int(year), int(month), int(day))
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_johnson_brothers_invoice(pdf_path):
    """
    Johnson Brothers of Florida -- puede venir en 1 o 2 páginas (el total
    real solo aparece en la última). El N° de invoice sale de "DOC {N}" al
    pie, salvo que el OCR lo lea como "boc" -- ahí se recurre a la fila de
    encabezado, anclada por la cuenta fija "146571", que trae fecha e
    invoice juntos (aunque a veces en orden invertido según la página).
    El total real está siempre pegado al texto fijo "APR 18%" del cargo
    por mora, nunca al "Gross Amount" de la tabla de arriba.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    date_match = re.search(r"146571.{0,30}?(\d{1,2}/\d{1,2}/\d{2,4})", text)
    invoice_match = re.search(r"DOC\s*(\d{6,7})", text, re.IGNORECASE)
    if not invoice_match:
        invoice_match = re.search(
            r"146571.{0,30}?\d{1,2}/\d{1,2}/\d{2,4}.{0,20}?(\d{6,7})", text
        )
    amount_match = re.search(r"APR\s*18%\s+([\d,]+)\s*\.?\s*(\d{2})\b", text)

    if not (date_match and invoice_match and amount_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "Johnson Brothers."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = _parse_mmdd_year_flexible(date_match.group(1))
    amount = float(f"{amount_match.group(1).replace(',', '')}.{amount_match.group(2)}")
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_flori_gas_invoice(pdf_path):
    """
    Flori-Gas -- formulario de matriz de punto (carbón), muy desgastado:
    el N° de invoice y la fecha impresos no se pueden leer de forma
    confiable ni recortando ni con más resolución. Por decisión del
    usuario, para ESTOS DOS CAMPOS se usa el nombre del archivo (formato
    "Invoice {N} {DD.MM.YYYY}.pdf"), que ya se viene usando como
    referencia confiable en todo este módulo.

    El monto sí se verifica contra el documento: sale del recibo de
    "Paid Out" grapado (Cash: $-X, igual al total de la factura en todos
    los casos vistos). Si no hay recibo de "Paid Out" en el PDF, falla
    con un error en vez de arriesgar leer mal el total escrito a mano.
    """
    filename_match = re.search(
        r"Invoice\s+(\d+)\s+(\d{1,2})\.(\d{1,2})\.(\d{4})",
        os.path.basename(pdf_path),
        re.IGNORECASE,
    )
    if not filename_match:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: el nombre del archivo no tiene el formato "
            '"Invoice {N} {DD.MM.YYYY}.pdf" esperado para Flori-Gas.'
        )
    invoice_no = int(filename_match.group(1))
    day, month, year = filename_match.group(2), filename_match.group(3), filename_match.group(4)
    invoice_date = datetime(int(year), int(month), int(day))

    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    # El diseño en 2 columnas (recibo de Paid Out + factura lado a lado)
    # a veces separa la etiqueta "Cash:" de su valor en la lectura del
    # OCR, así que se busca directamente el patrón "$-X.XX" del recibo.
    amount_match = re.search(r"\$-\s*([\d,]+\.\d{2})", text)
    if not amount_match:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró el recibo de \"Paid Out\" para leer "
            "el monto de esta factura de Flori-Gas -- cárguela manualmente."
        )
    amount = float(amount_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_airgas_invoice(pdf_path):
    """
    Airgas National Carbonation -- texto digital (no escaneado), dos tipos
    de factura ("STANDARD INVOICE" de gas y "CYLINDER RENTAL INVOICE" de
    alquiler de tanque), ambos con la misma tabla resumen "INVOICE DATE
    PAYER INVOICE NO. DUE DATE PAY THIS AMOUNT".

    Trampa real: en facturas de alquiler con débito automático, "PAY THIS
    AMOUNT" a veces muestra $0.00 (no queda nada por pagar aparte del
    débito), pero el importe real de la factura es el que aparece más
    abajo en el campo "AMOUNT" (junto a "Sales Tax"). Por eso se toma
    siempre la ÚLTIMA coincidencia de "AMOUNT" en el texto, nunca la
    primera.
    """
    _ensure_pdfplumber()
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""

    header_match = re.search(
        r"(\d{1,2}/\d{1,2}/\d{4})\s+\d{6,8}\s+(\d{8,10})\s+\d{1,2}/\d{1,2}/\d{4}\s+\$?\s*[\d,]+\.\d{2}",
        text,
    )
    amount_matches = re.findall(r"AMOUNT\D{0,6}?\$?\s*([\d,]+\.\d{2})", text)

    if not (header_match and amount_matches):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Airgas."
        )

    invoice_no = int(header_match.group(2))
    invoice_date = datetime.strptime(header_match.group(1), "%m/%d/%Y")
    amount = float(amount_matches[-1].replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_az_invoice(pdf_path):
    """
    AZ Southeast Distributors -- vale de entrega escaneado (no factura
    formal), con bastante ruido visual (grapas, dobleces, cheque o recibo
    de "Paid Out" superpuesto). El N° de factura es el "DELIVERY NO." de
    la cajita de arriba a la izquierda -- el CUST# de esa misma cajita
    (siempre el mismo número en todas las facturas del cliente) NO es el
    N° de factura, y el nombre de archivo a veces usa por error ese
    número en vez del DELIVERY NO. real.

    La fecha normalmente se lee bien del texto completo de la página, pero
    en algún escaneo esa cajita sale en blanco del todo -- ahí se recurre
    al recorte de la cajita con la grilla removida, donde sí aparece. El
    monto ("SUB TOTAL") casi nunca se lee del texto completo (queda
    perdido entre el cheque o el recibo de Paid Out superpuestos) -- se
    recorta la cajita de totales de abajo a la derecha con la grilla
    removida, ahí sí es consistente.

    LÍMITE CONOCIDO IMPORTANTE: a diferencia de los demás proveedores de
    este módulo, este extractor falla en la MAYORÍA de las facturas reales
    probadas (~5 de 32 en una validación amplia contra el archivo del
    Drive, sin importar el año) -- el recorte fijo de la cajita de totales
    no encuentra "SUB TOTAL" en muchos escaneos porque esa franja de la
    imagen varía de tamaño/posición según cómo quedó pegado el cheque o el
    recibo de "Paid Out" sobre el vale, o directamente el texto no es
    legible ahí. Cuando falla, falla limpio (no carga un valor incorrecto),
    pero la tasa de éxito real es baja -- conviene tratarlo como un
    complemento que a veces ahorra la carga manual, no como confiable en
    general.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        full_text = page.extract_text() or ""
        image = _extract_page_image(page)
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada del vale de "
            "AZ Southeast."
        )
    if not full_text.strip():
        full_text = pytesseract.image_to_string(image)

    top_crop = _remove_grid_lines(_crop_relative(image, 0.0, 0.0, 0.25, 0.10), upscale=4)
    top_text = pytesseract.image_to_string(top_crop)
    # Preferir el número anclado a la etiqueta "DELIVERY NO." -- el CUST#
    # (siempre el mismo número, no es la factura) convive en la misma
    # cajita recortada, así que un (\d{9}) suelto puede tomar cualquiera
    # de los dos según el orden de lectura del OCR. Si la etiqueta no se
    # lee (ruido/grapas), se cae al primer número de 9 dígitos como antes.
    invoice_match = re.search(r"DELIVERY\s*NO\.?\s*[:.]?\s*(\d{9})", top_text, re.IGNORECASE)
    if not invoice_match:
        invoice_match = re.search(r"(\d{9})", top_text)

    date_match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", full_text)
    if not date_match:
        date_match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", top_text)

    bottom_crop = _remove_grid_lines(_crop_relative(image, 0.55, 0.68, 1.0, 0.92), upscale=3)
    bottom_text = pytesseract.image_to_string(bottom_crop)
    total_match = re.search(r"TOTAL\D{0,10}?([\d,]+\.\d{2})", bottom_text)

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del vale de "
            "AZ Southeast."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime(int(date_match.group(3)), int(date_match.group(1)), int(date_match.group(2)))
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_express_beverage_invoice(pdf_path):
    """
    Express Beverage of Tampa -- escaneo de una sola página, impreso
    prolijo (no manuscrito). N°/fecha viven en la tabla "Date | Invoice #"
    de arriba a la derecha; el total en la celda "Total" de abajo a la
    derecha, justo arriba del cheque grapado. El N° de factura a veces
    sale con espacios sueltos entre dígitos por el OCR (ej. "01 371 1"),
    por eso se toma todo el resto de esa línea después de la fecha y se
    le quitan los espacios, en vez de exigir un ancho fijo de dígitos.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura "
            "Express Beverage."
        )

    top_crop = _crop_relative(image, 0.68, 0.08, 1.0, 0.17)
    top_text = pytesseract.image_to_string(top_crop)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", top_text)
    invoice_match = None
    if date_match:
        line_after_date = top_text[date_match.end():].split("\n", 1)[0]
        invoice_match = re.search(r"(\d[\d\s]{3,7}\d)", line_after_date)

    total_crop = _crop_relative(image, 0.55, 0.69, 1.0, 0.80)
    total_text = pytesseract.image_to_string(total_crop)
    total_match = re.search(r"([\d,]+\.\d{2})", total_text)
    if total_match is None:
        total_text = pytesseract.image_to_string(_remove_grid_lines(total_crop, upscale=3))
        total_match = re.search(r"([\d,]+\.\d{2})", total_text)

    if not (date_match and invoice_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "Express Beverage."
        )

    invoice_no = int(re.sub(r"\s+", "", invoice_match.group(1)))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_kooler_ice_invoice(pdf_path):
    """
    Kooler Ice, Inc. -- mezcla facturas con texto digital y escaneos según
    el pedido, y al menos dos plantillas: los pedidos normales/reparación/
    suscripción usan N° "INV{n}" y terminan la tabla resumen en "Applied
    Deposit" (que en esos casos coincide con el importe real -- se usa en
    vez de "Total", que en las facturas de suscripción queda separado y en
    $0.00, el saldo pendiente, no el importe de la factura); los pedidos
    de equipos usan N° "SO{n}" y no tienen "Applied Deposit" -- ahí el
    importe real sí está directamente en "Total".

    Validado contra el archivo histórico del Drive: ~75% de éxito (9/12).
    Las fallas restantes vistas fueron escaneos con el papel arrugado/
    doblado justo sobre el N° de factura ("#INV...") -- el resto del texto
    (fecha, importe) se lee bien, pero sin el N° no hay forma segura de
    cargar la fila; falla limpio en esos casos.
    """
    _ensure_pdfplumber()
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ""
        if not text.strip():
            _ensure_pytesseract()
            image = _extract_page_image(page)
            if image is None:
                raise ValueError(
                    f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la "
                    "factura de Kooler Ice."
                )
            text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"(?:INV|SO)\s*(\d{3,7})", text, re.IGNORECASE)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
    amount_match = re.search(r"Applied Deposit\D{0,6}\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)
    if not amount_match:
        amount_match = re.search(r"\bTotal\D{0,6}\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (invoice_match and date_match and amount_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "Kooler Ice."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(amount_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_sams_club_invoice(pdf_path):
    """
    Sam's Club -- recibo de caja escaneado (no factura formal). El N° de
    "factura" es el número de 4 dígitos que sigue a la hora en el
    encabezado del recibo (ej. "01/06/26 09:40 5495 08201 002 3909" -> N°
    5495), el mismo número que ya usa el usuario en el nombre de archivo.

    El monto se prefiere del recibo de "Paid Out" grapado ("Cash: $-X"),
    porque la columna de TOTAL del recibo original suele leerse mal (los
    dígitos de precio de la tabla de ítems quedan mezclados verticalmente
    por el OCR) -- pero no todas las compras tienen ese recibo (ej. una
    compra sin impuesto, solo alimentos, puede no llevarlo), así que si no
    aparece se cae al campo "TOTAL" leído directamente. El recibo y el
    "Paid Out" a veces quedan en páginas separadas del PDF (no una sola
    imagen combinada), por eso se juntan todas las páginas antes de buscar.

    Riesgo conocido (mismo tipo que Red Bull): en un caso Tesseract
    confundió un solo dígito del N° de recibo (5 leído como 6) de forma
    consistente sin importar el recorte o preprocesamiento probado --
    conviene revisar visualmente antes de guardar.

    Validado contra el archivo histórico del Drive: ~68% de éxito (13/19).
    Las fallas restantes fueron recibos viejos con ruido de OCR puntual
    (ej. la hora del encabezado ilegible, lo que rompe el patrón fecha+N°)
    -- fallan limpio, no insertan un valor incorrecto.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    if not pages_text:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada del recibo de "
            "Sam's Club."
        )
    text = "\n".join(pages_text)

    header_match = re.search(r"(\d{2}/\d{2}/\d{2})\s+\d{1,2}:\d{2}\s+(\d{4})\s+\d{5}", text)
    amount_match = re.search(r"\$-\s*([\d,]+\.\d{2})", text)
    if not amount_match:
        amount_match = re.search(r"\bTOTAL\D{0,10}?([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (header_match and amount_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del recibo de "
            "Sam's Club."
        )

    invoice_no = int(header_match.group(2))
    invoice_date = datetime.strptime(header_match.group(1), "%m/%d/%y")
    amount = float(amount_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_fs_wholesale_invoice(pdf_path):
    """
    FS Wholesale -- en realidad factura Florida Smokes Wholesale, LLC.
    Escaneo de 2-3 páginas (puede traer foto de cheque en una página
    aparte, se ignora): el N°/fecha viven en el encabezado de la primera
    página ("05 Nov 2025 98836 10 BRADENTON GAS STATION USA" -- fecha,
    invoice, cantidad de ítems, cliente, todo en la misma fila), pero el
    total real ("Tax Invoice Total (USD)") está en la página de Payments,
    no en la primera -- se buscan ambos patrones en el texto combinado de
    todas las páginas.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = []
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is not None:
                pages_text.append(pytesseract.image_to_string(image))
    text = "\n".join(pages_text)

    header_match = re.search(r"(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s+(\d{5,7})\s+\d+", text)
    total_match = re.search(
        r"(?:Tax\s+)?Invoice Total\s*\(USD\):\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE
    )

    if not (header_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de "
            "FS Wholesale."
        )

    invoice_no = int(header_match.group(2))
    invoice_date = datetime.strptime(header_match.group(1), "%d %b %Y")
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_lmt_invoice(pdf_path):
    """
    LMT Trading Company LLC -- escaneo, y el PDF casi siempre agrupa la
    factura real de LMT con una de J.J. Taylor Dist. FL, Inc. (comparten
    domicilio, teléfono y plantilla de reparto) en cualquier orden entre
    1-2 páginas. J.J. Taylor está pausado por pedido del usuario, así que
    se usa SOLO la página que trae "Paylink - LMT" (la de J.J. Taylor
    trae "Paylink - JT") -- el ancla que más limpio lee de todo el
    encabezado, incluso cuando "LMT Trading Company" sale muy garabateado.

    El N°/fecha del nombre de archivo ("[LMT. ]Invoice {N} {DD.MM.YYYY}.pdf")
    y los impresos en el encabezado del documento ("{fecha} {N} {load
    sheet}...") normalmente coinciden, pero NINGUNO de los dos es
    confiable al 100% por separado -- se encontró un caso real donde el
    nombre de archivo tenía un N°/fecha completamente distintos a los del
    documento y el libro real (archivo mal nombrado), y otro caso real
    donde el documento leyó un dígito de más en el N° por OCR. Por eso se
    cruzan ambas fuentes cuando el encabezado del documento se puede leer:
    si coinciden, hay confianza; si no, se falla con un error claro en vez
    de adivinar cuál de las dos está bien (si el documento no se puede
    leer, se sigue confiando solo en el nombre de archivo, como Flori-Gas).

    El total ("Sub Total"/"Total", mismo importe repetido) se lee del
    documento: se ubica la fila por posición vía pytesseract.image_to_data
    (ancla "Sub", con "Total" de respaldo si "Sub" no se detectó como
    palabra aparte) y se recorta esa franja para un OCR más limpio.
    """
    filename_match = re.search(
        r"Invoice\s+(\d+)\s+(\d{1,2})\.(\d{1,2})\.(\d{4})",
        os.path.basename(pdf_path),
        re.IGNORECASE,
    )
    if not filename_match:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: el nombre del archivo no tiene el formato "
            '"Invoice {N} {DD.MM.YYYY}.pdf" esperado para LMT.'
        )
    invoice_no = int(filename_match.group(1))
    day, month, year = filename_match.group(2), filename_match.group(3), filename_match.group(4)
    invoice_date = datetime(int(year), int(month), int(day))

    _ensure_pdfplumber()
    _ensure_pytesseract()
    lmt_image = None
    lmt_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is None:
                continue
            text = pytesseract.image_to_string(image)
            if re.search(r"paylink\s*-\s*lmt", text, re.IGNORECASE):
                lmt_image = image
                lmt_text = text
                break
    if lmt_image is None:
        raise ValueError(
            f'{os.path.basename(pdf_path)}: no se encontró la página de LMT (con "Paylink - LMT") '
            "en este PDF."
        )

    header_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\s+(\d{5,8})\b", lmt_text)
    if header_match:
        doc_invoice_no = int(header_match.group(2))
        if doc_invoice_no != invoice_no:
            raise ValueError(
                f"{os.path.basename(pdf_path)}: el N° de factura del nombre de archivo "
                f"({invoice_no}) no coincide con el N° leído del documento ({doc_invoice_no}) "
                "-- revise cuál es el correcto y cárguela manualmente."
            )
        try:
            doc_date = datetime.strptime(header_match.group(1), "%m/%d/%Y")
        except ValueError:
            doc_date = None
        if doc_date is not None and doc_date.date() != invoice_date.date():
            raise ValueError(
                f"{os.path.basename(pdf_path)}: la fecha del nombre de archivo "
                f"({invoice_date:%d/%m/%Y}) no coincide con la fecha leída del documento "
                f"({doc_date:%d/%m/%Y}) -- revise cuál es la correcta y cárguela manualmente."
            )

    upscaled = lmt_image.resize((lmt_image.width * 2, lmt_image.height * 2), Image.LANCZOS)
    data = pytesseract.image_to_data(upscaled, output_type=pytesseract.Output.DICT)

    sub_ys = [
        data["top"][i]
        for i, w in enumerate(data["text"])
        if w.strip().lower() in ("sub", "suo", "sud", "sob")
    ]
    total_ys = [
        data["top"][i]
        for i, w in enumerate(data["text"])
        if "otal" in w.lower() or w.strip().lower() in ("tol", "oul", "onl", "toul")
    ]

    amount = None
    if sub_ys:
        y = max(sub_ys)
        crop = upscaled.crop((0, max(0, y - 15), upscaled.width, min(upscaled.height, y + 150)))
        amount_match = re.search(r"([\d,]+\.\d{2})", pytesseract.image_to_string(crop))
        if amount_match:
            amount = float(amount_match.group(1).replace(",", ""))
    if amount is None and total_ys:
        y = max(total_ys)
        crop = upscaled.crop((0, max(0, y - 30), upscaled.width, min(upscaled.height, y + 250)))
        amount_match = re.search(r"([\d,]+\.\d{2})", pytesseract.image_to_string(crop))
        if amount_match:
            amount = float(amount_match.group(1).replace(",", ""))

    if amount is None:
        raise ValueError(f"{os.path.basename(pdf_path)}: no se pudo leer el total del PDF de LMT.")

    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_overflow_invoice(pdf_path):
    """
    Overflow Group Distribution -- escaneo de una sola página, texto
    limpio. N°/fecha en el encabezado ("Invoice #1500034491" /
    "11/13/2025 2:45:03 PM"), monto en "Total Sales:" (igual a "Balance:"
    en las facturas vistas, sin pagos parciales).
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada de la factura Overflow."
        )
    text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"Invoice\s*#\s*(\d+)", text, re.IGNORECASE)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\s+\d{1,2}:\d{2}:\d{2}\s*[AP]M", text)
    total_match = re.search(r"Total Sales:\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (invoice_match and date_match and total_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Overflow."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%Y")
    amount = float(total_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_swisher_invoice(pdf_path):
    """
    Swisher -- recibo de caja escaneado (no factura formal, como Sam's
    Club/AZ). El N° es el "Receipt #" del encabezado (a veces con un
    espacio de más en medio, se le saca). La fecha viene en dos formatos
    según la antigüedad de la factura ("8/28/24 10:14 AM" en las viejas,
    "2026-02-23" suelto en las nuevas) -- se prueban ambos. El monto se
    prefiere del recibo de "Paid Out" grapado ("Cash: $-X", con alguna
    coma suelta de OCR que se descarta) -- el "Total"/"Cash Due Now" de la
    tabla principal casi siempre sale con la parte decimal incompleta.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    with pdfplumber.open(pdf_path) as pdf:
        image = _extract_page_image(pdf.pages[0])
    if image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la imagen escaneada del recibo de Swisher."
        )
    text = pytesseract.image_to_string(image)

    invoice_match = re.search(r"Receipt\s*#?\s*(\d[\d\s]{5,20}\d)", text, re.IGNORECASE)
    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{2})\s+\d{1,2}:\d{2}\s*[AP]M", text)
    if date_match:
        invoice_date = datetime.strptime(date_match.group(1), "%m/%d/%y")
    else:
        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", text)
        invoice_date = datetime.strptime(date_match.group(1), "%Y-%m-%d") if date_match else None
    amount_match = re.search(r"Cash:\s*/?\s*\$-\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

    if not (invoice_match and date_match and amount_match):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del recibo de Swisher."
        )

    invoice_no = int(re.sub(r"\s+", "", invoice_match.group(1)))
    amount = float(amount_match.group(1).replace(",", ""))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


def _extract_signarama_invoice(pdf_path):
    """
    Signarama (Bradenton Signs) -- escaneo, puede traer foto de cheque en
    página aparte (se ignora). N°/fecha salen limpios del texto completo
    ("IN INV-8148", "Created Date: 7/17/2026" -- coincide con el nombre
    de archivo en casi todos los casos vistos; se prefiere sobre
    "Generated On", que puede ser muchos días posterior por una reimpresión).

    El "Grand Total" vive en una plantilla a 2 columnas (etiquetas a la
    izquierda, importes a la derecha) que la mayoría de las veces
    descoloca el valor del OCR de página completa -- se ubica la palabra
    "Grand" (o "Total" de respaldo) por posición con
    pytesseract.image_to_data y se recorta esa fila hacia la derecha, con
    más resolución, para leer el importe.
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    signarama_image = None
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            image = _extract_page_image(page)
            if image is None:
                continue
            page_text = pytesseract.image_to_string(image)
            if "bradentonsigns" in page_text.lower() or "inv-" in page_text.lower():
                signarama_image = image
                text = page_text
                break
    if signarama_image is None:
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se encontró la página de factura de Signarama."
        )

    invoice_match = re.search(r"INV[-\s]?(\d{3,6})", text, re.IGNORECASE)
    date_match = re.search(r"Created Date:\s*(\d{1,2})/(\d{1,2})/(\d{4})", text, re.IGNORECASE)

    data = pytesseract.image_to_data(signarama_image, output_type=pytesseract.Output.DICT)
    label_idx = next((i for i, w in enumerate(data["text"]) if w.strip().lower() == "grand"), None)
    if label_idx is None:
        label_idx = next((i for i, w in enumerate(data["text"]) if "total" in w.lower()), None)

    amount = None
    if label_idx is not None:
        top = data["top"][label_idx]
        height = data["height"][label_idx]
        left = data["left"][label_idx]
        crop = signarama_image.crop(
            (left, max(0, top - 10), signarama_image.width, top + height + 25)
        )
        upscaled = crop.resize((crop.width * 3, crop.height * 3), Image.LANCZOS)
        amount_match = re.search(r"([\d,]+\.\d{2})", pytesseract.image_to_string(upscaled))
        if amount_match:
            amount = float(amount_match.group(1).replace(",", ""))

    if not (invoice_match and date_match and amount is not None):
        raise ValueError(
            f"{os.path.basename(pdf_path)}: no se pudo leer invoice/fecha/total del PDF de Signarama."
        )

    invoice_no = int(invoice_match.group(1))
    invoice_date = datetime(int(date_match.group(3)), int(date_match.group(1)), int(date_match.group(2)))
    return {"invoice_no": invoice_no, "date": invoice_date, "amount": amount}


SUPPLIER_REGISTRY = {
    "ht_hackney": {
        "label": "H.T. Hackney",
        "sheet_name": "HT Hackney",
        "resumen_label": "HT HACKNEY",
        "detect": lambda text: "h.t. hackney" in text.lower(),
        "extract": _extract_ht_hackney_invoice,
    },
    "cec": {
        "label": "CEC (Chinook Enterprises Corp.)",
        "sheet_name": "Chinook CEC",
        "resumen_label": "CEC",
        "detect": lambda text: "cec distributing" in text.lower(),
        "extract": _extract_cec_invoice,
    },
    "colonial": {
        "label": "Colonial Wholesale Dist. LLC",
        "sheet_name": "Colonial",
        "resumen_label": "COLONIAL",
        "detect": lambda text: "colonial wholesale" in text.lower(),
        "extract": _extract_colonial_invoice,
    },
    "gce": {
        "label": "Gold Coast Eagle",
        "sheet_name": "GOLDCE",
        "resumen_label": "GCE",
        "detect": lambda text: "gold coast eagle" in text.lower(),
        "extract": _extract_gce_invoice,
    },
    "frito_lay": {
        "label": "Frito-Lay",
        "sheet_name": "FRITO-LAY",
        "resumen_label": "FRITO-LAY",
        "detect": lambda text: "frito" in text.lower() and "lay" in text.lower(),
        "extract": _extract_frito_lay_invoice,
    },
    "kings": {
        "label": "King's Wholesale Florists",
        "sheet_name": "KING'S",
        "resumen_label": "KING'S",
        "detect": lambda text: "wholesale florists" in text.lower(),
        "extract": _extract_kings_invoice,
    },
    "red_bull": {
        "label": "Red Bull Distribution Company",
        "sheet_name": "RED BULL",
        "resumen_label": "RED BULL",
        "detect": lambda text: "red bull distribution" in text.lower(),
        "extract": _extract_red_bull_invoice,
    },
    "sweetheart": {
        "label": "Sweetheart Ice Cream",
        "sheet_name": "SWEETHEART-ICE CREAM",
        "resumen_label": "SWT ICE CREAM",
        "detect": lambda text: "sweetheart" in text.lower(),
        "extract": _extract_sweetheart_invoice,
    },
    "bimbo": {
        "label": "Bimbo Bakeries USA, Inc.",
        "sheet_name": "BIMBO",
        "resumen_label": "BIMBO",
        "detect": lambda text: "bimbo bakeries" in text.lower(),
        "extract": _extract_bimbo_invoice,
    },
    "midtown": {
        "label": "Midtown Wholesale LLC",
        "sheet_name": "MIDTOWN",
        "resumen_label": "MIDTOWN",
        "detect": lambda text: "midtown wholesale" in text.lower(),
        "extract": _extract_midtown_invoice,
    },
    "johnson": {
        "label": "Johnson Brothers of Florida",
        "sheet_name": "JOHNSON",
        "resumen_label": "JOHNSON",
        "detect": lambda text: "johnson brothers" in text.lower(),
        "extract": _extract_johnson_brothers_invoice,
    },
    "flori_gas": {
        "label": "Flori-Gas",
        "sheet_name": "FLORI-GAS",
        "resumen_label": "FLORI GAS",
        "detect": lambda text: "305-637-9262" in text,
        "extract": _extract_flori_gas_invoice,
    },
    "airgas": {
        "label": "Airgas National Carbonation",
        "sheet_name": "AIRGAS",
        "resumen_label": "AIRGAS",
        "detect": lambda text: "airgas" in text.lower(),
        "extract": _extract_airgas_invoice,
    },
    "az": {
        "label": "AZ Southeast Distributors LLC",
        "sheet_name": "AZ Sout",
        "resumen_label": "AZ",
        "detect": lambda text: "az southeast distributors" in text.lower(),
        "extract": _extract_az_invoice,
    },
    "express": {
        "label": "Express Beverage of Tampa",
        "sheet_name": "EXPRESS ",
        "resumen_label": "EXPRESS",
        "detect": lambda text: "express beverage" in text.lower(),
        "extract": _extract_express_beverage_invoice,
    },
    "kooler_ice": {
        "label": "Kooler Ice, Inc.",
        "sheet_name": "KOOLER ICE",
        "resumen_label": "KOOLER ICE",
        "detect": lambda text: "kooler ice" in text.lower(),
        "extract": _extract_kooler_ice_invoice,
    },
    "sams_club": {
        "label": "Sam's Club",
        "sheet_name": "SAM'S",
        "resumen_label": "SAM'S",
        "detect": lambda text: re.search(r"sam.?s\s*club", text.lower()) is not None,
        "extract": _extract_sams_club_invoice,
    },
    "fs_wholesale": {
        "label": "FS Wholesale (Florida Smokes Wholesale, LLC)",
        "sheet_name": "FS WHOLESALE",
        "resumen_label": "FS WHOLESALE",
        "detect": lambda text: "florida smokes" in text.lower(),
        "extract": _extract_fs_wholesale_invoice,
    },
    "lmt": {
        "label": "LMT Trading Company LLC",
        "sheet_name": "LMT",
        "resumen_label": "LMT",
        "detect": lambda text: re.search(r"paylink\s*-\s*lmt", text.lower()) is not None,
        "extract": _extract_lmt_invoice,
    },
    "overflow": {
        "label": "Overflow Group Distribution",
        "sheet_name": "OVERFLOW",
        "resumen_label": "OVERFLOW",
        "detect": lambda text: "overflowgroupdistribution" in text.lower().replace(" ", ""),
        "extract": _extract_overflow_invoice,
    },
    "swisher": {
        "label": "Swisher",
        "sheet_name": "SWISHER",
        "resumen_label": "SWISHER",
        "detect": lambda text: "swisher" in text.lower(),
        "extract": _extract_swisher_invoice,
    },
    "signarama": {
        "label": "Signarama (Bradenton Signs)",
        "sheet_name": "SIGNARAMA",
        "resumen_label": "SIGNARAMA",
        "detect": lambda text: "bradentonsigns" in text.lower(),
        "extract": _extract_signarama_invoice,
    },
}


def _detect_supplier(pdf_path):
    """
    Detecta el proveedor por el contenido del PDF: primero intenta con el
    texto digital (rápido, sin OCR); si el PDF es un escaneo sin capa de
    texto, recién ahí hace OCR -- de cada página hasta encontrar una
    coincidencia, porque algunos proveedores meten la foto del cheque
    ANTES que la factura (la primera página sola no alcanza).
    """
    _ensure_pdfplumber()
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if not text.strip():
                image = _extract_page_image(page)
                if image is not None:
                    _ensure_pytesseract()
                    text = pytesseract.image_to_string(image)
            for key, config in SUPPLIER_REGISTRY.items():
                if config["detect"](text):
                    return key

    raise ValueError(f"{os.path.basename(pdf_path)}: proveedor no reconocido.")


def _get_supplier_sheet(workbook, sheet_name):
    target = sheet_name.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    raise ValueError(
        f'Hoja "{sheet_name}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _find_last_real_row(sheet):
    """
    Última fila con una fecha real (datetime) cargada en la columna A.

    Chequea el tipo, no solo que no esté vacía: el encabezado fijo de
    cada hoja real (filas 1-7: título, "PROVEEDOR:", "DESCRIPCION:", los
    nombres de columna) trae texto en columna A, y algunas hojas tienen
    sueltos años como entero (ej. 2025, 2026) marcando el cambio de año
    -- ninguno de los dos es una fecha real, y "not in (None, '')" no los
    filtraba.
    """
    last_real_row = None
    for row in range(1, sheet.max_row + 1):
        if isinstance(sheet.cell(row=row, column=COL_DATE).value, datetime):
            last_real_row = row
    return last_real_row


def _find_last_invoice_row(sheet, from_row):
    """
    Última fila de tipo "invoice" (COMPROB) hasta from_row inclusive -- una
    fila "OP" (pago) no sirve de referencia de estilo porque su columna N
    normalmente queda en blanco, sin la negrita que sí llevan las facturas.
    """
    for row in range(from_row, 0, -1):
        comprob = sheet.cell(row=row, column=COL_COMPROB).value
        if isinstance(comprob, str) and comprob.strip().lower() == "invoice":
            return row
    return from_row


def _find_last_month_color(sheet, from_row):
    for row in range(from_row, 0, -1):
        fill = sheet.cell(row=row, column=COL_NUMERO).fill
        if fill and fill.patternType == "solid" and fill.fgColor and fill.fgColor.rgb in MONTH_FILL_COLORS:
            return fill.fgColor.rgb
    return None


def _compute_sheet_balance(sheet):
    """
    Suma DEBE y HABER directamente en vez de leer la fórmula de BALANCE,
    porque openpyxl no evalúa fórmulas y una fila recién agregada todavía
    no tiene un valor calculado en caché.
    """
    last_row = _find_last_real_row(sheet)
    total_debe = 0.0
    total_haber = 0.0
    for row in range(1, (last_row or 0) + 1):
        debe = sheet.cell(row=row, column=COL_DEBE).value
        haber = sheet.cell(row=row, column=COL_HABER).value
        if isinstance(debe, (int, float)):
            total_debe += debe
        if isinstance(haber, (int, float)):
            total_haber += haber
    return total_debe - total_haber


def _update_sheet_tab_color(sheet):
    """
    Pinta la pestaña de la hoja de celeste si el proveedor tiene saldo
    pendiente (le debemos), y la despinta si el saldo llega a 0 (ej. al
    cargar un pago desde Chase en la Fase 2 del módulo).
    """
    balance = _compute_sheet_balance(sheet)
    sheet.sheet_properties.tabColor = DEBT_TAB_COLOR if balance > 0.005 else None


def _existing_invoice_numbers(sheet):
    numbers = set()
    for row in range(1, sheet.max_row + 1):
        comprob = sheet.cell(row=row, column=COL_COMPROB).value
        if not (isinstance(comprob, str) and comprob.strip().lower() == "invoice"):
            continue
        value = sheet.cell(row=row, column=COL_NUMERO).value
        if isinstance(value, (int, float)):
            numbers.add(int(value))
    return numbers


def _write_invoice_row(sheet, target_row, style_row, balance_ref_row, color, invoice):
    """
    Escribe los datos y el estilo de una factura en target_row -- código
    compartido entre el encadenado al final y la inserción cronológica en
    medio de la hoja (ver _find_invoice_insertion_point). El estilo
    (fuente/borde/alineación/formato) se copia de style_row -- una fila
    "invoice" real, nunca "OP" ni el separador -- en vez de confiar en lo
    que traiga target_row, que puede tener overrides viejos inconsistentes.

    balance_ref_row es la fila real inmediata anterior en la posición
    final de target_row (puede ser el separador entre meses, que también
    lleva su propia fórmula de arrastre) -- nunca target_row - 1 a
    ciegas, porque tras una inserción en medio de la hoja la fila de
    arriba física puede no ser la referencia correcta todavía en este
    punto (ver _find_invoice_insertion_point).
    """
    for col in INVOICE_ROW_COLUMNS:
        ref_cell = sheet.cell(row=style_row, column=col)
        new_cell = sheet.cell(row=target_row, column=col)
        new_cell.font = copy(ref_cell.font)
        new_cell.border = copy(ref_cell.border)
        new_cell.alignment = copy(ref_cell.alignment)
        new_cell.number_format = ref_cell.number_format

    # El N° de factura siempre va en negrita, sin importar el estilo que
    # haya traído style_row.
    number_cell = sheet.cell(row=target_row, column=COL_NUMERO)
    number_font = number_cell.font
    number_cell.font = Font(
        name=number_font.name,
        size=number_font.sz,
        bold=True,
        italic=number_font.italic,
        color=number_font.color,
        underline=number_font.underline,
    )

    sheet.cell(row=target_row, column=COL_DATE, value=invoice["date"])
    sheet.cell(row=target_row, column=COL_COMPROB, value="invoice")
    sheet.cell(row=target_row, column=COL_NUMERO, value=invoice["invoice_no"])
    sheet.cell(row=target_row, column=COL_DEBE, value=invoice["amount"])
    sheet.cell(
        row=target_row,
        column=COL_BALANCE,
        value=f"=+{sheet.cell(row=balance_ref_row, column=COL_BALANCE).coordinate}"
        f"+{sheet.cell(row=target_row, column=COL_DEBE).coordinate}"
        f"-{sheet.cell(row=target_row, column=COL_HABER).coordinate}",
    )
    sheet.cell(row=target_row, column=COL_DETALLE).value = None

    sheet.cell(row=target_row, column=COL_NUMERO).fill = PatternFill(
        start_color=color, end_color=color, fill_type="solid"
    )


def _resolve_invoice_style_row(sheet, target_row, last_row):
    """
    Ubica la fila "invoice" real más cercana a target_row para copiarle
    el estilo -- primero buscando hacia arriba (mismo criterio que
    _find_last_invoice_row), y si no hay ninguna (se está insertando
    antes de la primera factura de toda la hoja), hacia abajo.
    """
    candidate = _find_last_invoice_row(sheet, target_row - 1)
    comprob = sheet.cell(row=candidate, column=COL_COMPROB).value
    if isinstance(comprob, str) and comprob.strip().lower() == "invoice":
        return candidate
    for row in range(target_row, last_row + 1):
        comprob = sheet.cell(row=row, column=COL_COMPROB).value
        if isinstance(comprob, str) and comprob.strip().lower() == "invoice":
            return row
    return candidate


def _resolve_invoice_month_color(sheet, invoice_date, target_row, last_row):
    """
    Determina qué color (verde/amarillo) le corresponde a la columna N°
    de una factura que se va a insertar en target_row -- se calcula
    ANTES de abrir la fila, con los números de fila previos a la
    inserción.

    El color es por mes calendario, no por posición: todas las facturas
    del mismo mes comparten color sin importar en qué parte de la hoja
    estén, así que primero se busca si ya existe alguna factura real de
    ese mismo (año, mes) en cualquier lugar de la hoja y se reusa su
    color -- nunca se alterna dentro de un mismo mes. Si el mes es
    nuevo, se alterna respecto a la factura real más cercana arriba de
    target_row (o, si no hay ninguna arriba, la más cercana abajo) --
    mismo criterio de alternancia que ya usaba el encadenado al final.
    """
    for row in range(1, last_row + 1):
        row_date = sheet.cell(row=row, column=COL_DATE).value
        comprob = sheet.cell(row=row, column=COL_COMPROB).value
        if (
            isinstance(row_date, datetime)
            and isinstance(comprob, str)
            and comprob.strip().lower() == "invoice"
            and (row_date.year, row_date.month) == (invoice_date.year, invoice_date.month)
        ):
            fill = sheet.cell(row=row, column=COL_NUMERO).fill
            if fill and fill.patternType == "solid" and fill.fgColor and fill.fgColor.rgb in MONTH_FILL_COLORS:
                return fill.fgColor.rgb

    def _nearest_invoice_color(rows):
        for row in rows:
            comprob = sheet.cell(row=row, column=COL_COMPROB).value
            if isinstance(comprob, str) and comprob.strip().lower() == "invoice":
                fill = sheet.cell(row=row, column=COL_NUMERO).fill
                if fill and fill.patternType == "solid" and fill.fgColor and fill.fgColor.rgb in MONTH_FILL_COLORS:
                    return fill.fgColor.rgb
        return None

    nearest_color = _nearest_invoice_color(range(target_row - 1, 0, -1))
    if nearest_color is None:
        nearest_color = _nearest_invoice_color(range(target_row, last_row + 1))
    if nearest_color is None:
        return GREEN_FILL
    return GREEN_FILL if nearest_color == YELLOW_FILL else YELLOW_FILL


def _find_invoice_insertion_point(sheet, invoice_date):
    """
    Determina dónde debe ir una factura con fecha invoice_date para
    mantener la hoja en orden cronológico -- pedido explícito del
    usuario (2026-09-02): antes, Facturas siempre encadenaba al final
    sin importar la fecha, así que cargar una factura vieja después de
    una más nueva la dejaba fuera de orden (el BALANCE seguía siendo
    matemáticamente correcto, pero el orden visual de fechas quedaba
    mal). Mismo patrón que ya usaba _find_payment_insertion_point para
    Pagos (ver ese docstring para el porqué de cada paso: por qué se
    compara por tipo datetime y no "no vacío", por qué balance_ref_row
    nunca es un separador en blanco, etc.) -- acá además hay que resolver
    el estilo/color propios de una factura (ver _resolve_invoice_style_row
    / _resolve_invoice_month_color).

    Devuelve (target_row, needs_shift, style_row, balance_ref_row, last_row, color).
    - needs_shift=False: target_row es el próximo lugar libre al final de
      la hoja (saltando el separador si cambia el mes) -- no hace falta
      mover nada.
    - needs_shift=True: target_row ya tiene una fila real cargada -- hay
      que abrir espacio con _insert_row_preserving_merges(target_row),
      escribir la factura con _write_invoice_row, y recién ahí llamar a
      _reformulate_rows_below para las filas que quedaron abajo.

    Limitación conocida (igual que para Pagos): si la factura cae en un
    mes que no tiene NINGUNA fila cargada todavía (un hueco completo
    entre dos meses con datos), no se agrega una fila separadora nueva --
    el BALANCE sigue siendo correcto, pero el formato visual queda un
    poco distinto en ese caso puntual.
    """
    last_row = _find_last_real_row(sheet)
    if last_row is None:
        raise ValueError("La hoja no tiene ninguna fila real cargada.")
    last_date = sheet.cell(row=last_row, column=COL_DATE).value

    if invoice_date >= last_date:
        month_changed = (invoice_date.year, invoice_date.month) != (last_date.year, last_date.month)
        target_row = last_row + 2 if month_changed else last_row + 1
        style_row = _find_last_invoice_row(sheet, last_row)
        color = _find_last_month_color(sheet, last_row) or GREEN_FILL
        if month_changed:
            color = GREEN_FILL if color == YELLOW_FILL else YELLOW_FILL
        return target_row, False, style_row, last_row, last_row, color

    for row in range(1, last_row + 1):
        row_date = sheet.cell(row=row, column=COL_DATE).value
        if not isinstance(row_date, datetime):
            continue
        if row_date > invoice_date:
            balance_ref_row = _nearest_balance_row_at_or_above(sheet, row - 1)
            if balance_ref_row is None:
                raise ValueError(
                    f"La factura del {invoice_date:%d/%m/%Y} caería antes de que la hoja "
                    "tenga ninguna fila con un balance calculable (fórmula o número) de la "
                    "cual partir -- cargarla a mano."
                )
            style_row = _resolve_invoice_style_row(sheet, row, last_row)
            color = _resolve_invoice_month_color(sheet, invoice_date, row, last_row)
            return row, True, style_row, balance_ref_row, last_row, color

    # No debería llegar acá dado el chequeo de arriba, pero por las dudas.
    style_row = _find_last_invoice_row(sheet, last_row)
    color = _find_last_month_color(sheet, last_row) or GREEN_FILL
    return last_row + 1, False, style_row, last_row, last_row, color


# ---- Resumen Compras (hoja "RESUMEN COMPRAS") ----
#
# Cada vez que se carga una factura, además de la fila en la hoja del
# proveedor, se suma el monto en la hoja "RESUMEN COMPRAS": una fila por
# proveedor (columna A) x una columna por mes (B=Enero .. M=Diciembre),
# donde cada celda es una fórmula que suma las celdas D de las filas de
# factura de ese proveedor cargadas ese mes -- mismo criterio que usaba el
# usuario a mano. Si el proveedor todavía no tiene fila, se crea siguiendo
# el mismo formato que las demás. Nunca debe tumbar la carga de la factura:
# cualquier problema acá queda como advertencia en el resumen de la carga.

RESUMEN_SHEET_NAME = "RESUMEN COMPRAS"
RESUMEN_LABEL_COL = 1
RESUMEN_FIRST_MONTH_COL = 2  # B = Enero
MONTH_NAMES_ES = (
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
)


def _get_resumen_sheet(workbook):
    """
    Ubica la hoja "RESUMEN COMPRAS" -- tolera el espacio de más al final
    que tiene la pestaña real del libro (y variaciones de tipeo futuras,
    mismo criterio que get_worksheet en eft_cta_cte.py). Devuelve None si
    no se encuentra -- el llamador no debe fallar la carga por esto.
    """
    target = RESUMEN_SHEET_NAME.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    close = difflib.get_close_matches(RESUMEN_SHEET_NAME, workbook.sheetnames, n=1, cutoff=0.85)
    if close:
        return workbook[close[0]]
    return None


def _resumen_month_column(sheet, month):
    """
    Columna del mes (1=Enero..12=Diciembre) según el encabezado real de la
    fila 1 -- nunca asume la posición fija B..M sin chequearla.
    """
    expected = MONTH_NAMES_ES[month - 1]
    for col in range(RESUMEN_FIRST_MONTH_COL, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if isinstance(value, str) and value.strip().upper().startswith(expected):
            return col
    return None


def _resumen_find_supplier_row(sheet, label):
    target = label.strip().upper()
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=RESUMEN_LABEL_COL).value
        if isinstance(value, str) and value.strip().upper() == target:
            return row
    return None


def _resumen_quote_sheet_ref(sheet_name):
    """
    Rodea el nombre de hoja con comillas simples solo si hace falta
    (espacios, guiones, apóstrofes, etc.) -- mismo criterio que ya usan las
    fórmulas existentes de RESUMEN COMPRAS (ej. "GOLDCE!D320" sin comillas
    vs. "'AZ Sout'!D106" con comillas), duplicando un apóstrofe interno
    como exige Excel (ej. KING'S -> 'KING''S').
    """
    if re.search(r"[^A-Za-z0-9_]", sheet_name):
        return "'" + sheet_name.replace("'", "''") + "'"
    return sheet_name


def _resumen_create_supplier_row(sheet, label):
    """
    Agrega una fila nueva para un proveedor que todavía no aparece en
    RESUMEN COMPRAS, pegada debajo del último proveedor existente (antes de
    la fila "TOTAL"), copiando su estilo -- y corrige a mano las fórmulas
    de TOTAL y de la fila "diferencia con Chase" (openpyxl no traduce
    fórmulas al insertar filas como sí hace Excel). Levanta ValueError si
    la hoja no tiene la estructura esperada (columna A con proveedores +
    fila "TOTAL") -- mejor fallar claro que insertar a ciegas.

    Devuelve la fila nueva.
    """
    last_supplier_row = None
    row = 2
    while True:
        value = sheet.cell(row=row, column=RESUMEN_LABEL_COL).value
        if not isinstance(value, str) or not value.strip() or value.strip().upper() == "TOTAL":
            break
        last_supplier_row = row
        row += 1
    if last_supplier_row is None:
        raise ValueError('RESUMEN COMPRAS no tiene ningún proveedor cargado en la columna A todavía.')

    total_row = None
    for r in range(last_supplier_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=r, column=RESUMEN_LABEL_COL).value
        if isinstance(value, str) and value.strip().upper() == "TOTAL":
            total_row = r
            break
    if total_row is None:
        raise ValueError('RESUMEN COMPRAS no tiene una fila "TOTAL" -- no se puede agregar el proveedor de forma segura.')

    # Fila "CHASE" y la fila de diferencia (=+B{total}-B{chase}) debajo del
    # TOTAL -- opcionales, si no están simplemente no se tocan.
    chase_row = None
    for r in range(total_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=r, column=RESUMEN_LABEL_COL).value
        if isinstance(value, str) and value.strip().upper() == "CHASE":
            chase_row = r
            break

    diff_row = None
    diff_formulas = {}
    if chase_row is not None:
        diff_re = re.compile(r"^=\+([A-Z]+)" + str(total_row) + r"-([A-Z]+)" + str(chase_row) + r"$")
        for r in range(chase_row + 1, sheet.max_row + 1):
            probe = sheet.cell(row=r, column=RESUMEN_FIRST_MONTH_COL).value
            if isinstance(probe, str) and diff_re.match(probe):
                diff_row = r
                for col in range(RESUMEN_FIRST_MONTH_COL, sheet.max_column + 1):
                    value = sheet.cell(row=r, column=col).value
                    if isinstance(value, str) and diff_re.match(value):
                        diff_formulas[col] = value
                break

    total_formulas = {}
    sum_re = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$")
    for col in range(RESUMEN_FIRST_MONTH_COL, sheet.max_column + 1):
        value = sheet.cell(row=total_row, column=col).value
        if isinstance(value, str):
            match = sum_re.match(value)
            if match:
                total_formulas[col] = match

    insert_at = last_supplier_row + 1
    sheet.insert_rows(insert_at)

    for col in range(1, sheet.max_column + 1):
        ref_cell = sheet.cell(row=last_supplier_row, column=col)
        new_cell = sheet.cell(row=insert_at, column=col)
        new_cell.font = copy(ref_cell.font)
        new_cell.border = copy(ref_cell.border)
        new_cell.alignment = copy(ref_cell.alignment)
        new_cell.fill = copy(ref_cell.fill)
        new_cell.number_format = ref_cell.number_format
    sheet.cell(row=insert_at, column=RESUMEN_LABEL_COL, value=label)
    for col in range(RESUMEN_FIRST_MONTH_COL, sheet.max_column + 1):
        sheet.cell(row=insert_at, column=col).value = None

    new_total_row = total_row + 1
    for col, match in total_formulas.items():
        start_col, start_row, end_col, end_row = match.groups()
        new_end_row = int(end_row) + 1
        sheet.cell(row=new_total_row, column=col).value = f"=SUM({start_col}{start_row}:{end_col}{new_end_row})"

    if diff_row is not None and chase_row is not None:
        new_diff_row = diff_row + 1
        new_chase_row = chase_row + 1
        for col in diff_formulas:
            col_letter = get_column_letter(col)
            sheet.cell(row=new_diff_row, column=col).value = (
                f"=+{col_letter}{new_total_row}-{col_letter}{new_chase_row}"
            )

    return insert_at


def _shift_resumen_compras_refs(resumen_sheet, sheet_title, insert_at_row):
    """
    Corrige las referencias que RESUMEN COMPRAS ya tenía hacia esta hoja de
    proveedor después de insertar una fila en insert_at_row -- openpyxl no
    traduce fórmulas de OTRA hoja al insertar filas como sí hace Excel, así
    que cualquier referencia tipo "AIRGAS!D130" que ya apuntaba a
    insert_at_row o más abajo queda apuntando a la fila vieja (equivocada)
    tras el corrimiento. Bug real encontrado el 2026-09-02 insertando una
    factura cronológicamente antes de otra ya cargada y verificando que
    RESUMEN COMPRAS seguía sumando el monto correcto después.

    Escanea TODAS las filas de RESUMEN COMPRAS (no solo la de un proveedor
    puntual) buscando referencias a esta hoja -- necesario porque también
    se llama desde Pagos (append_supplier_payments), que identifica al
    proveedor por hoja de destino (proveedores_pago_rules.py), no por
    SUPPLIER_REGISTRY, así que no siempre tiene a mano el resumen_label
    para ubicar la fila de antemano. Segundo bug real encontrado el mismo
    día: un pago de Chase insertado en medio de una hoja (empujando
    facturas ya cargadas hacia abajo) dejaba estas mismas referencias
    desactualizadas exactamente igual que una factura insertada en medio
    -- Pagos nunca llamaba a este ajuste, solo Facturas.
    """
    if resumen_sheet is None:
        return

    quoted = _resumen_quote_sheet_ref(sheet_title)
    pattern = re.compile(re.escape(quoted) + r"!D(\d+)")

    def _shift(match):
        old_row = int(match.group(1))
        new_row = old_row + 1 if old_row >= insert_at_row else old_row
        return f"{quoted}!D{new_row}"

    for row in range(1, resumen_sheet.max_row + 1):
        for col in range(RESUMEN_FIRST_MONTH_COL, resumen_sheet.max_column + 1):
            cell = resumen_sheet.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                cell.value = pattern.sub(_shift, cell.value)


def _update_resumen_compras(resumen_sheet, sheet_title, resumen_label, invoice_row, invoice_date):
    """
    Suma la factura recién agregada en la celda proveedor x mes de RESUMEN
    COMPRAS -- crea la fila del proveedor si hace falta (ver
    _resumen_create_supplier_row) y agrega la referencia a la fórmula
    existente en vez de reemplazarla, igual que hacía el usuario a mano.

    Devuelve un dict {"status": ..., ...} para el resumen de la carga; nunca
    levanta una excepción hacia afuera -- la factura ya se cargó bien en la
    hoja del proveedor, esto es un agregado informativo.
    """
    try:
        year_cell = resumen_sheet.cell(row=1, column=RESUMEN_LABEL_COL).value
        try:
            resumen_year = int(str(year_cell).strip())
        except (TypeError, ValueError):
            resumen_year = None
        if resumen_year is not None and resumen_year != invoice_date.year:
            return {
                "status": "year_mismatch",
                "supplier": resumen_label,
                "detail": f"RESUMEN COMPRAS está en {resumen_year}, la factura es de {invoice_date.year}.",
            }

        month_col = _resumen_month_column(resumen_sheet, invoice_date.month)
        if month_col is None:
            return {
                "status": "month_not_found",
                "supplier": resumen_label,
                "detail": f"No se encontró la columna del mes {invoice_date.month} en RESUMEN COMPRAS.",
            }

        row = _resumen_find_supplier_row(resumen_sheet, resumen_label)
        created = False
        if row is None:
            row = _resumen_create_supplier_row(resumen_sheet, resumen_label)
            created = True

        cell = resumen_sheet.cell(row=row, column=month_col)
        ref = f"{_resumen_quote_sheet_ref(sheet_title)}!D{invoice_row}"
        current = cell.value
        if isinstance(current, str) and current.strip().startswith("="):
            cell.value = current + f"+{ref}"
        else:
            old_number = 0.0
            if isinstance(current, (int, float)):
                old_number = float(current)
            elif isinstance(current, str) and current.strip():
                try:
                    old_number = float(current.strip())
                except ValueError:
                    old_number = 0.0
            if old_number:
                cell.value = f"=+{old_number}+{ref}"
            else:
                cell.value = f"=+{ref}"

        return {"status": "created_row" if created else "updated", "supplier": resumen_label}
    except Exception as exc:  # noqa: BLE001 - nunca debe tumbar la carga de la factura
        return {"status": "error", "supplier": resumen_label, "detail": str(exc)}


def append_supplier_invoices(ledger_path, pdf_paths):
    """
    Lee cada PDF, detecta su proveedor y agrega una fila "invoice" a la
    hoja correspondiente del libro -- cada una en su posición cronológica
    correcta dentro de esa hoja (ver _find_invoice_insertion_point). Un
    mismo invoice ya cargado (por N° de factura) se omite. Guarda una
    copia temporal y la devuelve; quien llame se encarga de servirla para
    descargar (nunca se abre sola en la PC del servidor).
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Proveedores requiere openpyxl. Instale con: pip install openpyxl")

    ledger_path = os.path.abspath(ledger_path)
    if not os.path.isfile(ledger_path):
        raise FileNotFoundError(f"Excel no encontrado: {ledger_path}")

    by_supplier = {}
    failed = []
    for pdf_path in pdf_paths:
        supplier_key = None
        try:
            supplier_key = _detect_supplier(pdf_path)
            invoice = SUPPLIER_REGISTRY[supplier_key]["extract"](pdf_path)
        except (ValueError, TypeError, AttributeError, RuntimeError) as exc:
            # Antes solo se atrapaba ValueError -- varios extractores hacen
            # trabajo de imagen (pytesseract, Pillow, pdfplumber sobre un
            # PDF corrupto) que puede tirar otros tipos de excepción, y esas
            # se escapaban hasta el catch-all genérico de la ruta, tirando
            # el LOTE ENTERO (incluidas facturas de otros proveedores ya
            # extraídas bien) y mostrando el texto crudo de la excepción
            # -- que puede traer el nombre de archivo incrustado -- en vez
            # del aviso corto agrupado por proveedor de más abajo.
            supplier_label = SUPPLIER_REGISTRY[supplier_key]["label"] if supplier_key else None
            failed.append({"filename": os.path.basename(pdf_path), "error": str(exc), "supplier": supplier_label})
            continue
        invoice["filename"] = os.path.basename(pdf_path)
        by_supplier.setdefault(supplier_key, []).append(invoice)

    if not by_supplier:
        raise ValueError("No se pudo cargar ninguna factura. Revisalas a mano.")

    workbook = load_workbook(ledger_path, data_only=False)
    resumen_sheet = _get_resumen_sheet(workbook)
    resumen_warnings = []
    if resumen_sheet is None:
        resumen_warnings.append({"status": "sheet_not_found", "supplier": None})

    batch_results = []
    total_appended = 0
    for supplier_key, invoices in by_supplier.items():
        config = SUPPLIER_REGISTRY[supplier_key]
        try:
            sheet = _get_supplier_sheet(workbook, config["sheet_name"])
        except ValueError:
            # Mismo criterio que ya usa append_supplier_payments más abajo
            # -- antes esta búsqueda no estaba protegida acá, así que un
            # Ledger sin la hoja esperada (versión vieja, hoja renombrada)
            # tiraba el lote ENTERO de todos los proveedores, no solo el de
            # la hoja faltante.
            for invoice in invoices:
                failed.append({
                    "filename": invoice["filename"],
                    "error": f'La hoja "{config["sheet_name"]}" no existe en el Ledger.',
                    "supplier": config["label"],
                })
            continue
        invoices.sort(key=lambda inv: inv["date"])

        existing_numbers = _existing_invoice_numbers(sheet)

        appended = 0
        duplicates_skipped = []
        for invoice in invoices:
            if invoice["invoice_no"] in existing_numbers:
                duplicates_skipped.append(invoice["filename"])
                continue

            structural_mutation_done = False
            try:
                target_row, needs_shift, style_row, balance_ref_row, last_row, color = (
                    _find_invoice_insertion_point(sheet, invoice["date"])
                )
                if needs_shift:
                    _insert_row_preserving_merges(sheet, target_row)
                    _shift_resumen_compras_refs(resumen_sheet, sheet.title, target_row)
                    structural_mutation_done = True
                _write_invoice_row(sheet, target_row, style_row, balance_ref_row, color, invoice)
                if needs_shift:
                    _reformulate_rows_below(sheet, target_row, last_row)
            except (ValueError, TypeError, AttributeError) as exc:
                # No hay forma de deshacer limpio un insert_rows/repunteo de
                # RESUMEN COMPRAS ya aplicado -- si la falla ocurre DESPUÉS
                # de eso (estructuralmente posible aunque no se encontró un
                # disparador real en datos de producción), avisar que la
                # hoja quedó modificada de verdad, no solo "revisar a mano
                # esta factura" como si no hubiera pasado nada.
                failed.append({
                    "filename": invoice["filename"],
                    "error": str(exc),
                    "supplier": config["label"],
                    "partial_write": structural_mutation_done,
                })
                continue

            existing_numbers.add(invoice["invoice_no"])
            appended += 1

            if resumen_sheet is not None:
                resumen_result = _update_resumen_compras(
                    resumen_sheet, sheet.title, config["resumen_label"], target_row, invoice["date"]
                )
                if resumen_result["status"] not in ("updated", "created_row"):
                    resumen_warnings.append(resumen_result)

        _update_sheet_tab_color(sheet)

        total_appended += appended
        batch_results.append(
            {
                "supplier": config["label"],
                "sheet_name": config["sheet_name"],
                "invoices_appended": appended,
                "duplicates_skipped": duplicates_skipped,
            }
        )

    temp_path = _create_temp_workbook_path()
    workbook.save(os.path.abspath(temp_path))
    workbook.close()

    summary = {
        "files_processed": len(pdf_paths),
        "invoices_appended": total_appended,
        "batch_results": batch_results,
        "failed": failed,
        "resumen_warnings": resumen_warnings,
    }
    return temp_path, summary


# ---- Pagos (Fase 2, vía Chase) ----

# Columnas que se copian de estilo para una fila de pago (A,B,C,E,F,G) --
# a diferencia de INVOICE_ROW_COLUMNS, acá se copia E (HABER) en vez de D
# (DEBE), porque un pago escribe en HABER.
PAYMENT_ROW_COLUMNS = (1, 2, 3, 5, 6, 7)


def _find_bank_column(df, name_hint, fallback_index):
    """
    Ubica una columna del Excel del banco por nombre de encabezado
    (case-insensitive), con un índice fijo de respaldo si no aparece por
    nombre -- mismo patrón que find_chase_column en app.py.
    """
    hint = str(name_hint).strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == hint:
            return col
    if 0 <= fallback_index < len(df.columns):
        return df.columns[fallback_index]
    raise ValueError(f'No se encontró la columna "{name_hint}" en el Excel del banco.')


def _read_bank_payment_candidates(bank_path):
    """
    Lee el Excel/CSV del banco -- ya categorizado por la pestaña Chase Bank,
    con la columna Detalle ya rellenada -- y devuelve una lista de
    {"date", "amount", "description"} para cada fila cuyo Detalle sea
    "PROVEEDORES". date/amount quedan en None si no se pudieron leer; el
    llamador decide qué hacer con esas filas, nunca se descartan en
    silencio.
    """
    if pd is None:
        raise ImportError("Proveedores requiere pandas. Instale con: pip install pandas")

    ext = os.path.splitext(bank_path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(bank_path, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(bank_path, dtype=str, keep_default_na=False)

    description_col = _find_bank_column(df, "Description", 2)
    posting_col = _find_bank_column(df, "Posting Date", 1)
    amount_col = _find_bank_column(df, "Amount", 3)
    detalle_col = _find_bank_column(df, "Detalle", 7)

    candidates = []
    for _, row in df.iterrows():
        detalle = str(row.get(detalle_col, "")).strip().lower()
        if detalle != "proveedores":
            continue

        description = str(row.get(description_col, "")).strip()

        amount_text = str(row.get(amount_col, "")).strip().replace("$", "").replace(",", "")
        try:
            amount = abs(float(amount_text))
        except ValueError:
            amount = None

        date_text = str(row.get(posting_col, "")).strip()
        date_value = None
        if date_text:
            parsed_date = pd.to_datetime(date_text, errors="coerce")
            if not pd.isna(parsed_date):
                date_value = parsed_date.to_pydatetime()

        candidates.append({"date": date_value, "amount": amount, "description": description})

    return candidates


def _nearest_balance_row_at_or_above(sheet, row):
    """
    Camina hacia arriba desde row (inclusive) hasta encontrar una fila
    cuya columna F (BALANCE) tenga un valor -- fórmula o número -- para
    usar como referencia de saldo anterior.

    Ojo: NO busca por fecha real (columna A). En el libro real, las filas
    separadoras en blanco entre meses SÍ llevan su propia fórmula de
    arrastre de saldo (con D/E vacíos, simplemente repite el balance de
    la fila de arriba) aunque su columna A esté vacía -- confirmado
    inspeccionando el archivo real, no es una suposición. Referenciarlas
    directo es válido y es la misma convención que ya usa el resto de la
    hoja (cada fila real referencia la fila físicamente anterior, sea
    cual sea su tipo). Solo hace falta seguir subiendo cuando la fila de
    arriba está genuinamente vacía en F también (una hoja nueva, o un
    separador que todavía no tiene ninguna fórmula prellenada).

    El valor tiene que ser número o fórmula, no alcanza con "no vacío":
    la fila 7 del encabezado fijo de cada hoja trae el texto "Balance"
    literal en esta misma columna (es el título de la columna F), y
    algunas hojas reales tienen un tramo de facturas viejas cargadas sin
    ninguna fórmula de balance en absoluto (ej. AIRGAS entre 2023-10 y
    2023-12) -- en ninguno de los dos casos hay un saldo previo real del
    cual partir.
    """
    while row >= 1:
        value = sheet.cell(row=row, column=COL_BALANCE).value
        if isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith("=")):
            return row
        row -= 1
    return None


def _nearest_real_payment_style_row(sheet, from_row):
    """
    Ubica la fila real (con fecha datetime, invoice u OP) más cercana a
    from_row hacia arriba -- para copiarle el estilo a un pago nuevo.

    Nunca hay que confundir esto con _nearest_balance_row_at_or_above:
    esa función puede devolver una fila separadora de mes en blanco (que
    sí lleva su propia fórmula de arrastre de BALANCE, aunque columna A
    esté vacía), válida como referencia de saldo pero NO como referencia
    de estilo -- un separador puede estar combinado A:E con el texto del
    mes, o traer un number_format distinto en la columna de fecha. Bug
    real corregido 2026-09-03: _append_payment_row copiaba estilo
    directo de la misma fila que usaba para el balance, así que un pago
    insertado justo después de un separador heredaba su formato --
    causaba que la fecha del pago se pegara sin el formato de fecha
    correcto (se mostraba como texto crudo tipo "2026-08-03 0:00:00" en
    vez de con el formato esperado).
    """
    for row in range(from_row, 0, -1):
        if isinstance(sheet.cell(row=row, column=COL_DATE).value, datetime):
            return row
    return from_row


def _find_payment_insertion_point(sheet, payment_date):
    """
    Determina dónde debe ir un pago con fecha payment_date para mantener la
    hoja en orden cronológico.

    Devuelve (target_row, needs_shift, style_row, previous_balance_row, last_row):
    - needs_shift=False: target_row es el próximo lugar libre al final de
      la hoja (respetando el separador de mes existente), igual que
      _append_invoice_row -- no hace falta mover nada.
    - needs_shift=True: target_row ya tiene una fila real cargada -- hay
      que abrir espacio con _insert_row_preserving_merges(target_row),
      escribir el pago con _append_payment_row, y recién ahí llamar a
      _reformulate_rows_below para las filas que quedaron abajo.
    - previous_balance_row es la referencia para la fórmula de BALANCE --
      puede ser una fila separadora en blanco (ver
      _nearest_balance_row_at_or_above) -- nunca copiarle el estilo.
    - style_row es siempre una fila con fecha real (nunca un separador),
      para el estilo (ver _nearest_real_payment_style_row).
    """
    last_row = _find_last_real_row(sheet)
    if last_row is None:
        raise ValueError("La hoja no tiene ninguna fila real cargada.")
    last_date = sheet.cell(row=last_row, column=COL_DATE).value

    if payment_date >= last_date:
        month_changed = (payment_date.year, payment_date.month) != (last_date.year, last_date.month)
        target_row = last_row + 2 if month_changed else last_row + 1
        return target_row, False, last_row, last_row, last_row

    for row in range(1, last_row + 1):
        row_date = sheet.cell(row=row, column=COL_DATE).value
        # No alcanza con chequear None/"": el encabezado fijo de la hoja
        # (título, "PROVEEDOR:", nombres de columna) trae texto en esta
        # misma columna, y algunas hojas marcan el cambio de año con un
        # entero suelto (ej. 2025) -- comparar cualquiera de los dos
        # contra payment_date rompe con TypeError.
        if not isinstance(row_date, datetime):
            continue
        if row_date > payment_date:
            previous_balance_row = _nearest_balance_row_at_or_above(sheet, row - 1)
            if previous_balance_row is None:
                raise ValueError(
                    f"El pago del {payment_date:%d/%m/%Y} caería antes de que la hoja "
                    "tenga ninguna fila con un balance calculable (fórmula o número) de "
                    "la cual partir -- puede ser anterior a la primera factura, o caer "
                    "en un tramo viejo cargado sin fórmula de arrastre. Hay que cargarlo "
                    "a mano."
                )
            style_row = _nearest_real_payment_style_row(sheet, row - 1)
            return row, True, style_row, previous_balance_row, last_row

    # No debería llegar acá dado el chequeo de arriba, pero por las dudas.
    return last_row + 1, False, last_row, last_row, last_row


def _insert_row_preserving_merges(sheet, insert_at_row):
    """
    Abre una fila en insert_at_row igual que sheet.insert_rows(), pero
    sin dejar celdas combinadas "fantasma" en la posición equivocada.

    Las hojas reales de proveedores usan una celda combinada A:E de una
    sola fila como separador visual entre meses (con la etiqueta del mes
    en la celda ancla) -- en TODA la hoja, no solo en el punto donde
    justo cambia de mes al final. openpyxl.Worksheet.insert_rows mueve
    los valores de celda pero NO actualiza sheet.merged_cells, así que
    sin este paso cualquier separador debajo del punto de inserción
    queda combinado sobre la fila equivocada tras el corrimiento, y
    escribir ahí con _append_payment_row revienta con
    "MergedCell object attribute 'value' is read-only".
    """
    ranges_to_shift = [
        mcr.coord for mcr in list(sheet.merged_cells.ranges) if mcr.min_row >= insert_at_row
    ]
    for coord in ranges_to_shift:
        sheet.unmerge_cells(coord)

    sheet.insert_rows(insert_at_row, amount=1)

    for coord in ranges_to_shift:
        shifted = CellRange(coord)
        shifted.shift(row_shift=1)
        sheet.merge_cells(
            start_row=shifted.min_row,
            start_column=shifted.min_col,
            end_row=shifted.max_row,
            end_column=shifted.max_col,
        )


def _reformulate_rows_below(sheet, insert_at_row, last_row):
    """
    Después de insertar una fila nueva en insert_at_row -- y de haber
    escrito ya sus datos ahí, con _append_payment_row -- reconstruye la
    fórmula de BALANCE de cada fila real que quedó debajo, para que siga
    apuntando a la fila real inmediata anterior en su nueva posición.

    No alcanza con trasladar la fórmula vieja +1 fila (ej. con
    Translator): una fila justo después de un cambio de mes referencia la
    última fila real ANTES del separador, no la fila físicamente anterior
    -- ese salto tiene que recalcularse contra la nueva posición de la
    hoja, no simplemente correrse, o terminaría apuntando a la fila
    separadora en blanco (u otra fila equivocada) en vez de a la fila real
    que ahora quedó inmediatamente arriba. Por eso además esta función
    debe correr DESPUÉS de escribir la fila nueva, no antes: si insert_at_row
    todavía estuviera en blanco, "la fila real más cercana hacia arriba"
    la saltearía a ella también.

    No se salta las filas separadoras en blanco: en el libro real, esas
    filas también llevan su propia fórmula de BALANCE (arrastra el saldo
    sin sumar/restar nada, ya que D/E quedan vacíos) y esa fórmula queda
    igual de desactualizada tras el corrimiento. El chequeo real acá no
    es "tiene fecha" sino "tiene una fórmula en la columna F" -- eso
    cubre tanto facturas/pagos como separadores por igual.
    """
    for row in range(insert_at_row + 1, last_row + 2):
        cell = sheet.cell(row=row, column=COL_BALANCE)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            previous_row = _nearest_balance_row_at_or_above(sheet, row - 1)
            cell.value = (
                f"=+{sheet.cell(row=previous_row, column=COL_BALANCE).coordinate}"
                f"+{sheet.cell(row=row, column=COL_DEBE).coordinate}"
                f"-{sheet.cell(row=row, column=COL_HABER).coordinate}"
            )


def _existing_payment_keys(sheet):
    """
    Devuelve el conjunto de (fecha, monto) de las filas "OP" ya cargadas en
    la hoja -- para evitar duplicar un pago si se reprocesa el mismo
    período bancario.

    Se calcula UNA SOLA VEZ antes de procesar el lote entero de esa hoja,
    nunca fila por fila durante el loop: si se recalculara contra el
    estado ya mutado de la hoja en cada pago, un pago recién insertado en
    esta misma corrida contaminaría el chequeo del siguiente -- dos pagos
    legítimos y distintos que coinciden en fecha y monto (ej. dos cheques
    de $300 el mismo día, a facturas distintas) harían que el segundo se
    descarte como si fuera un duplicado del primero.
    """
    keys = set()
    for row in range(1, sheet.max_row + 1):
        comprob = sheet.cell(row=row, column=COL_COMPROB).value
        if not (isinstance(comprob, str) and comprob.strip().lower() == "op"):
            continue
        row_date = sheet.cell(row=row, column=COL_DATE).value
        row_amount = sheet.cell(row=row, column=COL_HABER).value
        if isinstance(row_amount, (int, float)):
            keys.add((row_date, round(row_amount, 2)))
    return keys


def _append_payment_row(sheet, target_row, style_row, previous_balance_row, payment):
    """
    Escribe una fila "OP" (pago) en target_row. A diferencia de
    _append_invoice_row: usa HABER (columna E) en vez de DEBE, no fuerza
    negrita ni aplica el fill mensual en la columna N° (un pago no tiene
    número de factura real). El estilo se copia de style_row -- una fila
    real con fecha, nunca un separador -- mientras que la fórmula de
    BALANCE referencia previous_balance_row, que sí puede ser un
    separador (ver _find_payment_insertion_point): son dos filas
    distintas a propósito, nunca asumir que son la misma.
    """
    for col in PAYMENT_ROW_COLUMNS:
        ref_cell = sheet.cell(row=style_row, column=col)
        new_cell = sheet.cell(row=target_row, column=col)
        new_cell.font = copy(ref_cell.font)
        new_cell.border = copy(ref_cell.border)
        new_cell.alignment = copy(ref_cell.alignment)
        new_cell.number_format = ref_cell.number_format

    sheet.cell(row=target_row, column=COL_DATE, value=payment["date"])
    sheet.cell(row=target_row, column=COL_COMPROB, value="OP")
    sheet.cell(row=target_row, column=COL_HABER, value=payment["amount"])
    sheet.cell(
        row=target_row,
        column=COL_BALANCE,
        value=f"=+{sheet.cell(row=previous_balance_row, column=COL_BALANCE).coordinate}"
        f"+{sheet.cell(row=target_row, column=COL_DEBE).coordinate}"
        f"-{sheet.cell(row=target_row, column=COL_HABER).coordinate}",
    )
    sheet.cell(row=target_row, column=COL_DETALLE).value = None


def append_supplier_payments(ledger_path, bank_path):
    """
    Lee el Excel del banco (ya categorizado por la pestaña Chase Bank),
    identifica los pagos a proveedores (Detalle="PROVEEDORES"), determina a
    qué hoja corresponde cada uno por la Descripción bancaria (reglas de
    proveedores_pago_rules.py) e inserta una fila "OP" en la posición
    cronológica correcta de esa hoja -- corriendo filas hacia abajo y
    trasladando fórmulas de BALANCE si el pago cae antes de facturas ya
    cargadas. Guarda una copia temporal y la devuelve; quien llame se
    encarga de servirla para descargar (nunca se abre sola en la PC del
    servidor).
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Proveedores requiere openpyxl. Instale con: pip install openpyxl")

    ledger_path = os.path.abspath(ledger_path)
    if not os.path.isfile(ledger_path):
        raise FileNotFoundError(f"Excel no encontrado: {ledger_path}")

    candidates = _read_bank_payment_candidates(bank_path)

    workbook = load_workbook(ledger_path, data_only=False)
    resumen_sheet = _get_resumen_sheet(workbook)

    by_sheet = {}
    unmatched = []
    for candidate in candidates:
        description = candidate["description"]
        if candidate["date"] is None or candidate["amount"] is None:
            unmatched.append({"supplier": None, "detail": f"{description!r}: no se pudo leer la fecha o el monto."})
            continue

        sheet_name = match_supplier_sheet(description)
        if not sheet_name:
            unmatched.append({"supplier": None, "detail": f"{description!r}: ningún proveedor conocido matchea esta descripción."})
            continue

        try:
            sheet = _get_supplier_sheet(workbook, sheet_name)
        except ValueError:
            unmatched.append({
                "supplier": sheet_name,
                "detail": f'{description!r}: la regla apunta a la hoja "{sheet_name}", que no existe en el libro.',
            })
            continue

        bucket = by_sheet.setdefault(sheet.title, {"sheet": sheet, "payments": []})
        bucket["payments"].append(candidate)

    batch_results = []
    total_appended = 0
    for sheet_title, bucket in by_sheet.items():
        sheet = bucket["sheet"]
        payments = sorted(bucket["payments"], key=lambda p: p["date"])

        appended = 0
        duplicates_skipped = []
        # Congelado ANTES del loop -- ver el docstring de _existing_payment_keys
        # sobre por qué no se puede recalcular fila por fila dentro del loop.
        existing_keys = _existing_payment_keys(sheet)
        for payment in payments:
            key = (payment["date"], round(payment["amount"], 2))
            if key in existing_keys:
                duplicates_skipped.append(payment["description"])
                continue

            # Aislado por pago con una excepción amplia a propósito: hojas
            # cargadas a mano durante años pueden traer de todo debajo del
            # punto de inserción -- una fecha tipeada como texto en vez de
            # datetime (TypeError al comparar), una celda combinada que
            # insert_rows corre sin avisar (AttributeError al escribir en
            # una MergedCell), etc. Sin esto, cualquiera de esos casos
            # aborta append_supplier_payments entero ANTES de workbook.save,
            # perdiendo también el trabajo ya hecho en otras hojas del lote.
            try:
                target_row, needs_shift, style_row, previous_balance_row, last_row = _find_payment_insertion_point(
                    sheet, payment["date"]
                )
                if needs_shift:
                    _insert_row_preserving_merges(sheet, target_row)
                    _shift_resumen_compras_refs(resumen_sheet, sheet.title, target_row)
                _append_payment_row(sheet, target_row, style_row, previous_balance_row, payment)
                if needs_shift:
                    _reformulate_rows_below(sheet, target_row, last_row)
            except (ValueError, TypeError, AttributeError) as exc:
                unmatched.append({"supplier": sheet_title, "detail": f"{payment['description']!r} ({sheet_title}): {exc}"})
                continue
            appended += 1

        _update_sheet_tab_color(sheet)

        total_appended += appended
        batch_results.append(
            {
                "sheet_name": sheet_title,
                "payments_appended": appended,
                "duplicates_skipped": duplicates_skipped,
            }
        )

    temp_path = _create_temp_workbook_path()
    workbook.save(os.path.abspath(temp_path))
    workbook.close()

    summary = {
        "candidates_found": len(candidates),
        "payments_appended": total_appended,
        "batch_results": batch_results,
        "unmatched": unmatched,
    }
    return temp_path, summary
