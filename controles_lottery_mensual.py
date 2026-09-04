"""
Controles: Lottery mensual — segundo módulo de la sección Controles.

Cruza el "Monthly Sales Report" del portal de Florida Lottery (mismo
formato exacto que el "Daily Sales Report" que ya lee Gettel/Lottery día a
día -- Terminal Sales/Pay/Commission online, Instant Sales/Pay/Commission
de raspaditas -- solo que acumulado para el mes completo) contra el total
del mes ya cargado en la hoja de Lottery del Excel "LOTTERY. Analisis
{mes}.{año}.xlsx" -- columnas F, G, H, I, K, P, Q, R, S.

Al ser el mismo formato/labels que el reporte diario, se reusa
`extract_lottery_receipt_fields_from_sales_report` tal cual (sin ningún
cambio) -- el propio total mensual que imprime el portal ya viene en las
mismas unidades/signo que usa la hoja para cada columna, así que no hace
falta re-derivar nada.

Es de solo lectura: nunca escribe nada en el Excel ni genera un archivo
para descargar -- el resultado se muestra en pantalla (verde/rojo por
chequeo), según lo ya decidido para toda la sección Controles.
"""

import calendar
import os
from datetime import datetime

from openpyxl import load_workbook

from controles_utils import eval_literal_sum_cell
from reporte_diario import (
    LOTTERY_COL_CASH_BALANCE,
    LOTTERY_COL_COMIS,
    LOTTERY_COL_DATE_B,
    LOTTERY_COL_PAGOS,
    LOTTERY_COL_PRIZE_FREE_PLAYS,
    LOTTERY_COL_SALES,
    LOTTERY_COL_SALES_COMM,
    LOTTERY_COL_SKOFF_PAYS_AMOUNT,
    LOTTERY_COL_SKOFF_PAYS_UNITS,
    LOTTERY_COL_SKOFF_SALES_AMOUNT,
    LOTTERY_DATA_START_ROW,
    _find_lottery_sheet,
    extract_lottery_receipt_fields_from_sales_report,
)

TOLERANCE = 0.01

# Nombre de campo (igual al que ya devuelve
# extract_lottery_receipt_fields_from_sales_report) -> columna de la hoja
# de Lottery -- una sola lista maneja tanto la suma mensual del Excel como
# el armado de los 9 chequeos de abajo, sin duplicar la lista de columnas
# dos veces.
_MONTHLY_SUM_COLUMNS = {
    "sales": LOTTERY_COL_SALES,
    "pagos": LOTTERY_COL_PAGOS,
    "cash_balance": LOTTERY_COL_CASH_BALANCE,
    "comis": LOTTERY_COL_COMIS,
    "prize_free_plays": LOTTERY_COL_PRIZE_FREE_PLAYS,
    "pays_units": LOTTERY_COL_SKOFF_PAYS_UNITS,
    "pays_amount": LOTTERY_COL_SKOFF_PAYS_AMOUNT,
    "skoff_sales_amount": LOTTERY_COL_SKOFF_SALES_AMOUNT,
    "sales_comm": LOTTERY_COL_SALES_COMM,
}


def _eval_literal_sum_cell(value):
    return eval_literal_sum_cell(value, "Lottery")


def _sum_lottery_month(sheet, year, month):
    """
    Suma, para cada fila real del mes (year, month) -- encontrada por
    columna B, nunca por posición fija, porque la hoja de Lottery tiene
    filas separadoras intercaladas -- cada columna listada en
    `_MONTHLY_SUM_COLUMNS`.

    Devuelve (totales, missing_days, periods_seen) -- un día del calendario
    sin ninguna fila con esa fecha en columna B se reporta en missing_days
    en vez de aportar 0 en silencio, para que el usuario sepa que ese día
    todavía no se cargó (y no confundir "no cargado" con "cargado en
    cero"). `periods_seen` es el conjunto de todos los (año, mes) que
    aparecen en columna B de la hoja completa, hayan matcheado (year,
    month) o no -- permite detectar en `check_lottery_monthly` que el
    Excel cargado pertenece a otro período distinto al del PDF (nada
    impide subir esa combinación, pero hay que avisarlo).
    """
    days_in_month = calendar.monthrange(year, month)[1]
    totals = {field: 0.0 for field in _MONTHLY_SUM_COLUMNS}
    found_days = set()
    periods_seen = set()
    max_row = max(sheet.max_row, LOTTERY_DATA_START_ROW)
    for row in range(LOTTERY_DATA_START_ROW, max_row + 1):
        date_value = sheet.cell(row=row, column=LOTTERY_COL_DATE_B).value
        if not isinstance(date_value, datetime):
            continue
        periods_seen.add((date_value.year, date_value.month))
        if date_value.year != year or date_value.month != month:
            continue
        found_days.add(date_value.day)
        for field, column in _MONTHLY_SUM_COLUMNS.items():
            totals[field] += _eval_literal_sum_cell(sheet.cell(row=row, column=column).value)
    missing_days = [day for day in range(1, days_in_month + 1) if day not in found_days]
    return totals, missing_days, periods_seen


def _build_check(label, pdf_value, excel_value, unit="$", tolerance=TOLERANCE):
    diff = round(pdf_value - excel_value, 2)
    return {
        "label": label,
        "pdf_value": round(pdf_value, 2),
        "excel_value": round(excel_value, 2),
        "diff": diff,
        "ok": abs(diff) <= tolerance,
        "unit": unit,
    }


def check_lottery_monthly(lottery_path, monthly_pdf_path):
    """
    Cruza el "Monthly Sales Report" del portal de Florida Lottery contra
    la hoja del mes correspondiente en el Excel de Lottery para el mismo
    mes/año (tomado del propio "Start Date" del PDF). Devuelve un dict
    con un chequeo por cada columna que ese reporte permite verificar --
    Ventas Terminal, Pagos Terminal, Balance Efectivo, Comisión Terminal,
    Prize/Promo Free Plays, Boletos Instantáneos Pagados, Pagos
    Instantáneos, Ventas Instantáneas y Comisión Instantánea -- listo
    para mostrar en pantalla. Nunca escribe ni descarga nada.
    """
    lottery_path = os.path.abspath(str(lottery_path).strip())
    if not os.path.isfile(lottery_path):
        raise FileNotFoundError(f"Excel de Lottery no encontrado: {lottery_path}")

    pdf_fields = extract_lottery_receipt_fields_from_sales_report(monthly_pdf_path)
    if pdf_fields["warning"]:
        raise ValueError(pdf_fields["warning"])
    report_date = pdf_fields["report_date"]

    workbook = load_workbook(lottery_path, data_only=False)
    try:
        sheet = _find_lottery_sheet(workbook)
        totals, missing_days, periods_seen = _sum_lottery_month(sheet, report_date.year, report_date.month)
    finally:
        workbook.close()

    # A diferencia de Store Info (Cierre Mensual), acá SIEMPRE es normal
    # encontrar alguna fila de otro mes -- el ciclo de liquidación real de
    # Florida Lottery no corta exacto en el límite del mes calendario, así
    # que un puñado de días de fin/comienzo del mes vecino en la misma hoja
    # no significa que se cargó el Excel equivocado. Por eso el aviso solo
    # se muestra cuando NINGÚN día real del mes pedido apareció en la hoja
    # (missing_days cubre el mes entero) -- ahí sí es una señal fuerte de
    # que el Excel o el reporte cargado pertenece a otro período. Si se leyó
    # bien aunque sea un solo día del mes correcto, se ignora cualquier otro
    # período que además esté presente.
    days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
    if len(missing_days) < days_in_month:
        period_mismatch = []
    else:
        other_periods = sorted(periods_seen - {(report_date.year, report_date.month)})
        period_mismatch = [f"{month:02d}/{year}" for year, month in other_periods]

    checks = [
        _build_check("Ventas Terminal (Online)", pdf_fields["sales"], totals["sales"]),
        _build_check("Pagos Terminal (Online)", pdf_fields["pagos"], totals["pagos"]),
        _build_check("Balance Efectivo Terminal", pdf_fields["cash_balance"], totals["cash_balance"]),
        _build_check("Comisión Terminal (Online)", pdf_fields["comis"], totals["comis"]),
        _build_check("Prize/Promo Free Plays", pdf_fields["prize_free_plays"], totals["prize_free_plays"]),
        _build_check(
            "Boletos Instantáneos Pagados",
            pdf_fields["pays_units"],
            totals["pays_units"],
            unit="u",
        ),
        _build_check("Pagos Instantáneos", pdf_fields["pays_amount"], totals["pays_amount"]),
        _build_check("Ventas Instantáneas", pdf_fields["skoff_sales_amount"], totals["skoff_sales_amount"]),
        _build_check("Comisión Instantánea", pdf_fields["sales_comm"], totals["sales_comm"]),
    ]

    return {
        "period_year": report_date.year,
        "period_month": report_date.month,
        "missing_days": missing_days,
        "period_mismatch": period_mismatch,
        "checks": checks,
        "all_ok": all(check["ok"] for check in checks) and not missing_days,
    }
