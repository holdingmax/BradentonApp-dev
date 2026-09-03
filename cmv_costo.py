"""
CMV / Elistars department cost automation for the master COSTO.TODOS sheet.

Parses department files with pandas, updates the master workbook with openpyxl
(style-preserving), and saves a temp preview only.
"""

import logging
import os
import tempfile

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]
    Border = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    Side = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False

try:
    import xlrd  # noqa: F401

    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)

COSTO_TODOS_SHEET_NAME = "COSTO.TODOS"
COSTO_TODOS_SHEET_INDEX = 3  # 1-based sheet position in workbook
DATA_START_ROW = 2  # Row 1 is header; product grid overwrite begins at A2
COSTO_DATA_COLUMNS = 7  # A through G (UPC … DeptName)
PRICE_CHANGE_COLUMN = 8  # H — Price this run minus Price last run, same UPC
PRICE_CHANGE_HEADER = "Cambio Precio"

if OPENPYXL_AVAILABLE and Side is not None and Border is not None:
    _THIN_SIDE = Side(style="thin", color="000000")
    COSTO_GRID_BORDER = Border(
        left=_THIN_SIDE,
        right=_THIN_SIDE,
        top=_THIN_SIDE,
        bottom=_THIN_SIDE,
    )
else:
    COSTO_GRID_BORDER = None

if OPENPYXL_AVAILABLE and PatternFill is not None:
    PRICE_INCREASE_FILL = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )  # light green
    PRICE_DECREASE_FILL = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )  # light red
else:
    PRICE_INCREASE_FILL = None
    PRICE_DECREASE_FILL = None

RAW_COLUMN_NAMES = [
    "UPC",
    "UPCMod",
    "Name",
    "Cost",
    "BuyDown",
    "Price",
    "DeptID",
    "DeptName",
    "SellUnit",
]
DROP_COLUMNS = frozenset({"BuyDown", "SellUnit"})
EXPECTED_FIELD_COUNT = len(RAW_COLUMN_NAMES)

PATH_SEPARATORS = (";", "|")


def join_paths(paths):
    """Join absolute department paths for the Entry widget (multi-select display)."""
    normalized = [
        os.path.abspath(str(path).strip())
        for path in paths
        if path is not None and str(path).strip()
    ]
    return "; ".join(normalized)


def split_paths(text):
    if not text or not str(text).strip():
        return []
    raw = str(text).strip()
    for separator in PATH_SEPARATORS:
        if separator in raw:
            return [part.strip() for part in raw.split(separator) if part.strip()]
    return [raw]


def _ensure_openpyxl_available():
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "La fusión de CMV requiere openpyxl. Instale con: pip install openpyxl"
        )


def clean_upc(value):
    """Cast to string, strip whitespace, and remove trailing '.0' artifacts."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, bool):
        return ""

    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if value == int(value):
            text = str(int(value))
        else:
            text = format(value, "f").rstrip("0").rstrip(".")
    else:
        text = str(value).strip()

    if not text or text.lower() in {"nan", "none"}:
        return ""

    if "." in text:
        whole, fractional = text.split(".", 1)
        if fractional == "" or set(fractional) <= {"0"}:
            text = whole

    return text.strip()


def _strip_cell(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _split_raw_fields(raw_text):
    parts = [_strip_cell(part) for part in str(raw_text).split(",")]
    if len(parts) < EXPECTED_FIELD_COUNT:
        parts.extend([""] * (EXPECTED_FIELD_COUNT - len(parts)))
    elif len(parts) > EXPECTED_FIELD_COUNT:
        name = ",".join(parts[2 : len(parts) - 6])
        parts = parts[:2] + [name] + parts[-6:]
    return parts[:EXPECTED_FIELD_COUNT]


def _row_to_fields(row):
    cells = [_strip_cell(value) for value in row]
    non_empty = [cell for cell in cells if cell]
    if not non_empty:
        return None

    if len(non_empty) == 1:
        return _split_raw_fields(non_empty[0])

    if len(cells) >= EXPECTED_FIELD_COUNT and any(cells[2:]):
        return cells[:EXPECTED_FIELD_COUNT]

    return _split_raw_fields(",".join(cells))


def _read_raw_table(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        return pd.read_csv(file_path, header=None, dtype=str, keep_default_na=False)
    if extension in {".xlsx", ".xlsm"}:
        return pd.read_excel(file_path, header=None, dtype=str, keep_default_na=False)
    if extension == ".xls":
        if not XLRD_AVAILABLE:
            raise ImportError(
                "Leer archivos de departamento .xls requiere xlrd. "
                "Instale las dependencias: pip install -r requirements.txt"
            )
        return pd.read_excel(file_path, header=None, dtype=str, engine="xlrd")
    raise ValueError(
        f"Tipo de archivo no soportado '{extension}'. Use .csv, .xlsx, .xlsm o .xls."
    )


def read_elistars_department_file(file_path):
    df = _read_raw_table(file_path)
    records = []
    for _, row in df.iterrows():
        fields = _row_to_fields(row.tolist())
        if fields is None or all(not field for field in fields):
            continue
        joined = ",".join(fields).lower()
        if joined.startswith("upc,") and "buydown" in joined:
            continue
        records.append(dict(zip(RAW_COLUMN_NAMES, fields)))

    if not records:
        raise ValueError(
            f"No hay filas de producto en {os.path.basename(file_path)}. "
            "Se esperaban datos separados por comas en una columna o en las columnas A–I."
        )

    result = pd.DataFrame(records)
    result = result.drop(columns=[c for c in DROP_COLUMNS if c in result.columns])
    return _finalize_department_frame(result)


def _parse_decimal(value):
    text = _strip_cell(value)
    if not text:
        return None
    text = text.replace("$", "").replace(" ", "")
    if text.count(",") == 1 and "." not in text:
        # Un unico "," sin punto es ambiguo: decimal europeo ("12,50") vs.
        # separador de miles US ("1,200"). Un costo/precio en moneda usa 1-2
        # decimales -- 3 digitos despues de la coma es la senal de que es
        # agrupamiento de miles (1,200 = mil doscientos, no 1.2).
        _, _, after_comma = text.partition(",")
        if len(after_comma) == 3:
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _finalize_department_frame(df):
    if "UPC" in df.columns:
        df["UPC"] = df["UPC"].astype(str).str.strip().apply(clean_upc)
        df = df[df["UPC"] != ""]

    for col in ("Cost", "Price"):
        if col in df.columns:
            df[col] = df[col].apply(_parse_decimal)

    for col in ("UPCMod", "Name", "DeptName"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df


def _department_sort_column(df):
    """Resolve department label column ('dept name' / DeptName) for grouping."""
    for col in df.columns:
        normalized = str(col).strip().lower().replace("_", " ")
        if normalized in {"dept name", "deptname"}:
            return col
    if "DeptName" in df.columns:
        return "DeptName"
    return None


def _sort_by_department(df):
    """Group products by department before writing to COSTO.TODOS."""
    dept_col = _department_sort_column(df)
    if dept_col is None:
        return df

    sorted_df = df.copy()
    sorted_df[dept_col] = sorted_df[dept_col].astype(str).str.strip()
    return sorted_df.sort_values(by=dept_col, ascending=True, kind="stable")


def _consolidate_department_files(department_paths):
    """
    Read all Elistars department exports, concatenate, clean, and sort by dept.
    """
    frames = []
    file_stats = []
    ordered_paths = [os.path.abspath(path) for path in department_paths]

    for department_path in ordered_paths:
        dept_df = read_elistars_department_file(department_path)
        label = infer_department_label(dept_df, department_path)
        frames.append(dept_df)
        file_stats.append(
            {
                "path": department_path,
                "label": label,
                "parsed": len(dept_df),
                "mapped": len(dept_df),
            }
        )

    if not frames:
        raise ValueError("No se pudo leer ninguna fila de departamento de los archivos seleccionados.")

    combined = pd.concat(frames, ignore_index=True)
    combined = _finalize_department_frame(combined)
    combined = _sort_by_department(combined)
    return combined, file_stats


def infer_department_label(df, source_path):
    if "DeptName" in df.columns:
        names = df["DeptName"].replace("", pd.NA).dropna().astype(str).str.strip()
        names = names[names != ""]
        if not names.empty:
            return names.mode().iloc[0]
    return os.path.splitext(os.path.basename(source_path))[0]


def _cell_is_empty(value):
    if value is None:
        return True
    text = str(value).strip()
    return text == "" or text.lower() in {"nan", "none"}




def _coerce_cost_price_float(value):
    """Cast Cost/Price to native float for Excel numeric cells."""
    if value is None:
        return None
    try:
        if isinstance(value, float) and pd.isna(value):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        parsed = _parse_decimal(value)
        if parsed is None:
            return None
        return float(parsed)
    except (TypeError, ValueError):
        return None


def _validate_master_path(master_path):
    extension = os.path.splitext(master_path)[1].lower()
    if extension not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError("El archivo maestro CMV debe ser .xls, .xlsx o .xlsm.")
    if extension == ".xls":
        raise ValueError(
            "Los maestros .xls antiguos no se pueden actualizar directamente con openpyxl. "
            "Abra el maestro en Excel y use Guardar como .xlsx o .xlsm, luego intente de nuevo."
        )
    return extension


def _create_temp_xlsx_path():
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="cmv_preview_")
    os.close(fd)
    return temp_path


def _get_costo_todos_sheet(workbook):
    """Target COSTO.TODOS by name; fall back to the third worksheet."""
    names = list(workbook.sheetnames)
    target = COSTO_TODOS_SHEET_NAME.strip().lower()
    for name in names:
        if name.strip().lower() == target:
            return workbook[name]

    if len(workbook.worksheets) >= COSTO_TODOS_SHEET_INDEX:
        return workbook.worksheets[COSTO_TODOS_SHEET_INDEX - 1]

    raise ValueError(
        f'Hoja "{COSTO_TODOS_SHEET_NAME}" no encontrada. Disponibles: {", ".join(names)}'
    )


def _load_master_workbook(master_path):
    _ensure_openpyxl_available()
    extension = os.path.splitext(master_path)[1].lower()
    keep_vba = extension == ".xlsm"
    return load_workbook(master_path, data_only=False, keep_vba=keep_vba)


def _hide_column_b(sheet):
    sheet.column_dimensions["B"].hidden = True


def _apply_costo_row_grid_border(sheet, row_idx):
    """Thin black border on all four sides for every cell in columns A–G."""
    if COSTO_GRID_BORDER is None:
        return
    for col in range(1, COSTO_DATA_COLUMNS + 1):
        sheet.cell(row=row_idx, column=col).border = COSTO_GRID_BORDER


def _write_upc_text_cell(sheet, row_idx, column, value, default="0"):
    """
    Write a UPC as text, never as a number.

    A leading zero only survives round-tripping through Excel if the cell is
    text — writing it as a number (even with a "0" display format) silently
    drops it, since an int can't represent one. This never invents digits:
    whatever clean_upc() already preserved (or didn't) is written as-is.
    """
    text = clean_upc(value) if value is not None else ""
    if not text:
        text = default
    cell = sheet.cell(row=row_idx, column=column, value=text)
    cell.number_format = "@"


def _write_currency_cell(sheet, row_idx, column, value):
    try:
        if isinstance(value, (int, float)):
            amount = float(value)
        else:
            coerced = _coerce_cost_price_float(value)
            if coerced is None:
                return
            amount = float(coerced)
    except (TypeError, ValueError):
        return

    cell = sheet.cell(row=row_idx, column=column, value=amount)
    cell.number_format = "0.00"


def _write_department_rows(sheet, dept_df, start_row, previous_prices=None):
    """Write consolidated department export starting at start_row; return next empty row."""
    previous_prices = previous_prices or {}
    row_idx = start_row
    rows_appended = 0

    for _, row in dept_df.iterrows():
        upc = clean_upc(row.get("UPC"))
        if not upc:
            continue

        _write_upc_text_cell(sheet, row_idx, 1, upc)

        upc_mod = row.get("UPCMod")
        if upc_mod is not None and not _cell_is_empty(upc_mod):
            _write_upc_text_cell(sheet, row_idx, 2, upc_mod, default="0")
        else:
            _write_upc_text_cell(sheet, row_idx, 2, "0", default="0")

        name = row.get("Name")
        sheet.cell(
            row=row_idx,
            column=3,
            value="" if name is None else str(name).strip(),
        )

        cost = row.get("Cost")
        if cost is not None:
            try:
                cost_float = (
                    float(cost)
                    if isinstance(cost, (int, float))
                    else _coerce_cost_price_float(cost)
                )
                if cost_float is not None:
                    _write_currency_cell(sheet, row_idx, 4, cost_float)
            except (TypeError, ValueError):
                pass

        price_float = None
        price = row.get("Price")
        if price is not None:
            try:
                price_float = (
                    float(price)
                    if isinstance(price, (int, float))
                    else _coerce_cost_price_float(price)
                )
                if price_float is not None:
                    _write_currency_cell(sheet, row_idx, 5, price_float)
            except (TypeError, ValueError):
                price_float = None

        dept_id = row.get("DeptID")
        if dept_id is not None and not _cell_is_empty(dept_id):
            try:
                sheet.cell(row=row_idx, column=6, value=int(float(str(dept_id).strip())))
            except (TypeError, ValueError):
                sheet.cell(row=row_idx, column=6, value=str(dept_id).strip())

        dept_name = row.get("DeptName")
        sheet.cell(
            row=row_idx,
            column=7,
            value="" if dept_name is None else str(dept_name).strip(),
        )

        _write_price_change_cell(sheet, row_idx, upc, price_float, previous_prices)

        _apply_costo_row_grid_border(sheet, row_idx)

        row_idx += 1
        rows_appended += 1

    return row_idx, rows_appended


def _row_has_costo_data(sheet, row_idx):
    """True when any product column (A–G) on the row contains a value."""
    return any(
        not _cell_is_empty(sheet.cell(row=row_idx, column=col).value)
        for col in range(1, COSTO_DATA_COLUMNS + 1)
    )


def _find_last_costo_data_row(sheet):
    """Last populated product row at or below DATA_START_ROW."""
    last_used = DATA_START_ROW - 1
    max_scan = max(sheet.max_row or 0, DATA_START_ROW)
    for row_idx in range(DATA_START_ROW, max_scan + 1):
        if _row_has_costo_data(sheet, row_idx):
            last_used = row_idx
    return last_used


def _clear_costo_data_grid(sheet):
    """
    Wipe existing product rows from DATA_START_ROW through the last used row.

    Uses delete_rows so stale UPCs, costs, prices, and cell styles are removed
    before the fresh import overwrites the grid from the top entry row.
    """
    last_row = _find_last_costo_data_row(sheet)
    if last_row < DATA_START_ROW:
        return 0

    rows_to_delete = last_row - DATA_START_ROW + 1
    sheet.delete_rows(DATA_START_ROW, amount=rows_to_delete)
    return rows_to_delete


def _snapshot_costo_prices(sheet):
    """
    Map UPC -> current Price before the grid is wiped, so the fresh import
    can show each product's price change against last run in column H.
    """
    prices = {}
    last_row = _find_last_costo_data_row(sheet)
    for row_idx in range(DATA_START_ROW, last_row + 1):
        upc = clean_upc(sheet.cell(row=row_idx, column=1).value)
        if not upc:
            continue
        price = sheet.cell(row=row_idx, column=5).value
        if isinstance(price, (int, float)):
            prices[upc] = float(price)
    return prices


def _read_existing_costo_rows(sheet):
    """
    Read every product row already in COSTO.TODOS before the grid is wiped.

    Lets a partial upload (e.g. only FLOWER's export) replace just the
    department(s) it contains, instead of the previous behavior of clearing
    the whole grid and keeping only what came in that particular upload —
    which silently dropped every other department's costs.
    """
    last_row = _find_last_costo_data_row(sheet)
    records = []
    for row_idx in range(DATA_START_ROW, last_row + 1):
        upc = clean_upc(sheet.cell(row=row_idx, column=1).value)
        if not upc:
            continue
        upc_mod = sheet.cell(row=row_idx, column=2).value
        name = sheet.cell(row=row_idx, column=3).value
        cost = sheet.cell(row=row_idx, column=4).value
        price = sheet.cell(row=row_idx, column=5).value
        dept_id = sheet.cell(row=row_idx, column=6).value
        dept_name = sheet.cell(row=row_idx, column=7).value
        records.append(
            {
                "UPC": upc,
                "UPCMod": "" if _cell_is_empty(upc_mod) else str(upc_mod).strip(),
                "Name": "" if name is None else str(name).strip(),
                "Cost": cost if isinstance(cost, (int, float)) else _parse_decimal(cost),
                "Price": price if isinstance(price, (int, float)) else _parse_decimal(price),
                "DeptID": "" if _cell_is_empty(dept_id) else str(dept_id).strip(),
                "DeptName": "" if dept_name is None else str(dept_name).strip(),
            }
        )
    return pd.DataFrame(
        records,
        columns=["UPC", "UPCMod", "Name", "Cost", "Price", "DeptID", "DeptName"],
    )


def _merge_with_existing_departments(sheet, new_df):
    """
    Merge freshly-parsed department rows into what's already in COSTO.TODOS.

    Only the department(s) present in new_df are replaced — every other
    department's existing rows are carried over untouched, so uploading a
    single department's export can never wipe out the rest of the grid.
    """
    existing_df = _read_existing_costo_rows(sheet)
    if existing_df.empty:
        return new_df

    touched = set(new_df["DeptName"].astype(str).str.strip().str.casefold()) - {""}
    if touched:
        keep_mask = ~existing_df["DeptName"].astype(str).str.strip().str.casefold().isin(touched)
        existing_df = existing_df[keep_mask]

    if existing_df.empty:
        return new_df
    return pd.concat([existing_df, new_df], ignore_index=True)


def _ensure_price_change_header(sheet):
    cell = sheet.cell(row=1, column=PRICE_CHANGE_COLUMN)
    if _cell_is_empty(cell.value):
        cell.value = PRICE_CHANGE_HEADER


def _write_price_change_cell(sheet, row_idx, upc, new_price, previous_prices):
    """
    Column H: new_price - last run's price for the same UPC.

    Blank for a brand-new UPC (nothing to compare against) and for an
    unchanged price — light green fill when it went up, red when it went
    down, matching the "cambio de precio" the user actually wants to see.
    """
    if new_price is None:
        return
    old_price = previous_prices.get(upc)
    if old_price is None:
        return

    diff = round(new_price - old_price, 2)
    if diff == 0:
        return

    cell = sheet.cell(row=row_idx, column=PRICE_CHANGE_COLUMN, value=diff)
    cell.number_format = "+0.00;-0.00"
    if PRICE_INCREASE_FILL is None:
        return
    cell.fill = PRICE_INCREASE_FILL if diff > 0 else PRICE_DECREASE_FILL


def _replace_costo_departments(sheet, department_paths):
    """
    Consolidate department exports and rewrite the COSTO.TODOS data grid.

    Only the department(s) present in department_paths are replaced -- every
    other department already in the sheet is preserved (see
    _merge_with_existing_departments). Uploading every department's export
    together still overwrites the whole grid, same as before.
    """
    _hide_column_b(sheet)
    combined_df, file_stats = _consolidate_department_files(department_paths)
    previous_prices = _snapshot_costo_prices(sheet)
    merged_df = _merge_with_existing_departments(sheet, combined_df)
    merged_df = _sort_by_department(merged_df)
    _clear_costo_data_grid(sheet)
    _ensure_price_change_header(sheet)
    _, rows_written = _write_department_rows(
        sheet, merged_df, DATA_START_ROW, previous_prices=previous_prices
    )
    _hide_column_b(sheet)
    return file_stats, rows_written


def _openpyxl_merge_and_save(master_path, department_paths, temp_xlsx_path):
    workbook = None
    try:
        workbook = _load_master_workbook(master_path)
        sheet = _get_costo_todos_sheet(workbook)
        file_stats, rows_appended = _replace_costo_departments(sheet, department_paths)
        workbook.save(os.path.abspath(temp_xlsx_path))
        return file_stats, rows_appended
    finally:
        if workbook is not None:
            workbook.close()


def update_master_costo_todos_bulk(master_path, department_paths):
    """
    Overwrite COSTO.TODOS product rows via openpyxl (clean-then-write).

    Reads all selected department files into one pandas DataFrame, cleans UPC
    and Cost/Price fields, sorts by department name, deletes existing rows from
    row 2 downward (columns A–G), then writes the consolidated import from A2.
    Saves a transient .xlsx preview and returns its path.

    Returns:
        tuple: (temp_xlsx_path, file_stats, total_parsed, rows_appended,
                upcs_not_in_master, master_row_count)
    """
    if not department_paths:
        raise ValueError("No se proporcionaron archivos de departamento.")

    master_path = os.path.abspath(master_path)
    _validate_master_path(master_path)

    temp_xlsx_path = _create_temp_xlsx_path()

    file_stats, rows_appended = _openpyxl_merge_and_save(
        master_path, department_paths, temp_xlsx_path
    )

    if rows_appended == 0:
        raise ValueError("No se pudo escribir ninguna fila de departamento en COSTO.TODOS.")

    total_parsed = sum(item["parsed"] for item in file_stats)
    return (
        temp_xlsx_path,
        file_stats,
        total_parsed,
        rows_appended,
        0,
        rows_appended,
    )


def workbook_has_costo_todos_sheet(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in {".xlsx", ".xlsm", ".xls"}:
        return False

    if extension in {".xlsx", ".xlsm"} and OPENPYXL_AVAILABLE:
        workbook = None
        try:
            workbook = load_workbook(
                os.path.abspath(file_path), read_only=True, data_only=False
            )
            _get_costo_todos_sheet(workbook)
            return True
        except ValueError:
            return False
        except Exception:
            pass
        finally:
            if workbook is not None:
                workbook.close()

    try:
        if extension == ".xls":
            if not XLRD_AVAILABLE:
                return False
            book = pd.ExcelFile(file_path, engine="xlrd")
        else:
            book = pd.ExcelFile(file_path)
        try:
            target = COSTO_TODOS_SHEET_NAME.lower()
            names = [name.strip().lower() for name in book.sheet_names]
            if target in names:
                return True
            return len(book.sheet_names) > COSTO_TODOS_SHEET_INDEX - 1
        finally:
            book.close()
    except Exception:
        return False
