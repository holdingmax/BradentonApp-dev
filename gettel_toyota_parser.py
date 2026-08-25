"""
GETTEL_TOYOTA_PARSER — Fuel coupon PDF extraction into a unified Excel workbook.

The source PDF holds an 8-column layout that is really two identical mirrored
blocks of 4 columns each (Date | Amount | Tracking | Gallons). The report rows
are mapped directly into a strict 8-column matrix — the left and right blocks
stay side by side (no vertical stacking) and the 6-digit tracking values are
preserved in their own columns for full mapping verification. Amount and gallons
are parsed into native floats with a '0.00' number format so a Spanish Excel
renders them with a decimal comma. Batch runs produce one workbook with one
worksheet per PDF (sheet title from the filename date range, e.g. 04-05 al 15-05).
"""

import io
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime

try:
    import pdfplumber
except ImportError:  # pragma: no cover - environment guard
    pdfplumber = None  # type: ignore[assignment]

try:
    import pandas as pd
except ImportError:  # pragma: no cover - environment guard
    pd = None  # type: ignore[assignment]

try:
    import pytesseract
    from PIL import Image
except ImportError:  # pragma: no cover - environment guard
    pytesseract = None  # type: ignore[assignment]
    Image = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils.datetime import from_excel

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - environment guard
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Border = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    Side = None  # type: ignore[assignment,misc]
    from_excel = None  # type: ignore[assignment,misc]
    OPENPYXL_AVAILABLE = False

# Vendor detection keywords -> (vendor label, output sheet name).
VENDOR_GETTEL = ("Gettel", "Gettel Report")
VENDOR_TOYOTA = ("Toyota", "Toyota Report")
GETTEL_TOKENS = ("GETTEL", "GETTLE")
TOYOTA_TOKENS = ("TOYOTA",)

DECIMAL_NUMBER_FORMAT = "0.00"
DATE_OUTPUT_FORMAT = "%m/%d/%Y"

# A standalone 6-digit integer is a tracking number (Col 3 / Col 7) and is dropped.
TRACKING_PATTERN = re.compile(r"^\d{6}$")
# Date tokens such as 03/14/2025 or 3/14/25.
DATE_PATTERN = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b")
# Filename date-range e.g. "04-05 al 15-05" from "Gettel report 04-05 al 15-05.pdf"
FILENAME_DATE_RANGE_PATTERN = re.compile(
    r"(\d{1,2}\s*[-/.]\s*\d{1,2}\s+al\s+\d{1,2}\s*[-/.]\s*\d{1,2})",
    re.IGNORECASE,
)
_VENDOR_PREFIX_PATTERN = re.compile(
    r"^(?:(?:gettel|gettle|toyota|report|reporte|cupon|cupón|fuel)[\s_\-.]*)+",
    re.IGNORECASE,
)

_DATE_INPUT_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d/%m/%Y",
    "%d/%m/%y",
)

LEFT_ALIGNMENT = Alignment(horizontal="left") if Alignment is not None else None
RIGHT_ALIGNMENT = Alignment(horizontal="right") if Alignment is not None else None
HEADER_FONT = Font(bold=True) if Font is not None else None


def _ensure_pdfplumber():
    if pdfplumber is None:
        raise ImportError(
            "GETTEL_TOYOTA_PARSER requiere pdfplumber. Instale con: pip install pdfplumber"
        )


_TESSERACT_CANDIDATE_PATHS = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
)
_TESSERACT_CONFIGURED = False


def _ensure_pytesseract():
    global _TESSERACT_CONFIGURED
    if pytesseract is None or Image is None:
        raise ImportError(
            "Leer reportes escaneados/fotografiados requiere pytesseract y Pillow. "
            "Instale con: pip install pytesseract pillow"
        )
    if pd is None:
        raise ImportError(
            "Leer reportes escaneados/fotografiados requiere pandas. "
            "Instale con: pip install pandas"
        )
    if _TESSERACT_CONFIGURED:
        return
    try:
        pytesseract.get_tesseract_version()
        _TESSERACT_CONFIGURED = True
        return
    except Exception:
        pass
    for candidate in _TESSERACT_CANDIDATE_PATHS:
        if os.path.isfile(candidate):
            pytesseract.pytesseract.tesseract_cmd = candidate
            _TESSERACT_CONFIGURED = True
            return
    raise ImportError(
        "No se encontró el motor Tesseract OCR. Instálelo (ej. con 'winget install "
        "UB-Mannheim.TesseractOCR') para poder leer reportes escaneados/fotografiados."
    )


def _ensure_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "GETTEL_TOYOTA_PARSER requiere openpyxl. Instale con: pip install openpyxl"
        )


def detect_vendor(page_texts):
    """
    Identify the report vendor from the header text of the pages.

    Returns:
        tuple[str, str]: (vendor label, output sheet name). Falls back to Gettel
        when no recognizable header token is present.
    """
    haystack = "\n".join(text for text in page_texts if text).upper()
    if any(token in haystack for token in TOYOTA_TOKENS):
        return VENDOR_TOYOTA
    if any(token in haystack for token in GETTEL_TOKENS):
        return VENDOR_GETTEL
    return VENDOR_GETTEL


def _parse_date(token):
    """Return a datetime for a MM/DD/YYYY-style token, or None when unparseable."""
    text = token.strip()
    for fmt in _DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _clean_number_token(token):
    """Strip currency noise ($, backslashes, commas, spaces) from a raw token."""
    return token.replace("$", "").replace("\\", "").replace(",", "").strip()


def _coerce_float(cleaned):
    """Return a float for a cleaned numeric token, or None when not numeric."""
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _correct_amount(value):
    """
    Enforce at most 3 integer digits on a Coupon Amount (real value < 1000.00).

    A PDF extraction that drops the decimal point yields a 4+ integer-digit value
    (e.g., 3868.00); divide by 100.0 to reposition the decimal (-> 38.68).
    """
    if value is None:
        return None
    while abs(value) >= 1000:
        value /= 100.0
    return value


def _correct_gallons(value):
    """
    Enforce at most 2 integer digits on a Gallons value (real value < 100.00).

    A 3+ integer-digit value (e.g., 1555.00 or 150.00) means the decimal point was
    lost; divide by 100.0 to restore the true reading (-> 15.55 or 1.50).
    """
    if value is None:
        return None
    while abs(value) >= 100:
        value /= 100.0
    return value


def _parse_block(block_text):
    """
    Parse one 4-column block (Date | Amount | Tracking | Gallons) into a record.

    The 6-digit tracking integer is kept in its own column for full mapping
    verification. A valid block yields a date plus two numeric values (amount
    first, gallons last). Returns a dict with keys date (datetime), amount
    (float), tracking (str), gallons (float), or None for incomplete/total rows.
    """
    date_match = DATE_PATTERN.search(block_text)
    if not date_match:
        return None
    parsed_date = _parse_date(date_match.group(1))
    if parsed_date is None:
        return None

    remainder = block_text[date_match.end():]
    numeric_values = []
    tracking = ""
    for raw_token in remainder.split():
        cleaned = _clean_number_token(raw_token)
        if not cleaned:
            continue
        if TRACKING_PATTERN.match(cleaned):
            tracking = cleaned  # 6-digit tracking — keep in its own column.
            continue
        value = _coerce_float(cleaned)
        if value is not None:
            numeric_values.append(value)

    if len(numeric_values) < 2:
        return None  # Incomplete trailing text / total row — discard.

    return {
        "date": parsed_date,
        "amount": _correct_amount(numeric_values[0]),
        "tracking": tracking,
        "gallons": _correct_gallons(numeric_values[-1]),
    }


def _split_mirrored_blocks(line):
    """
    Split a row into its left and right blocks using the two date anchors.

    The mirrored layout places one date at the start of each 4-column block, so
    the second date marks the boundary between the left and right halves.
    """
    matches = list(DATE_PATTERN.finditer(line))
    if not matches:
        return []
    if len(matches) == 1:
        return [line[matches[0].start():]]
    second = matches[1]
    left = line[matches[0].start():second.start()]
    right = line[second.start():]
    return [left, right]


def _parse_row(line):
    """
    Parse one PDF line into an 8-column row, keeping the left and right blocks
    side by side (no vertical stacking).

    Returns a dict with keys left and right, each either a parsed block dict or
    None. Returns None when neither block is valid.
    """
    blocks = _split_mirrored_blocks(line)
    if not blocks:
        return None
    left = _parse_block(blocks[0]) if len(blocks) >= 1 else None
    right = _parse_block(blocks[1]) if len(blocks) >= 2 else None
    if left is None and right is None:
        return None
    return {"left": left, "right": right}


def parse_fuel_coupons_pdf(pdf_path):
    """
    Extract fuel coupon rows from a Gettel/Toyota PDF report, grouped by page.

    Each row preserves the report's mirrored 8-column layout (left block and
    right block side by side) in original document order. Rows are grouped per
    PDF page so the writer can insert a blank separator row between pages.

    Returns:
        tuple[str, str, list[list[dict]]]: (vendor label, sheet name, pages).
        Each page is a list of row dicts with keys left and right; each block
        (or None) carries date, amount, tracking, and gallons.
    """
    _ensure_pdfplumber()
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    page_texts = []
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            page_texts.append(text)
            page_rows = []
            for line in text.splitlines():
                if not line.strip():
                    continue
                row = _parse_row(line)
                if row is not None:
                    page_rows.append(row)
            if page_rows:  # Skip empty pages so no spurious separator is added.
                pages.append(page_rows)

    vendor_label, sheet_name = detect_vendor(page_texts)
    return vendor_label, sheet_name, pages, page_texts


def _sanitize_excel_sheet_name(name):
    """Excel sheet titles: max 31 chars; no : \\ / ? * [ ]."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", str(name).strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return (cleaned[:31] if cleaned else "Report").strip()


def extract_sheet_title_from_filename(pdf_path):
    """
    Derive a worksheet title from the PDF filename.

    Example: "Gettel report 04-05 al 15-05.pdf" -> "04-05 al 15-05"
    """
    base = os.path.splitext(os.path.basename(pdf_path))[0].strip()
    range_match = FILENAME_DATE_RANGE_PATTERN.search(base)
    if range_match:
        title = range_match.group(1)
        title = re.sub(r"\s*[-/.]\s*", "-", title)
        title = re.sub(r"\s+al\s+", " al ", title, flags=re.IGNORECASE)
        return _sanitize_excel_sheet_name(title.strip())

    title = base
    while True:
        stripped = _VENDOR_PREFIX_PATTERN.sub("", title).strip(" _-.")
        if stripped == title:
            break
        title = stripped
    return _sanitize_excel_sheet_name(title or base)


def _unique_sheet_title(workbook, desired_title, used_titles):
    """Return a sheet title unique within the workbook (Excel + used set)."""
    base = _sanitize_excel_sheet_name(desired_title)
    candidate = base
    suffix = 2
    while candidate in used_titles or candidate in workbook.sheetnames:
        tail = f"_{suffix}"
        candidate = f"{base[: 31 - len(tail)]}{tail}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _normalize_pdf_paths(pdf_paths):
    normalized = []
    seen = set()
    for raw in pdf_paths:
        path = os.path.abspath(str(raw).strip())
        if not path or path in seen:
            continue
        seen.add(path)
        normalized.append(path)
    return normalized


def _build_output_path(pdf_paths, vendor_label):
    base_dir = os.path.dirname(os.path.abspath(pdf_paths[0]))
    if len(pdf_paths) == 1:
        filename = f"{vendor_label} Fuel Coupons.xlsx"
    else:
        filename = f"{vendor_label} Fuel Coupons Batch.xlsx"
    return os.path.join(base_dir, filename)


def _launch_workbook(path):
    """Open the saved workbook with a deferred, platform-aware launch."""
    abs_path = os.path.abspath(path)
    if sys.platform == "win32":
        time.sleep(0.35)
        os.startfile(abs_path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        os.system(f'open "{abs_path}"')
    else:
        os.system(f'xdg-open "{abs_path}"')


def _write_date_cell(worksheet, row, column, block):
    """Write a left-aligned MM/DD/YYYY date string (blank when the block is empty)."""
    cell = worksheet.cell(row=row, column=column)
    if block is not None and block.get("date") is not None:
        cell.value = block["date"].strftime(DATE_OUTPUT_FORMAT)
    if LEFT_ALIGNMENT is not None:
        cell.alignment = LEFT_ALIGNMENT


def _write_number_cell(worksheet, row, column, value):
    """Write a right-aligned native float with the '0.00' decimal format."""
    cell = worksheet.cell(row=row, column=column)
    if value is not None:
        cell.value = float(value)
        cell.number_format = DECIMAL_NUMBER_FORMAT
    if RIGHT_ALIGNMENT is not None:
        cell.alignment = RIGHT_ALIGNMENT


def _write_tracking_cell(worksheet, row, column, block):
    """Write the right-aligned 6-digit tracking value (kept for verification)."""
    cell = worksheet.cell(row=row, column=column)
    tracking = block.get("tracking") if block is not None else ""
    if tracking:
        cell.value = tracking
    if RIGHT_ALIGNMENT is not None:
        cell.alignment = RIGHT_ALIGNMENT


def _write_layout_row(worksheet, row, row_data):
    """Write one PDF line into its strict 8-column Excel row (A-H)."""
    left = row_data.get("left")
    right = row_data.get("right")

    # Left block: A Date | B Amount | C Tracking | D Gallons
    _write_date_cell(worksheet, row, 1, left)
    _write_number_cell(worksheet, row, 2, left.get("amount") if left else None)
    _write_tracking_cell(worksheet, row, 3, left)
    _write_number_cell(worksheet, row, 4, left.get("gallons") if left else None)

    # Right block: E Date | F Amount | G Tracking | H Gallons
    _write_date_cell(worksheet, row, 5, right)
    _write_number_cell(worksheet, row, 6, right.get("amount") if right else None)
    _write_tracking_cell(worksheet, row, 7, right)
    _write_number_cell(worksheet, row, 8, right.get("gallons") if right else None)


_LAYOUT_HEADERS = (
    "Date",
    "Amount",
    "Tracking",
    "Gallons",
    "Date",
    "Amount",
    "Tracking",
    "Gallons",
)


def _write_sheet_headers(worksheet):
    for column, label in enumerate(_LAYOUT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=label)
        if HEADER_FONT is not None:
            cell.font = HEADER_FONT


def _write_pages_to_worksheet(worksheet, pages):
    """
    Write parsed pages to one worksheet (8-column A-H, 1 blank row between pages).

    Returns:
        int: number of data rows written.
    """
    current_row = 2
    record_count = 0
    for page_index, page_rows in enumerate(pages):
        if page_index > 0:
            current_row += 1
        for row_data in page_rows:
            _write_layout_row(worksheet, current_row, row_data)
            current_row += 1
            record_count += 1
    return record_count


def write_multi_sheet_fuel_coupon_workbook(pdf_paths, output_path):
    """
    Parse each PDF into its own worksheet (hoja por archivo).

    Sheet names are derived from each file's date-range text in the filename.
    Returns:
        tuple[int, int]: (total data rows written, number of worksheets created)
    """
    _ensure_openpyxl()
    workbook = Workbook()
    used_titles = set()
    record_count = 0
    sheets_written = 0
    first_sheet = True
    all_page_texts = []

    for pdf_path in pdf_paths:
        _vendor, _sheet_name, pages, page_texts = parse_fuel_coupons_pdf(pdf_path)
        all_page_texts.extend(page_texts)
        if not pages:
            continue

        desired_title = extract_sheet_title_from_filename(pdf_path)
        sheet_title = _unique_sheet_title(workbook, desired_title, used_titles)

        if first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_title
            first_sheet = False
        else:
            worksheet = workbook.create_sheet(title=sheet_title)

        _write_sheet_headers(worksheet)
        record_count += _write_pages_to_worksheet(worksheet, pages)
        sheets_written += 1

    if record_count == 0:
        workbook.close()
        raise ValueError("No se encontraron registros de cupones de combustible en el/los PDF seleccionado(s).")

    workbook.save(output_path)
    workbook.close()
    vendor_label, _ = detect_vendor(all_page_texts)
    return record_count, sheets_written, vendor_label


def generate_fuel_coupon_workbook_from_pdfs(pdf_paths, output_path=None, launch=True):
    """
    Parse multiple PDFs into one workbook (one sheet per file).

    Returns:
        tuple[str, int, int]: (saved path, total data rows, sheet count)
    """
    pdf_paths = _normalize_pdf_paths(pdf_paths)
    if not pdf_paths:
        raise ValueError("No se proporcionaron rutas de PDF.")

    if output_path is None:
        probe_texts = []
        for pdf_path in pdf_paths:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    probe_texts.append(page.extract_text() or "")
        vendor_label, _ = detect_vendor(probe_texts)
        output_path = _build_output_path(pdf_paths, vendor_label)

    record_count, sheet_count, _vendor_label = write_multi_sheet_fuel_coupon_workbook(
        pdf_paths, output_path
    )
    if launch:
        _launch_workbook(output_path)
    return output_path, record_count, sheet_count


_SOURCE_FIRST_DATA_ROW = 3
_GETTEL_TOYOTA_DEST_SHEET_PATTERN = re.compile(
    r"^Gettel-Toyota\s+\d{2}\.\d{4}$",
    re.IGNORECASE,
)
_DAILY_SHEET_HEADER_SCAN_ROWS = 5
_DAILY_SHEET_COLUMN_TOKENS = {
    "date": ("date",),
    "amount": ("amount",),
    "gallons": ("gallon",),
}


def _parse_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(value, (int, float)) and from_excel is not None:
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.replace(hour=0, minute=0, second=0, microsecond=0)
        except (ValueError, OSError, TypeError):
            pass
    text = str(value).strip()
    if not text:
        return None
    return _parse_date(text)


def _date_dict_key(value):
    parsed = _parse_excel_date(value)
    if parsed is None:
        return None
    return parsed.date()


def _identify_vendor_sheets(workbook):
    """
    Find the Gettel and Toyota daily fuel-log sheets by name.

    The real source file (copy/pasted by hand from the station manager's
    reports) names its tabs "GETTLE"/"GETTEL"/"Gettel..." and "Toyota...",
    not the Excel-default "Sheet1"/"Sheet2" — so this matches by substring,
    tolerant of the "GETTLE" typo, instead of requiring exact default names.
    """
    gettel_sheet = None
    toyota_sheet = None
    for worksheet in workbook.worksheets:
        title = worksheet.title.strip().upper()
        if gettel_sheet is None and any(token in title for token in GETTEL_TOKENS):
            gettel_sheet = worksheet
        elif toyota_sheet is None and any(token in title for token in TOYOTA_TOKENS):
            toyota_sheet = worksheet

    if gettel_sheet is None or toyota_sheet is None:
        available = ", ".join(workbook.sheetnames) or "(none)"
        raise ValueError(
            "No se encontraron las hojas de Gettel y Toyota en el Excel de "
            f"origen. Hojas disponibles: {available}"
        )
    if gettel_sheet.title == toyota_sheet.title:
        raise ValueError("Gettel y Toyota resolvieron a la misma hoja.")
    return gettel_sheet, toyota_sheet


def _find_daily_sheet_header_row(worksheet):
    """Locate the header row carrying Date/Amount/Gallons labels."""
    max_col = max(worksheet.max_column, 4)
    for row in range(1, _DAILY_SHEET_HEADER_SCAN_ROWS + 1):
        labels = [
            str(worksheet.cell(row=row, column=col).value or "").strip().lower()
            for col in range(1, max_col + 1)
        ]
        if any("date" in label for label in labels) and any(
            "amount" in label for label in labels
        ):
            return row
    raise ValueError(
        f"No se encontró una fila de encabezado Date/Amount en la hoja '{worksheet.title}'."
    )


def _map_daily_sheet_columns(worksheet, header_row):
    """Map Date/Amount/Gallons to column indices (Tracking is dropped)."""
    columns = {}
    max_col = max(worksheet.max_column, 4)
    for col in range(1, max_col + 1):
        label = str(worksheet.cell(row=header_row, column=col).value or "").strip().lower()
        for key, tokens in _DAILY_SHEET_COLUMN_TOKENS.items():
            if key in columns:
                continue
            if any(token in label for token in tokens):
                columns[key] = col
                break
    missing = [key for key in ("date", "amount", "gallons") if key not in columns]
    if missing:
        raise ValueError(
            f"A la hoja '{worksheet.title}' le faltan columna(s): {', '.join(missing)}."
        )
    return columns


def _parse_daily_sheet_number(value):
    """
    Parse an Amount/Gallons cell from the daily fuel-log sheet.

    Amounts are pasted as Latin-formatted text like "$42,52" (comma is the
    decimal separator); Gallons are plain native numbers. Neither needs the
    PDF-OCR lost-decimal correction used elsewhere in this module — this
    data comes from a direct copy/paste into Excel, not text extraction.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(" ", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _summarize_daily_sheet(worksheet):
    """Sum Amount/Gallons per calendar day from a Gettel/Toyota daily-entry sheet."""
    header_row = _find_daily_sheet_header_row(worksheet)
    columns = _map_daily_sheet_columns(worksheet, header_row)
    totals_by_date = {}
    max_row = worksheet.max_row or header_row
    for row in range(header_row + 1, max_row + 1):
        day = _date_dict_key(worksheet.cell(row=row, column=columns["date"]).value)
        if day is None:
            continue
        amount = _parse_daily_sheet_number(
            worksheet.cell(row=row, column=columns["amount"]).value
        )
        gallons = _parse_daily_sheet_number(
            worksheet.cell(row=row, column=columns["gallons"]).value
        )
        if amount is None and gallons is None:
            continue
        bucket = totals_by_date.setdefault(day, {"amount": 0.0, "gallons": 0.0})
        bucket["amount"] += amount or 0.0
        bucket["gallons"] += gallons or 0.0
    return totals_by_date


def _summarize_origin_workbook(source_workbook):
    """Summarize the Gettel and Toyota daily fuel-log sheets independently."""
    gettel_sheet, toyota_sheet = _identify_vendor_sheets(source_workbook)
    gettel_totals = _summarize_daily_sheet(gettel_sheet)
    toyota_totals = _summarize_daily_sheet(toyota_sheet)
    return gettel_totals, toyota_totals


# ---------------------------------------------------------------------------
# Scanned/photographed report parsing (OCR)
#
# The manager sometimes sends a phone photo of the same Date/Amount/Tracking/
# Gallons report instead of an Excel export, saved as a PDF with no text
# layer (each page is one embedded photo). Printed on a landscape page, the
# report packs two mirrored copies of the 4-column layout side by side to
# fit more rows. Any handwritten totals/adjustments scrawled below the
# printed table are ignored entirely — only OCR text from the printed table
# is trusted, and only rows with a recognizable printed date are kept.
# ---------------------------------------------------------------------------

_OCR_MIN_WORD_CONFIDENCE = 40
_OCR_ROW_HEIGHT_TOLERANCE_FACTOR = 0.6
_OCR_COLUMNS_PER_BLOCK = 4  # Date, Amount, Tracking (dropped), Gallons


_OSD_ROTATE_TO_TRANSPOSE = {
    90: Image.ROTATE_270 if Image is not None else None,
    180: Image.ROTATE_180 if Image is not None else None,
    270: Image.ROTATE_90 if Image is not None else None,
}


def _correct_image_orientation(image):
    """
    Detect and undo whole-page rotation (phone photos taken upside-down or
    sideways — common when a multi-page report is snapped in a hurry, and
    inconsistent from page to page within the same PDF) via Tesseract's own
    orientation/script detection, before the main OCR pass ever runs.

    Uses Image.transpose (exact 90°-multiple remap, no interpolation)
    instead of Image.rotate — .rotate() resamples every pixel even for
    right-angle turns, which was blurring characters (especially the "/" in
    dates) just enough to break column detection on rotated pages.
    Never raises — a page OSD can't confidently read is left as-is.
    """
    try:
        osd = pytesseract.image_to_osd(image, output_type=pytesseract.Output.DICT)
        rotate = int(osd.get("rotate", 0) or 0)
    except Exception:
        return image
    transpose_const = _OSD_ROTATE_TO_TRANSPOSE.get(rotate)
    if transpose_const is not None:
        image = image.transpose(transpose_const)
    return image


def _extract_pdf_page_images(pdf_path):
    """
    Return one PIL Image per PDF page — the largest embedded photo on that
    page, rotated upright — without needing poppler/ghostscript (the JPEG
    bytes are pulled directly from the PDF's own image XObject).
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()
    images = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if not page.images:
                continue
            biggest = max(page.images, key=lambda im: im["width"] * im["height"])
            raw = biggest["stream"].get_data()
            image = Image.open(io.BytesIO(raw)).convert("RGB")
            images.append(_correct_image_orientation(image))
    if not images:
        raise ValueError(f"No se encontraron imágenes embebidas en {pdf_path}.")
    return images


# No single Tesseract page-segmentation mode is reliably best across every
# photo — PSM 12 (sparse text + OSD) read one report's photos almost
# perfectly, but read ~30% of dates as garbage on a different, lower-quality
# photo of the same report format, while other modes did better there. So
# every page is tried under each of these and scored (see
# _best_effort_extract_page) rather than trusting one fixed mode.
_OCR_TESSERACT_CONFIGS = ("--psm 12", "--psm 11", "--psm 3", "--psm 6", "--psm 4")


def _ocr_words(image, config=_OCR_TESSERACT_CONFIGS[0]):
    """OCR one page image into a cleaned word-level DataFrame (left/top/text)."""
    raw = pytesseract.image_to_data(
        image, output_type=pytesseract.Output.DATAFRAME, config=config
    )
    words = raw[raw.text.notna() & (raw.text.str.strip() != "")]
    words = words[words.conf > _OCR_MIN_WORD_CONFIDENCE]
    return words


def _cluster_words_into_rows(words):
    """Group OCR words into text rows by vertical (y) proximity."""
    if words.empty:
        return []
    working = words.copy()
    working["center_y"] = working.top + working.height / 2
    tolerance = max(8.0, working.height.median() * _OCR_ROW_HEIGHT_TOLERANCE_FACTOR)

    rows = []
    current = []
    last_y = None
    for _, word in working.sort_values("center_y").iterrows():
        if last_y is not None and abs(word.center_y - last_y) > tolerance:
            rows.append(current)
            current = []
        current.append(word)
        last_y = word.center_y
    if current:
        rows.append(current)
    return rows


def _row_date_token(row_words):
    """Return the first printed date-looking token in a row, if any."""
    for word in row_words:
        text = str(word.text).strip()
        if DATE_PATTERN.fullmatch(text) or re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
            return text
    return None


def _detect_block_count(rows):
    """
    Count how many mirrored Date/Amount/Tracking/Gallons blocks the page
    has, by finding the header row and counting its "Date" labels.

    More reliable than gap-detection on data-row positions: the printed
    Amount/Gallons values are right-aligned and can sit much further from
    their header label than the label-to-label gaps would suggest, so
    counting header labels (not guessing from data-row spacing) is what
    tells us how many blocks to expect.
    """
    for row in rows:
        date_labels = sum(1 for w in row if str(w.text).strip().lower() == "date")
        if date_labels >= 1:
            return date_labels
    return None


def _kmeans_1d(values, k, iterations=50):
    """Minimal 1-D k-means — no external dependency needed for this."""
    values = sorted(values)
    if k <= 1 or len(values) <= k:
        return [sum(values) / len(values)] if values else []
    lo, hi = values[0], values[-1]
    centers = [lo + (hi - lo) * i / (k - 1) for i in range(k)]
    for _ in range(iterations):
        buckets = [[] for _ in range(k)]
        for v in values:
            idx = min(range(k), key=lambda i: abs(v - centers[i]))
            buckets[idx].append(v)
        new_centers = [
            (sum(bucket) / len(bucket)) if bucket else centers[i]
            for i, bucket in enumerate(buckets)
        ]
        if new_centers == centers:
            break
        centers = new_centers
    return sorted(centers)


def _detect_ocr_column_boundaries(data_rows, block_count):
    """
    Cluster every data-row word's x-position into exactly
    block_count * _OCR_COLUMNS_PER_BLOCK column centers (k-means), then
    return the midpoints between consecutive centers as boundaries.

    Clustering the data itself (rather than assuming a fixed gap size)
    adapts to whatever resolution/crop a given photo happens to have.
    """
    lefts = [word.left for row in data_rows for word in row]
    if not lefts:
        return []
    k = block_count * _OCR_COLUMNS_PER_BLOCK
    centers = _kmeans_1d(lefts, k)
    return [
        (a + b) / 2 for a, b in zip(centers, centers[1:])
    ]


def _assign_row_columns(row_words, boundaries):
    """Bucket one row's words into column-index groups using x boundaries."""
    bounds = [0.0] + sorted(boundaries) + [float("inf")]
    buckets = ["" for _ in range(len(bounds) - 1)]
    for word in sorted(row_words, key=lambda w: w.left):
        for idx in range(len(bounds) - 1):
            if bounds[idx] <= word.left < bounds[idx + 1]:
                text = str(word.text).strip()
                buckets[idx] = (buckets[idx] + " " + text).strip() if buckets[idx] else text
                break
    return buckets


def _parse_ocr_amount_or_number(text, max_plausible=None):
    """
    Parse a printed $Amount or plain decimal, tolerant of stray OCR
    punctuation. If max_plausible is given and the parsed value exceeds it,
    returns None instead — a single fuel fill-up is never $500+ or 150+
    gallons, so a value that large means OCR merged/misread digits (e.g.
    the tracking number bleeding into the amount), not a real reading.
    """
    if not text:
        return None
    cleaned = text.replace("$", "").replace(",", "").strip().strip(".").strip("_")
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if max_plausible is not None and abs(value) > max_plausible:
        return None
    return value


_MAX_PLAUSIBLE_AMOUNT = 500.0
_MAX_PLAUSIBLE_GALLONS = 150.0


def _estimate_block_count_from_data_rows(data_rows):
    """
    Fallback block-count guess when no page in the PDF has a header row
    (e.g. every page was cropped to just the data). Uses the typical
    number of words per data row, which should be a multiple of 4.
    """
    if not data_rows:
        return 1
    counts = sorted(len(row) for row in data_rows)
    median_count = counts[len(counts) // 2]
    return max(1, round(median_count / _OCR_COLUMNS_PER_BLOCK))


def _extract_page_transactions(image, block_count, config):
    """
    OCR one report page and return a list of (date_str, amount, gallons)
    tuples for every printed transaction row, across however many mirrored
    Date/Amount/Tracking/Gallons blocks the page has (Tracking is dropped).

    block_count is detected once per PDF (see summarize_pdf_report) since
    only the first page of a multi-page report repeats the header row.
    """
    words = _ocr_words(image, config)
    rows = _cluster_words_into_rows(words)

    data_rows = [row for row in rows if _row_date_token(row) is not None]
    if not data_rows:
        return []

    boundaries = _detect_ocr_column_boundaries(data_rows, block_count)

    transactions = []
    clean_date_matches = 0
    last_valid_date = [None] * block_count
    for row in data_rows:
        columns = _assign_row_columns(row, boundaries)
        for block in range(block_count):
            offset = block * _OCR_COLUMNS_PER_BLOCK
            if offset + 3 >= len(columns):
                break
            date_text = columns[offset]
            amount_text = columns[offset + 1]
            gallons_text = columns[offset + 3]

            match = DATE_PATTERN.search(date_text or "")
            if match:
                date_text = match.group(1)
                last_valid_date[block] = date_text
                clean_date_matches += 1
            elif last_valid_date[block] is not None:
                # Date cell OCR'd blank/garbled (e.g. a stray "_" prefix) but
                # the row itself is real — same-block rows run chronologically,
                # so reuse the last date this block actually read rather than
                # silently dropping a transaction whose amount did come through.
                date_text = last_valid_date[block]
            else:
                continue

            amount = _parse_ocr_amount_or_number(amount_text, _MAX_PLAUSIBLE_AMOUNT)
            gallons = _parse_ocr_amount_or_number(gallons_text, _MAX_PLAUSIBLE_GALLONS)
            if amount is None and gallons is None:
                continue
            transactions.append((date_text, amount, gallons))
    return transactions, clean_date_matches


def _score_page_transactions(transactions, clean_date_matches):
    """
    Higher is better. Cleanly-recognized date cells (not carried forward
    from a previous row) dominate the score — a config that reads more
    complete Amount/Gallons pairs but garbles a lot of dates is actually
    worse, since it silently misattributes those rows to the wrong day.
    Complete (Amount+Gallons both present) rows break ties.
    """
    complete = sum(1 for _date, amount, gallons in transactions if amount is not None and gallons is not None)
    return clean_date_matches * 100 + complete


def _best_effort_extract_page(image, block_count):
    """
    Try every candidate Tesseract config for this one page and keep
    whichever reads the most complete/plausible transactions.

    Returns:
        tuple[list, str]: (winning transactions, the config that won)
    """
    best_transactions = []
    best_score = -1
    best_config = _OCR_TESSERACT_CONFIGS[0]
    for config in _OCR_TESSERACT_CONFIGS:
        transactions, clean_date_matches = _extract_page_transactions(image, block_count, config)
        score = _score_page_transactions(transactions, clean_date_matches)
        if score > best_score:
            best_score = score
            best_transactions = transactions
            best_config = config
    return best_transactions, best_config


def _detect_ocr_subtotal(image, block_count, config):
    """
    Best-effort read of a printed subtotal row (Amount/Gallons per block,
    no date) for cross-checking the OCR'd sum — never blocks on failure.
    """
    try:
        words = _ocr_words(image, config)
        rows = _cluster_words_into_rows(words)
        data_rows = [row for row in rows if _row_date_token(row) is not None]
        boundaries = _detect_ocr_column_boundaries(data_rows, block_count)
        for row in rows:
            if _row_date_token(row) is not None:
                continue
            columns = _assign_row_columns(row, boundaries)
            totals = []
            for block in range(block_count):
                offset = block * _OCR_COLUMNS_PER_BLOCK
                if offset + 3 >= len(columns):
                    continue
                amount = _parse_ocr_amount_or_number(columns[offset + 1])
                gallons = _parse_ocr_amount_or_number(columns[offset + 3])
                if amount is not None or gallons is not None:
                    totals.append((amount or 0.0, gallons or 0.0))
            if len(totals) >= 2:
                return {
                    "amount": sum(t[0] for t in totals),
                    "gallons": sum(t[1] for t in totals),
                }
    except Exception:
        return None
    return None


def summarize_pdf_report(pdf_path):
    """
    OCR a photographed Date/Amount/Tracking/Gallons report (PDF with no text
    layer) into the same {date: {"amount", "gallons"}} shape produced by the
    Excel-based summarizer, plus a diagnostic summary for the caller to
    surface to the user.

    Returns:
        tuple[dict, dict]: (totals_by_date, diagnostics)
    """
    _ensure_pdfplumber()
    _ensure_pytesseract()

    images = _extract_pdf_page_images(pdf_path)

    # Only the first page of a multi-page report repeats the header row, so
    # detect the block count once — trying every OCR config on every page
    # until one finds a legible header — and reuse that single count for
    # every page. This keeps the block layout consistent across pages even
    # when different pages end up needing different winning OCR configs.
    block_count = None
    for image in images:
        for config in _OCR_TESSERACT_CONFIGS:
            rows = _cluster_words_into_rows(_ocr_words(image, config))
            block_count = _detect_block_count(rows)
            if block_count:
                break
        if block_count:
            break

    all_transactions = []
    printed_subtotal = {"amount": 0.0, "gallons": 0.0}
    found_subtotal = False
    page_configs_used = []

    for image in images:
        if block_count is None:
            # No page anywhere had a legible header — fall back to guessing
            # from this page's own data-row word counts.
            default_rows = _cluster_words_into_rows(
                _ocr_words(image, _OCR_TESSERACT_CONFIGS[0])
            )
            data_rows = [row for row in default_rows if _row_date_token(row) is not None]
            block_count = _estimate_block_count_from_data_rows(data_rows)

        page_transactions, winning_config = _best_effort_extract_page(image, block_count)
        all_transactions.extend(page_transactions)
        page_configs_used.append(winning_config)

        page_subtotal = _detect_ocr_subtotal(image, block_count, winning_config)
        if page_subtotal:
            printed_subtotal["amount"] += page_subtotal["amount"]
            printed_subtotal["gallons"] += page_subtotal["gallons"]
            found_subtotal = True

    totals_by_date = {}
    for date_text, amount, gallons in all_transactions:
        day = _date_dict_key(date_text)
        if day is None:
            continue
        bucket = totals_by_date.setdefault(day, {"amount": 0.0, "gallons": 0.0})
        bucket["amount"] += amount or 0.0
        bucket["gallons"] += gallons or 0.0

    computed_amount = sum(v["amount"] for v in totals_by_date.values())
    computed_gallons = sum(v["gallons"] for v in totals_by_date.values())

    diagnostics = {
        "pages": len(images),
        "transactions_read": len(all_transactions),
        "days_found": len(totals_by_date),
        "computed_amount": computed_amount,
        "computed_gallons": computed_gallons,
        "printed_subtotal_found": found_subtotal,
        "printed_subtotal_amount": printed_subtotal["amount"] if found_subtotal else None,
        "printed_subtotal_gallons": printed_subtotal["gallons"] if found_subtotal else None,
        "ocr_configs_used": page_configs_used,
    }
    if found_subtotal:
        diagnostics["amount_matches_subtotal"] = (
            abs(computed_amount - printed_subtotal["amount"]) <= 0.05
        )
        diagnostics["gallons_matches_subtotal"] = (
            abs(computed_gallons - printed_subtotal["gallons"]) <= 0.05
        )

    return totals_by_date, diagnostics


def detect_vendor_from_ocr_text(pdf_path):
    """
    Detect Gettel vs Toyota for a photographed report.

    Checks the filename first (the manager's naming convention, e.g. "Toyota
    report 03-08 al 13-08.pdf", is a much more reliable signal than OCR here
    — the small gray title banner at the top of the photo is exactly the
    kind of low-contrast text PSM 12 sparse-text mode isn't tuned for, since
    that mode is chosen for the printed data table instead). Falls back to
    OCR'd text across every page if the filename doesn't say.
    """
    basename = os.path.basename(pdf_path).upper()
    if any(token in basename for token in GETTEL_TOKENS):
        return VENDOR_GETTEL[0]
    if any(token in basename for token in TOYOTA_TOKENS):
        return VENDOR_TOYOTA[0]

    images = _extract_pdf_page_images(pdf_path)
    blob = " ".join(
        str(t).upper() for image in images for t in _ocr_words(image).text.tolist()
    )
    if any(token in blob for token in GETTEL_TOKENS):
        return VENDOR_GETTEL[0]
    if any(token in blob for token in TOYOTA_TOKENS):
        return VENDOR_TOYOTA[0]
    return None


def _find_gettel_toyota_destination_sheet(workbook):
    if len(workbook.worksheets) >= 4:
        fourth = workbook.worksheets[3]
        if _GETTEL_TOYOTA_DEST_SHEET_PATTERN.match(fourth.title.strip()):
            return fourth
    for worksheet in workbook.worksheets:
        if _GETTEL_TOYOTA_DEST_SHEET_PATTERN.match(worksheet.title.strip()):
            return worksheet
    raise ValueError(
        "El Excel de destino no tiene ninguna hoja que coincida con 'Gettel-Toyota MM.YYYY'."
    )


def _write_summary_number_cell(worksheet, row, column, value):
    """
    Write only the numeric value — the master template already carries the
    correct number format and alignment for this cell, and must not be
    overridden.
    """
    worksheet.cell(row=row, column=column).value = float(value)


def _populate_gettel_toyota_destination_sheet(worksheet, gettel_totals, toyota_totals):
    """
    Match both Gettel and Toyota daily totals against Column A's date.

    The DIF formula on this sheet (=C{row}-E{row}-G{row}) compares Local
    Account, Gettel, and Toyota all on the same row, so both vendors must
    land on the row identified by Column A — Column B is a separate,
    unrelated date and is never used for matching.

    A source day with no matching Column A row (the report covers a date
    the master sheet doesn't have — wrong month, or a day past its range)
    is skipped safely rather than raising: it's collected into
    unmatched_days for the caller to surface, never silently dropped.

    Returns:
        tuple[int, list]: (rows_matched, sorted list of unmatched dates)
    """
    rows_matched = 0
    matched_days = set()
    max_row = worksheet.max_row or _SOURCE_FIRST_DATA_ROW
    for row in range(_SOURCE_FIRST_DATA_ROW, max_row + 1):
        day = _date_dict_key(worksheet.cell(row=row, column=1).value)
        if day is None:
            continue
        row_matched = False

        if day in gettel_totals:
            totals = gettel_totals[day]
            _write_summary_number_cell(worksheet, row, 5, totals["amount"])
            _write_summary_number_cell(worksheet, row, 6, totals["gallons"])
            row_matched = True

        if day in toyota_totals:
            totals = toyota_totals[day]
            _write_summary_number_cell(worksheet, row, 7, totals["amount"])
            _write_summary_number_cell(worksheet, row, 8, totals["gallons"])
            row_matched = True

        if row_matched:
            rows_matched += 1
            matched_days.add(day)

    all_source_days = set(gettel_totals) | set(toyota_totals)
    unmatched_days = sorted(all_source_days - matched_days)
    return rows_matched, unmatched_days


def merge_gettel_toyota_into_master(source_path, destination_path, launch=True):
    """
    Summarize the origin Excel's Gettel and Toyota daily fuel-log sheets and
    merge the daily totals into the destination sheet (Gettel-Toyota
    MM.YYYY). Opens a temp preview copy.

    Returns:
        tuple[str, int, int, int, list]: (preview path, rows matched, gettel
        days, toyota days, unmatched dates — source days with no matching
        Column A row in the master, safely skipped rather than erroring)
    """
    import shutil
    import tempfile

    _ensure_openpyxl()
    source_path = os.path.abspath(source_path)
    destination_path = os.path.abspath(destination_path)
    if not os.path.isfile(source_path):
        raise FileNotFoundError(f"Excel de origen no encontrado: {source_path}")
    if not os.path.isfile(destination_path):
        raise FileNotFoundError(f"Excel de destino no encontrado: {destination_path}")

    source_workbook = load_workbook(source_path, data_only=True)
    try:
        gettel_totals, toyota_totals = _summarize_origin_workbook(source_workbook)
    finally:
        source_workbook.close()

    if not gettel_totals and not toyota_totals:
        raise ValueError("No se encontraron datos de cupones en el Excel de origen.")

    preview_handle = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix="GettelToyotaMasterPreview_",
        delete=False,
    )
    preview_path = preview_handle.name
    preview_handle.close()
    shutil.copy2(destination_path, preview_path)

    destination_workbook = load_workbook(preview_path)
    try:
        target_sheet = _find_gettel_toyota_destination_sheet(destination_workbook)
        rows_matched, unmatched_days = _populate_gettel_toyota_destination_sheet(
            target_sheet, gettel_totals, toyota_totals
        )
        destination_workbook.save(preview_path)
    finally:
        destination_workbook.close()

    if launch:
        _launch_workbook(preview_path)

    return preview_path, rows_matched, len(gettel_totals), len(toyota_totals), unmatched_days


def merge_gettel_toyota_pdf_into_master(pdf_path, destination_path, launch=True):
    """
    OCR a photographed Date/Amount/Tracking/Gallons report (one vendor per
    PDF) and merge its daily totals into the destination sheet (Gettel-
    Toyota MM.YYYY), the same way the Excel-based pipeline does. Opens a
    temp preview copy — the destination file itself is never modified.

    Returns:
        tuple[str, int, str, int, dict]: (preview path, rows matched,
        detected vendor label, days found, diagnostics dict from
        summarize_pdf_report — includes the printed-subtotal cross-check
        when the report has one)
    """
    import shutil
    import tempfile

    _ensure_openpyxl()
    _ensure_pdfplumber()
    _ensure_pytesseract()
    pdf_path = os.path.abspath(pdf_path)
    destination_path = os.path.abspath(destination_path)
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF de origen no encontrado: {pdf_path}")
    if not os.path.isfile(destination_path):
        raise FileNotFoundError(f"Excel de destino no encontrado: {destination_path}")

    vendor = detect_vendor_from_ocr_text(pdf_path)
    if vendor is None:
        raise ValueError(
            "No se pudo determinar si el reporte es de Gettel o Toyota a partir de su "
            "encabezado impreso. Verifique que la fila de encabezado sea legible."
        )

    totals_by_date, diagnostics = summarize_pdf_report(pdf_path)
    if not totals_by_date:
        raise ValueError("No se pudo leer ninguna fila de transacción del reporte PDF.")
    diagnostics["vendor"] = vendor

    gettel_totals = totals_by_date if vendor == VENDOR_GETTEL[0] else {}
    toyota_totals = totals_by_date if vendor == VENDOR_TOYOTA[0] else {}

    preview_handle = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix="GettelToyotaMasterPreview_",
        delete=False,
    )
    preview_path = preview_handle.name
    preview_handle.close()
    shutil.copy2(destination_path, preview_path)

    destination_workbook = load_workbook(preview_path)
    try:
        target_sheet = _find_gettel_toyota_destination_sheet(destination_workbook)
        rows_matched, unmatched_days = _populate_gettel_toyota_destination_sheet(
            target_sheet, gettel_totals, toyota_totals
        )
        destination_workbook.save(preview_path)
    finally:
        destination_workbook.close()

    diagnostics["unmatched_days"] = [d.isoformat() for d in unmatched_days]

    if launch:
        _launch_workbook(preview_path)

    return preview_path, rows_matched, vendor, len(totals_by_date), diagnostics


# ---- Pagos (payment receipts) OCR -> "PAGO Cupones" sheet ----
#
# Each "PagosN (Empresa).pdf" is one payment run: every page is a single
# register receipt, and the whole PDF shares one payment date and one
# payment number (the sheet merges those down column A/B, one value for
# the whole run, exactly like the user already does by hand). Column C
# (Trans #) and D (Total) are per-receipt, one value per row.

PAGO_CUPONES_SHEET_NAME = "PAGO Cupones"
PAGO_CUPONES_DATA_START_ROW = 4
PAGO_COL_FECHA = 1  # A
PAGO_COL_NUMERO = 2  # B
PAGO_COL_TRANS = 3  # C
PAGO_COL_TOTAL = 4  # D
PAGO_COL_EMPRESA = 5  # E

_PAGOS_FILENAME_RE = re.compile(r"pagos?\s*(\d+)\s*\(([^)]+)\)", re.IGNORECASE)
_PAGOS_TRANS_RE = re.compile(r"trans\s*#\s*[:;,.]*\s*(\d+)", re.IGNORECASE)
_PAGOS_TOTAL_RE = re.compile(r"total\s*=?\s*\$?\s*([\d,]+\.\d{2})", re.IGNORECASE)
_PAGOS_DATE_RE = re.compile(r"(\d{1,2}/\d{1,2}/\d{2,4})")

_PAGO_ROW_STYLE_SPEC = {
    PAGO_COL_FECHA: {"size": 11, "bold": False, "format": "mm-dd-yy", "align": "center"},
    PAGO_COL_NUMERO: {"size": 10, "bold": False, "format": "0", "align": "center"},
    PAGO_COL_TRANS: {"size": 11, "bold": False, "format": "0", "align": "center"},
    PAGO_COL_TOTAL: {"size": 11, "bold": True, "format": '"$"\\ #,##0.00', "align": "center"},
}


def _parse_pagos_filename(pdf_path):
    """
    "Pagos2 (Toyota).pdf" -> (2, "Toyota").

    The payment number and company come from the filename, not the receipt
    photos: a cash receipt never names which company the payment is being
    tracked under (only some card receipts show a store name, and that's a
    different thing), so the filename is the only reliable source.
    """
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    match = _PAGOS_FILENAME_RE.search(base)
    if not match:
        raise ValueError(
            f'No se pudo leer el número de pago ni la empresa del nombre "{os.path.basename(pdf_path)}" '
            '(se espera algo como "Pagos2 (Toyota).pdf").'
        )
    return int(match.group(1)), match.group(2).strip()


def _ocr_pago_receipt_text(image):
    """
    Best-effort OCR of one payment receipt photo, scored by how many of the
    three fields this module needs (date/trans#/total) come through — these
    are clean register receipts (not a cramped/watermarked report), so a
    plain image_to_string per PSM mode is enough without needing multiple
    OCR passes' output merged together.
    """
    if image is None:
        return ""
    _ensure_pytesseract()
    best_text = ""
    best_score = -1
    for config in _OCR_TESSERACT_CONFIGS:
        try:
            text = pytesseract.image_to_string(image, config=config) or ""
        except Exception:
            continue
        score = sum(
            1
            for pattern in (_PAGOS_TRANS_RE, _PAGOS_TOTAL_RE, _PAGOS_DATE_RE)
            if pattern.search(text)
        )
        if score > best_score:
            best_score = score
            best_text = text
    return best_text


def extract_pago_batch_from_pdf(pdf_path):
    """
    OCR every page of one "PagosN (Empresa).pdf" into an ordered list of
    {trans_number, total} receipts (in page order), plus the batch's shared
    date (the most common date read across its pages) and the payment
    number/company parsed from the filename.
    """
    payment_number, company = _parse_pagos_filename(pdf_path)
    images = _extract_pdf_page_images(pdf_path)

    receipts = []
    dates_found = []
    warnings = []
    for index, image in enumerate(images):
        text = _ocr_pago_receipt_text(image)
        trans_match = _PAGOS_TRANS_RE.search(text)
        total_match = _PAGOS_TOTAL_RE.search(text)
        date_match = _PAGOS_DATE_RE.search(text)

        trans_number = int(trans_match.group(1)) if trans_match else None
        total = _coerce_float(_clean_number_token(total_match.group(1))) if total_match else None
        date_value = _parse_date(date_match.group(1)) if date_match else None

        if date_value is not None:
            dates_found.append(date_value)
        if trans_number is None:
            warnings.append(f'Página {index + 1}: no se pudo leer "Trans #" — revisar manualmente.')
        if total is None:
            warnings.append(f'Página {index + 1}: no se pudo leer "Total = $" — revisar manualmente.')
        receipts.append({"trans_number": trans_number, "total": total})

    if not dates_found:
        raise ValueError(f'No se pudo leer ninguna fecha en "{os.path.basename(pdf_path)}".')
    batch_date = Counter(dates_found).most_common(1)[0][0]

    return {
        "payment_number": payment_number,
        "company": company,
        "date": batch_date,
        "receipts": receipts,
        "warning": " | ".join(warnings) if warnings else None,
    }


def _find_pago_cupones_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower() == PAGO_CUPONES_SHEET_NAME.lower():
            return workbook[name]
    raise ValueError(
        f'Hoja "{PAGO_CUPONES_SHEET_NAME}" no encontrada. Disponibles: {", ".join(workbook.sheetnames)}'
    )


def _merge_range_at(sheet, row, column):
    """The MergedCellRange covering (row, column), or None if that cell isn't merged."""
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range
    return None


def _pago_block_row_span(sheet, start_row):
    """How many rows the block at `start_row` currently occupies, from column A's merge (1 if it isn't merged)."""
    merged = _merge_range_at(sheet, start_row, PAGO_COL_FECHA)
    if merged is not None and merged.min_row == start_row:
        return merged.max_row - merged.min_row + 1
    return 1


def _find_pago_block_start_row(sheet, payment_number):
    """The row where column B already holds `payment_number` (top of an existing block), or None."""
    for row in range(PAGO_CUPONES_DATA_START_ROW, sheet.max_row + 1):
        value = sheet.cell(row=row, column=PAGO_COL_NUMERO).value
        if value is not None and _coerce_float(str(value)) == payment_number:
            return row
    return None


def _find_pago_insertion_row(sheet, payment_number):
    """
    Where a brand-new Pago N block (no existing stub row for it) should
    start: right after the highest existing payment number below N, or at
    the very first data row if there's no earlier block at all.
    """
    best_row = None
    best_number = None
    for row in range(PAGO_CUPONES_DATA_START_ROW, sheet.max_row + 1):
        value = sheet.cell(row=row, column=PAGO_COL_NUMERO).value
        if value is None:
            continue
        existing_number = _coerce_float(str(value))
        if existing_number is not None and existing_number < payment_number:
            if best_number is None or existing_number > best_number:
                best_number = existing_number
                best_row = row
    if best_row is None:
        return PAGO_CUPONES_DATA_START_ROW
    return best_row + _pago_block_row_span(sheet, best_row)


def _resize_pago_block(sheet, start_row, old_size, new_size):
    """
    Grow or shrink the block at `start_row` from `old_size` to `new_size`
    rows, relocating every merged range in the sheet by hand.

    openpyxl's insert_rows/delete_rows shift cell values and styles down or
    up correctly, but do NOT adjust merged-cell ranges — left alone, every
    block below the one being resized (and the summary rows further down)
    would keep pointing at their old row numbers instead of following the
    shift, corrupting the sheet's structure. Unmerging everything first,
    doing the plain row insert/delete, then recomputing each range against
    the same delta keeps every block (and the fixed 2-row gap before
    whatever comes after the last one) lined up automatically.

    Deliberately leaves the resized block's own A/B range unmerged on
    return — every cell in a merged range other than its top-left anchor is
    a read-only MergedCell in openpyxl, so styling it after re-merging
    silently does nothing. The caller re-merges only after restyling each
    row.
    """
    delta = new_size - old_size
    if delta == 0:
        return

    boundary = start_row + old_size
    existing_merges = [
        (r.min_row, r.min_col, r.max_row, r.max_col) for r in list(sheet.merged_cells.ranges)
    ]
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))

    if delta > 0:
        sheet.insert_rows(boundary, delta)
    else:
        sheet.delete_rows(boundary + delta, -delta)

    for min_row, min_col, max_row, max_col in existing_merges:
        if min_row == start_row and old_size > 0 and max_row == start_row + old_size - 1:
            continue  # the block being resized -- left unmerged, see docstring
        elif min_row >= boundary:
            min_row += delta
            max_row += delta
        if max_row < min_row or (min_row == max_row and min_col == max_col):
            continue
        sheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)


def _apply_pago_row_style(sheet, row, is_first_row_of_block):
    thin = Side(style="thin")
    for column, spec in _PAGO_ROW_STYLE_SPEC.items():
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(name="Calibri", size=spec["size"], bold=spec["bold"])
        cell.number_format = spec["format"]
        cell.alignment = Alignment(horizontal=spec["align"])
        if column in (PAGO_COL_FECHA, PAGO_COL_NUMERO):
            cell.border = Border(top=thin if is_first_row_of_block else None, left=thin, right=thin)
        else:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _write_pago_batch(sheet, batch):
    payment_number = batch["payment_number"]
    receipts = batch["receipts"]
    new_size = len(receipts)

    start_row = _find_pago_block_start_row(sheet, payment_number)
    if start_row is not None:
        old_size = _pago_block_row_span(sheet, start_row)
    else:
        start_row = _find_pago_insertion_row(sheet, payment_number)
        old_size = 0

    _resize_pago_block(sheet, start_row, old_size, new_size)

    # Style every row while its A/B cells are still plain (unmerged) Cell
    # objects -- every cell but a merge's top-left anchor becomes a
    # read-only MergedCell once merged, silently dropping style changes.
    for offset in range(new_size):
        _apply_pago_row_style(sheet, start_row + offset, is_first_row_of_block=(offset == 0))

    if new_size > 1:
        sheet.merge_cells(
            start_row=start_row, start_column=PAGO_COL_FECHA,
            end_row=start_row + new_size - 1, end_column=PAGO_COL_FECHA,
        )
        sheet.merge_cells(
            start_row=start_row, start_column=PAGO_COL_NUMERO,
            end_row=start_row + new_size - 1, end_column=PAGO_COL_NUMERO,
        )

    sheet.cell(row=start_row, column=PAGO_COL_FECHA, value=batch["date"])
    sheet.cell(row=start_row, column=PAGO_COL_NUMERO, value=payment_number)
    company_cell = sheet.cell(row=start_row, column=PAGO_COL_EMPRESA, value=f"({batch['company']})")
    company_cell.font = Font(name="Calibri", size=11, bold=True)

    for offset, receipt in enumerate(receipts):
        row = start_row + offset
        if receipt["trans_number"] is not None:
            sheet.cell(row=row, column=PAGO_COL_TRANS, value=receipt["trans_number"])
        if receipt["total"] is not None:
            sheet.cell(row=row, column=PAGO_COL_TOTAL, value=receipt["total"])

    return start_row


def process_gettel_pagos(master_path, pdf_paths):
    """
    OCR one or more "PagosN (Empresa).pdf" payment-receipt batches into the
    "PAGO Cupones" sheet's A-E columns.

    Each batch's row block is resized to match its real receipt count
    exactly (growing or shrinking as needed), which shifts every block
    below it — and the summary rows further down the sheet — by however
    many rows that adds or removes, so the fixed 2-row gap after the last
    block is preserved automatically regardless of how many receipts came
    in this month. Opens a temp preview copy — the destination file itself
    is never modified.

    Returns:
        tuple[str, dict]: (preview path, summary dict with batch_results)
    """
    import shutil
    import tempfile

    _ensure_openpyxl()
    _ensure_pdfplumber()
    _ensure_pytesseract()
    master_path = os.path.abspath(str(master_path).strip())
    paths = _normalize_pdf_paths(pdf_paths)

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Excel no encontrado: {master_path}")
    if not paths:
        raise ValueError("No se proporcionaron PDF de pagos.")
    for pdf_path in paths:
        if not os.path.isfile(pdf_path):
            raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

    batches = [extract_pago_batch_from_pdf(pdf_path) for pdf_path in paths]
    batches.sort(key=lambda b: b["payment_number"])

    preview_handle = tempfile.NamedTemporaryFile(
        suffix=".xlsx", prefix="GettelPagosPreview_", delete=False
    )
    preview_path = preview_handle.name
    preview_handle.close()
    shutil.copy2(master_path, preview_path)

    workbook = load_workbook(preview_path)
    try:
        sheet = _find_pago_cupones_sheet(workbook)
        batch_results = []
        for batch in batches:
            row = _write_pago_batch(sheet, batch)
            batch_results.append(
                {
                    "payment_number": batch["payment_number"],
                    "company": batch["company"],
                    "date": batch["date"].strftime("%d/%m/%Y"),
                    "row": row,
                    "receipt_count": len(batch["receipts"]),
                    "warning": batch["warning"],
                }
            )
        workbook.save(preview_path)
    finally:
        workbook.close()

    _launch_workbook(preview_path)

    return preview_path, {"batch_results": batch_results}


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    if argv:
        pdf_paths = argv
    else:
        try:
            import tkinter as tk
            from tkinter import filedialog

            root = tk.Tk()
            root.withdraw()
            pdf_paths = filedialog.askopenfilenames(
                title="Select fuel coupon PDF report(s)",
                filetypes=[("PDF Files", "*.pdf")],
            )
            root.destroy()
        except Exception:
            return 1

    if not pdf_paths:
        return 1

    generate_fuel_coupon_workbook_from_pdfs(list(pdf_paths))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
